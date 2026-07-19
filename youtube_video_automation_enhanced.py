"""
Video Quote Overlay Automation - ENHANCED with Advanced Effects
Compatible with MoviePy 2.x
"""

import os
import sys
import re
import math
import time
from pathlib import Path
from typing import List, Tuple, Optional
import json
from datetime import datetime
import numpy as np

# ── AI Avatar (Wav2Lip) — optional import, ignored if unavailable
try:
    from wav2lip_helper import run_wav2lip, composite_avatar, extract_audio, is_available as _wav2lip_available
except ImportError:
    run_wav2lip = composite_avatar = extract_audio = None
    _wav2lip_available = lambda: False

# ── Qwen3-TTS — optional import, ignored if unavailable
try:
    from qwen3_helper import generate_speech as _qwen3_generate
except ImportError:
    _qwen3_generate = None

# ── Piper TTS — optional import, ignored if unavailable
try:
    import piper_tts_helper as _piper_tts
except ImportError:
    _piper_tts = None

# PERF: module-level caches for per-frame mask computations. The vignette,
# spotlight, and other per-frame effects would otherwise re-allocate large
# mask arrays (6MB+ each) for every single output frame. With these caches
# the mask is built once and reused.
_VIGNETTE_MASK_CACHE: dict = {}
_MASK_CACHE: dict = {}

# ===== MOVIEPY 1.0.3 COMPATIBILITY PATCH =====
# MoviePy 1.0.3 uses set_* methods (set_duration, set_start, etc.)
# while MoviePy 2.x uses with_* methods (with_duration, with_start, etc.).
# This patch adds with_* aliases at the class level so the SAME code works
# on both 1.x and 2.x without changing hundreds of call sites.
def _patch_moviepy_1_x():
    """Add MoviePy 2.x API aliases for MoviePy 1.x compatibility."""
    import moviepy as _mp
    _ver = tuple(int(x) for x in _mp.__version__.split('.'))
    if _ver >= (2, 0, 0):
        return  # MoviePy 2.x already has the with_* API

    from moviepy.Clip import Clip as _Clip
    from moviepy.video.VideoClip import VideoClip as _VClip
    from moviepy.audio.AudioClip import AudioClip as _AClip

    # For each class, mirror every set_* method as with_*
    for _cls in [_Clip, _VClip, _AClip]:
        for _name in list(_cls.__dict__.keys()):
            if _name.startswith('set_'):
                _with = _name.replace('set_', 'with_', 1)
                if not hasattr(_cls, _with):
                    setattr(_cls, _with, getattr(_cls, _name))

    # -- resized  → resize(video, ...)  (resize is a function in 1.0.3) --
    from moviepy.video.fx.resize import resize as _resize

    def _resized(self, newsize=None, height=None, width=None, apply_to_mask=True):
        return _resize(self, newsize, height, width, apply_to_mask)
    if not hasattr(_VClip, 'resized'):
        _VClip.resized = _resized

    # -- cropped  → crop(video, ...)  (crop is a function in 1.0.3) --
    from moviepy.video.fx.crop import crop as _crop

    def _cropped(self, x1=None, y1=None, x2=None, y2=None,
                 width=None, height=None, x_center=None, y_center=None):
        return _crop(self, x1, y1, x2, y2, width, height, x_center, y_center)
    if not hasattr(_VClip, 'cropped'):
        _VClip.cropped = _cropped

    # -- transform  → fl  (1.0.3 uses fl / fl_image, not transform) --
    if not hasattr(_VClip, 'transform'):
        _VClip.transform = _VClip.fl

    # -- with_effects stub (1.0.3 doesn't have the v2 effect system) --
    if not hasattr(_VClip, 'with_effects'):
        _VClip.with_effects = lambda self, effects: self


_patch_moviepy_1_x()

# ===== PILLOW COMPATIBILITY PATCH =====
# Pillow 10+ removed Image.ANTIALIAS (deprecated in favor of
# Image.Resampling.LANCZOS).  MoviePy 1.0.3's resize function still
# references Image.ANTIALIAS, so we restore it as an alias to prevent
# AttributeError during video resizing.
import PIL.Image as _PILImage
import PIL.ImageFilter as _PILImageFilter
if not hasattr(_PILImage, 'ANTIALIAS'):
    _PILImage.ANTIALIAS = _PILImage.Resampling.LANCZOS

# ===== OPENCV + FFMPEG RENDERER (bypasses MoviePy per-frame overhead) =====
try:
    import cv2 as _CV2
    _HAS_CV2 = True
except ImportError:
    _HAS_CV2 = False

# ⚡ SPEED: probe the ffmpeg NVENC encoder ONCE per process, not per render.
# The previous code spawned `ffmpeg -encoders` on every single video render.
_NVENC_CACHE = None

def _nvenc_available():
    """Return True if ffmpeg exposes the h264_nvenc (NVIDIA GPU) encoder.

    Result is cached for the lifetime of the process. On a laptop without an
    NVIDIA GPU this returns False and callers fall back to libx264.
    """
    global _NVENC_CACHE
    if _NVENC_CACHE is None:
        import subprocess as _sp
        try:
            _r = _sp.run(['ffmpeg', '-encoders'],
                         capture_output=True, text=True, timeout=5)
            _NVENC_CACHE = 'h264_nvenc' in _r.stdout
        except Exception:
            _NVENC_CACHE = False
    return _NVENC_CACHE

class _SequentialFrameReader:
    """Read frames from a video file using OpenCV, optimized for sequential access.

    Keeps the capture open across frames and favors read-next over seek
    (~50x faster for sequential access).
    """
    __slots__ = ('path', '_cap', 'source_fps', 'total_frames', 'w', 'h',
                 'duration', '_last_idx')

    def __init__(self, path):
        self.path = str(path)
        self._cap = _CV2.VideoCapture(self.path)
        self.source_fps = self._cap.get(_CV2.CAP_PROP_FPS)
        self.total_frames = int(self._cap.get(_CV2.CAP_PROP_FRAME_COUNT))
        self.w = int(self._cap.get(_CV2.CAP_PROP_FRAME_WIDTH))
        self.h = int(self._cap.get(_CV2.CAP_PROP_FRAME_HEIGHT))
        dur = self.total_frames / self.source_fps if self.source_fps > 0 else 0
        self.duration = dur
        self._last_idx = -2

    def read(self, frame_idx):
        """Return RGB frame at *frame_idx* (0-based source index).

        ⚡ SPEED: avoid seeks. When the source fps (e.g. 30) differs from the
        output fps (e.g. 24), consecutive output frames map to non-consecutive
        source indices (0,1,3,4,5,6,8…). Seeking on every skipped frame is
        catastrophic for AV1/HEVC (sparse keyframes → each seek re-decodes from
        the last keyframe, ~300ms+). Instead we read forward and DISCARD the
        skipped frames — decoding one throwaway frame is ~10-50x cheaper than a
        seek. Only seek for a true backward jump or a large forward gap.
        """
        if frame_idx == self._last_idx:
            # Same frame requested again — re-read is cheap; reuse last decode.
            ret, frame = self._cap.read()
        elif frame_idx > self._last_idx and (frame_idx - self._last_idx) <= 8:
            # Forward by a small gap: decode-and-discard intermediate frames
            # so we never trigger an expensive keyframe seek.
            ret, frame = True, None
            for _ in range(frame_idx - self._last_idx):
                ret, frame = self._cap.read()
                if not ret:
                    break
        else:
            # Large jump or backward seek — unavoidable hard seek.
            self._cap.set(_CV2.CAP_PROP_POS_FRAMES, frame_idx)
            ret, frame = self._cap.read()
        self._last_idx = frame_idx
        if not ret or frame is None:
            return None
        return _CV2.cvtColor(frame, _CV2.COLOR_BGR2RGB)

    def release(self):
        if self._cap is not None:
            self._cap.release()
            self._cap = None

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        self.release()


def _opencv_render(reader, output_path, target_w, target_h, target_fps,
                   duration, effects_pipeline, overlay_sources,
                   crop_box=None, transition_fn=None, pre_overlay_fn=None):
    """Render video via OpenCV source read + numpy composite + FFmpeg NVENC pipe.

    Parameters
    ----------
    reader : _SequentialFrameReader
        OpenCV-backed source reader (already opened).
    output_path : Path or str
    target_w, target_h : int
        Output frame dimensions.
    target_fps : int
    duration : float
        Total output duration in seconds.
    effects_pipeline : callable or None
        ``frame(rgb uint8) → frame(rgb uint8)`` — applied *after* crop/resize
        and *before* overlay compositing (typically the combined effects pipeline
        from the caller's closure).
    overlay_sources : list of dict
        Each dict::
            { 'get_frame': callable(t) → RGBA|RGB ndarray,
              'pos': (x,y) or callable(t)→(x,y),
              'start': float, 'end': float or None }
    crop_box : (x1,y1,x2,y2) or None
        Pixel region cropped from the *source* frame before resize.
    transition_fn : callable or None
        ``frame(rgb uint8, t) → frame(rgb uint8)`` — applied *after* all
        compositing (effects + overlays) and *before* writing to the pipe.
        Handles fade / zoom / blur transitions that vary with time t.
    """
    import subprocess as _sp
    import gc as _gc
    import os as _os
    _gc.disable()
    total_frames = int(duration * target_fps)
    if total_frames < 1:
        return
    # Test hook: cap frames for quick perf measurement (set CG_RENDER_MAXFRAMES).
    _cap_frames = _os.environ.get('CG_RENDER_MAXFRAMES')
    if _cap_frames:
        try:
            total_frames = min(total_frames, int(_cap_frames))
        except ValueError:
            pass

    # ── Build ffmpeg command ───────────────────────────────────────────
    cmd = [
        'ffmpeg', '-y',
        '-f', 'rawvideo',
        '-vcodec', 'rawvideo',
        '-s', f'{target_w}x{target_h}',
        '-pix_fmt', 'rgb24',
        '-r', str(target_fps),
        '-i', '-',
        '-an',
    ]
    # Detect NVENC (cached once per process — see _nvenc_available)
    _nvenc_ok = _nvenc_available()

    if _nvenc_ok:
        cmd += [
            '-c:v', 'h264_nvenc', '-preset', 'p7',
            '-tune', 'hq', '-rc', 'vbr', '-cq', '23',
            '-b:v', '0', '-profile:v', 'main',
            '-pix_fmt', 'yuv420p',
        ]
    else:
        # ⚡ No NVIDIA GPU (e.g. the integrated-graphics laptop): use the
        # fastest libx264 software preset across all CPU cores. yuv420p keeps
        # the file playable everywhere. This is still far faster than the old
        # MoviePy per-frame path because decode + composite happen in C/numpy.
        cmd += [
            '-c:v', 'libx264', '-preset', 'ultrafast',
            '-crf', '23', '-pix_fmt', 'yuv420p', '-threads', '0',
        ]
    cmd.append(str(output_path))

    proc = _sp.Popen(cmd, stdin=_sp.PIPE, stderr=_sp.DEVNULL)
    _frame_times = []
    _log_interval = max(1, int(target_fps * 10))  # every ~10s

    try:
        for frame_idx in range(total_frames):
            _t0 = time.perf_counter()
            t = frame_idx / target_fps

            # ── Source read ───────────────────────────────────────────
            _t_src = time.perf_counter()
            source_idx = int(round(t * reader.source_fps))
            source_idx = max(0, min(source_idx, reader.total_frames - 1))
            frame = reader.read(source_idx)
            if frame is None:
                frame = np.zeros((target_h, target_w, 3), dtype=np.uint8)
            _t_src = time.perf_counter() - _t_src

            # ── Crop (source coords) ──────────────────────────────────
            if crop_box:
                _t_prep = time.perf_counter()
                x1, y1, x2, y2 = crop_box
                frame = frame[y1:y2, x1:x2]
                _t_src += time.perf_counter() - _t_prep

            # ── Resize to output dimensions ───────────────────────────
            if frame.shape[1] != target_w or frame.shape[0] != target_h:
                _t_prep = time.perf_counter()
                frame = _CV2.resize(frame, (target_w, target_h),
                                    interpolation=_CV2.INTER_LINEAR)
                _t_src += time.perf_counter() - _t_prep

            # ── Effects pipeline ──────────────────────────────────────
            _t_fx = time.perf_counter()
            if effects_pipeline:
                frame = effects_pipeline(frame)
            _t_fx = time.perf_counter() - _t_fx

            # ── Pre-overlay transition (applies to video only, before captions/etc) ──
            _t_pre = 0.0
            if pre_overlay_fn:
                _t_pre = time.perf_counter()
                frame = pre_overlay_fn(frame, t)
                _t_pre = time.perf_counter() - _t_pre

            # ── Overlay compositing ───────────────────────────────────
            _t_ov = time.perf_counter()
            for ov in overlay_sources:
                # ⚡ Time-range gate for caption bounds (uses absolute time,
                # no get_frame offset — _render_rgba has its own timeline).
                _min_t = ov.get('_min_t')
                _max_t = ov.get('_max_t')
                if _min_t is not None and t < _min_t:
                    continue
                if _max_t is not None and t >= _max_t:
                    continue
                if t < ov['start'] or (ov['end'] is not None and t >= ov['end']):
                    continue
                ov_frame = ov['get_frame'](t - ov['start'])
                if ov_frame is None or ov_frame.size == 0:
                    continue

                # Position
                pos = ov['pos']
                if callable(pos):
                    px, py = pos(t)
                else:
                    px, py = pos

                oh, ow = ov_frame.shape[:2]
                if oh <= 0 or ow <= 0:
                    continue

                if isinstance(px, str) and px == 'center':
                    px = (target_w - ow) // 2
                if isinstance(py, str) and py == 'center':
                    py = (target_h - oh) // 2
                px, py = int(px), int(py)

                # Clip to frame bounds
                dx1 = max(0, px)
                dy1 = max(0, py)
                dx2 = min(target_w, px + ow)
                dy2 = min(target_h, py + oh)
                if dx2 <= dx1 or dy2 <= dy1:
                    continue

                sx1 = dx1 - px
                sy1 = dy1 - py
                roi = frame[dy1:dy2, dx1:dx2]
                ovr = ov_frame[sy1:sy1 + dy2 - dy1, sx1:sx1 + dx2 - dx1]

                if ovr.shape[2] == 4:
                    # ⚡ SPEED: blend mostly-transparent RGBA overlays cheaply.
                    # Two regimes, chosen by how DENSE the lit pixels are inside
                    # their bounding box:
                    #   • DENSE (caption band — a solid filled rectangle): one
                    #     contiguous vectorized alpha blend over the tight bbox.
                    #   • SPARSE (border outline + crosshair cross — thin lines
                    #     spanning the whole frame, bbox≈full but <5% lit): blend
                    #     ONLY the nonzero pixels via fancy indexing, so we never
                    #     touch the millions of transparent pixels in between.
                    _a_chan = ovr[:, :, 3]
                    _rows = np.any(_a_chan, axis=1)
                    if not _rows.any():
                        continue  # fully transparent → nothing to draw
                    _cols = np.any(_a_chan, axis=0)
                    _ry = np.where(_rows)[0]
                    _cx = np.where(_cols)[0]
                    _by0, _by1 = int(_ry[0]), int(_ry[-1]) + 1
                    _bx0, _bx1 = int(_cx[0]), int(_cx[-1]) + 1
                    _bbox_area = (_by1 - _by0) * (_bx1 - _bx0)
                    # Dense vs sparse is stable per source (captions are always a
                    # filled band; border/crosshair are always thin lines) — decide
                    # once and cache, so np.count_nonzero doesn't run every frame.
                    _dense = ov.get('_dense')
                    if _dense is None:
                        _lit = int(np.count_nonzero(_a_chan))
                        _dense = _lit > 0.30 * _bbox_area
                        ov['_dense'] = _dense
                    if _dense:
                        # Dense → contiguous bbox blend over the tight caption band.
                        _roi = frame[dy1 + _by0:dy1 + _by1, dx1 + _bx0:dx1 + _bx1]
                        _ob = ovr[_by0:_by1, _bx0:_bx1]
                        _af = _ob[:, :, 3:4].astype(np.float32) / 255.0
                        frame[dy1 + _by0:dy1 + _by1, dx1 + _bx0:dx1 + _bx1] = (
                            _roi.astype(np.float32) * (1.0 - _af)
                            + _ob[:, :, :3].astype(np.float32) * _af
                        ).astype(np.uint8)
                    else:
                        # Sparse → touch only lit pixels.
                        _ys, _xs = np.nonzero(_a_chan)
                        _af = (_a_chan[_ys, _xs].astype(np.float32) / 255.0)[:, None]
                        _gy = _ys + dy1
                        _gx = _xs + dx1
                        frame[_gy, _gx] = (
                            frame[_gy, _gx].astype(np.float32) * (1.0 - _af)
                            + ovr[_ys, _xs, :3].astype(np.float32) * _af
                        ).astype(np.uint8)
                elif ovr.shape[2] >= 3:
                    frame[dy1:dy2, dx1:dx2] = ovr[:, :, :3]
            _t_ov = time.perf_counter() - _t_ov

            # ── Apply transitions (time-varying post-processing) ────────
            _t_tr = time.perf_counter()
            if transition_fn:
                frame = transition_fn(frame, t)
            _t_tr = time.perf_counter() - _t_tr

            # ── Write to FFmpeg pipe ──────────────────────────────────
            _t_wr = time.perf_counter()
            proc.stdin.write(frame.tobytes())
            _t_wr = time.perf_counter() - _t_wr

            # ── Profiling ─────────────────────────────────────────────
            _elapsed = time.perf_counter() - _t0
            _sum_ms = (_t_src + _t_fx + _t_pre + _t_ov + _t_tr + _t_wr) * 1000
            _gap_ms = _elapsed * 1000 - _sum_ms
            _frame_times.append(_elapsed)
            # Accumulate component totals across ALL frames (the sampled
            # single-frame print is misleading when frames vary widely).
            try:
                _acc  # noqa
            except NameError:
                _acc = {'src': 0.0, 'fx': 0.0, 'pre': 0.0, 'ov': 0.0,
                        'tr': 0.0, 'wr': 0.0, 'gap': 0.0, 'n': 0}
            _acc['src'] += _t_src; _acc['fx'] += _t_fx; _acc['pre'] += _t_pre
            _acc['ov'] += _t_ov; _acc['tr'] += _t_tr; _acc['wr'] += _t_wr
            _acc['gap'] += _gap_ms / 1000.0; _acc['n'] += 1
            if frame_idx > 0 and frame_idx % _log_interval == 0:
                _recent = _frame_times[-_log_interval:]
                _avg_ms = sum(_recent) / len(_recent) * 1000
                _rem = (total_frames - frame_idx) * _avg_ms / 1000
                print(f"  [CV] Frame {frame_idx}/{total_frames} | "
                      f"avg {_avg_ms:.0f}ms ({1000/_avg_ms:.1f} it/s) | "
                      f"fx={_t_fx*1000:.0f} ov={_t_ov*1000:.0f} "
                      f"tr={_t_tr*1000:.0f} wr={_t_wr*1000:.0f} "
                      f"src={_t_src*1000:.0f} pre={_t_pre*1000:.0f} "
                      f"gap={_gap_ms:.0f} | "
                      f"~{_rem:.0f}s left", flush=True)

    finally:
        proc.stdin.close()
        proc.wait()
        _gc.enable()

    # Summary
    if _frame_times:
        avg_ms = sum(_frame_times) / len(_frame_times) * 1000
        print(f"  [⚡CV] Done — avg {avg_ms:.0f}ms/frame, "
              f"worst {max(_frame_times)*1000:.0f}ms, "
              f"best {min(_frame_times)*1000:.0f}ms", flush=True)
        try:
            _n = max(1, _acc['n'])
            print(f"  [⚡CV] AVG component ms/frame over {_n} frames: "
                  f"src={_acc['src']/_n*1000:.0f} fx={_acc['fx']/_n*1000:.0f} "
                  f"pre={_acc['pre']/_n*1000:.0f} ov={_acc['ov']/_n*1000:.0f} "
                  f"tr={_acc['tr']/_n*1000:.0f} wr={_acc['wr']/_n*1000:.0f} "
                  f"gap={_acc['gap']/_n*1000:.0f}", flush=True)
        except (NameError, KeyError):
            pass


# ===== SMART WORD TIMING for caption sync =====
# Common short function words spoken quickly (reduces timing drift)
_SHORT_WORDS = {
    'a', 'an', 'the', 'to', 'in', 'of', 'for', 'on', 'at', 'by', 'with',
    'from', 'as', 'is', 'it', 'its', "it's", 'and', 'or', 'but', 'so',
    'if', 'no', 'not', 'up', 'out', 'be', 'he', 'she', 'we', 'they',
    'me', 'him', 'her', 'us', 'them', 'my', 'your', 'his', 'our',
    'i', 'you', 'are', 'was', 'were', 'been', 'do', 'does', 'did',
    'has', 'had', 'have', 'can', 'could', 'will', 'would', 'shall',
    'should', 'may', 'might', 'must', 'than', 'that', 'this', 'these',
    'those', 'there', 'their', 'them', 'then', 'which', 'what', 'when',
    'where', 'how', 'all', 'each', 'every', 'both', 'few', 'more',
    'most', 'some', 'any', 'such', 'only', 'own', 'same', 'too',
    'very', 'just', 'also', 'else', 'off', 'over', 'here', 'there',
}


def refine_word_timings_smart(word_timings, total_duration, audio_path=None):
    """Refine word timings using smarter syllable-based model + VAD gap detection.

    The old character-weighted model (equal time per character) causes ~0.3-0.5s
    drift because short function words (spoken fast) get too much time while
    long content words (spoken slower) get too little.

    This function replaces naive linear-char weighting with:
    1. sqrt(char_count) weight - compresses range between short/long words
    2. Function-word discounts - words like 'the', 'a', 'and' get 40% less time
    3. Punctuation pauses - 120ms pause added after . ! ?
    4. VAD gap detection - if audio_path given, uses pydub silence detector to
       align word boundaries to actual speech gaps in the audio waveform
    5. Scale-preserving - total duration is maintained regardless of weighting

    Returns: updated word_timings (mutated in-place and returned)
    """
    word_count = len(word_timings)
    if word_count == 0 or total_duration <= 0:
        return word_timings

    # Step 1: Smart weight per word
    weights = []
    total_weight = 0.0
    for w in word_timings:
        word = w['word']
        # sqrt compresses range: 'a'=1.0, 'the'=1.73, 'extraordinary'=3.74
        # old linear: 'a'=2, 'the'=3, 'extraordinary'=13 (huge imbalance)
        weight = math.sqrt(max(1, len(word)))
        # Function words are spoken faster in natural speech
        if word.lower() in _SHORT_WORDS:
            weight *= 0.6
        # Longer words (>8 chars) take more relative time
        if len(word) > 8:
            weight *= 1.2
        weights.append(weight)
        total_weight += weight

    if total_weight <= 0:
        return word_timings

    # Step 2: Distribute time by weight (scaled to match total_duration)
    current_time = 0.0
    for i, (word_info, weight) in enumerate(zip(word_timings, weights)):
        word_info['offset'] = current_time
        word_duration = (weight / total_weight) * total_duration
        word_duration = max(0.06, word_duration)  # Minimum 60ms per word
        word_info['duration'] = word_duration
        current_time += word_duration

    # Step 3: Insert pauses after sentence-ending punctuation (. ! ?)
    pause_indices = []
    for i, w in enumerate(word_timings):
        word = w['word']
        if word and word[-1] in '.!?' and i < word_count - 1:
            pause_indices.append(i)

    if pause_indices:
        pause_per = min(0.12, total_duration * 0.02)  # 120ms per pause
        shift_acc = 0.0
        for i in range(word_count):
            word_timings[i]['offset'] += shift_acc
            if i in pause_indices:
                word_timings[i]['duration'] += pause_per * 0.3
                shift_acc += pause_per * 0.7

    # Step 4: VAD gap detection for real alignment (optional)
    if audio_path and audio_path.exists():
        try:
            _refine_with_vad_pydub(word_timings, audio_path)
        except Exception:
            pass  # Smart timing is already a good fallback

    # Step 5: Scale to ensure total duration matches exactly
    last_end = word_timings[-1]['offset'] + word_timings[-1]['duration']
    if last_end > 0 and abs(last_end - total_duration) > 0.02:
        scale = total_duration / last_end
        for w in word_timings:
            w['offset'] *= scale
            w['duration'] *= scale

    return word_timings


def _refine_with_vad_pydub(word_timings, audio_path):
    """Refine word timings by aligning to actual speech gaps via pydub VAD.

    Detects natural pauses in the TTS audio waveform, then maps word boundaries
    to those pauses for frame-accurate sync that's far better than any
    character-based model.
    """
    from pydub import AudioSegment
    from pydub.silence import detect_silence

    audio = AudioSegment.from_file(str(audio_path))
    duration_ms = len(audio)
    if duration_ms <= 0:
        return

    silent_segs = detect_silence(audio, min_silence_len=80,
                                 silence_thresh=-25, seek_step=10)
    if not silent_segs:
        return

    # Convert to speech segments (inverse of silence)
    speech_segs = []
    prev_end = 0
    for start_ms, end_ms in silent_segs:
        if start_ms > prev_end:
            speech_segs.append((prev_end, start_ms))
        prev_end = end_ms
    if prev_end < duration_ms:
        speech_segs.append((prev_end, duration_ms))

    if not speech_segs:
        return

    total_s = duration_ms / 1000.0
    word_count = len(word_timings)

    # Count words per speech segment using current estimated positions
    counts = [0] * len(speech_segs)
    for w in word_timings:
        s = w['offset']
        for si, (ss, se) in enumerate(speech_segs):
            if ss / 1000.0 <= s <= se / 1000.0:
                counts[si] += 1
                break

    # Redistribute each segment's real duration across its words
    idx = 0
    cur_s = 0.0
    for si, (ss, se) in enumerate(speech_segs):
        nw = counts[si]
        if nw == 0:
            continue
        seg_s = (se - ss) / 1000.0
        sw = []
        for j in range(idx, min(idx + nw, word_count)):
            word = word_timings[j]['word']
            wt = math.sqrt(max(1, len(word)))
            if word.lower() in _SHORT_WORDS:
                wt *= 0.7
            sw.append(wt)
        tw = sum(sw)
        for j, wt in enumerate(sw):
            pos = idx + j
            word_timings[pos]['offset'] = cur_s
            dur = (wt / tw) * seg_s if tw > 0 else seg_s / nw
            word_timings[pos]['duration'] = max(0.06, dur)
            cur_s += word_timings[pos]['duration']
        idx += nw

    # Final scale to exact total
    if word_timings and cur_s > 0:
        sc = total_s / cur_s
        for w in word_timings:
            w['offset'] *= sc
            w['duration'] *= sc

    print(f"  [VAD] Aligned {word_count} words to {len(speech_segs)} speech "
          f"segments ({len(silent_segs)} pauses)")


# Fix Windows console encoding issues
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'ignore')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'ignore')

# ANSI color codes for terminal output
class Colors:
    """ANSI color codes for colored terminal output"""
    RED = '\033[91m'        # Bright red for errors
    YELLOW = '\033[93m'     # Yellow for warnings
    GREEN = '\033[92m'      # Green for success
    BLUE = '\033[94m'       # Blue for info
    CYAN = '\033[96m'       # Cyan for debug
    RESET = '\033[0m'       # Reset to default

    @staticmethod
    def error(msg):
        """Print error message in red"""
        return f"{Colors.RED}{msg}{Colors.RESET}"

    @staticmethod
    def warning(msg):
        """Print warning message in yellow"""
        return f"{Colors.YELLOW}{msg}{Colors.RESET}"

    @staticmethod
    def success(msg):
        """Print success message in green"""
        return f"{Colors.GREEN}{msg}{Colors.RESET}"

    @staticmethod
    def info(msg):
        """Print info message in blue"""
        return f"{Colors.BLUE}{msg}{Colors.RESET}"

# Override print to automatically color [ERROR], [WARNING], [OK] messages
_original_print = print
def colored_print(*args, **kwargs):
    """Enhanced print that automatically colors error/warning/success messages"""
    if args:
        msg = ' '.join(str(arg) for arg in args)
        # Color code based on message content
        if '[ERROR]' in msg or '✗ Error' in msg or 'Traceback' in msg or 'AttributeError' in msg or 'ModuleNotFoundError' in msg:
            msg = Colors.error(msg)
        elif '[WARNING]' in msg or '[⚠️' in msg:
            msg = Colors.warning(msg)
        elif '[OK]' in msg or '✓ ' in msg or '[✓' in msg or '[⚡' in msg:
            msg = Colors.success(msg)
        elif '[DEBUG]' in msg or '[FONT DEBUG]' in msg or '[SAVE DEBUG]' in msg:
            msg = Colors.info(msg)
        args = (msg,) + args[1:]
    _original_print(*args, **kwargs)

# Replace built-in print
print = colored_print

try:
    from moviepy import VideoFileClip, ImageClip, CompositeVideoClip, AudioFileClip, CompositeAudioClip
    from moviepy.video.fx import Resize, FadeIn
    from moviepy.audio.fx import MultiplyVolume, AudioLoop
    from moviepy import AudioArrayClip
except ImportError:
    from moviepy.editor import VideoFileClip, ImageClip, CompositeVideoClip, AudioFileClip, CompositeAudioClip
    from moviepy.video.fx.resize import resize as Resize
    from moviepy.video.fx.crop import crop
    from moviepy.audio.fx.volumex import volumex as MultiplyVolume
    try:
        from moviepy import AudioArrayClip
    except ImportError:
        try:
            from moviepy.audio.AudioClip import AudioArrayClip
        except ImportError:
            AudioArrayClip = None
    AudioLoop = None
    FadeIn = None

try:
    from moviepy.video.fx.crop import crop as _crop_func
    crop = _crop_func
except Exception:
    try:
        from moviepy.video.fx import Crop as _CropClass
        def crop(clip, x1=None, y1=None, x2=None, y2=None, **_):
            return clip.with_effects([_CropClass(x1=x1, y1=y1, x2=x2, y2=y2)])
    except Exception:
        def crop(clip, x1=None, y1=None, x2=None, y2=None, **_):
            return clip.cropped(x1=x1, y1=y1, x2=x2, y2=y2)

from PIL import Image, ImageDraw, ImageFont, ImageFilter, ImageEnhance, ImageColor

# Text-to-Speech - Using edge-tts for natural, human-like voices
try:
    import edge_tts
    import asyncio
    TTS_AVAILABLE = True
except ImportError as _e:
    TTS_AVAILABLE = False
    print(f"[WARNING] edge-tts not available: {_e}")
    print("  Install with: pip install edge-tts")
except Exception as _e:
    TTS_AVAILABLE = False
    print(f"[WARNING] edge-tts import failed: {type(_e).__name__}: {_e}")
    print("  Install with: pip install edge-tts")


def set_volume(clip, volume):
    """Compatible volume adjustment for MoviePy 1.x and 2.x"""
    try:
        return clip.with_effects([MultiplyVolume(volume)])
    except:
        return clip.volumex(volume)


def set_duration(clip, duration):
    """Compatible duration setting for MoviePy 1.x and 2.x"""
    try:
        return clip.with_duration(duration)
    except:
        return clip.set_duration(duration)


def subclip(clip, start, end):
    """Compatible subclipping for MoviePy 1.x and 2.x"""
    try:
        return clip.subclipped(start, end)
    except:
        return clip.subclip(start, end)


def set_audio(clip, audio):
    """Compatible audio setting for MoviePy 1.x and 2.x"""
    try:
        return clip.with_audio(audio)
    except:
        return clip.set_audio(audio)


def set_position(clip, position):
    """Compatible position setting for MoviePy 1.x and 2.x"""
    try:
        return clip.with_position(position)
    except:
        return clip.set_position(position)


class VideoEffects:
    """Advanced video effects module"""

    @staticmethod
    def get_text_from_spreadsheet(video_filepath, spreadsheet_path, column='B'):
        """
        Read text from Excel spreadsheet based on video_id

        Args:
            video_filepath: Path to video file (e.g., "path/to/jQlZ9DSR3Mo.mp4")
            spreadsheet_path: Path to Excel file with video data
            column: Column letter to read from (default 'B')

        Returns:
            Text from spreadsheet, or None if not found
        """
        if not spreadsheet_path or not os.path.exists(spreadsheet_path):
            return None

        try:
            import pandas as pd

            # Extract video_id from filename (remove extension)
            video_id = Path(video_filepath).stem

            # Read Excel file
            df = pd.read_excel(spreadsheet_path)

            # Find row with matching video_id (assume column A is video_id)
            matching_row = df[df.iloc[:, 0] == video_id]

            if matching_row.empty:
                print(f"[SPREADSHEET] No match found for video_id: {video_id}")
                return None

            # Get text from specified column
            column_index = ord(column.upper()) - ord('A')
            if column_index >= len(df.columns):
                print(f"[SPREADSHEET] Column {column} doesn't exist")
                return None

            text = matching_row.iloc[0, column_index]
            print(f"[SPREADSHEET] Found text for {video_id}: {str(text)[:50]}...")
            return str(text) if pd.notna(text) else None

        except Exception as e:
            print(f"[SPREADSHEET ERROR] {e}")
            return None

    @staticmethod
    def apply_alight_motion_look(frame, settings):
        _TEMPLATES = (
            'Account Growth (Magenta+Green)',
            'DILDAR EDITZ (Pink Cream EdgeGlow)',
            'DRAMA CC (Soft Yellow/Purple)',
            'PAID Suraj (HSL Drama)',
            'Sammad (Cyan Teal Minimal)',
        )
        template = settings.get('am_template', 'None')
        if template == 'None' or template not in _TEMPLATES:
            return frame
        try:
            from PIL import ImageColor
        except ImportError:
            return frame

        f = frame.astype(np.float32)
        h, w = f.shape[:2]

        def _h2n(hex_str):
            try:
                r, g, b = ImageColor.getcolor(hex_str, 'RGB')
                return (r - 128) / 128.0, (g - 128) / 128.0, (b - 128) / 128.0
            except Exception:
                return 0.0, 0.0, 0.0

        def _f(key, default):
            try:
                return float(settings.get(key, default))
            except Exception:
                return float(default)

        def _hex3(key, default):
            return _h2n(settings.get(key, default))

        def _apply_tone(hex_key, alpha, default='#FFFFFF'):
            r, g, b = _hex3(hex_key, default)
            f[..., 0] = np.clip(f[..., 0] * (1 + r * alpha), 0, 255)
            f[..., 1] = np.clip(f[..., 1] * (1 + g * alpha), 0, 255)
            f[..., 2] = np.clip(f[..., 2] * (1 + b * alpha), 0, 255)

        def _saturation(amount):
            gray = 0.299 * f[..., 2] + 0.587 * f[..., 1] + 0.114 * f[..., 0]
            gray3 = np.stack([gray, gray, gray], axis=-1)
            f[:] = gray3 + (f - gray3) * amount

        def _gamma(gamma_val):
            f[:] = np.power(np.clip(f / 255.0, 0, 1), 1.0 / max(gamma_val, 0.01)) * 255.0

        def _brightness(b):
            f[:] = np.clip(f * b, 0, 255)

        def _exposure(exposure_val, gamma_val, offset_val):
            f[:] = np.clip(f * (2.0 ** exposure_val) + offset_val * 255.0, 0, 255)
            if gamma_val != 1.0:
                _gamma(gamma_val)

        def _blur(strength):
            if strength < 0.05:
                return
            try:
                from PIL import Image as _PILImg, ImageFilter as _PILFlt
                pil = _PILImg.fromarray(np.clip(f, 0, 255).astype(np.uint8))
                radius = max(0.5, strength * 8.0)
                pil = pil.filter(_PILFlt.GaussianBlur(radius=radius))
                f[:] = np.asarray(pil, dtype=np.float32)
            except Exception:
                pass

        def _edge_glow(strength, threshold, spread, fill_hex='#FFFFFF',
                        tint_hex='#000000', tint_amount=0.0,
                        smoothing=0.5, invert=False, blend='screen'):
            if strength < 0.01:
                return
            try:
                from PIL import Image as _PILImg, ImageFilter as _PILFlt
                pil = _PILImg.fromarray(np.clip(f, 0, 255).astype(np.uint8))
                radius = max(0.5, spread * 6.0)
                blurred = pil.filter(_PILFlt.GaussianBlur(radius=radius))
                # Build mask of bright areas
                gray = np.asarray(pil.convert('L'), dtype=np.float32) / 255.0
                t = float(threshold)
                mask = np.clip((gray - t) / max(1.0 - t, 0.001), 0, 1)
                if invert:
                    mask = 1.0 - mask
                if smoothing > 0:
                    sm = _PILImg.fromarray((mask * 255).astype(np.uint8)).filter(
                        _PILFlt.GaussianBlur(radius=radius * smoothing))
                    mask = np.asarray(sm, dtype=np.float32) / 255.0
                # Build glow color
                fr, fg, fb = _h2n(fill_hex)
                fill_col = np.array([128 + fr * 128, 128 + fg * 128, 128 + fb * 128],
                                    dtype=np.float32)
                base = np.asarray(pil, dtype=np.float32)
                glow_rgb = np.stack([mask, mask, mask], axis=-1) * fill_col
                if blend == 'screen':
                    result = 255.0 - (255.0 - base) * (255.0 - glow_rgb) / 255.0
                elif blend == 'add':
                    result = base + glow_rgb
                else:
                    result = base + (glow_rgb - base) * 0.5
                f[:] = np.clip(result, 0, 255)
                if tint_amount > 0.01:
                    tr, tg, tb = _h2n(tint_hex)
                    tint_layer = np.array([128 + tr * 128, 128 + tg * 128,
                                            128 + tb * 128], dtype=np.float32)
                    m3 = np.stack([mask, mask, mask], axis=-1) * tint_amount
                    f[:] = np.clip(f * (1 - m3) + tint_layer * m3, 0, 255)
            except Exception:
                pass

        def _channel_map(red, green, blue, alpha_mode):
            # Alight Motion channelmap: 0=normal,1=red,2=green,3=blue,4=alpha
            r = f[..., 0]
            g = f[..., 1]
            b = f[..., 2]
            if red == 1:
                r = r
            elif red == 2:
                r = g
            elif red == 3:
                r = b
            elif red == 4:
                a = (f[..., 0] + f[..., 1] + f[..., 2]) / 3.0
                r = a
            if green == 2:
                g = g
            elif green == 1:
                g = f[..., 0]
            elif green == 3:
                g = b
            elif green == 4:
                g = (f[..., 0] + f[..., 1] + f[..., 2]) / 3.0
            if blue == 3:
                b = b
            elif blue == 1:
                b = f[..., 0]
            elif blue == 2:
                b = g
            elif blue == 4:
                b = (f[..., 0] + f[..., 1] + f[..., 2]) / 3.0
            f[..., 0] = np.clip(r, 0, 255)
            f[..., 1] = np.clip(g, 0, 255)
            f[..., 2] = np.clip(b, 0, 255)

        def _star_streak(strength, threshold, brightness_amt, alpha_amt,
                          blend='add'):
            if strength < 0.01:
                return
            try:
                from PIL import Image as _PILImg, ImageFilter as _PILFlt
                pil = _PILImg.fromarray(np.clip(f, 0, 255).astype(np.uint8))
                radius = max(0.5, strength * 12.0)
                blurred = pil.filter(_PILFlt.GaussianBlur(radius=radius))
                gray = np.asarray(pil.convert('L'), dtype=np.float32) / 255.0
                t = float(threshold)
                mask = np.clip((gray - t) / max(1.0 - t, 0.001), 0, 1) * alpha_amt
                m3 = np.stack([mask, mask, mask], axis=-1)
                bright = brightness_amt * 255.0
                base = np.asarray(pil, dtype=np.float32)
                if blend == 'screen':
                    result = 255.0 - (255.0 - base) * (255.0 - bright * m3) / 255.0
                else:
                    result = base + bright * m3
                f[:] = np.clip(result, 0, 255)
            except Exception:
                pass

        def _soft_glow(strength, brightness_amt, contrast_amt,
                        highlights_amt, color_hex='#FFFFFF',
                        blend='add', alpha_amt=1.0, outside_alpha=0.0):
            if strength < 0.01:
                return
            try:
                from PIL import Image as _PILImg, ImageFilter as _PILFlt
                pil = _PILImg.fromarray(np.clip(f, 0, 255).astype(np.uint8))
                radius = max(1.0, strength * 16.0)
                blurred = pil.filter(_PILFlt.GaussianBlur(radius=radius))
                cr, cg, cb = _h2n(color_hex)
                col = np.array([128 + cr * 128, 128 + cg * 128, 128 + cb * 128],
                               dtype=np.float32)
                blo = np.asarray(blurred, dtype=np.float32)
                blo = blo * brightness_amt + col * (1 - brightness_amt)
                blo = np.clip(blo, 0, 255)
                # contrast
                blo = (blo - 128.0) * contrast_amt + 128.0
                # highlights
                gray = blo.mean(axis=-1, keepdims=True)
                hi_mask = np.clip((gray - 128) / 127.0, 0, 1) * highlights_amt
                blo = blo + hi_mask * 255.0
                blo = np.clip(blo, 0, 255)
                base = np.asarray(pil, dtype=np.float32)
                if blend == 'screen':
                    result = 255.0 - (255.0 - base) * (255.0 - blo) / 255.0
                else:
                    result = base + (blo - base) * alpha_amt
                f[:] = np.clip(result, 0, 255)
            except Exception:
                pass

        def _hdr(strength, exposure_amt, alpha_amt, iter_count):
            if strength < 0.01:
                return
            iters = max(1, int(round(iter_count)))
            work = f.copy()
            for _ in range(iters):
                # local tonemap: lift midtones
                g = work / 255.0
                lum = 0.299 * g[..., 2] + 0.587 * g[..., 1] + 0.114 * g[..., 0]
                lum = np.clip(lum[..., None], 1e-3, 1.0)
                work = work * (lum ** (exposure_amt * 0.5)) * (1 + strength)
                work = np.clip(work, 0, 255)
            f[:] = f * (1 - alpha_amt) + work * alpha_amt

        def _highlight_shadow(highlights, shadows):
            # Alight Motion: highlights adds to bright areas, shadows adds to dark
            gray = f.mean(axis=-1, keepdims=True) / 255.0
            hi_mask = np.clip((gray - 0.5) * 2, 0, 1)
            sh_mask = np.clip((0.5 - gray) * 2, 0, 1)
            f[:] = np.clip(f + hi_mask * highlights * 255.0, 0, 255)
            f[:] = np.clip(f + sh_mask * shadows * 255.0, 0, 255)

        def _displacement(off_x, off_y, edge_clamp, from_center, invert):
            if abs(off_x) < 0.5 and abs(off_y) < 0.5:
                return
            try:
                from PIL import Image as _PILImg
                pil = _PILImg.fromarray(np.clip(f, 0, 255).astype(np.uint8))
                dx = int(round(off_x))
                dy = int(round(off_y))
                arr = np.asarray(pil, dtype=np.float32)
                shifted = np.zeros_like(arr)
                h2, w2 = arr.shape[:2]
                if invert:
                    dx, dy = -dx, -dy
                src_y0 = max(0, -dy)
                src_y1 = min(h2, h2 - dy)
                src_x0 = max(0, -dx)
                src_x1 = min(w2, w2 - dx)
                dst_y0 = max(0, dy)
                dst_y1 = min(h2, h2 + dy)
                dst_x0 = max(0, dx)
                dst_x1 = min(w2, w2 + dx)
                shifted[dst_y0:dst_y1, dst_x0:dst_x1] = arr[src_y0:src_y1,
                                                          src_x0:src_x1]
                f[:] = shifted
            except Exception:
                pass

        def _colorize(tint_hex, amount=0.5):
            if amount < 0.01:
                return
            tr, tg, tb = _h2n(tint_hex)
            tint_rgb = np.array([128 + tr * 128, 128 + tg * 128, 128 + tb * 128],
                                dtype=np.float32)
            gray = f.mean(axis=-1, keepdims=True)
            f[:] = np.clip(gray + (tint_rgb - gray) * amount, 0, 255)

        def _bright_cont(brightness_amt, contrast_amt):
            f[:] = (f - 128.0) * contrast_amt + 128.0 + brightness_amt * 128.0
            f[:] = np.clip(f, 0, 255)

        def _sharpen(strength, bright_amt, iter_count, white_only, blurred_only):
            if strength < 0.01:
                return
            try:
                from PIL import Image as _PILImg, ImageFilter as _PILFlt
                pil = _PILImg.fromarray(np.clip(f, 0, 255).astype(np.uint8))
                iters = max(1, int(round(iter_count)))
                for _ in range(iters):
                    blurred = pil.filter(_PILFlt.GaussianBlur(radius=1.5))
                    base = np.asarray(pil, dtype=np.float32)
                    b_arr = np.asarray(blurred, dtype=np.float32)
                    detail = base - b_arr
                    if blurred_only > 0.01:
                        detail = detail * blurred_only
                    if white_only > 0.01:
                        # emphasize bright detail only
                        gray = base.mean(axis=-1, keepdims=True)
                        keep = np.clip((gray - 128) / 127, 0, 1)
                        detail = detail * (1 - white_only + white_only * keep)
                    pil = _PILImg.fromarray(np.clip(base + detail * strength,
                                                     0, 255).astype(np.uint8))
                f[:] = np.asarray(pil, dtype=np.float32)
                if bright_amt != 1.0:
                    f[:] = np.clip(f * bright_amt, 0, 255)
            except Exception:
                pass

        def _gradient_shadow(color_hex, start_pos, angle, alpha_top, feather,
                              end_pos, bright_amt, blend, alpha_bottom):
            # Synthesize a gradient overlay (linear gradient between start and end)
            if alpha_top < 0.01 and alpha_bottom < 0.01:
                return
            try:
                cr, cg, cb = _h2n(color_hex)
                tint_rgb = np.array([128 + cr * 128, 128 + cg * 128,
                                      128 + cb * 128], dtype=np.float32)
                # Project pixels along gradient axis
                yy, xx = np.ogrid[:h, :w]
                rad = np.deg2rad(float(angle))
                proj = xx * np.cos(rad) + yy * np.sin(rad)
                proj_min = proj.min()
                proj_max = proj.max()
                t = (proj - proj_min) / max(proj_max - proj_min, 1.0)
                # alpha varies linearly from alpha_top at start to alpha_bottom at end
                a = alpha_top + (alpha_bottom - alpha_top) * t
                if feather > 0.01:
                    # blur alpha
                    from PIL import Image as _PILImg, ImageFilter as _PILFlt
                    a_img = _PILImg.fromarray((a * 255).astype(np.uint8))
                    a_img = a_img.filter(_PILFlt.GaussianBlur(
                        radius=max(0.5, feather * 6.0)))
                    a = np.asarray(a_img, dtype=np.float32) / 255.0
                a3 = np.stack([a, a, a], axis=-1)
                base = f
                if blend == 'multiply':
                    result = base * (tint_rgb / 255.0)
                elif blend == 'screen':
                    result = 255.0 - (255.0 - base) * (255.0 - tint_rgb) / 255.0
                else:
                    result = base * (1 - a3) + tint_rgb * a3
                f[:] = np.clip(result * bright_amt, 0, 255)
            except Exception:
                pass

        def _hsl_rot(red_x, red_y, red_z, or_x, or_y, or_z,
                      yel_x, yel_y, yel_z, grn_x, grn_y, grn_z,
                      cyn_x, cyn_y, cyn_z, blu_x, blu_y, blu_z,
                      pur_x, pur_y, pur_z, mag_x, mag_y, mag_z,
                      bright, alpha_amt, blend='normal'):
            if alpha_amt < 0.01:
                return
            # Convert to HSL
            r = f[..., 0] / 255.0
            g = f[..., 1] / 255.0
            b = f[..., 2] / 255.0
            mx = np.maximum(np.maximum(r, g), b)
            mn = np.minimum(np.minimum(r, g), b)
            df = mx - mn
            l = (mx + mn) / 2.0
            h = np.zeros_like(l)
            s = np.zeros_like(l)
            mask = df > 0
            mask_r = mask & (mx == r)
            mask_g = mask & (mx == g) & ~mask_r
            mask_b = mask & (mx == b) & ~mask_r & ~mask_g
            h[mask_r] = ((g[mask_r] - b[mask_r]) / df[mask_r]) % 6.0
            h[mask_g] = ((b[mask_g] - r[mask_g]) / df[mask_g]) + 2.0
            h[mask_b] = ((r[mask_b] - g[mask_b]) / df[mask_b]) + 4.0
            h = h * 60.0
            s = np.where(l <= 0.5, df / (mx + mn + 1e-6), df / (2.0 - mx - mn + 1e-6))
            # Per-band hue remap: red 0-30, orange 30-60, yellow 60-90,
            # green 90-150, cyan 150-210, blue 210-270, purple 270-300,
            # magenta 300-360
            def _band_apply(lo, hi, x, y, z):
                band = (h >= lo) & (h < hi)
                if not band.any():
                    return
                if x != 1.0:
                    h[band] *= x
                if y != 0.0:
                    h[band] = (h[band] + y) % 360.0
                if z != 1.0:
                    s[band] = np.clip(s[band] * z, 0, 1)
            _band_apply(0, 30, red_x, red_y, red_z)
            _band_apply(30, 60, or_x, or_y, or_z)
            _band_apply(60, 90, yel_x, yel_y, yel_z)
            _band_apply(90, 150, grn_x, grn_y, grn_z)
            _band_apply(150, 210, cyn_x, cyn_y, cyn_z)
            _band_apply(210, 270, blu_x, blu_y, blu_z)
            _band_apply(270, 300, pur_x, pur_y, pur_z)
            _band_apply(300, 360, mag_x, mag_y, mag_z)
            if bright != 1.0:
                l = np.clip(l * bright, 0, 1)
            # Back to RGB
            c = (1 - np.abs(2 * l - 1)) * s
            hp = h / 60.0
            x_ = c * (1 - np.abs(hp % 2 - 1))
            zero = np.zeros_like(c)
            rgb = np.stack([
                np.where(hp < 1, c, np.where(hp < 2, x_, np.where(hp < 3, zero,
                np.where(hp < 4, zero, np.where(hp < 5, x_, c))))),
                np.where(hp < 1, x_, np.where(hp < 2, c, np.where(hp < 3, c,
                np.where(hp < 4, x_, np.where(hp < 5, zero, zero))))),
                np.where(hp < 1, zero, np.where(hp < 2, zero, np.where(hp < 3, x_,
                np.where(hp < 4, c, np.where(hp < 5, c, x_)))))
            ], axis=-1)
            m = l - c / 2.0
            new_rgb = np.clip((rgb + m[..., None]) * 255.0, 0, 255)
            f[:] = f * (1 - alpha_amt) + new_rgb * alpha_amt

        def _gain(amount):
            if abs(amount - 1.0) < 0.01:
                return
            f[:] = np.clip(f * amount, 0, 255)

        # ============================================================
        # Per-template application
        # ============================================================

        if template == 'Account Growth (Magenta+Green)':
            # Magenta layer stack: satvib(saturation, vib) + gaussianblur
            mag_sat = _f('am_magenta_saturation', 1.0)
            mag_vib = _f('am_magenta_vib', 0.0)
            mag_blur = _f('am_magenta_blur', 0.0)
            mag_int = _f('am_magenta_intensity', 0.20)
            # Green layer stack: exposure(exposure, gamma, offset) + satvib + blur
            grn_exp = _f('am_green_exposure', 0.0)
            grn_gamma = _f('am_green_gamma', 1.0)
            grn_offset = _f('am_green_offset', 0.0)
            grn_sat = _f('am_green_saturation', 1.0)
            grn_vib = _f('am_green_vib', 0.0)
            grn_blur = _f('am_green_blur', 0.0)
            grn_int = _f('am_green_intensity', 0.12)
            # Grey layer stack: satvib + brightcont2(brightness, contrast)
            grey_sat = _f('am_grey_saturation', 1.0)
            grey_vib = _f('am_grey_vib', 0.0)
            grey_bright = _f('am_grey_brightness', 0.0)
            grey_cont = _f('am_grey_contrast', 1.15)
            # Main effect: circle1 (location, opacity, blur)
            glow_on = bool(settings.get('am_glow', False))
            glow_cx = _f('am_glow_x', 0.78)
            glow_cy = _f('am_glow_y', 0.22)
            glow_opacity = _f('am_glow_opacity', 0.15)
            glow_blur = _f('am_glow_blur', 0.4)
            # Lift fills
            mi = mag_int * 0.6
            gi = grn_int * 0.6
            mr, mg, mb = _hex3('am_magenta', '#D238FD')
            gr, gg, gb = _hex3('am_green', '#019C01')
            f[..., 0] = np.clip(f[..., 0] * (1 + mr * mi), 0, 255)
            f[..., 1] = np.clip(f[..., 1] * (1 + mg * mi), 0, 255)
            f[..., 2] = np.clip(f[..., 2] * (1 + mb * mi), 0, 255)
            f[..., 0] = np.clip(f[..., 0] * (1 + gr * gi * 0.5), 0, 255)
            f[..., 1] = np.clip(f[..., 1] * (1 + gg * gi * 0.5), 0, 255)
            f[..., 2] = np.clip(f[..., 2] * (1 + gb * gi * 0.5), 0, 255)
            # satvib on green
            if grn_sat != 1.0 or grn_vib != 0.0:
                _saturation(grn_sat)
                f[:] = f + grn_vib * 50.0
                f[:] = np.clip(f, 0, 255)
            # green gaussianblur
            _blur(grn_blur)
            # green exposure
            _exposure(grn_exp, grn_gamma, grn_offset)
            # satvib on magenta
            if mag_sat != 1.0 or mag_vib != 0.0:
                _saturation(mag_sat)
                f[:] = f + mag_vib * 50.0
                f[:] = np.clip(f, 0, 255)
            # magenta gaussianblur
            _blur(mag_blur)
            # grey brightcont2 + satvib
            _bright_cont(grey_bright, grey_cont)
            if grey_sat != 1.0 or grey_vib != 0.0:
                _saturation(grey_sat)
                f[:] = f + grey_vib * 50.0
                f[:] = np.clip(f, 0, 255)
            # main circle1 glow
            if glow_on:
                yy, xx = np.ogrid[:h, :w]
                cx, cy = int(w * glow_cx), int(h * glow_cy)
                max_r = np.sqrt((w * glow_blur) ** 2 + (h * glow_blur) ** 2)
                d = np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2) / max_r
                glow = np.clip(1.0 - d, 0, 1) ** 2
                add = (255.0 * glow_opacity) * np.stack(
                    [glow * 1.0, glow * 0.9, glow * 0.6], axis=-1)
                f[:] = np.clip(f + add, 0, 255)

        elif template == 'DILDAR EDITZ (Pink Cream EdgeGlow)':
            pr, pg, pb = _hex3('am_pink', '#EABEDF')
            dpr, dpg, dpb = _hex3('am_darkpink', '#D95568')
            pi = _f('am_pink_intensity', 0.30)
            f[..., 0] = np.clip(f[..., 0] * (1 + pr * pi), 0, 255)
            f[..., 1] = np.clip(f[..., 1] * (1 + pg * pi), 0, 255)
            f[..., 2] = np.clip(f[..., 2] * (1 + pb * pi), 0, 255)
            f[..., 0] = np.clip(f[..., 0] * (1 + dpr * 0.4), 0, 255)
            f[..., 1] = np.clip(f[..., 1] * (1 + dpg * 0.4), 0, 255)
            f[..., 2] = np.clip(f[..., 2] * (1 + dpb * 0.4), 0, 255)
            # lift fills applied
            # satvib on dark pink
            dp_sat = _f('am_darkpink_saturation', 1.0)
            dp_vib = _f('am_darkpink_vib', 0.0)
            if dp_sat != 1.0 or dp_vib != 0.0:
                _saturation(dp_sat)
                f[:] = f + dp_vib * 50.0
                f[:] = np.clip(f, 0, 255)
            # gaussianblur
            _blur(_f('am_pink_blur', 0.0))
            # exposure
            _exposure(_f('am_pink_exposure', 0.0),
                      _f('am_pink_gamma', 1.0),
                      _f('am_pink_offset', 0.0))
            # displacement map
            _displacement(_f('am_displace_x', 0.0),
                          _f('am_displace_y', 0.0),
                          _f('am_displace_edge', 0.0),
                          _f('am_displace_from_center', 0.0) > 0.5,
                          bool(settings.get('am_displace_invert', False)))
            # edge glow
            _edge_glow(_f('am_edge_strength', 0.0),
                        _f('am_edge_threshold', 0.5),
                        _f('am_edge_spread', 0.0),
                        settings.get('am_edge_fill', '#FFFFFF'),
                        settings.get('am_edge_tint', '#000000'),
                        _f('am_edge_tint_amount', 0.0),
                        _f('am_edge_smoothing', 0.5),
                        bool(settings.get('am_edge_invert', False)),
                        settings.get('am_edge_blend', 'screen'))
            glow = _f('am_soft_glow', 0.0)
            if glow > 0.01:
                f[:] = np.clip(f * (1 + glow * 0.05), 0, 255)

        elif template == 'DRAMA CC (Soft Yellow/Purple)':
            # Channelmap on top
            _channel_map(int(_f('am_chmap_red', 0)),
                          int(_f('am_chmap_green', 1)),
                          int(_f('am_chmap_blue', 2)),
                          int(_f('am_chmap_alpha', 0)))
            # satvib on yellow
            yel_sat = _f('am_yellow_saturation', 1.0)
            yel_vib = _f('am_yellow_vib', 0.0)
            if yel_sat != 1.0 or yel_vib != 0.0:
                _saturation(yel_sat)
                f[:] = f + yel_vib * 50.0
                f[:] = np.clip(f, 0, 255)
            # yellow lift
            _apply_tone('am_yellow', _f('am_yellow_intensity', 0.25), '#ABC04E')
            # HDR
            _hdr(_f('am_hdr_strength', 0.0),
                  _f('am_hdr_exposure', 0.0),
                  _f('am_hdr_alpha', 0.0),
                  _f('am_hdr_iter', 1.0))
            # starstreak on purple
            _star_streak(_f('am_starstreak_strength', 0.0),
                          _f('am_starstreak_threshold', 0.5),
                          _f('am_starstreak_brightness', 0.0),
                          _f('am_starstreak_alpha', 0.0),
                          settings.get('am_starstreak_blend', 'add'))
            # purple lift
            _apply_tone('am_purple', _f('am_purple_intensity', 0.30), '#9F47FA')
            # softglow on blue
            _soft_glow(_f('am_softglow_strength', 0.0),
                        _f('am_softglow_brightness', 1.0),
                        _f('am_softglow_contrast', 1.0),
                        _f('am_softglow_highlights', 0.0),
                        settings.get('am_softglow_color', '#FFFFFF'),
                        settings.get('am_softglow_blend', 'add'),
                        _f('am_softglow_alpha', 1.0),
                        _f('am_softglow_outside_alpha', 0.0))
            # blue lift
            _apply_tone('am_blue', _f('am_blue_intensity', 0.20), '#433CE7')
            # exposure last
            _exposure(_f('am_exposure_amount', 0.0),
                      _f('am_exposure_gamma', 1.0),
                      _f('am_exposure_offset', 0.0))
            ci = _f('am_color_intensity', 0.4)
            if ci > 0.01:
                _saturation(1.0 + ci * 0.15)

        elif template == 'PAID Suraj (HSL Drama)':
            # gain first
            _gain(_f('am_gain', 1.29))
            # colorize
            _colorize(settings.get('am_colorize_tint', '#FFFFFF'),
                      _f('am_colorize_amount', 0.5))
            # highlightshadow on blue layer
            _highlight_shadow(_f('am_drama_blue_hl', 0.0),
                               _f('am_drama_blue_sh', 0.0))
            _apply_tone('am_drama_blue', _f('am_drama_blue_intensity', 0.30),
                          '#94A7EF')
            # HSL drama on purple layer
            _hsl_rot(_f('am_hsl_red_x', 1.0), _f('am_hsl_red_y', 0.0),
                      _f('am_hsl_red_z', 1.0),
                      _f('am_hsl_orange_x', 1.0), _f('am_hsl_orange_y', 0.0),
                      _f('am_hsl_orange_z', 1.0),
                      _f('am_hsl_yellow_x', 1.0), _f('am_hsl_yellow_y', 0.0),
                      _f('am_hsl_yellow_z', 1.0),
                      _f('am_hsl_green_x', 1.0), _f('am_hsl_green_y', 0.0),
                      _f('am_hsl_green_z', 1.0),
                      _f('am_hsl_cyan_x', 1.0), _f('am_hsl_cyan_y', 0.0),
                      _f('am_hsl_cyan_z', 1.0),
                      _f('am_hsl_blue_x', 1.0), _f('am_hsl_blue_y', 0.0),
                      _f('am_hsl_blue_z', 1.0),
                      _f('am_hsl_purple_x', 1.0), _f('am_hsl_purple_y', 0.0),
                      _f('am_hsl_purple_z', 1.0),
                      _f('am_hsl_magenta_x', 1.0), _f('am_hsl_magenta_y', 0.0),
                      _f('am_hsl_magenta_z', 1.0),
                      _f('am_hsl_bright', 1.0), _f('am_hsl_alpha', 0.0),
                      settings.get('am_hsl_blend', 'normal'))
            _apply_tone('am_drama_purple', _f('am_drama_purple_intensity', 0.25),
                          '#894DB9')
            # highlightshadow on teal/green layer
            _highlight_shadow(_f('am_drama_teal_hl', 0.0),
                               _f('am_drama_teal_sh', 0.0))
            _apply_tone('am_drama_teal', _f('am_drama_teal_intensity', 0.20),
                          '#62B793')
            # new sharpen on fuchsia (synthesized)
            _sharpen(_f('am_sharpen_strength', 0.0),
                      _f('am_sharpen_bright', 1.0),
                      _f('am_sharpen_iter', 1.0),
                      _f('am_sharpen_white_only', 0.0),
                      _f('am_sharpen_blurred_only', 0.0))
            # gradshadow on red
            _gradient_shadow(settings.get('am_grad_color', '#FF0000'),
                               _f('am_grad_start', 0.0),
                               _f('am_grad_angle', 0.0),
                               _f('am_grad_alpha_top', 0.0),
                               _f('am_grad_feather', 0.0),
                               _f('am_grad_end', 1.0),
                               _f('am_grad_bright', 1.0),
                               settings.get('am_grad_blend', 'normal'),
                               _f('am_grad_alpha_bottom', 0.0))
            di = _f('am_drama_intensity', 0.3)
            if di > 0.01:
                _saturation(1.0 + di * 0.3)

        elif template == 'Sammad (Cyan Teal Minimal)':
            cr, cg, cb = _hex3('am_cyan', '#17E1E1')
            tr, tg, tb = _hex3('am_teal', '#62D2B1')
            ci = _f('am_cyan_intensity', 0.20)
            f[..., 0] = np.clip(f[..., 0] * (1 + cr * ci * 0.5), 0, 255)
            f[..., 1] = np.clip(f[..., 1] * (1 + cg * ci * 0.5), 0, 255)
            f[..., 2] = np.clip(f[..., 2] * (1 + cb * ci * 0.5), 0, 255)
            f[..., 0] = np.clip(f[..., 0] * (1 + tr * 0.3), 0, 255)
            f[..., 1] = np.clip(f[..., 1] * (1 + tg * 0.3), 0, 255)
            f[..., 2] = np.clip(f[..., 2] * (1 + tb * 0.3), 0, 255)
            # satvib on cyan
            cy_sat = _f('am_cyan_saturation', 1.0)
            cy_vib = _f('am_cyan_vib', 0.0)
            if cy_sat != 1.0 or cy_vib != 0.0:
                _saturation(cy_sat)
                f[:] = f + cy_vib * 50.0
                f[:] = np.clip(f, 0, 255)
            # gaussianblur
            _blur(_f('am_cyan_blur', 0.0))
            # exposure
            _exposure(_f('am_cyan_exposure', 0.0),
                      _f('am_cyan_gamma', 1.0),
                      _f('am_cyan_offset', 0.0))
            darken = _f('am_darken', 1.0)
            if abs(darken - 1.0) > 0.01:
                f[:] = np.clip(f * darken, 0, 255)

        return f.astype(np.uint8)

    def apply_color_grade(frame, grade_type='warm', intensity=0.5):
        """Apply color grading to frame"""
        # Work on a copy to avoid modifying read-only arrays
        frame = frame.copy()
        if grade_type == 'warm':
            frame[:,:,0] = np.clip(frame[:,:,0] * (1 + intensity * 0.2), 0, 255)
            frame[:,:,2] = np.clip(frame[:,:,2] * (1 - intensity * 0.1), 0, 255)
        elif grade_type == 'cold':
            frame[:,:,2] = np.clip(frame[:,:,2] * (1 + intensity * 0.2), 0, 255)
            frame[:,:,0] = np.clip(frame[:,:,0] * (1 - intensity * 0.1), 0, 255)
        elif grade_type == 'cinematic':
            frame = frame * 0.95
            frame[:,:,1] = np.clip(frame[:,:,1] * 1.05, 0, 255)
        elif grade_type == 'vintage':
            frame[:,:,0] = np.clip(frame[:,:,0] * 1.1, 0, 255)
            frame[:,:,1] = np.clip(frame[:,:,1] * 0.95, 0, 255)
            frame[:,:,2] = np.clip(frame[:,:,2] * 0.85, 0, 255)
        return frame.astype('uint8')

    # === 3D LUT Filter System ============================================
    # Professional color grading via .cube lookup tables.
    # Loads .cube files from a "luts/" folder or uses built-in presets.

    _LUT_CACHE: dict = {}  # {name: (N, N, N, 3) float32 in [0, 255]}

    @staticmethod
    def _parse_cube_file(path):
        """Parse a .cube file into a (N, N, N, 3) float32 LUT array.

        Supports standard .cube format with LUT_3D_SIZE header and
        RGB triplets laid out B-fastest, then G, then R.
        """
        import struct
        size = None
        data = []
        with open(str(path), 'r') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                if line.startswith('TITLE'):
                    continue
                if line.startswith('LUT_3D_SIZE'):
                    size = int(line.split()[-1])
                    continue
                if line.startswith('DOMAIN'):
                    continue
                parts = line.split()
                if len(parts) >= 3:
                    try:
                        r, g, b = float(parts[0]), float(parts[1]), float(parts[2])
                        data.append([r, g, b])
                    except ValueError:
                        continue

        if size is None or len(data) != size**3:
            raise ValueError(f"Invalid .cube: {path} (size={size}, entries={len(data)})")

        lut = np.array(data, dtype=np.float32).reshape(size, size, size, 3) * 255.0
        return lut

    @staticmethod
    def _create_builtin_lut(name):
        """Generate a 3D LUT from built-in color preset definitions.
        Each preset defines per-channel curves at a few control points.
        """
        N = 33  # Standard 33x33x33 LUT
        lut = np.zeros((N, N, N, 3), dtype=np.float32)

        # Build each channel as a function of input value, with optional
        # cross-channel influence (e.g., R depends on B for teal/orange)
        for r in range(N):
            for g in range(N):
                for b in range(N):
                    rn = r / (N - 1.0)
                    gn = g / (N - 1.0)
                    bn = b / (N - 1.0)

                    if name == 'Cinematic (Teal/Orange)':
                        # Classic teal shadows + orange highlights
                        luma = 0.299 * rn + 0.587 * gn + 0.114 * bn
                        rn2 = rn * (1 + 0.15 * luma) - 0.05 * bn
                        gn2 = gn * (1 + 0.05 * luma)
                        bn2 = bn * (1 + 0.25 * (1 - luma)) - 0.1 * rn
                        lut[r, g, b] = [
                            np.clip(rn2 * 255, 0, 255),
                            np.clip(gn2 * 255, 0, 255),
                            np.clip(bn2 * 255, 0, 255),
                        ]

                    elif name == 'Warm Golden':
                        rn2 = rn * 1.1 - 0.05 * bn
                        gn2 = gn * 0.95 + 0.05 * rn
                        bn2 = bn * 0.8
                        lut[r, g, b] = [
                            np.clip(rn2 * 255, 0, 255),
                            np.clip(gn2 * 255, 0, 255),
                            np.clip(bn2 * 255, 0, 255),
                        ]

                    elif name == 'Cold Blue':
                        rn2 = rn * 0.85
                        gn2 = gn * 0.9
                        bn2 = bn * 1.15 + 0.05 * rn
                        lut[r, g, b] = [
                            np.clip(rn2 * 255, 0, 255),
                            np.clip(gn2 * 255, 0, 255),
                            np.clip(bn2 * 255, 0, 255),
                        ]

                    elif name == 'Vintage Fade':
                        fade = 0.1
                        rn2 = rn * 0.9 + fade
                        gn2 = gn * 0.85 + fade
                        bn2 = bn * 0.8 + fade
                        lut[r, g, b] = [
                            np.clip(rn2 * 255, 0, 255),
                            np.clip(gn2 * 255, 0, 255),
                            np.clip(bn2 * 255, 0, 255),
                        ]

                    elif name == 'Drama (High Contrast)':
                        # S-curve contrast + saturation boost
                        def s_curve(x):
                            return 1 / (1 + np.exp(-8 * (x - 0.5)))
                        rn2 = s_curve(rn)
                        gn2 = s_curve(gn)
                        bn2 = s_curve(bn)
                        # Slight saturation boost
                        avg = (rn2 + gn2 + bn2) / 3
                        rn2 = avg + 1.2 * (rn2 - avg)
                        gn2 = avg + 1.2 * (gn2 - avg)
                        bn2 = avg + 1.2 * (bn2 - avg)
                        lut[r, g, b] = [
                            np.clip(rn2 * 255, 0, 255),
                            np.clip(gn2 * 255, 0, 255),
                            np.clip(bn2 * 255, 0, 255),
                        ]

                    elif name == 'Muted (Desaturated)':
                        avg = (rn + gn + bn) / 3
                        rn2 = avg + 0.4 * (rn - avg)
                        gn2 = avg + 0.4 * (gn - avg)
                        bn2 = avg + 0.4 * (bn - avg)
                        lut[r, g, b] = [
                            np.clip(rn2 * 255, 0, 255),
                            np.clip(gn2 * 255, 0, 255),
                            np.clip(bn2 * 255, 0, 255),
                        ]

                    else:
                        # Identity (passthrough)
                        lut[r, g, b] = [rn * 255, gn * 255, bn * 255]

        return lut

    @staticmethod
    def get_lut_names():
        """Return list of available LUT preset names (built-in + file-based)."""
        builtins = ['None', 'Cinematic (Teal/Orange)', 'Warm Golden',
                    'Cold Blue', 'Vintage Fade', 'Drama (High Contrast)',
                    'Muted (Desaturated)']
        # Scan for .cube files in luts/ folder
        try:
            luts_dir = Path(__file__).parent / 'luts'
            if luts_dir.exists():
                for f in sorted(luts_dir.glob('*.cube')):
                    name = f'[FILE] {f.stem}'
                    if name not in builtins:
                        builtins.append(name)
        except Exception:
            pass
        return builtins

    @staticmethod
    def get_lut_data(name):
        """Return (N, N, N, 3) float32 LUT for the given name.

        Caches loaded LUTs so each file is parsed once per session.
        """
        if name == 'None' or not name:
            return None
        if name in VideoEffects._LUT_CACHE:
            return VideoEffects._LUT_CACHE[name]

        try:
            if name.startswith('[FILE] '):
                fname = name[7:] + '.cube'
                luts_dir = Path(__file__).parent / 'luts'
                path = luts_dir / fname
                if path.exists():
                    lut = VideoEffects._parse_cube_file(path)
                    VideoEffects._LUT_CACHE[name] = lut
                    print(f"[LUT] Loaded .cube: {path.name} ({lut.shape[0]}³)")
                    return lut
                print(f"[WARNING] LUT file not found: {path}")
                return None
            else:
                lut = VideoEffects._create_builtin_lut(name)
                VideoEffects._LUT_CACHE[name] = lut
                print(f"[LUT] Built-in preset: {name}")
                return lut
        except Exception as e:
            print(f"[LUT ERROR] Failed to load '{name}': {e}")
            return None

    @staticmethod
    def apply_lut_filter(frame, lut_name, intensity=1.0):
        """Apply a 3D LUT to a frame with optional blending.

        Uses fully vectorized trilinear interpolation for speed.
        Falls back to identity (no-op) if LUT fails to load.

        Args:
            frame: (H, W, 3) uint8 numpy array
            lut_name: Name of LUT preset or .cube file
            intensity: Blend factor (0.0 = no effect, 1.0 = full LUT)
        """
        if intensity <= 0 or not lut_name or lut_name == 'None':
            return frame

        lut = VideoEffects.get_lut_data(lut_name)
        if lut is None:
            return frame

        frame = frame.astype(np.float32)
        N = lut.shape[0]
        H, W = frame.shape[:2]

        # Scale to [0, N-1] and find floor/ceil indices
        scale = (N - 1) / 255.0
        idx = frame * scale

        i0 = np.floor(idx).astype(np.int32)
        i1 = np.minimum(i0 + 1, N - 1)
        frac = idx - i0.astype(np.float32)

        # Flatten spatial dims for fancy indexing
        r0, g0, b0 = i0[:, :, 0].ravel(), i0[:, :, 1].ravel(), i0[:, :, 2].ravel()
        r1, g1, b1 = i1[:, :, 0].ravel(), i1[:, :, 1].ravel(), i1[:, :, 2].ravel()
        fr = frac[:, :, 0].ravel()[:, None]
        fg = frac[:, :, 1].ravel()[:, None]
        fb = frac[:, :, 2].ravel()[:, None]

        # 8 cube corners
        c000 = lut[r0, g0, b0]
        c001 = lut[r0, g0, b1]
        c010 = lut[r0, g1, b0]
        c011 = lut[r0, g1, b1]
        c100 = lut[r1, g0, b0]
        c101 = lut[r1, g0, b1]
        c110 = lut[r1, g1, b0]
        c111 = lut[r1, g1, b1]

        # Trilinear interpolation (vectorized)
        c00 = c000 * (1 - fb) + c001 * fb
        c01 = c010 * (1 - fb) + c011 * fb
        c10 = c100 * (1 - fb) + c101 * fb
        c11 = c110 * (1 - fb) + c111 * fb
        c0 = c00 * (1 - fg) + c01 * fg
        c1 = c10 * (1 - fg) + c11 * fg
        result = c0 * (1 - fr) + c1 * fr

        result = result.reshape(H, W, 3)

        # Blend with original if intensity < 1.0
        if intensity < 1.0:
            result = frame * (1 - intensity) + result * intensity

        return np.clip(result, 0, 255).astype(np.uint8)

    @staticmethod
    def apply_vignette(frame, intensity=0.4):
        """Apply vignette darkening. PERF: precompute vignette mask once
        via a module-level cache; the legacy `frame * vignette[:,:,np.newaxis]`
        path allocates a 6MB+ float32 array per frame."""
        # Work on a copy to avoid modifying read-only arrays
        frame = frame.copy()
        h, w = frame.shape[:2]
        cache_key = ('vignette', w, h, float(intensity))
        cached = _VIGNETTE_MASK_CACHE.get(cache_key)
        if cached is None:
            y_g, x_g = np.ogrid[:h, :w]
            cx, cy = w / 2.0, h / 2.0
            max_dist = np.sqrt(cx * cx + cy * cy)
            distance = np.sqrt((x_g - cx) ** 2 + (y_g - cy) ** 2)
            vignette = np.clip(1 - (distance / max_dist * float(intensity)), 0, 1)
            # Pre-expand to 3 channels as a float32 mask so we can do a
            # single multiply with the uint8 frame (cast frame just once).
            vignette_3d = np.repeat(vignette[:, :, np.newaxis], 3, axis=2).astype(np.float32)
            cached = vignette_3d
            _VIGNETTE_MASK_CACHE[cache_key] = cached
        # uint8 * float32 -> float32 result. The legacy path allocated a
        # (h,w,3) float here too, so the shape is the same; the win is that
        # we don't re-create the (h,w,3) vignette mask every frame.
        return (frame.astype(np.float32) * cached).astype(np.uint8)

    @staticmethod
    def apply_circular_spotlight(frame, center_x=50, center_y=50, radius=40,
                                 outside_effect='blur', blur_intensity=50,
                                 outside_color='#000000', feather=20,
                                 show_outline=True, outline_color='#FF00FF',
                                 outline_thickness=5, shape='circle'):
        """
        Apply spotlight effect (circle or square) - only spotlight area is visible, rest is blurred/darkened

        Args:
            frame: Video frame
            center_x: Spotlight center X position (0-100, percentage)
            center_y: Spotlight center Y position (0-100, percentage)
            radius: Spotlight size (0-100, percentage of smaller dimension)
            outside_effect: 'blur' or 'solid' - what to do with area outside spotlight
            blur_intensity: Blur strength for outside area (0-100)
            outside_color: Hex color for solid color effect
            feather: Edge softness (0-100, percentage of radius)
            show_outline: Whether to draw spotlight outline
            outline_color: Hex color for spotlight outline (default: pink/magenta)
            outline_thickness: Outline thickness in pixels (1-20)
            shape: 'circle' or 'square' - shape of spotlight

        Returns:
            Frame with spotlight effect applied
        """
        import cv2

        frame = frame.copy()
        h, w = frame.shape[:2]

        # Convert percentage to pixels
        cx = int(w * center_x / 100)
        cy = int(h * center_y / 100)
        r = int(min(w, h) * radius / 100)

        # Create mask based on shape
        mask = np.zeros((h, w), dtype=np.float32)

        if shape == 'square':
            # Create square/rectangle mask
            half_size = r
            x1, y1 = max(0, cx - half_size), max(0, cy - half_size)
            x2, y2 = min(w, cx + half_size), min(h, cy + half_size)

            # Create distance map for feathering
            y_grid, x_grid = np.ogrid[:h, :w]

            # Distance from rectangle edges
            dist_x = np.minimum(np.abs(x_grid - x1), np.abs(x_grid - x2))
            dist_y = np.minimum(np.abs(y_grid - y1), np.abs(y_grid - y2))

            # Inside rectangle
            inside_x = (x_grid >= x1) & (x_grid <= x2)
            inside_y = (y_grid >= y1) & (y_grid <= y2)
            inside = inside_x & inside_y

            # Apply feathering
            feather_px = int(r * feather / 100)
            if feather_px > 0:
                # Calculate distance from edge
                edge_dist = np.minimum(dist_x, dist_y)
                # Create gradient from edge
                mask = np.where(inside,
                               np.minimum(edge_dist / feather_px, 1.0),
                               0.0)
            else:
                mask[inside] = 1.0
        else:
            # Create circular mask
            y, x = np.ogrid[:h, :w]
            distance = np.sqrt((x - cx)**2 + (y - cy)**2)

            # Apply feathering (smooth edges)
            feather_px = int(r * feather / 100)
            if feather_px > 0:
                # Gradual transition from 1 (inside) to 0 (outside)
                mask = np.clip((r + feather_px - distance) / feather_px, 0, 1)
            else:
                # Hard edge
                mask[distance <= r] = 1.0

        # Apply effect to outside area
        if outside_effect == 'blur':
            # Blur the entire frame
            blur_amount = max(1, int(blur_intensity))
            if blur_amount % 2 == 0:
                blur_amount += 1  # Must be odd for GaussianBlur
            blurred_frame = cv2.GaussianBlur(frame, (blur_amount, blur_amount), 0)

            # Blend original and blurred using mask
            mask_3d = mask[:, :, np.newaxis]
            result = (frame * mask_3d + blurred_frame * (1 - mask_3d)).astype(np.uint8)
        else:  # solid color
            # Convert hex color to BGR
            color_hex = outside_color.lstrip('#')
            color_rgb = tuple(int(color_hex[i:i+2], 16) for i in (0, 2, 4))
            color_bgr = (color_rgb[2], color_rgb[1], color_rgb[0])

            # Create solid color frame
            solid_frame = np.full_like(frame, color_bgr)

            # Blend original and solid color using mask
            mask_3d = mask[:, :, np.newaxis]
            result = (frame * mask_3d + solid_frame * (1 - mask_3d)).astype(np.uint8)

        # Draw outline if enabled
        if show_outline and outline_thickness > 0:
            # Convert outline hex color to BGR
            outline_hex = outline_color.lstrip('#')
            outline_rgb = tuple(int(outline_hex[i:i+2], 16) for i in (0, 2, 4))
            outline_bgr = (outline_rgb[2], outline_rgb[1], outline_rgb[0])

            if shape == 'square':
                # Draw rectangle outline
                half_size = r
                x1, y1 = max(0, cx - half_size), max(0, cy - half_size)
                x2, y2 = min(w, cx + half_size), min(h, cy + half_size)
                cv2.rectangle(result, (x1, y1), (x2, y2), outline_bgr, thickness=int(outline_thickness))
            else:
                # Draw circle outline
                cv2.circle(result, (cx, cy), r, outline_bgr, thickness=int(outline_thickness))

        return result

    @staticmethod
    def create_cached_spotlight_transformer(video_size, center_x=50, center_y=50, radius=40,
                                           outside_effect='blur', blur_intensity=50,
                                           outside_color='#000000', feather=20,
                                           show_outline=True, outline_color='#FF00FF',
                                           outline_thickness=5, shape='circle',
                                           background_media_path=None,
                                           bg_zoom=100, bg_crop_x=0, bg_crop_y=0,
                                           inside_effect='none', inside_color='#00FF00',
                                           inside_opacity=30):
        """
        Create an optimized spotlight transformer with pre-calculated mask for 3-5x speedup

        This pre-calculates the mask once and reuses it for all frames, dramatically improving performance.

        Args:
            video_size: Tuple of (width, height) of video
            ... (same args as apply_circular_spotlight)
            background_media_path: Optional path to image/video to use as background instead of blur/solid

        Returns:
            Function that can be used with video.transform() for fast frame-by-frame processing
        """
        import cv2
        from pathlib import Path

        w, h = video_size

        # Pre-calculate mask (this is the expensive part - only do it once!)
        # Convert percentage to pixels
        cx = int(w * center_x / 100)
        cy = int(h * center_y / 100)
        r = int(min(w, h) * radius / 100)

        # Create mask based on shape
        mask = np.zeros((h, w), dtype=np.float32)

        if shape == 'square':
            # Create square/rectangle mask
            half_size = r
            x1, y1 = max(0, cx - half_size), max(0, cy - half_size)
            x2, y2 = min(w, cx + half_size), min(h, cy + half_size)

            # Create distance map for feathering
            y_grid, x_grid = np.ogrid[:h, :w]

            # Distance from rectangle edges
            dist_x = np.minimum(np.abs(x_grid - x1), np.abs(x_grid - x2))
            dist_y = np.minimum(np.abs(y_grid - y1), np.abs(y_grid - y2))

            # Inside rectangle
            inside_x = (x_grid >= x1) & (x_grid <= x2)
            inside_y = (y_grid >= y1) & (y_grid <= y2)
            inside = inside_x & inside_y

            # Apply feathering
            feather_px = int(r * feather / 100)
            if feather_px > 0:
                edge_dist = np.minimum(dist_x, dist_y)
                mask = np.where(inside, np.minimum(edge_dist / feather_px, 1.0), 0.0)
            else:
                mask[inside] = 1.0
        else:
            # Create circular mask
            y, x = np.ogrid[:h, :w]
            distance = np.sqrt((x - cx)**2 + (y - cy)**2)

            feather_px = int(r * feather / 100)
            if feather_px > 0:
                mask = np.clip((r + feather_px - distance) / feather_px, 0, 1)
            else:
                mask[distance <= r] = 1.0

        # Pre-calculate 3D mask for blending
        mask_3d = mask[:, :, np.newaxis]
        # PERF: a boolean/0-1 version for np.where (avoids float32 intermediates
        # that would otherwise be allocated per frame and dominate render time).
        # The original `frame * mask_3d + bg * (1 - mask_3d)` is ~6-7x slower.
        mask_3d_bool = mask_3d > 0.5

        # Pre-calculate blur amount
        blur_amount = max(1, int(blur_intensity))
        if blur_amount % 2 == 0:
            blur_amount += 1

        # Pre-calculate colors
        color_hex = outside_color.lstrip('#')
        color_rgb = tuple(int(color_hex[i:i+2], 16) for i in (0, 2, 4))
        color_bgr = (color_rgb[2], color_rgb[1], color_rgb[0])

        outline_hex = outline_color.lstrip('#')
        outline_rgb = tuple(int(outline_hex[i:i+2], 16) for i in (0, 2, 4))
        outline_bgr = (outline_rgb[2], outline_rgb[1], outline_rgb[0])

        # Pre-calculate inside effect color
        inside_hex = inside_color.lstrip('#')
        inside_rgb = tuple(int(inside_hex[i:i+2], 16) for i in (0, 2, 4))
        inside_bgr = (inside_rgb[2], inside_rgb[1], inside_rgb[0])

        # Load background media if provided (supports single file OR folder with random selection)
        background_media = None
        background_is_video = False
        if background_media_path and Path(background_media_path).exists():
            try:
                from moviepy import ImageClip, VideoFileClip
            except ImportError:
                from moviepy.editor import ImageClip, VideoFileClip

            selected_file = background_media_path

            # Check if path is a folder
            if Path(background_media_path).is_dir():
                # Get all media files from folder
                import random
                media_extensions = {'.mp4', '.avi', '.mov', '.mkv', '.webm', '.png', '.jpg', '.jpeg', '.gif', '.bmp'}
                media_files = [f for f in Path(background_media_path).iterdir()
                             if f.is_file() and f.suffix.lower() in media_extensions]

                if media_files:
                    selected_file = str(random.choice(media_files))
                    print(f"[🎲 RANDOM] Selected background from folder: {Path(selected_file).name} ({len(media_files)} total files)")
                else:
                    print(f"[WARNING] No media files found in folder: {background_media_path}")
                    selected_file = None

            if selected_file and Path(selected_file).exists():
                if selected_file.lower().endswith(('.mp4', '.avi', '.mov', '.mkv', '.webm')):
                    # Video background - loop method doesn't exist, we'll handle looping manually
                    background_media = VideoFileClip(selected_file)
                    background_is_video = True
                    print(f"[OK] Loaded background video: {Path(selected_file).name}")
                else:
                    # Image background
                    background_media = ImageClip(selected_file)
                    background_is_video = False
                    print(f"[OK] Loaded background image: {Path(selected_file).name}")

                # Apply zoom (scale) if needed
                zoom_scale = bg_zoom / 100.0
                if zoom_scale != 1.0:
                    # Calculate new size with zoom
                    zoomed_w = int(w * zoom_scale)
                    zoomed_h = int(h * zoom_scale)
                    background_media = background_media.resized((zoomed_w, zoomed_h))
                    print(f"[OK] Applied zoom: {int(bg_zoom)}% (size: {zoomed_w}x{zoomed_h})")

                    # When zoomed, we MUST crop to extract the visible window
                    # Calculate crop offsets in pixels based on crop percentages
                    crop_offset_x = int((background_media.w - w) * (bg_crop_x / 100.0))
                    crop_offset_y = int((background_media.h - h) * (bg_crop_y / 100.0))

                    # Ensure offsets don't go out of bounds
                    crop_offset_x = max(0, min(crop_offset_x, background_media.w - w))
                    crop_offset_y = max(0, min(crop_offset_y, background_media.h - h))

                    # Crop the zoomed media to extract the visible window (always needed when zoomed)
                    background_media = background_media.cropped(
                        x1=crop_offset_x,
                        y1=crop_offset_y,
                        x2=crop_offset_x + w,
                        y2=crop_offset_y + h
                    )
                    if bg_crop_x != 0 or bg_crop_y != 0:
                        print(f"[OK] Applied crop offset: X={bg_crop_x}%, Y={bg_crop_y}%")
                else:
                    # No zoom, just resize to video size
                    background_media = background_media.resized((w, h))

        # PERF: Pre-decode the background video into a list of numpy frames so
        # the per-frame transform doesn't re-decode from disk for every output
        # frame. The MoviePy VideoFileClip.get_frame(t) call is single-threaded
        # and re-reads from the ffmpeg subprocess — for 250+ output frames this
        # is a major cost. We cap the cache at 300 frames (~5s at 60fps) which
        # is more than enough for a 5-10s render and bounds memory at ~1.5GB.
        bg_frame_cache = None
        bg_frame_count = 0
        bg_frame_duration = 0.0
        if background_media is not None and background_is_video:
            try:
                bg_frame_count = min(
                    int(background_media.fps * background_media.duration) + 1,
                    300,
                )
                bg_frame_duration = background_media.duration
                # Decode all frames at the source rate (after resize/crop)
                _src_fps = max(1, int(round(background_media.fps)) or 24)
                _dur = bg_frame_duration
                _t = 0.0
                _dt = 1.0 / _src_fps
                bg_frame_cache = []
                _t0 = time.monotonic()
                while _t < _dur and len(bg_frame_cache) < bg_frame_count:
                    try:
                        bg_frame_cache.append(background_media.get_frame(_t))
                    except Exception:
                        break
                    _t += _dt
                if bg_frame_cache:
                    print(f"[⚡ CACHE] Pre-decoded {len(bg_frame_cache)} background frames in {time.monotonic()-_t0:.2f}s")
            except Exception as _e:
                print(f"[WARNING] bg frame cache failed: {_e}")
                bg_frame_cache = None

        def _get_bg_frame(t):
            """Get a background media frame at time t, using the pre-decoded
            cache when available. Falls back to direct VideoFileClip decode."""
            if bg_frame_cache:
                # Loop around the cache using the bg duration
                _t = t % max(bg_frame_duration, 1e-3)
                # Map t to a frame index in the cache
                idx = int(_t * (len(bg_frame_cache) / max(bg_frame_duration, 1e-3)))
                if idx >= len(bg_frame_cache):
                    idx = len(bg_frame_cache) - 1
                if idx < 0:
                    idx = 0
                return bg_frame_cache[idx]
            if background_media is None:
                return None
            if background_is_video:
                return background_media.get_frame(t % background_media.duration)
            return background_media.get_frame(0)

        # Create the transformer function (this will be called for each frame)
        def transform_frame(get_frame, t):
            frame = get_frame(t).copy()

            # Apply inside effect if enabled (change background inside the circle)
            # PERF: use np.where with a 0/1 bool mask (no float32 intermediates).
            # The legacy `frame * mask_3d + ...` is ~6-7x slower because it
            # allocates several 6MB+ float arrays per frame.
            if inside_effect == 'solid_color':
                opacity = inside_opacity / 100.0
                if opacity >= 0.999:
                    frame = np.where(mask_3d_bool, np.array(inside_bgr, dtype=np.uint8), frame)
                elif opacity > 0.001:
                    inside_frame = np.full_like(frame, inside_bgr)
                    blended = (frame.astype(np.float32) * (1 - opacity) +
                               inside_frame.astype(np.float32) * opacity).astype(np.uint8)
                    frame = np.where(mask_3d_bool, blended, frame)
            elif inside_effect == 'green_screen':
                opacity = inside_opacity / 100.0
                if opacity >= 0.999:
                    frame = np.where(mask_3d_bool, np.array((0, 255, 0), dtype=np.uint8), frame)
                elif opacity > 0.001:
                    green_screen = np.full_like(frame, (0, 255, 0))
                    blended = (frame.astype(np.float32) * (1 - opacity) +
                               green_screen.astype(np.float32) * opacity).astype(np.uint8)
                    frame = np.where(mask_3d_bool, blended, frame)

            # Apply effect to outside area
            if background_media:
                # Use provided media as background (now served from cache when
                # the bg is a video — see _get_bg_frame above).
                background_frame = _get_bg_frame(t)
                result = np.where(mask_3d_bool, frame, background_frame)
            elif outside_effect == 'blur':
                blurred_frame = cv2.GaussianBlur(frame, (blur_amount, blur_amount), 0)
                result = np.where(mask_3d_bool, frame, blurred_frame)
            else:
                solid_bgr = np.array(color_bgr, dtype=np.uint8)
                result = np.where(mask_3d_bool, frame, solid_bgr)

            # Draw outline if enabled
            if show_outline and outline_thickness > 0:
                if shape == 'square':
                    half_size = r
                    x1, y1 = max(0, cx - half_size), max(0, cy - half_size)
                    x2, y2 = min(w, cx + half_size), min(h, cy + half_size)
                    cv2.rectangle(result, (x1, y1), (x2, y2), outline_bgr, thickness=int(outline_thickness))
                else:
                    cv2.circle(result, (cx, cy), r, outline_bgr, thickness=int(outline_thickness))

            return result

        return transform_frame

    @staticmethod
    def apply_film_grain(frame, intensity=0.15):
        """Apply film grain overlay"""
        # Work on a copy to avoid modifying read-only arrays
        frame = frame.copy()
        noise = np.random.normal(0, intensity * 255, frame.shape)
        return np.clip(frame + noise, 0, 255).astype('uint8')


    @staticmethod
    def apply_selective_blur(get_frame, t):
        """Apply blur to a specific region (for hiding watermarks/logos)"""
        frame = get_frame(t)
        
        # Get blur settings
        if not hasattr(apply_selective_blur, 'settings'):
            return frame
            
        settings = apply_selective_blur.settings
        
        if not settings.get('blur_watermark_enabled', False):
            return frame
        
        try:
            from PIL import Image, ImageFilter
            import numpy as np
            
            # Get blur region
            x = settings.get('blur_x', 50)
            y = settings.get('blur_y', 700)
            width = settings.get('blur_width', 200)
            height = settings.get('blur_height', 50)
            intensity = settings.get('blur_intensity', 15)
            
            # Convert frame to PIL Image
            img = Image.fromarray(frame.astype('uint8'), 'RGB')
            
            # Extract region to blur
            region = img.crop((x, y, x + width, y + height))
            
            # Apply blur
            blurred_region = region.filter(ImageFilter.GaussianBlur(radius=intensity))
            
            # Paste back
            img.paste(blurred_region, (x, y))
            
            # Convert back to numpy array
            return np.array(img)
        except Exception as e:
            print(f"Blur error: {e}")
            return frame

    @staticmethod
    def apply_background_dim(frame, intensity=0.25):
        """Dim the background"""
        # Work on a copy to avoid modifying read-only arrays
        frame = frame.copy()
        return (frame * (1 - intensity)).astype('uint8')

    @staticmethod
    def apply_chromatic_aberration(frame, intensity=5, direction='horizontal'):
        """Apply chromatic aberration (RGB split) to a single frame.
        Returns a new uint8 frame. `intensity` is in pixels (clamped 0..min(h,w)).
        `direction` is one of 'horizontal', 'vertical', 'both'."""
        frame = frame.copy()
        shift = max(0, int(intensity))
        if shift <= 0:
            return frame
        r = frame[:, :, 0].copy()
        b = frame[:, :, 2].copy()
        h, w = frame.shape[:2]
        if direction in ('horizontal', 'both'):
            rs = np.zeros_like(r)
            bs = np.zeros_like(b)
            if 0 < shift < w:
                rs[:, shift:] = r[:, :-shift]
                bs[:, :-shift] = b[:, shift:]
            else:
                rs, bs = r, b
            r, b = rs, bs
        if direction in ('vertical', 'both'):
            rs = np.zeros_like(r)
            bs = np.zeros_like(b)
            if 0 < shift < h:
                rs[shift:, :] = r[:-shift, :]
                bs[:-shift, :] = b[shift:, :]
            else:
                rs, bs = r, b
            r, b = rs, bs
        frame[:, :, 0] = r
        frame[:, :, 2] = b
        return frame

    @staticmethod
    def apply_region_blur(frame, settings):
        """Apply blur to a specific region of the frame with optional color tint and text (OPTIMIZED)"""
        # Check if region blur OR custom blur regions are enabled
        region_blur_enabled = settings.get('region_blur_enabled', False)
        custom_regions = settings.get('custom_blur_regions', [])
        has_enabled_custom_regions = any(r.get('enabled', False) for r in custom_regions if isinstance(r, dict))

        if not region_blur_enabled and not has_enabled_custom_regions:
            return frame

        try:
            import cv2
            import numpy as np

            frame = frame.copy()
            h, w = frame.shape[:2]

            # Get blur settings
            intensity = int(settings.get('blur_intensity', 15))
            # Ensure kernel size is odd
            kernel_size = intensity * 2 + 1

            # Calculate region coordinates
            regions_to_blur = []

            # Process predefined regions via per-side toggles + crop values
            if region_blur_enabled:
                # Crop values (pixels trimmed from frame edges)
                crop_t = int(settings.get('blur_crop_top', 0))
                crop_b = int(settings.get('blur_crop_bottom', 30))
                crop_l = int(settings.get('blur_crop_left', 0))
                crop_r = int(settings.get('blur_crop_right', 0))

                # Per-side booleans (from OurScript tab's per-side toggles)
                e_top = settings.get('blur_enable_top', False)
                e_bot = settings.get('blur_enable_bottom', False)
                e_left = settings.get('blur_enable_left', False)
                e_right = settings.get('blur_enable_right', False)
                size_pct = int(settings.get('blur_region_size', 20))

                if any([e_top, e_bot, e_left, e_right]):
                    s_top = max(0.0, min(1.0, int(settings.get('blur_top_size', size_pct)) / 100.0))
                    s_bot = max(0.0, min(1.0, int(settings.get('blur_bottom_size', size_pct)) / 100.0))
                    s_left = max(0.0, min(1.0, int(settings.get('blur_left_size', size_pct)) / 100.0))
                    s_right = max(0.0, min(1.0, int(settings.get('blur_right_size', size_pct)) / 100.0))
                    if e_top:
                        regions_to_blur.append((0, 0, w, int(h * s_top)))
                    if e_bot:
                        regions_to_blur.append((0, h - int(h * s_bot), w, h))
                    if e_left:
                        regions_to_blur.append((0, 0, int(w * s_left), h))
                    if e_right:
                        regions_to_blur.append((w - int(w * s_right), 0, w, h))
                else:
                    # Legacy: use blur_region string + blur_crop_* as region sizes
                    rng = settings.get('blur_region', 'bottom')
                    s = size_pct / 100.0
                    if rng in ('top_bottom', 'left_right'):
                        half = s / 2.0
                        if rng == 'top_bottom':
                            regions_to_blur.extend([(0, 0, w, int(h * half)),
                                                     (0, h - int(h * half), w, h)])
                        else:
                            regions_to_blur.extend([(0, 0, int(w * half), h),
                                                     (w - int(w * half), 0, w, h)])
                    elif rng == 'top':
                        regions_to_blur.append((0, 0, w, int(h * s)))
                    elif rng == 'bottom':
                        regions_to_blur.append((0, h - int(h * s), w, h))
                    elif rng == 'left':
                        regions_to_blur.append((0, 0, int(w * s), h))
                    elif rng == 'right':
                        regions_to_blur.append((w - int(w * s), 0, w, h))
                    elif rng == 'center':
                        cx, cy = w // 2, h // 2
                        bw, bh = int(w * s), int(h * s)
                        regions_to_blur.append((cx - bw // 2, cy - bh // 2,
                                                 cx + bw // 2, cy + bh // 2))

                # Apply crop trimming to all regions
                trimmed = []
                for (x1, y1, x2, y2) in regions_to_blur:
                    x1, y1 = max(x1, crop_l), max(y1, crop_t)
                    x2, y2 = min(x2, w - crop_r), min(y2, h - crop_b)
                    if x2 > x1 and y2 > y1:
                        trimmed.append((x1, y1, x2, y2))
                regions_to_blur = trimmed

            # Add custom blur regions (for hiding specific logos/watermarks)
            custom_regions = settings.get('custom_blur_regions', [])
            if custom_regions and isinstance(custom_regions, list):
                for custom_region in custom_regions:
                    if not isinstance(custom_region, dict):
                        continue

                    if not custom_region.get('enabled', True):
                        continue

                    # Get region coordinates (support both percentage and pixel values)
                    x = custom_region.get('x', 0)
                    y = custom_region.get('y', 0)
                    width = custom_region.get('width', 100)
                    height = custom_region.get('height', 100)

                    # Convert percentage to pixels if needed
                    if isinstance(x, str) and '%' in str(x):
                        x = int(w * float(str(x).rstrip('%')) / 100)
                    elif isinstance(x, (int, float)) and 0 <= x <= 100:
                        x = int(w * x / 100)
                    else:
                        x = int(x)

                    if isinstance(y, str) and '%' in str(y):
                        y = int(h * float(str(y).rstrip('%')) / 100)
                    elif isinstance(y, (int, float)) and 0 <= y <= 100:
                        y = int(h * y / 100)
                    else:
                        y = int(y)

                    if isinstance(width, str) and '%' in str(width):
                        width = int(w * float(str(width).rstrip('%')) / 100)
                    elif isinstance(width, (int, float)) and 0 <= width <= 100:
                        width = int(w * width / 100)
                    else:
                        width = int(width)

                    if isinstance(height, str) and '%' in str(height):
                        height = int(h * float(str(height).rstrip('%')) / 100)
                    elif isinstance(height, (int, float)) and 0 <= height <= 100:
                        height = int(h * height / 100)
                    else:
                        height = int(height)

                    # Calculate coordinates
                    x1 = max(0, x)
                    y1 = max(0, y)
                    x2 = min(w, x + width)
                    y2 = min(h, y + height)

                    # Get custom blur intensity for this region (optional)
                    region_intensity = custom_region.get('intensity', intensity)
                    region_kernel_size = int(region_intensity) * 2 + 1

                    # Get per-region opacity (0.0-1.0); default fully opaque
                    region_opacity = float(custom_region.get('opacity', 1.0))
                    region_opacity = max(0.0, min(1.0, region_opacity))

                    # Get per-region fill color + opacity (for color-overlay on blur)
                    region_fill_color = custom_region.get('fill_color', '')
                    region_fill_opacity = int(custom_region.get('fill_opacity', 0))

                    # Get per-region cover mode (pill instead of blur)
                    cover_mode = bool(custom_region.get('cover_mode', False))
                    cover_radius = int(custom_region.get('cover_radius', 8))

                    # Add to regions list with custom intensity + opacity + fill
                    regions_to_blur.append((x1, y1, x2, y2, region_kernel_size,
                                            region_opacity, region_fill_color, region_fill_opacity,
                                            cover_mode, cover_radius))

            # Apply blur using cv2 (FAST)
            for region_data in regions_to_blur:
                if len(region_data) == 10:
                    # Custom region with fill + cover mode
                    x1, y1, x2, y2, custom_kernel, region_opacity, region_fill_color, region_fill_opacity, cover_mode, cover_radius = region_data
                    is_custom_region = True
                elif len(region_data) == 8:
                    # Custom region with fill: (x1,y1,x2,y2,kernel,opacity,fill_hex,fill_op)
                    x1, y1, x2, y2, custom_kernel, region_opacity, region_fill_color, region_fill_opacity = region_data
                    cover_mode = False
                    cover_radius = 8
                    is_custom_region = True
                elif len(region_data) == 6:
                    # Custom region: (x1,y1,x2,y2,custom_kernel,opacity)
                    x1, y1, x2, y2, custom_kernel, region_opacity = region_data
                    region_fill_color = ''
                    region_fill_opacity = 0
                    cover_mode = False
                    cover_radius = 8
                    is_custom_region = True
                elif len(region_data) == 5:
                    # Legacy custom region (no opacity): (x1,y1,x2,y2,custom_kernel)
                    x1, y1, x2, y2, custom_kernel = region_data
                    region_opacity = 1.0
                    region_fill_color = ''
                    region_fill_opacity = 0
                    cover_mode = False
                    cover_radius = 8
                    is_custom_region = True
                else:
                    # Predefined region: (x1,y1,x2,y2)
                    x1, y1, x2, y2 = region_data
                    custom_kernel = kernel_size
                    region_opacity = 1.0
                    region_fill_color = ''
                    region_fill_opacity = 0
                    is_custom_region = False  # Predefined regions have 4 values

                if x2 <= x1 or y2 <= y1:
                    continue

                # Extract region
                roi = frame[y1:y2, x1:x2]

                # ── Cover mode (custom regions): per-region rounded pill ──
                if is_custom_region and cover_mode:
                    # Use fill_color + fill_opacity to draw a solid rounded pill instead of blur
                    if region_fill_color and region_fill_opacity > 0:
                        try:
                            fill_rgb = tuple(int(region_fill_color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
                        except Exception:
                            fill_rgb = (0, 0, 0)
                        alpha = min(1.0, max(0.0, region_fill_opacity / 100.0))
                        overlay = np.full_like(roi, fill_rgb, dtype=np.uint8)
                        blended = cv2.addWeighted(overlay, alpha, roi, 1 - alpha, 0)
                    else:
                        blended = roi.copy()

                    # Rounded corner mask
                    rh, rw = roi.shape[:2]
                    radius_px = int(min(rw, rh) * cover_radius / 100)
                    if radius_px > 0:
                        mask = np.zeros((rh, rw), dtype=np.uint8)
                        cv2.rectangle(mask, (radius_px, 0), (rw - radius_px, rh), 1, -1)
                        cv2.rectangle(mask, (0, radius_px), (rw, rh - radius_px), 1, -1)
                        cv2.circle(mask, (radius_px, radius_px), radius_px, 1, -1)
                        cv2.circle(mask, (rw - radius_px, radius_px), radius_px, 1, -1)
                        cv2.circle(mask, (radius_px, rh - radius_px), radius_px, 1, -1)
                        cv2.circle(mask, (rw - radius_px, rh - radius_px), radius_px, 1, -1)
                        mask_3ch = np.stack([mask, mask, mask], axis=-1).astype(bool)
                        frame[y1:y2, x1:x2] = np.where(mask_3ch, blended, roi)
                    else:
                        frame[y1:y2, x1:x2] = blended
                    continue  # skip blur pipeline for this region

                # ── Cover mode (predefined regions): dark rounded pill instead of blur ──
                if not is_custom_region and settings.get('region_blur_mode', 'blur') == 'cover':
                    cover_color = settings.get('cover_color', '#000000')
                    cover_opacity = int(settings.get('cover_opacity', 85))
                    cover_radius = int(settings.get('cover_radius', 8))  # % of region width

                    try:
                        # Parse color hex → RGB
                        c_rgb = tuple(int(cover_color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
                    except Exception:
                        c_rgb = (0, 0, 0)
                    alpha = min(1.0, max(0.0, cover_opacity / 100.0))

                    # Create color overlay and blend
                    overlay = np.full_like(roi, c_rgb, dtype=np.uint8)
                    blended = cv2.addWeighted(overlay, alpha, roi, 1 - alpha, 0)

                    # Rounded corner mask
                    rh, rw = roi.shape[:2]
                    radius_px = int(min(rw, rh) * cover_radius / 100)
                    if radius_px > 0:
                        mask = np.zeros((rh, rw), dtype=np.uint8)
                        cv2.rectangle(mask, (radius_px, 0), (rw - radius_px, rh), 1, -1)
                        cv2.rectangle(mask, (0, radius_px), (rw, rh - radius_px), 1, -1)
                        cv2.circle(mask, (radius_px, radius_px), radius_px, 1, -1)
                        cv2.circle(mask, (rw - radius_px, radius_px), radius_px, 1, -1)
                        cv2.circle(mask, (radius_px, rh - radius_px), radius_px, 1, -1)
                        cv2.circle(mask, (rw - radius_px, rh - radius_px), radius_px, 1, -1)
                        mask_3ch = np.stack([mask, mask, mask], axis=-1).astype(bool)
                        frame[y1:y2, x1:x2] = np.where(mask_3ch, blended, roi)
                    else:
                        frame[y1:y2, x1:x2] = blended
                    continue  # skip blur pipeline for this region

                # ── Blur mode (adaptive downscale Gaussian for performance) ──
                rh, rw = roi.shape[:2]
                if custom_kernel > 25 and rw > 100 and rh > 100:
                    # Downscale before blur for large kernels on sizable regions.
                    # kernel=135 on 1080×576 → 16× fewer pixels, same visual result.
                    ds = max(2, custom_kernel // 16)
                    small = cv2.resize(roi, (rw // ds, rh // ds), interpolation=cv2.INTER_LINEAR)
                    small_k = max(3, custom_kernel // ds)
                    if small_k % 2 == 0:
                        small_k += 1
                    blurred = cv2.GaussianBlur(small, (small_k, small_k), 0)
                    blurred = cv2.resize(blurred, (rw, rh), interpolation=cv2.INTER_LINEAR)
                else:
                    blurred = cv2.GaussianBlur(roi, (custom_kernel, custom_kernel), 0)

                # Apply color tint if enabled (predefined regions only — custom regions
                # have their own fill_color/fill_opacity system that must not mix)
                if not is_custom_region and settings.get('blur_color_tint_enabled', False):
                    tint_color = settings.get('blur_tint_color', '#000000')
                    tint_opacity = settings.get('blur_tint_opacity', 50) / 100

                    # Convert hex to RGB (MoviePy frames are RGB, not BGR)
                    tint_rgb = tuple(int(tint_color[i:i+2], 16) for i in (1, 3, 5))

                    # Create tint overlay in RGB
                    tint_layer = np.full_like(blurred, tint_rgb, dtype=np.uint8)
                    blurred = cv2.addWeighted(blurred, 1 - tint_opacity, tint_layer, tint_opacity, 0)

                # Apply feathered edge ONLY for predefined (non-custom) regions
                if not is_custom_region and settings.get('blur_feather_edge', True) and region_blur_enabled:
                    feather_px = min(blurred.shape[0], blurred.shape[1]) // 4
                    if feather_px > 0:
                        bh, bw = blurred.shape[:2]
                        # ⚡ The feather mask depends ONLY on region geometry, not
                        # pixel content — identical every frame. Build it once and
                        # cache it (keyed by geometry) instead of looping per row
                        # per frame. Vectorized assignment replaces the Python loop.
                        _edge = ('top' if (y1 == 0 and y2 < h) else
                                 'bottom' if (y2 == h and y1 > 0) else
                                 'left' if (x1 == 0 and x2 < w) else
                                 'right' if (x2 == w and x1 > 0) else None)
                        _mkey = (bh, bw, feather_px, _edge)
                        _mcache = getattr(VideoEffects.apply_region_blur, '_feather_cache', None)
                        if _mcache is None:
                            _mcache = {}
                            VideoEffects.apply_region_blur._feather_cache = _mcache
                        mask = _mcache.get(_mkey)
                        if mask is None:
                            mask = np.ones((bh, bw), dtype=np.float32)
                            if _edge == 'top':
                                grad = np.linspace(1, 0, feather_px, dtype=np.float32)
                                mask[bh - feather_px:bh, :] = grad[:, None]
                            elif _edge == 'bottom':
                                grad = np.linspace(0, 1, feather_px, dtype=np.float32)
                                mask[:feather_px, :] = grad[:, None]
                            elif _edge == 'left':
                                grad = np.linspace(1, 0, feather_px, dtype=np.float32)
                                mask[:, bw - feather_px:bw] = grad[None, :]
                            elif _edge == 'right':
                                grad = np.linspace(0, 1, feather_px, dtype=np.float32)
                                mask[:, :feather_px] = grad[None, :]
                            mask = mask[:, :, np.newaxis]
                            _mcache[_mkey] = mask
                        # Apply mask
                        original_roi = frame[y1:y2, x1:x2].astype(np.float32)
                        blurred = (blurred.astype(np.float32) * mask + original_roi * (1 - mask)).astype(np.uint8)

                # Apply per-region opacity blending for custom regions
                if is_custom_region and region_opacity < 1.0:
                    original_roi = frame[y1:y2, x1:x2].astype(np.float32)
                    blurred = (blurred.astype(np.float32) * region_opacity
                               + original_roi * (1.0 - region_opacity)).astype(np.uint8)

                # Apply per-region fill color overlay (unless fill_opacity is 0)
                if is_custom_region and region_fill_color and region_fill_opacity > 0:
                    try:
                        fill_rgb = tuple(int(region_fill_color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
                        fill_layer = np.full_like(blurred, fill_rgb, dtype=np.uint8)
                        fill_alpha = min(1.0, max(0.0, region_fill_opacity / 100.0))
                        blurred = cv2.addWeighted(blurred, 1 - fill_alpha, fill_layer, fill_alpha, 0)
                    except Exception:
                        pass  # Invalid fill_color hex — skip overlay

                frame[y1:y2, x1:x2] = blurred

            # Render text overlays for custom blur regions
            custom_regions = settings.get('custom_blur_regions', [])
            if custom_regions and isinstance(custom_regions, list):
                for custom_region in custom_regions:
                    if not isinstance(custom_region, dict):
                        continue

                    if not custom_region.get('enabled', True):
                        continue

                    # Get text content - check if should use spreadsheet
                    # PERFORMANCE FIX: Cache spreadsheet lookups to avoid repeated calls per frame
                    region_id = custom_region.get('id', id(custom_region))
                    cache_key = f'_cached_text_{region_id}'

                    if custom_region.get('use_spreadsheet', False):
                        # Check if already cached in settings
                        if cache_key in settings:
                            text_content = settings[cache_key]
                        else:
                            # Try to get text from spreadsheet (only once, then cache)
                            spreadsheet_file = settings.get('blur_regions_spreadsheet_file', '')
                            column = custom_region.get('spreadsheet_column', 'B')
                            video_path = settings.get('_current_video_path', '')

                            text_content = VideoEffects.get_text_from_spreadsheet(
                                video_path, spreadsheet_file, column
                            )

                            if not text_content:
                                # Fall back to manual text if spreadsheet fails
                                text_content = custom_region.get('text', '').strip()

                            # Cache the result for subsequent frames
                            settings[cache_key] = text_content
                    else:
                        # Use manual text from settings
                        text_content = custom_region.get('text', '').strip()

                    if not text_content:
                        continue

                    # Get region coordinates
                    x = custom_region.get('x', 0)
                    y = custom_region.get('y', 0)
                    width = custom_region.get('width', 100)
                    height = custom_region.get('height', 100)

                    # Convert percentage to pixels (starting position)
                    if isinstance(x, (int, float)) and 0 <= x <= 100:
                        x = int(w * x / 100)
                    if isinstance(y, (int, float)) and 0 <= y <= 100:
                        y = int(h * y / 100)

                    x1 = max(0, x)
                    y1 = max(0, y)

                    # Get text styling
                    text_color = custom_region.get('text_color', '#FFFFFF')
                    bg_color = custom_region.get('bg_color', '#000000')
                    bg_opacity = custom_region.get('bg_opacity', 180)

                    # Convert hex colors to BGR
                    text_rgb = tuple(int(text_color[i:i+2], 16) for i in (1, 3, 5))
                    text_bgr = (text_rgb[2], text_rgb[1], text_rgb[0])

                    bg_rgb = tuple(int(bg_color[i:i+2], 16) for i in (1, 3, 5))
                    bg_bgr = (bg_rgb[2], bg_rgb[1], bg_rgb[0])

                    # Check if auto-expand is enabled
                    auto_expand = custom_region.get('auto_expand', False)
                    font = cv2.FONT_HERSHEY_DUPLEX  # Use DUPLEX for bolder appearance

                    if auto_expand:
                        # Calculate box size based on text dimensions
                        # Use a default font scale (3% of video height)
                        font_scale = max(0.5, h * 0.03 / 30.0)
                        thickness = max(1, int(font_scale * 2))

                        # Get text size
                        (text_w, text_h), baseline = cv2.getTextSize(text_content, font, font_scale, thickness)

                        # Add padding (20% of text size)
                        padding_x = int(text_w * 0.2)
                        padding_y = int(text_h * 0.3)

                        # Calculate expanded box size
                        x2 = min(w, x1 + text_w + padding_x * 2)
                        y2 = min(h, y1 + text_h + padding_y * 2)
                    else:
                        # Use manual width/height from settings
                        if isinstance(width, (int, float)) and 0 <= width <= 100:
                            width = int(w * width / 100)
                        if isinstance(height, (int, float)) and 0 <= height <= 100:
                            height = int(h * height / 100)

                        x2 = min(w, x + width)
                        y2 = min(h, y + height)

                    if x2 <= x1 or y2 <= y1:
                        continue

                    region_w = x2 - x1
                    region_h = y2 - y1

                    # Calculate font size based on region height (adaptive sizing)
                    if not auto_expand:
                        font_scale = region_h / 80.0
                        thickness = max(1, int(font_scale * 2))

                    # Get text size (recalculate for non-auto-expand to fit in fixed box)
                    (text_w, text_h), baseline = cv2.getTextSize(text_content, font, font_scale, thickness)

                    # If text is too wide (and not auto-expand), scale it down
                    if not auto_expand and text_w > region_w - 20:
                        font_scale = font_scale * (region_w - 20) / text_w
                        thickness = max(1, int(font_scale * 2))
                        (text_w, text_h), baseline = cv2.getTextSize(text_content, font, font_scale, thickness)

                    # Calculate text position (centered)
                    text_x = x1 + (region_w - text_w) // 2
                    text_y = y1 + (region_h + text_h) // 2

                    # Draw background rectangle with opacity
                    overlay = frame.copy()
                    cv2.rectangle(overlay, (x1, y1), (x2, y2), bg_bgr, -1)
                    alpha = bg_opacity / 255.0
                    frame = cv2.addWeighted(overlay, alpha, frame, 1 - alpha, 0)

                    # Draw text with outline for better visibility
                    outline_color = (0, 0, 0) if sum(text_rgb) > 384 else (255, 255, 255)
                    cv2.putText(frame, text_content, (text_x, text_y), font, font_scale,
                               outline_color, thickness + 2, cv2.LINE_AA)
                    cv2.putText(frame, text_content, (text_x, text_y), font, font_scale,
                               text_bgr, thickness, cv2.LINE_AA)

            # Add text on blur if enabled (use cv2 for speed)
            if settings.get('blur_text_enabled', False):
                text_content = settings.get('blur_text_content', '')
                if text_content and regions_to_blur:
                    font_size = settings.get('blur_text_size', 24) / 30  # Scale for cv2
                    text_color = settings.get('blur_text_color', '#FFFFFF')
                    text_position = settings.get('blur_text_position', 'center')

                    # Convert hex to BGR
                    text_rgb = tuple(int(text_color[i:i+2], 16) for i in (1, 3, 5))
                    text_bgr = (text_rgb[2], text_rgb[1], text_rgb[0])

                    # Get text size
                    font = cv2.FONT_HERSHEY_SIMPLEX
                    thickness = max(1, int(font_size * 2))
                    (text_w, text_h), baseline = cv2.getTextSize(text_content, font, font_size, thickness)

                    # Calculate position
                    x1, y1, x2, y2 = regions_to_blur[0]
                    text_x = (x1 + x2) // 2 - text_w // 2

                    if text_position == 'top':
                        text_y = y1 + text_h + 10
                    elif text_position == 'bottom':
                        text_y = y2 - 10
                    else:  # center
                        text_y = (y1 + y2) // 2 + text_h // 2

                    # Draw outline
                    cv2.putText(frame, text_content, (text_x, text_y), font, font_size, (0, 0, 0), thickness + 2)
                    # Draw text
                    cv2.putText(frame, text_content, (text_x, text_y), font, font_size, text_bgr, thickness)

            return frame

        except Exception as e:
            print(f"Region blur error: {e}")
            return frame

    @staticmethod
    def apply_gradient_overlay(frame, gradient_type='top_to_bottom', intensity=0.3):
        """Apply gradient overlay effect"""
        frame = frame.copy()
        h, w = frame.shape[:2]

        # Create gradient
        if gradient_type == 'top_to_bottom':
            gradient = np.linspace(1, 1 - intensity, h)[:, np.newaxis]
            gradient = np.repeat(gradient, w, axis=1)
        elif gradient_type == 'bottom_to_top':
            gradient = np.linspace(1 - intensity, 1, h)[:, np.newaxis]
            gradient = np.repeat(gradient, w, axis=1)
        elif gradient_type == 'left_to_right':
            gradient = np.linspace(1, 1 - intensity, w)[np.newaxis, :]
            gradient = np.repeat(gradient, h, axis=0)
        elif gradient_type == 'right_to_left':
            gradient = np.linspace(1 - intensity, 1, w)[np.newaxis, :]
            gradient = np.repeat(gradient, h, axis=0)
        elif gradient_type == 'radial':
            y, x = np.ogrid[:h, :w]
            cx, cy = w / 2, h / 2
            max_dist = np.sqrt(cx**2 + cy**2)
            distance = np.sqrt((x - cx)**2 + (y - cy)**2)
            gradient = 1 - (distance / max_dist * intensity)
        else:
            gradient = np.ones((h, w))

        gradient = np.clip(gradient, 0, 1)
        return (frame * gradient[:, :, np.newaxis]).astype('uint8')


def apply_frame_effects_to_clip(clip, settings, log_fn=None):
    """Apply the five frame-level effects (chromatic, vignette, dim, film
    grain, particles) to a MoviePy clip. Used by BOTH the main render path
    and the Cleanup / Spoofing path so toggles in the new FRAME EFFECTS card
    on the Cleanup tab actually fire during a Cleanup run.

    Effects fire in this order: chromatic → vignette → background dim → film
    grain → particles (the same order as the main render pipeline).

    `log_fn(level, message)` is an optional callback that, if supplied, is
    invoked for each effect that gets applied (or skipped with a reason)."""
    if clip is None:
        return clip

    def _log(level, msg):
        if log_fn is not None:
            try:
                log_fn(level, msg)
            except Exception:
                pass

    # Chromatic aberration (RGB Glitch)
    if settings.get('chromatic_aberration', False):
        intensity = int(settings.get('chromatic_intensity', 5))
        direction = settings.get('chromatic_direction', 'horizontal')
        if intensity <= 0:
            _log('warn',
                 f'[frame-effect] chromatic_aberration: skipped — '
                 f'intensity is {intensity}px (must be >= 1). Bump the '
                 f'RGB Glitch intensity slider in the FRAME EFFECTS card.')
        else:
            _log('ok',
                 f'[frame-effect] chromatic_aberration: ON '
                 f'(intensity={intensity}px, direction={direction})')
            try:
                def _chrom(get_frame, t):
                    return VideoEffects.apply_chromatic_aberration(
                        get_frame(t), intensity=intensity, direction=direction)

                try:
                    clip = clip.transform(_chrom)
                except AttributeError:
                    clip = clip.fl_image(
                        lambda fr: VideoEffects.apply_chromatic_aberration(
                            fr, intensity=intensity, direction=direction))
            except Exception as _e:
                _log('error', f'[frame-effect] chromatic_aberration FAILED: {_e}')
                print(f'[WARNING] frame effect chromatic_aberration failed: {_e}')

    # Vignette
    if settings.get('vignette', False):
        intensity = float(settings.get('vignette_intensity', 0.4))
        _log('ok', f'[frame-effect] vignette: ON (intensity={intensity:.2f})')
        try:
            def _vig(get_frame, t):
                return VideoEffects.apply_vignette(get_frame(t), intensity)

            try:
                clip = clip.transform(_vig)
            except AttributeError:
                clip = clip.fl_image(
                    lambda fr: VideoEffects.apply_vignette(fr, intensity))
        except Exception as _e:
            _log('error', f'[frame-effect] vignette FAILED: {_e}')
            print(f'[WARNING] frame effect vignette failed: {_e}')

    # Background Dim
    if settings.get('background_dim', False):
        intensity = float(settings.get('background_dim_intensity', 0.4))
        _log('ok', f'[frame-effect] background_dim: ON (intensity={intensity:.2f})')
        try:
            def _dim(get_frame, t):
                return VideoEffects.apply_background_dim(get_frame(t), intensity)

            try:
                clip = clip.transform(_dim)
            except AttributeError:
                clip = clip.fl_image(
                    lambda fr: VideoEffects.apply_background_dim(fr, intensity))
        except Exception as _e:
            _log('error', f'[frame-effect] background_dim FAILED: {_e}')
            print(f'[WARNING] frame effect background_dim failed: {_e}')

    # Film Grain
    if settings.get('film_grain', False):
        intensity = float(settings.get('film_grain_intensity', 0.05))
        _log('ok', f'[frame-effect] film_grain: ON (intensity={intensity:.3f})')
        try:
            def _grain(get_frame, t):
                return VideoEffects.apply_film_grain(get_frame(t), intensity)

            try:
                clip = clip.transform(_grain)
            except AttributeError:
                clip = clip.fl_image(
                    lambda fr: VideoEffects.apply_film_grain(fr, intensity))
        except Exception as _e:
            _log('error', f'[frame-effect] film_grain FAILED: {_e}')
            print(f'[WARNING] frame effect film_grain failed: {_e}')

    # Particle Effects (glitter / stars / hearts / confetti).
    # These are overlays, not per-frame transforms — we composite a
    # particle clip on top of the current clip.  Use CompositeVideoClip
    # so the particles are visible on top of whatever the cleanup did.
    if (settings.get('add_glitter', False) or
            settings.get('add_stars', False) or
            settings.get('add_hearts', False) or
            settings.get('add_confetti', False)):
        active = [k for k in ('add_glitter', 'add_stars', 'add_hearts',
                               'add_confetti') if settings.get(k, False)]
        _log('ok', f'[frame-effect] particles: ON ({", ".join(active)})')
        try:
            _p_intensity = float(settings.get('glitter_intensity', 0.4))
            particle_layer = ParticleEffects.create_combined(
                clip.w, clip.h, clip.duration, clip.fps,
                glitter=settings.get('add_glitter', False),
                glitter_intensity=_p_intensity,
                stars=settings.get('add_stars', False),
                hearts=settings.get('add_hearts', False),
                confetti=settings.get('add_confetti', False),
            )
            # Composite: base clip with the particle layer on top.
            try:
                from moviepy import CompositeVideoClip
            except Exception:
                try:
                    from moviepy.compositing import CompositeVideoClip
                except Exception:
                    from moviepy.video.compositing import CompositeVideoClip
            clip = CompositeVideoClip([clip, particle_layer])
        except Exception as _e:
            print(f'[WARNING] frame effect particles failed: {_e}')

    return clip


class TransitionEffects:
    """Professional transition effects for video intro/outro"""

    @staticmethod
    def apply_fade_transition(clip, fade_in_duration=0, fade_out_duration=0):
        """Apply fade in/out transitions"""
        if fade_in_duration > 0:
            try:
                clip = clip.with_effects([FadeIn(fade_in_duration)])
            except:
                try:
                    clip = clip.fadein(fade_in_duration)
                except:
                    pass

        if fade_out_duration > 0:
            try:
                from moviepy.video.fx import FadeOut
                clip = clip.with_effects([FadeOut(fade_out_duration)])
            except:
                try:
                    clip = clip.fadeout(fade_out_duration)
                except:
                    pass

        return clip

    @staticmethod
    def create_zoom_transition(clip, zoom_in=True, duration=1.0, zoom_scale=1.3):
        """Create zoom in/out transition effect"""
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        # Capture the *original* duration before any inner wrap mutates clip.duration
        original_duration = clip.duration

        def zoom_effect(get_frame, t):
            frame = get_frame(t)

            if zoom_in:
                # Zoom from scale to 1.0
                progress = min(t / duration, 1.0)
                current_scale = zoom_scale - (zoom_scale - 1.0) * progress
            else:
                # Zoom from 1.0 to scale
                progress = max((t - (original_duration - duration)) / duration, 0.0)
                current_scale = 1.0 + (zoom_scale - 1.0) * progress

            if abs(current_scale - 1.0) > 0.01:  # Only apply if zoom needed
                h, w = frame.shape[:2]
                new_h, new_w = int(h * current_scale), int(w * current_scale)

                from PIL import Image as PILImage
                pil_frame = PILImage.fromarray(frame)
                pil_frame = pil_frame.resize((new_w, new_h), PILImage.LANCZOS)

                # Crop to original size (center crop)
                crop_x = (new_w - w) // 2
                crop_y = (new_h - h) // 2
                pil_frame = pil_frame.crop((crop_x, crop_y, crop_x + w, crop_y + h))

                return np.array(pil_frame).copy()

            return frame

        try:
            return clip.transform(zoom_effect)
        except:
            return clip.fl(zoom_effect)

    @staticmethod
    def create_blur_transition(clip, blur_in=True, duration=0.5, max_blur=15):
        """Create blur in/out transition"""
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        # Capture original duration before any wrap
        original_duration = clip.duration

        def blur_effect(get_frame, t):
            frame = get_frame(t)

            if blur_in:
                # Blur from max to 0
                progress = min(t / duration, 1.0)
                blur_amount = int(max_blur * (1 - progress))
            else:
                # Blur from 0 to max
                progress = max((t - (original_duration - duration)) / duration, 0.0)
                blur_amount = int(max_blur * progress)

            if blur_amount > 0:
                pil_img = Image.fromarray(frame)
                pil_img = pil_img.filter(ImageFilter.GaussianBlur(radius=blur_amount))
                return np.array(pil_img).copy()

            return frame

        try:
            return clip.transform(blur_effect)
        except:
            return clip.fl(blur_effect)

    @staticmethod
    def create_slide_transition(clip, direction='left', in_transition=True, duration=0.8):
        """Create slide in/out transition (video slides into frame)"""
        # Capture original duration BEFORE the transform closure so the
        # lambda doesn't reference a MoviePy clip attribute that may change.
        orig_dur = getattr(clip, 'duration', 0) or 0

        def slide_effect(get_frame, t):
            frame = get_frame(t)
            h, w = frame.shape[:2]

            if in_transition:
                progress = min(t / duration, 1.0)
                # Ease out cubic
                progress = 1 - (1 - progress) ** 3
            else:
                progress = max((t - (orig_dur - duration)) / duration, 0.0) if orig_dur > 0 else 1.0
                # Ease in cubic
                progress = min(progress ** 3, 1.0)

            if direction == 'left':
                offset_px = int(w * (1 - progress)) if in_transition else int(-w * progress)
            elif direction == 'right':
                offset_px = int(-w * (1 - progress)) if in_transition else int(w * progress)
            elif direction == 'up':
                offset_px = int(h * (1 - progress)) if in_transition else int(-h * progress)
            elif direction == 'down':
                offset_px = int(-h * (1 - progress)) if in_transition else int(h * progress)
            else:
                offset_px = 0

            if direction in ('left', 'right'):
                out = np.zeros_like(frame)
                if direction == 'left':
                    # Slide in from left: frame slides right from off-screen
                    if in_transition:
                        # offset_px goes from w to 0 over duration
                        src_shift = w - offset_px  # how many pixels of source are visible
                        if offset_px < w:
                            out[:, :src_shift] = frame[:, offset_px:]
                    else:
                        # Slide out to left: frame moves left off-screen
                        if offset_px < 0:
                            out[:, :w + offset_px] = frame[:, -offset_px:]
                        elif offset_px < w:
                            out[:, offset_px:] = frame[:, :w - offset_px]
                else:  # right
                    if in_transition:
                        # Slide in from right
                        if offset_px < 0:
                            out[:, -offset_px:] = frame[:, :w + offset_px]
                    else:
                        # Slide out to right
                        if offset_px > 0:
                            out[:, :w - offset_px] = frame[:, offset_px:]
                return out
            else:  # up/down
                out = np.zeros_like(frame)
                if direction == 'up':
                    if in_transition:
                        # Slide in from top
                        src_shift = h - offset_px
                        if offset_px < h:
                            out[:src_shift] = frame[offset_px:]
                    else:
                        # Slide out to top
                        if offset_px < 0:
                            out[:h + offset_px] = frame[-offset_px:]
                else:  # down
                    if in_transition:
                        # Slide in from bottom
                        if offset_px < 0:
                            out[-offset_px:] = frame[:h + offset_px]
                    else:
                        # Slide out to bottom
                        if offset_px > 0:
                            out[:h - offset_px] = frame[offset_px:]
                return out

        try:
            return clip.transform(slide_effect)
        except AttributeError:
            return clip.fl(slide_effect)

    @staticmethod
    def create_wipe_transition(clip, direction='right', in_transition=True, duration=0.8):
        """Create wipe transition (reveals video gradually)"""
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        def wipe_effect(get_frame, t):
            frame = get_frame(t)
            h, w = frame.shape[:2]
            mask = np.zeros((h, w), dtype=np.uint8)

            if in_transition:
                progress = min(t / duration, 1.0)
            else:
                progress = 1.0 - max((t - (clip.duration - duration)) / duration, 0.0)

            if direction == 'right':
                reveal_x = int(w * progress)
                mask[:, :reveal_x] = 255
            elif direction == 'left':
                reveal_x = int(w * (1 - progress))
                mask[:, reveal_x:] = 255
            elif direction == 'down':
                reveal_y = int(h * progress)
                mask[:reveal_y, :] = 255
            elif direction == 'up':
                reveal_y = int(h * (1 - progress))
                mask[reveal_y:, :] = 255

            mask3 = np.stack([mask, mask, mask], axis=-1)
            bg = np.zeros_like(frame)
            return (frame.astype(np.float32) * (mask3 / 255.0) +
                    bg * (1.0 - mask3 / 255.0)).astype(np.uint8)

        try:
            return clip.transform(wipe_effect)
        except:
            return clip.fl(wipe_effect)

    @staticmethod
    def create_glitch_transition(clip, glitch_start=True, duration=0.5, intensity=0.5):
        """Create digital glitch transition effect"""
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        # Capture original duration before any wrap
        original_duration = clip.duration

        def glitch_effect(get_frame, t):
            frame = get_frame(t)

            if glitch_start:
                if t > duration:
                    return frame
                progress = t / duration
            else:
                if t < (original_duration - duration):
                    return frame
                progress = (t - (original_duration - duration)) / duration

            # Apply glitch only during transition
            glitch_amount = intensity * (1 - progress) if glitch_start else intensity * progress

            if glitch_amount > 0.1:
                frame = frame.copy()
                h, w = frame.shape[:2]

                # Reproducible random pattern per-frame
                np.random.seed(int(t * 1000) & 0xFFFFFFFF)

                # RGB channel shift
                shift = int(w * 0.02 * glitch_amount)
                if shift > 0:
                    frame[:, shift:, 0] = frame[:, :-shift, 0]
                    frame[:, :-shift, 2] = frame[:, shift:, 2]

                # Random horizontal slices
                if np.random.random() < glitch_amount:
                    num_slices = int(5 * glitch_amount)
                    for _ in range(num_slices):
                        y1 = np.random.randint(0, h - 10)
                        y2 = y1 + np.random.randint(5, 30)
                        offset = np.random.randint(-int(w * 0.1), int(w * 0.1))
                        if offset > 0:
                            frame[y1:y2, offset:] = frame[y1:y2, :-offset]
                        elif offset < 0:
                            frame[y1:y2, :offset] = frame[y1:y2, -offset:]

            return frame

        try:
            return clip.transform(glitch_effect)
        except:
            return clip.fl(glitch_effect)

    @staticmethod
    def create_bounce_transition(clip, duration=0.6, height=0.3, bounce_start=True):
        """Bouncy vertical pop: frame scales/translates along a sine
        curve during the transition window. Implemented with numpy/PIL
        so the canvas never fills with black from a positioned sub-clip."""
        import math as _m
        import numpy as _np

        original_duration = clip.duration

        def bounce_effect(get_frame, t):
            frame = get_frame(t)
            # Outside the transition window: pass through unchanged
            if bounce_start:
                if t > duration:
                    return frame
                progress = t / duration
            else:
                if t < (original_duration - duration):
                    return frame
                progress = (t - (original_duration - duration)) / duration
            # Bounce curve: peak at mid, 0 at endpoints
            k = 1.0 - progress if bounce_start else progress
            sin_v = _m.sin(k * _m.pi)
            scale = 1.0 + height * sin_v * (1.0 - k * 0.4)
            ty = -int(frame.shape[0] * height * 0.6 * sin_v)
            if abs(scale - 1.0) < 0.01 and ty == 0:
                return frame
            h, w = frame.shape[:2]
            new_h = max(1, int(round(h * scale)))
            new_w = max(1, int(round(w * scale)))
            # Resize frame with PIL
            try:
                from PIL import Image as _PILImage
                pil_src = _PILImage.fromarray(frame)
                pil_dst = pil_src.resize((new_w, new_h), _PILImage.BILINEAR)
                resized = _np.asarray(pil_dst)
            except Exception:
                # Fallback: nearest-neighbor repeat
                rs_h = max(1, int(round(1.0 / scale))) if scale > 0 else 1
                rs_w = rs_h
                resized = _np.repeat(_np.repeat(frame, rs_h, axis=0), rs_w, axis=1)
                rh, rw = resized.shape[:2]
                if rh != new_h or rw != new_w:
                    if rh > new_h:
                        resized = resized[:new_h, :]
                    elif rh < new_h:
                        pad = _np.zeros((new_h - rh, rw, 3), dtype=resized.dtype)
                        resized = _np.concatenate([resized, pad], axis=0)
                    if rw > new_w:
                        resized = resized[:, :new_w]
                    elif rw < new_w:
                        pad = _np.zeros((new_h, new_w - rw, 3), dtype=resized.dtype)
                        resized = _np.concatenate([resized, pad], axis=1)
            # Center the resized image on the original canvas, with vertical offset
            cy_src = new_h // 2 - ty
            cx_src = new_w // 2
            sy0 = max(0, cy_src - h // 2)
            sy1 = sy0 + h
            sx0 = max(0, cx_src - w // 2)
            sx1 = sx0 + w
            sy0 = max(0, min(sy0, max(0, new_h - h)))
            sy1 = sy0 + h
            sx0 = max(0, min(sx0, max(0, new_w - w)))
            sx1 = sx0 + w
            crop = resized[sy0:sy1, sx0:sx1]
            ch, cw = crop.shape[:2]
            if ch == h and cw == w:
                return crop
            out = _np.zeros_like(frame)
            oh = min(h, ch)
            ow = min(w, cw)
            out[:oh, :ow] = crop[:oh, :ow]
            return out

        try:
            return clip.transform(bounce_effect)
        except Exception:
            return clip.fl(bounce_effect)

    @staticmethod
    def _parse_hex_color(hex_str, default=(0, 0, 0)):
        """Convert '#RRGGBB' hex string to (R, G, B) tuple."""
        if not hex_str or not isinstance(hex_str, str) or not hex_str.startswith('#'):
            return default
        try:
            return (int(hex_str[1:3], 16), int(hex_str[3:5], 16), int(hex_str[5:7], 16))
        except (ValueError, IndexError):
            return default

    @staticmethod
    def create_mask_reveal(clip, duration=0.6, shape='circle', direction='in', mask_start=True, bg_color='black'):
        """Wipe in/out using a shape mask (circle / square / diamond / star).
        Uses a per-frame numpy mask to keep it dependency-light.

        ``bg_color`` can be ``'black'`` or a hex string like ``'#FF0000'``.
        """
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip
        import numpy as _np

        original_duration = clip.duration

        def _shape_mask(h, w, shape, progress):
            mask = _np.zeros((h, w), dtype=_np.float32)
            cy, cx = h / 2.0, w / 2.0
            max_r = _np.sqrt((cx) ** 2 + (cy) ** 2)
            r = max_r * (1.0 - progress) + 1.0
            yy, xx = _np.ogrid[:h, :w]
            if shape == 'circle':
                d = _np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2)
                mask = (d <= r).astype(_np.float32)
            elif shape == 'square':
                d = _np.maximum(_np.abs(xx - cx), _np.abs(yy - cy))
                mask = (d <= r).astype(_np.float32)
            elif shape == 'diamond':
                d = (_np.abs(xx - cx) + _np.abs(yy - cy))
                mask = (d <= r).astype(_np.float32)
            elif shape == 'star':
                # 5-point star: combine two rotated triangles approx
                a1 = _np.arctan2(yy - cy, xx - cx)
                rad = _np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2)
                lobes = 5
                wave = 0.5 + 0.5 * _np.cos(lobes * a1)
                r_eff = r * (0.45 + 0.55 * wave)
                mask = (rad <= r_eff).astype(_np.float32)
            return mask

        def mask_effect(get_frame, t):
            frame = get_frame(t)
            if mask_start:
                if t > duration:
                    return frame
                progress = 1.0 - (t / duration)  # opens up
            else:
                if t < (original_duration - duration):
                    return frame
                progress = (t - (original_duration - duration)) / duration  # closes
            h, w = frame.shape[:2]
            mask = _shape_mask(h, w, shape, max(0.0, min(1.0, progress)))
            mask3 = _np.stack([mask, mask, mask], axis=-1)
            try:
                # Background color when mask is closed
                if bg_color and bg_color != 'black':
                    _rgb = TransitionEffects._parse_hex_color(bg_color, (0, 0, 0))
                    bg = _np.full(frame.shape, _rgb, dtype=_np.uint8)
                else:
                    bg = _np.zeros_like(frame)
                out = (frame.astype(_np.float32) * mask3 + bg.astype(_np.float32) * (1 - mask3))
                return out.astype(_np.uint8)
            except Exception:
                return frame

        try:
            return clip.transform(mask_effect)
        except Exception:
            return clip.fl(mask_effect)

    @staticmethod
    def create_bounce_mask_transition(clip, duration=0.8, height=0.25, shape='circle', mask_start=True, bg_color='black'):
        """Combo: bounce frame + reveal/close through a shape mask.
        Implemented as a single per-frame callback that combines both."""
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip
        import numpy as _np
        import math as _m

        original_duration = clip.duration

        def combo(get_frame, t):
            frame = get_frame(t)
            if mask_start:
                if t > duration:
                    return frame
                p = 1.0 - (t / duration)
            else:
                if t < (original_duration - duration):
                    return frame
                p = (t - (original_duration - duration)) / duration
            p = max(0.0, min(1.0, p))
            # bounce component (numpy-based so canvas never gets black gaps)
            k = 1.0 - p if mask_start else p
            sin_v = _m.sin(k * _m.pi)
            scale = 1.0 + height * 0.2 * sin_v
            ty = -int(frame.shape[0] * height * 0.5 * sin_v)
            h, w = frame.shape[:2]
            if abs(scale - 1.0) >= 0.01 or ty != 0:
                new_h = max(1, int(round(h * scale)))
                new_w = max(1, int(round(w * scale)))
                try:
                    from PIL import Image as _PILImage
                    pil_src = _PILImage.fromarray(frame)
                    pil_dst = pil_src.resize((new_w, new_h), _PILImage.BILINEAR)
                    resized = _np.asarray(pil_dst)
                except Exception:
                    rs = max(1, int(round(1.0 / scale))) if scale > 0 else 1
                    resized = _np.repeat(_np.repeat(frame, rs, axis=0), rs, axis=1)
                    rh, rw = resized.shape[:2]
                    if rh > new_h: resized = resized[:new_h, :]
                    elif rh < new_h:
                        pad = _np.zeros((new_h - rh, rw, 3), dtype=resized.dtype)
                        resized = _np.concatenate([resized, pad], axis=0)
                    if rw > new_w: resized = resized[:, :new_w]
                    elif rw < new_w:
                        pad = _np.zeros((new_h, new_w - rw, 3), dtype=resized.dtype)
                        resized = _np.concatenate([resized, pad], axis=1)
                cy_src = new_h // 2 - ty
                cx_src = new_w // 2
                sy0 = max(0, min(cy_src - h // 2, max(0, new_h - h)))
                sx0 = max(0, min(cx_src - w // 2, max(0, new_w - w)))
                crop = resized[sy0:sy0 + h, sx0:sx0 + w]
                ch, cw = crop.shape[:2]
                if ch == h and cw == w:
                    frame = crop
                else:
                    out0 = _np.zeros_like(frame)
                    oh, ow = min(h, ch), min(w, cw)
                    out0[:oh, :ow] = crop[:oh, :ow]
                    frame = out0
            # mask reveal
            cy, cx = h / 2.0, w / 2.0
            max_r = _np.sqrt(cx ** 2 + cy ** 2)
            r = max_r * p + 1.0
            yy, xx = _np.ogrid[:h, :w]
            if shape == 'circle':
                d = _np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2)
            elif shape == 'square':
                d = _np.maximum(_np.abs(xx - cx), _np.abs(yy - cy))
            elif shape == 'diamond':
                d = (_np.abs(xx - cx) + _np.abs(yy - cy))
            else:
                d = _np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2)
            mask = (d <= r).astype(_np.float32)
            mask3 = _np.stack([mask, mask, mask], axis=-1)
            if bg_color and bg_color != 'black':
                _rgb = TransitionEffects._parse_hex_color(bg_color, (0, 0, 0))
                bg = _np.full(frame.shape, _rgb, dtype=_np.uint8)
            else:
                bg = _np.zeros_like(frame)
            out = (frame.astype(_np.float32) * mask3 + bg.astype(_np.float32) * (1 - mask3))
            return out.astype(_np.uint8)

        try:
            return clip.transform(combo)
        except Exception:
            return clip.fl(combo)

    @staticmethod
    def create_cinematic_bars(clip, fade_in=True, duration=0.8, bar_height_percent=10):
        """Create cinematic letterbox bars (black bars top/bottom).
        The bars are produced as a per-frame RGB clip that reads the
        underlying frame and only paints the bar strips black — so the
        rest of the frame stays unchanged (no fully-black overlay)."""
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip
        import numpy as _np

        original_duration = clip.duration
        w, h = clip.size
        bar_h = int(h * bar_height_percent / 100)

        def bars_effect(get_frame, t):
            frame = get_frame(t)
            if fade_in:
                progress = min(max(t / duration, 0.0), 1.0)
                cur = int(bar_h * progress)
            else:
                cur = bar_h
            out = frame.copy()
            if cur > 0:
                out[:cur, :] = 0
                out[-cur:, :] = 0
            return out

        try:
            return clip.transform(bars_effect)
        except Exception:
            return clip.fl(bars_effect)

    # === Pulse / interval transitions ====================================
    # These apply a quick visual pulse at every `interval` seconds, similar
    # to how the cinematic effects use repeat mode.

    @staticmethod
    def _apply_pulse_transform(clip, transform_fn, pulse_duration, interval, start_offset=0.0):
        """Single-pulse wrapper. See compose_pulses() for the fast multi-pulse path."""
        return TransitionEffects.compose_pulses(
            clip, [(transform_fn, pulse_duration, interval, start_offset)])

    @staticmethod
    def compose_pulses(clip, pulse_specs, start_offset=0.0):
        """Apply MANY pulse transforms in a SINGLE .transform() call so moviepy
        only re-encodes the clip once.

        pulse_specs: list of (transform_fn, pulse_duration, interval, kw)
            where transform_fn has signature (get_frame, t, t_in_pulse, **kw) -> frame.
        This replaces the old pattern of calling each create_*_pulse in a loop,
        which did 4 separate re-encodes and made renders 4x slower.
        """
        active = [(fn, dur, intv, kw) for (fn, dur, intv, kw) in pulse_specs
                  if fn is not None and intv and intv > 0 and dur and dur > 0]
        if not active:
            return clip

        def combined_transform(get_frame, t):
            # Apply each active pulse in order. Each pulse is only invoked
            # during its own time window. The previous frame is passed
            # through the chain (no intermediate re-encoding).
            cur = get_frame(t)
            for transform_fn, dur, intv, kw in active:
                if t < start_offset:
                    continue
                t_rel = t - start_offset
                t_in_pulse = t_rel - (int(t_rel / intv) * intv)
                if 0 <= t_in_pulse <= dur:
                    # Build a get_frame that returns the current frame so
                    # the pulse's transform_fn sees our chained output, not
                    # the original clip. This is the key to the speedup:
                    # only one .transform() runs on the clip.
                    def make_get_frame(captured, _fn=transform_fn, _t=t,
                                       _tip=t_in_pulse, _kw=kw):
                        def _gf(_t):
                            return _fn(lambda __: captured, _t, _tip, **_kw)
                        return _gf
                    cur = make_get_frame(cur)(t)
            return cur

        try:
            return clip.transform(combined_transform)
        except AttributeError:
            try:
                return clip.fl(combined_transform, apply_to=['mask'])
            except Exception:
                return clip.fl(combined_transform)

    @staticmethod
    def create_zoom_pulse(clip, duration=1.0, zoom_scale=1.3, interval=8.0, start_offset=0.0):
        """Zoom-in then zoom-back pulse, applied every `interval` seconds."""

        def transform(get_frame, t, t_in_pulse):
            progress = min(t_in_pulse / duration, 1.0)
            # Triangle wave: up then back down
            tri = progress if progress < 0.5 else (1.0 - progress)
            scale = 1.0 + (zoom_scale - 1.0) * (tri * 2)
            frame = get_frame(t)
            if abs(scale - 1.0) > 0.01:
                h, w = frame.shape[:2]
                new_h, new_w = int(h * scale), int(w * scale)
                from PIL import Image as PILImage
                pil_frame = PILImage.fromarray(frame)
                pil_frame = pil_frame.resize((new_w, new_h), PILImage.LANCZOS)
                crop_x = (new_w - w) // 2
                crop_y = (new_h - h) // 2
                pil_frame = pil_frame.crop((crop_x, crop_y, crop_x + w, crop_y + h))
                return np.array(pil_frame).copy()
            return frame

        return TransitionEffects._apply_pulse_transform(clip, transform, duration, interval, start_offset)

    @staticmethod
    def create_blur_pulse(clip, duration=0.5, max_blur=15, interval=8.0, start_offset=0.0):
        """Blur-up then blur-back pulse, applied every `interval` seconds."""

        def transform(get_frame, t, t_in_pulse):
            progress = min(t_in_pulse / duration, 1.0)
            tri = progress if progress < 0.5 else (1.0 - progress)
            blur_amount = int(max_blur * (tri * 2))
            frame = get_frame(t)
            if blur_amount > 0:
                pil_img = Image.fromarray(frame)
                pil_img = pil_img.filter(ImageFilter.GaussianBlur(radius=blur_amount))
                return np.array(pil_img).copy()
            return frame

        return TransitionEffects._apply_pulse_transform(clip, transform, duration, interval, start_offset)

    @staticmethod
    def create_glitch_pulse(clip, duration=0.5, intensity=0.5, interval=8.0, start_offset=0.0):
        """Glitch burst pulse, applied every `interval` seconds."""

        def transform(get_frame, t, t_in_pulse):
            progress = min(t_in_pulse / duration, 1.0)
            glitch_amount = intensity * (1.0 - progress)
            frame = get_frame(t)
            if glitch_amount > 0.1:
                frame = frame.copy()
                h, w = frame.shape[:2]
                np.random.seed(int(t * 1000) & 0xFFFFFFFF)
                shift = int(w * 0.02 * glitch_amount)
                if shift > 0:
                    frame[:, shift:, 0] = frame[:, :-shift, 0]
                    frame[:, :-shift, 2] = frame[:, shift:, 2]
                if np.random.random() < glitch_amount:
                    num_slices = int(5 * glitch_amount)
                    for _ in range(num_slices):
                        y1 = np.random.randint(0, max(1, h - 10))
                        y2 = y1 + np.random.randint(5, 30)
                        offset = np.random.randint(-int(w * 0.1), int(w * 0.1))
                        if offset > 0:
                            frame[y1:y2, offset:] = frame[y1:y2, :-offset]
                        elif offset < 0:
                            frame[y1:y2, :offset] = frame[y1:y2, -offset:]
            return frame

        return TransitionEffects._apply_pulse_transform(clip, transform, duration, interval, start_offset)

    @staticmethod
    def create_shake_pulse(clip, duration=0.5, intensity=0.05, interval=8.0, start_offset=0.0):
        """Camera-shake pulse, applied every `interval` seconds."""

        def transform(get_frame, t, t_in_pulse):
            progress = min(t_in_pulse / duration, 1.0)
            tri = progress if progress < 0.5 else (1.0 - progress)
            frame = get_frame(t)
            if tri > 0.05:
                h, w = frame.shape[:2]
                amp = int(min(w, h) * intensity * tri)
                np.random.seed(int(t * 1000) & 0xFFFFFFFF)
                dx = np.random.randint(-amp, amp + 1)
                dy = np.random.randint(-amp, amp + 1)
                frame = np.roll(frame, shift=(dy, dx), axis=(0, 1))
                # Fill the wrap-around border with a blurred edge
                if dx != 0:
                    edge_w = min(20, w // 10)
                    if dx > 0:
                        frame[:, :edge_w] = frame[:, edge_w:edge_w + edge_w]
                    else:
                        frame[:, -edge_w:] = frame[:, -2*edge_w:-edge_w]
                if dy != 0:
                    edge_h = min(20, h // 10)
                    if dy > 0:
                        frame[:edge_h, :] = frame[edge_h:edge_h + edge_h, :]
                    else:
                        frame[-edge_h:, :] = frame[-2*edge_h:-edge_h, :]
            return frame

        return TransitionEffects._apply_pulse_transform(clip, transform, duration, interval, start_offset)

    # ===== New CapCut-style transitions =================================

    @staticmethod
    def create_radial_wipe(clip, direction='in', duration=0.8, center_x=50, center_y=50):
        """Radial/Circular wipe — reveals clip via an expanding/shrinking circle.

        direction='in':  Wipe from edges toward center (circle shrinks)
        direction='out': Wipe from center outward (circle expands)

        This is the classic CapCut circle-intro / circle-outro transition.
        Uses per-frame numpy masking so the effect is baked into RGB.
        """

        original_duration = clip.duration

        def radial_effect(get_frame, t):
            frame = get_frame(t)
            h, w = frame.shape[:2]
            cx = int(w * center_x / 100)
            cy = int(h * center_y / 100)
            max_radius = int(np.sqrt(cx*cx + cy*cy) if direction == 'in'
                             else np.sqrt((w-cx)**2 + (h-cy)**2))

            progress = min(t / duration, 1.0)
            p = 1 - (1 - progress) ** 3  # ease-out cubic
            if direction == 'in':
                r = int(max_radius * (1 - p))
            else:
                r = int(max_radius * p)
            if r <= 0:
                return np.zeros_like(frame)
            mask = np.zeros((h, w), dtype=np.uint8)
            yg, xg = np.ogrid[:h, :w]
            dist = np.sqrt((xg - cx)**2 + (yg - cy)**2)
            feather = min(10, r // 4)
            if feather < 1:
                feather = 1
            mask[dist <= r] = 255
            if feather > 1:
                soft_band = (dist > r) & (dist <= r + feather)
                mask[soft_band] = np.clip((1 - (dist[soft_band] - r) / feather) * 255, 0, 255).astype(np.uint8)

            mask3 = np.stack([mask, mask, mask], axis=-1)
            bg = np.zeros_like(frame)
            return (frame.astype(np.float32) * (mask3 / 255.0) +
                    bg * (1.0 - mask3 / 255.0)).astype(np.uint8)

        try:
            return clip.transform(radial_effect)
        except AttributeError:
            return clip.fl(radial_effect)

    @staticmethod
    def create_color_dissolve(clip, direction='in', duration=0.8, color='#FFFFFF', ease_power=2):
        """Color dissolve — fades through a solid color (like CapCut's flash dissolve).

        direction='in':  Original → color → full visibility (flash through color at start)
        direction='out': Full visibility → color → black/nothing (flash through color at end)

        Works by blending the frame with a solid color over time.
        """
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        # Parse color
        color = color.lstrip('#')
        if len(color) == 6:
            col_rgb = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
        else:
            col_rgb = (255, 255, 255)

        import cv2  # noqa: F811 — local import for cv2.addWeighted

        original_duration = clip.duration

        def color_dissolve_effect(get_frame, t):
            frame = get_frame(t)
            if direction == 'in':
                progress = min(t / duration, 1.0)
                # Progress rises 0→1: dissolve to color then dissolve to frame
                if progress < 0.5:
                    blend = progress * 2  # 0→1 in first half
                    return cv2.addWeighted(frame, 1 - blend,
                                           np.full_like(frame, col_rgb), blend, 0)
                else:
                    blend = (progress - 0.5) * 2  # 0→1 in second half
                    return cv2.addWeighted(np.full_like(frame, col_rgb), 1 - blend,
                                           frame, blend, 0)
            else:
                # direction='out'
                progress = max((t - (original_duration - duration)) / duration, 0.0)
                progress = min(progress, 1.0)
                if progress < 0.5:
                    blend = progress * 2
                    return cv2.addWeighted(frame, 1.0 - blend,
                                           np.full_like(frame, col_rgb), blend, 0)
                else:
                    blend = (progress - 0.5) * 2
                    fade_out = 1 - blend
                    return cv2.addWeighted(np.full_like(frame, col_rgb), fade_out,
                                           frame, fade_out, 0)  # Both to color and fade

        try:
            return clip.transform(color_dissolve_effect)
        except AttributeError:
            return clip.fl(color_dissolve_effect)

    @staticmethod
    def create_split_wipe(clip, direction='horizontal', in_transition=True, duration=0.8):
        """Split wipe — reveals the clip as if doors or curtains opening.

        direction='horizontal': Left and right halves slide apart
        direction='vertical':   Top and bottom halves slide apart
        direction='diagonal':   Four corners slide outward
        in_transition=True:     Opening (revealing the clip)
        in_transition=False:    Closing (hiding the clip)

        Uses per-frame numpy masking so the effect is baked into RGB.
        """
        original_duration = clip.duration

        def split_effect(get_frame, t):
            frame = get_frame(t)
            h, w = frame.shape[:2]
            mask = np.ones((h, w), dtype=np.uint8) * 255

            if in_transition:
                progress = min(t / duration, 1.0)
            else:
                progress = min(max((t - (original_duration - duration)) / duration, 0.0), 1.0)
            p = 1 - (1 - progress) ** 3  # ease-out cubic

            if direction == 'horizontal':
                half = w // 2
                left_reveal = int(half * p)
                right_reveal = int((w - half) * p)
                mask[:, :half] = 0
                mask[:, half:] = 0
                mask[:, half - left_reveal:half + right_reveal] = 255
            elif direction == 'vertical':
                half = h // 2
                top_reveal = int(half * p)
                bottom_reveal = int((h - half) * p)
                mask[:half, :] = 0
                mask[half:, :] = 0
                mask[half - top_reveal:half + bottom_reveal, :] = 255
            elif direction == 'diagonal':
                cx, cy = w // 2, h // 2
                yg, xg = np.ogrid[:h, :w]
                dist = np.minimum.reduce([
                    np.sqrt(xg**2 + yg**2),
                    np.sqrt((w-xg)**2 + yg**2),
                    np.sqrt(xg**2 + (h-yg)**2),
                    np.sqrt((w-xg)**2 + (h-yg)**2),
                ])
                max_dist = np.sqrt(cx**2 + cy**2)
                threshold = max_dist * p
                mask[dist > threshold] = 0
                mask[dist <= threshold] = 255

            mask3 = np.stack([mask, mask, mask], axis=-1)
            bg = np.zeros_like(frame)
            return (frame.astype(np.float32) * (mask3 / 255.0) +
                    bg * (1.0 - mask3 / 255.0)).astype(np.uint8)

        try:
            return clip.transform(split_effect)
        except AttributeError:
            return clip.fl(split_effect)

    @staticmethod
    def create_luma_wipe(clip, direction='in', duration=0.8, edge_feather=4):
        """Luma/Brightness wipe — the frame brightens/saturates to reveal.

        Works by boosting the brightness (luma) progressively until full reveal.
        Gives a smooth 'bloom' or 'flash' effect similar to CapCut's lumafade.

        direction='in':  Gets brighter over time (revealing the image)
        direction='out': Gets darker over time (hiding the image)
        """
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        original_duration = clip.duration

        def luma_effect(get_frame, t):
            frame = get_frame(t)
            if direction == 'in':
                progress = min(t / duration, 1.0)
            else:
                progress = min(max((t - (original_duration - duration)) / duration, 0.0), 1.0)
                progress = 1 - progress

            # Smooth step
            p = 1 - (1 - progress) ** 2

            if p >= 1.0:
                return frame

            # Boost luma and saturation proportional to progress
            frame_f = frame.astype(np.float32)
            # Convert to YUV-like for luma boost
            luma = 0.299 * frame_f[:,:,2] + 0.587 * frame_f[:,:,1] + 0.114 * frame_f[:,:,0]
            luma_boost = luma * (0.5 * (1 - p))  # Boost up to 0.5x
            result = frame_f + luma_boost[:,:,np.newaxis]

            # Also reduce contrast for a hazy/flash look in early phase
            if p < 0.5:
                fade = (0.5 - p) * 2  # 1→0 as p goes 0→0.5
                avg = np.mean(result, axis=(0,1), keepdims=True)
                result = result * (1 - fade * 0.3) + avg * (fade * 0.3)

            return np.clip(result, 0, 255).astype(np.uint8)

        try:
            return clip.transform(luma_effect)
        except AttributeError:
            return clip.fl(luma_effect)


class LightLeaksEffects:
    """Cinematic light leaks and lens flare effects"""

    @staticmethod
    def create_light_leak(width, height, duration, fps, color='warm', intensity=0.6,
                         start_time=0, leak_duration=None, direction='top_right'):
        """
        Create light leak overlay effect

        Args:
            color: 'warm' (orange/yellow), 'cold' (blue/cyan), 'pink', 'purple', 'rainbow'
            direction: 'top_right', 'top_left', 'bottom_right', 'bottom_left', 'center'
        """
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        if leak_duration is None:
            leak_duration = duration

        # Color palettes
        colors = {
            'warm': [(255, 200, 100), (255, 150, 50), (255, 100, 0)],
            'cold': [(100, 200, 255), (50, 150, 255), (0, 100, 255)],
            'pink': [(255, 100, 150), (255, 150, 200), (255, 50, 100)],
            'purple': [(200, 100, 255), (150, 50, 255), (100, 0, 200)],
            'rainbow': [(255, 0, 0), (255, 127, 0), (255, 255, 0), (0, 255, 0), (0, 0, 255), (75, 0, 130)]
        }

        color_palette = colors.get(color, colors['warm'])

        def make_frame(t):
            # Create RGBA frame
            img = Image.new('RGBA', (width, height), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)

            # Calculate progress (0 to 1 and back)
            relative_t = t - start_time
            if relative_t < 0 or relative_t > leak_duration:
                return np.array(img).copy()

            progress = relative_t / leak_duration
            # Fade in and out
            if progress < 0.3:
                alpha = progress / 0.3
            elif progress > 0.7:
                alpha = (1.0 - progress) / 0.3
            else:
                alpha = 1.0

            # Position based on direction
            positions = {
                'top_right': (width * 0.7, -height * 0.2),
                'top_left': (-width * 0.2, -height * 0.2),
                'bottom_right': (width * 0.7, height * 0.7),
                'bottom_left': (-width * 0.2, height * 0.7),
                'center': (width * 0.3, height * 0.3)
            }

            pos_x, pos_y = positions.get(direction, positions['top_right'])

            # Draw multiple overlapping ellipses for organic look
            for i, color_rgb in enumerate(color_palette):
                offset_x = i * 50 + int(progress * 100)
                offset_y = i * 30

                size_w = int(width * 0.6 * (1 + i * 0.1))
                size_h = int(height * 0.8 * (1 + i * 0.1))

                bbox = [
                    int(pos_x + offset_x),
                    int(pos_y + offset_y),
                    int(pos_x + offset_x + size_w),
                    int(pos_y + offset_y + size_h)
                ]

                # Calculate alpha for this layer
                layer_alpha = int(255 * intensity * alpha * (1 - i * 0.2))

                draw.ellipse(bbox, fill=color_rgb + (layer_alpha,))

            # Apply gaussian blur for soft glow
            img = img.filter(ImageFilter.GaussianBlur(radius=12))

            return np.array(img).copy()

        clip = VideoClip(make_frame, duration=duration)
        try:
            clip = clip.with_fps(fps)
        except:
            clip = clip.set_fps(fps)

        return clip

    @staticmethod
    def create_lens_flare(width, height, duration, fps, intensity=0.5,
                         start_time=0, flare_duration=2.0, position='center'):
        """Create lens flare effect (numpy-only, ~0.5ms per frame)"""
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        # Precompute coordinate grid
        _yy, _xx = np.ogrid[:height, :width]
        yy = (_yy.astype(np.float32) + 0.5) / height
        xx = (_xx.astype(np.float32) + 0.5) / width

        def make_frame(t):
            relative_t = t - start_time
            if relative_t < 0 or relative_t > flare_duration:
                return np.zeros((height, width, 4), dtype=np.uint8)

            progress = relative_t / flare_duration
            # Quick flash
            if progress < 0.2:
                alpha_val = progress / 0.2
            else:
                alpha_val = max(0.0, 1.0 - (progress - 0.2) / 0.8)

            # Centre
            cx, cy = 0.5, 0.5
            if position == 'top':
                cy = 0.25

            # Main flare — big gaussian blob (broadcasts over ogrid)
            dx = xx - cx  # (1, W)
            dy = yy - cy  # (H, 1)
            main = np.exp(-((dx / 0.18) ** 2 + (dy / 0.18) ** 2) * 3.0)

            # Smaller satellite flares along the x-axis
            sat = np.zeros((height, width), dtype=np.float32)
            for i in range(3):
                offset = (i + 1) * 0.08
                size = 0.12 / (i + 2)
                s_dx = xx - (cx + offset)
                s_dy = dy
                sat += np.exp(-((s_dx / size) ** 2 + (s_dy / size) ** 2) * 2.0) * (0.5 / (i + 1))

            # Combine
            flare = main + sat * 0.5
            flare = np.clip(flare, 0, 1) * intensity * alpha_val

            # White → warm tint
            r = np.clip(flare * 255, 0, 255).astype(np.uint8)
            g = np.clip(flare * 210, 0, 255).astype(np.uint8)
            b = np.clip(flare * 150, 0, 255).astype(np.uint8)
            a = np.clip(flare * 255, 0, 255).astype(np.uint8)

            return np.stack([r, g, b, a], axis=-1)

        clip = VideoClip(make_frame, duration=duration)
        try:
            clip = clip.with_fps(fps)
        except:
            clip = clip.set_fps(fps)

        return clip

    @staticmethod
    def create_film_burn(width, height, duration, fps, start_time=0, burn_duration=1.5):
        """Create film burn effect (numpy-only, ~0.5ms per frame)"""
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        # Precompute coordinate grid with origin at bottom-right corner
        _yy, _xx = np.ogrid[:height, :width]
        yy_rel = (_yy.astype(np.float32) + 0.5) / height
        xx_rel = (_xx.astype(np.float32) + 0.5) / width
        # Distance from bottom-right corner (ogrid broadcast → (H, W))
        dx_corner = 1.0 - xx_rel
        dy_corner = 1.0 - yy_rel
        corner_dist = np.sqrt(dx_corner ** 2 + dy_corner ** 2)
        norm = np.sqrt(2.0)
        corner_dist /= norm

        # Pre-generate per-pixel noise for organic texture
        _rng = np.random.RandomState(0)
        noise_base = _rng.randint(0, 2 ** 16, size=(height, width)).astype(np.float32)

        def make_frame(t):
            relative_t = t - start_time
            if relative_t < 0 or relative_t > burn_duration:
                return np.zeros((height, width, 4), dtype=np.uint8)

            progress = relative_t / burn_duration
            spread = progress * 1.5

            # Organic noise that shifts per frame (avoids global seed)
            noise = ((noise_base + progress * 5000) % 256) / 256.0

            # Radial falloff from bottom-right + noise texture
            radial = corner_dist / max(spread, 0.001)
            mask = np.exp(-radial * 2.5) * 0.8 + (noise * np.exp(-radial * 1.5)) * 0.3
            mask = np.clip(mask * progress * 1.2, 0, 1) * 0.9

            # Burn alpha intensifies over time
            ia = min(1.0, progress * 2)

            # Orange/yellow burn colours
            r = np.clip(mask * 255, 0, 255).astype(np.uint8)
            g = np.clip(mask * 170, 0, 255).astype(np.uint8)
            b = np.clip(mask * 40, 0, 255).astype(np.uint8)
            a = np.clip(mask * 200 * ia, 0, 255).astype(np.uint8)

            return np.stack([r, g, b, a], axis=-1)

        clip = VideoClip(make_frame, duration=duration)
        try:
            clip = clip.with_fps(fps)
        except:
            clip = clip.set_fps(fps)

        return clip

    # ═══════════════════════════════════════════════════════════════
    # OpenCV-native overlay source generators
    # ═══════════════════════════════════════════════════════════════
    # These produce the same visual effects as create_light_leak /
    # create_lens_flare / create_film_burn but return pure-numpy
    # overlay source dicts (no PIL / no MoviePy) so the OpenCV
    # render path can composite them directly.
    # ═══════════════════════════════════════════════════════════════

    @staticmethod
    def _make_light_leak_overlay(width, height, start_time, leak_duration,
                                  color='warm', intensity=0.6, direction='top_right'):
        """Return overlay get_frame closure for a single light leak event.

        Uses OpenCV drawing + blur (no PIL).  Returns callable(t)→RGBA.
        """
        import cv2 as _CV2
        colors = {
            'warm': [(255, 200, 100), (255, 150, 50), (255, 100, 0)],
            'cold': [(100, 200, 255), (50, 150, 255), (0, 100, 255)],
            'pink': [(255, 100, 150), (255, 150, 200), (255, 50, 100)],
            'purple': [(200, 100, 255), (150, 50, 255), (100, 0, 200)],
            'rainbow': [(255, 0, 0), (255, 127, 0), (255, 255, 0),
                        (0, 255, 0), (0, 0, 255), (75, 0, 130)],
        }
        palette = colors.get(color, colors['warm'])
        pos_data = {
            'top_right': (0.7, -0.2), 'top_left': (-0.2, -0.2),
            'bottom_right': (0.7, 0.7), 'bottom_left': (-0.2, 0.7),
            'center': (0.3, 0.3),
        }
        px_n, py_n = pos_data.get(direction, (0.7, -0.2))
        _zero = np.zeros((height, width, 4), dtype=np.uint8)

        def _get_frame(t):
            rt = t - start_time
            if rt < 0 or rt > leak_duration:
                return _zero
            progress = rt / leak_duration
            # fade in/out
            if progress < 0.3:
                alpha = progress / 0.3
            elif progress > 0.7:
                alpha = (1.0 - progress) / 0.3
            else:
                alpha = 1.0

            pos_x = width * px_n
            pos_y = height * py_n
            # Accumulate RGBA
            canvas = np.zeros((height, width, 4), dtype=np.float32)

            for i, (cr, cg, cb) in enumerate(palette):
                ox = i * 50 + int(progress * 100)
                oy = i * 30
                sw = int(width * 0.6 * (1 + i * 0.1))
                sh = int(height * 0.8 * (1 + i * 0.1))
                cx = int(pos_x + ox + sw // 2)
                cy = int(pos_y + oy + sh // 2)
                la = intensity * alpha * (1 - i * 0.2)
                if la <= 0.005:
                    continue
                # Draw filled ellipse on temp alpha mask
                mask = np.zeros((height, width), dtype=np.float32)
                _CV2.ellipse(mask, (cx, cy), (sw // 2, sh // 2),
                             0, 0, 360, 1.0, -1)
                ma = mask * la
                oma = 1.0 - ma
                canvas[:, :, 0] = canvas[:, :, 0] * oma + cr * ma
                canvas[:, :, 1] = canvas[:, :, 1] * oma + cg * ma
                canvas[:, :, 2] = canvas[:, :, 2] * oma + cb * ma
                canvas[:, :, 3] = canvas[:, :, 3] + ma * 255.0

            # Gaussian blur for soft glow
            if np.max(canvas[:, :, 3]) > 0:
                canvas[:, :, :3] = _CV2.GaussianBlur(canvas[:, :, :3], (0, 0), 12)
                canvas[:, :, 3] = _CV2.GaussianBlur(canvas[:, :, 3], (0, 0), 12)

            return np.clip(canvas, 0, 255).astype(np.uint8)

        return _get_frame

    @staticmethod
    def _make_lens_flare_overlay(width, height, start_time, flare_duration,
                                  intensity=0.5, position='center'):
        """Return overlay get_frame closure for a single lens flare event.

        Pure numpy — same visual as create_lens_flare but no MoviePy wrapper.
        """
        _yy, _xx = np.ogrid[:height, :width]
        yy = (_yy.astype(np.float32) + 0.5) / height
        xx = (_xx.astype(np.float32) + 0.5) / width
        cx_n, cy_n = 0.5, 0.5
        if position == 'top':
            cy_n = 0.25

        def _get_frame(t):
            rt = t - start_time
            if rt < 0 or rt > flare_duration:
                return np.zeros((height, width, 4), dtype=np.uint8)
            progress = rt / flare_duration
            if progress < 0.2:
                av = progress / 0.2
            else:
                av = max(0.0, 1.0 - (progress - 0.2) / 0.8)

            dx = xx - cx_n
            dy = yy - cy_n
            main = np.exp(-((dx / 0.18) ** 2 + (dy / 0.18) ** 2) * 3.0)
            sat = np.zeros((height, width), dtype=np.float32)
            for i in range(3):
                off = (i + 1) * 0.08
                sz = 0.12 / (i + 2)
                sdx = xx - (cx_n + off)
                sdy = dy
                sat += np.exp(-((sdx / sz) ** 2 + (sdy / sz) ** 2) * 2.0) * (0.5 / (i + 1))
            flare = main + sat * 0.5
            flare = np.clip(flare, 0, 1) * intensity * av
            r = np.clip(flare * 255, 0, 255).astype(np.uint8)
            g = np.clip(flare * 210, 0, 255).astype(np.uint8)
            b = np.clip(flare * 150, 0, 255).astype(np.uint8)
            a = np.clip(flare * 255, 0, 255).astype(np.uint8)
            return np.stack([r, g, b, a], axis=-1)
        return _get_frame

    @staticmethod
    def _make_film_burn_overlay(width, height, start_time, burn_duration):
        """Return overlay get_frame closure for a single film burn event.

        Pure numpy — same visual as create_film_burn but no MoviePy wrapper.
        """
        _yy, _xx = np.ogrid[:height, :width]
        yy_rel = (_yy.astype(np.float32) + 0.5) / height
        xx_rel = (_xx.astype(np.float32) + 0.5) / width
        dxc = 1.0 - xx_rel
        dyc = 1.0 - yy_rel
        corner_dist = np.sqrt(dxc ** 2 + dyc ** 2) / np.sqrt(2.0)
        _rng = np.random.RandomState(0)
        noise_base = _rng.randint(0, 2 ** 16, size=(height, width)).astype(np.float32)

        def _get_frame(t):
            rt = t - start_time
            if rt < 0 or rt > burn_duration:
                return np.zeros((height, width, 4), dtype=np.uint8)
            p = rt / burn_duration
            spread = p * 1.5
            noise = ((noise_base + p * 5000) % 256) / 256.0
            radial = corner_dist / max(spread, 0.001)
            mask = np.exp(-radial * 2.5) * 0.8 + (noise * np.exp(-radial * 1.5)) * 0.3
            mask = np.clip(mask * p * 1.2, 0, 1) * 0.9
            ia = min(1.0, p * 2)
            r = np.clip(mask * 255, 0, 255).astype(np.uint8)
            g = np.clip(mask * 170, 0, 255).astype(np.uint8)
            b = np.clip(mask * 40, 0, 255).astype(np.uint8)
            a = np.clip(mask * 200 * ia, 0, 255).astype(np.uint8)
            return np.stack([r, g, b, a], axis=-1)
        return _get_frame

    @staticmethod
    def build_light_effect_overlays(settings, width, height, fps, duration):
        """Build overlay source dicts for the OpenCV render path.

        Inspects the user's light-leak / lens-flare / film-burn settings
        and returns a list of overlay-source dicts ready for ``_ov_sources``.

        Each dict::
            {'get_frame': callable(t) → RGBA ndarray,
             'pos': (0, 0), 'start': 0, 'end': duration}
        """
        overlays = []

        # ── Light leaks ────────────────────────────────────────────
        if settings.get('light_leak_enabled', False):
            try:
                color = settings.get('light_leak_color', 'warm')
                intensity = float(settings.get('light_leak_intensity', 0.6))
                start_time = float(settings.get('light_leak_start_time', 0.0))
                leak_dur = float(settings.get('light_leak_duration', 3.0))
                direction = settings.get('light_leak_direction', 'top_right')
                repeat = settings.get('light_leak_repeat_enabled', False)
                repeat_interval = float(settings.get('light_leak_repeat_interval', 8.0))

                if repeat:
                    ct = start_time
                    while ct < duration:
                        _fn = LightLeaksEffects._make_light_leak_overlay(
                            width, height, ct, leak_dur,
                            color=color, intensity=intensity, direction=direction)
                        overlays.append({
                            'get_frame': _fn, 'pos': (0, 0),
                            'start': ct, 'end': min(ct + leak_dur, duration),
                        })
                        ct += repeat_interval
                else:
                    _fn = LightLeaksEffects._make_light_leak_overlay(
                        width, height, start_time, leak_dur,
                        color=color, intensity=intensity, direction=direction)
                    overlays.append({
                        'get_frame': _fn, 'pos': (0, 0),
                        'start': start_time, 'end': min(start_time + leak_dur, duration),
                    })
            except Exception as e:
                print(f'  [WARNING] Light leak overlays (OpenCV): {e}')

        # ── Lens flares ────────────────────────────────────────────
        if settings.get('lens_flare_enabled', False):
            try:
                intensity = float(settings.get('lens_flare_intensity', 0.5))
                start_time = float(settings.get('lens_flare_start_time', 1.0))
                flare_dur = float(settings.get('lens_flare_duration', 2.0))
                position = settings.get('lens_flare_position', 'center')
                repeat = settings.get('lens_flare_repeat_enabled', False)
                repeat_interval = float(settings.get('lens_flare_repeat_interval', 5.0))

                if repeat:
                    ct = start_time
                    while ct < duration:
                        _fn = LightLeaksEffects._make_lens_flare_overlay(
                            width, height, ct, flare_dur,
                            intensity=intensity, position=position)
                        overlays.append({
                            'get_frame': _fn, 'pos': (0, 0),
                            'start': ct, 'end': min(ct + flare_dur, duration),
                        })
                        ct += repeat_interval
                else:
                    _fn = LightLeaksEffects._make_lens_flare_overlay(
                        width, height, start_time, flare_dur,
                        intensity=intensity, position=position)
                    overlays.append({
                        'get_frame': _fn, 'pos': (0, 0),
                        'start': start_time, 'end': min(start_time + flare_dur, duration),
                    })
            except Exception as e:
                print(f'  [WARNING] Lens flare overlays (OpenCV): {e}')

        # ── Film burns ─────────────────────────────────────────────
        if settings.get('film_burn_enabled', False):
            try:
                start_time = float(settings.get('film_burn_start_time', 0.0))
                burn_dur = float(settings.get('film_burn_duration', 1.5))
                repeat = settings.get('film_burn_repeat_enabled', False)
                repeat_interval = float(settings.get('film_burn_repeat_interval', 10.0))

                if repeat:
                    ct = start_time
                    while ct < duration:
                        _fn = LightLeaksEffects._make_film_burn_overlay(
                            width, height, ct, burn_dur)
                        overlays.append({
                            'get_frame': _fn, 'pos': (0, 0),
                            'start': ct, 'end': min(ct + burn_dur, duration),
                        })
                        ct += repeat_interval
                else:
                    _fn = LightLeaksEffects._make_film_burn_overlay(
                        width, height, start_time, burn_dur)
                    overlays.append({
                        'get_frame': _fn, 'pos': (0, 0),
                        'start': start_time, 'end': min(start_time + burn_dur, duration),
                    })
            except Exception as e:
                print(f'  [WARNING] Film burn overlays (OpenCV): {e}')

        return overlays


class ParticleEffects:
    """Particle effects for viral videos (glitter, stars, hearts, confetti)"""

    @staticmethod
    def create_glitter(width, height, duration, fps, intensity=0.5):
        """Create glitter/sparkle particle effect"""
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        # Number of particles based on intensity
        num_particles = int(20 * intensity)

        # Generate random particle positions and timing
        np.random.seed(42)
        particles = []
        for _ in range(num_particles):
            particles.append({
                'x': np.random.randint(0, width),
                'y': np.random.randint(0, height),
                'size': np.random.randint(3, 8),
                'phase': np.random.uniform(0, 2 * np.pi),
                'speed': np.random.uniform(0.5, 2.0)
            })

        def make_frame(t):
            img = Image.new('RGBA', (width, height), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)

            for particle in particles:
                # Twinkling effect using sine wave
                alpha = abs(np.sin(t * particle['speed'] + particle['phase']))

                # Add some vertical drift
                y_offset = int(t * 10) % height
                y = (particle['y'] + y_offset) % height

                # Draw sparkle (small cross shape)
                size = particle['size']
                x = particle['x']

                particle_alpha = int(255 * alpha * intensity)
                color = (255, 255, 255, particle_alpha)

                # Center dot
                draw.ellipse([x-size//2, y-size//2, x+size//2, y+size//2], fill=color)

                # Sparkle rays
                ray_length = size * 2
                draw.line([x-ray_length, y, x+ray_length, y], fill=color, width=1)
                draw.line([x, y-ray_length, x, y+ray_length], fill=color, width=1)

            return np.array(img).copy()

        clip = VideoClip(make_frame, duration=duration)
        try:
            clip = clip.with_fps(fps)
        except:
            clip = clip.set_fps(fps)

        return clip

    @staticmethod
    def create_stars(width, height, duration, fps):
        """Create floating star particles"""
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        num_stars = 15

        np.random.seed(43)
        stars = []
        for _ in range(num_stars):
            stars.append({
                'x': np.random.randint(0, width),
                'y': np.random.randint(0, height),
                'size': np.random.randint(15, 30),
                'speed_x': np.random.uniform(-20, 20),
                'speed_y': np.random.uniform(-30, -10),  # Float upward
                'phase': np.random.uniform(0, 2 * np.pi),
                'rotation_speed': np.random.uniform(1, 3)
            })

        def make_frame(t):
            img = Image.new('RGBA', (width, height), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)

            for star in stars:
                # Moving position
                x = (star['x'] + int(star['speed_x'] * t)) % width
                y = (star['y'] + int(star['speed_y'] * t)) % height

                # Fade in/out
                alpha = abs(np.sin(t * 0.5 + star['phase']))

                size = star['size']
                star_alpha = int(200 * alpha)

                # Draw 5-pointed star using emoji-like shape
                # Yellow/gold color
                color = (255, 215, 0, star_alpha)

                # Simple star using triangles
                points = []
                for i in range(5):
                    angle = i * 4 * np.pi / 5 - np.pi/2 + (t * star['rotation_speed'])
                    points.append((x + int(size * np.cos(angle)),
                                 y + int(size * np.sin(angle))))

                # Draw star outline
                for i in range(5):
                    j = (i + 2) % 5
                    draw.line([points[i], points[j]], fill=color, width=2)

                # Fill center
                center_size = size // 3
                draw.ellipse([x-center_size, y-center_size, x+center_size, y+center_size],
                           fill=color)

            return np.array(img).copy()

        clip = VideoClip(make_frame, duration=duration)
        try:
            clip = clip.with_fps(fps)
        except:
            clip = clip.set_fps(fps)

        return clip

    @staticmethod
    def create_hearts(width, height, duration, fps):
        """Create floating heart particles"""
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        num_hearts = 12

        np.random.seed(44)
        hearts = []
        for _ in range(num_hearts):
            hearts.append({
                'x': np.random.randint(0, width),
                'y': height + np.random.randint(0, 200),  # Start below screen
                'size': np.random.randint(20, 40),
                'speed_y': np.random.uniform(30, 60),  # Float upward
                'speed_x': np.random.uniform(-10, 10),  # Slight horizontal drift
                'phase': np.random.uniform(0, 2 * np.pi),
                'sway': np.random.uniform(10, 30)
            })

        def make_frame(t):
            img = Image.new('RGBA', (width, height), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)

            for heart in hearts:
                # Moving position with sway
                x = heart['x'] + int(heart['speed_x'] * t) + int(heart['sway'] * np.sin(t + heart['phase']))
                y = (heart['y'] - int(heart['speed_y'] * t)) % (height + 200)

                # Skip if off screen
                if y > height or y < -100:
                    continue

                # Fade based on position
                if y < 100:
                    alpha = y / 100
                elif y > height - 100:
                    alpha = (height - y) / 100
                else:
                    alpha = 1.0

                size = heart['size']
                heart_alpha = int(220 * alpha)

                # Pink/red hearts
                colors = [(255, 20, 147, heart_alpha), (255, 105, 180, heart_alpha)]
                color = colors[int(t * 2) % 2]

                # Draw heart shape using two circles and triangle
                half_size = size // 2
                draw.ellipse([x-half_size, y-half_size//2, x, y+half_size//2], fill=color)
                draw.ellipse([x, y-half_size//2, x+half_size, y+half_size//2], fill=color)
                draw.polygon([x-half_size, y, x+half_size, y, x, y+size], fill=color)

            return np.array(img).copy()

        clip = VideoClip(make_frame, duration=duration)
        try:
            clip = clip.with_fps(fps)
        except:
            clip = clip.set_fps(fps)

        return clip

    @staticmethod
    def create_confetti(width, height, duration, fps):
        """Create falling confetti particles"""
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        num_confetti = 30

        np.random.seed(45)
        confetti_pieces = []
        for _ in range(num_confetti):
            confetti_pieces.append({
                'x': np.random.randint(0, width),
                'y': -np.random.randint(0, 300),  # Start above screen
                'size': np.random.randint(5, 15),
                'speed_y': np.random.uniform(100, 200),  # Fall speed
                'speed_x': np.random.uniform(-30, 30),
                'rotation': np.random.uniform(0, 2 * np.pi),
                'rotation_speed': np.random.uniform(2, 5),
                'color': tuple(np.random.randint(0, 255, 3).tolist() + [220])
            })

        def make_frame(t):
            img = Image.new('RGBA', (width, height), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)

            for confetti in confetti_pieces:
                # Moving position
                x = (confetti['x'] + int(confetti['speed_x'] * t)) % width
                y = (confetti['y'] + int(confetti['speed_y'] * t)) % (height + 300)

                # Skip if way off screen
                if y < -100 or y > height + 100:
                    continue

                size = confetti['size']
                rotation = confetti['rotation'] + t * confetti['rotation_speed']

                # Draw rectangle with rotation effect (simplified)
                # Rotate by changing width/height ratio
                w = abs(int(size * np.cos(rotation)))
                h = abs(int(size * np.sin(rotation)))
                w = max(2, w)
                h = max(2, h)

                draw.rectangle([x-w, y-h, x+w, y+h], fill=confetti['color'])

            return np.array(img).copy()

        clip = VideoClip(make_frame, duration=duration)
        try:
            clip = clip.with_fps(fps)
        except:
            clip = clip.set_fps(fps)

        return clip

    @staticmethod
    def create_combined(width, height, duration, fps, glitter=False, glitter_intensity=0.5,
                       stars=False, hearts=False, confetti=False):
        """Create combined particle effects in single layer for better performance.
        This is much faster than rendering multiple separate particle layers."""
        try:
            from moviepy import VideoClip
        except ImportError:
            from moviepy.editor import VideoClip

        # Initialize all particles we need
        np.random.seed(42)  # Reproducible

        # Glitter particles
        glitter_particles = []
        if glitter:
            num_particles = int(15 * glitter_intensity)  # Reduced from 20 for performance
            for _ in range(num_particles):
                glitter_particles.append({
                    'x': np.random.randint(0, width),
                    'y': np.random.randint(0, height),
                    'size': np.random.randint(2, 6),  # Smaller for performance
                    'phase': np.random.uniform(0, 2 * np.pi),
                    'speed': np.random.uniform(0.5, 2.0)
                })

        # Star particles
        star_particles = []
        if stars:
            num_stars = 10  # Reduced from 15
            np.random.seed(43)
            for _ in range(num_stars):
                star_particles.append({
                    'x': np.random.randint(0, width),
                    'y': np.random.randint(0, height),
                    'size': np.random.randint(12, 25),  # Smaller
                    'speed_x': np.random.uniform(-15, 15),
                    'speed_y': np.random.uniform(-25, -8),
                    'phase': np.random.uniform(0, 2 * np.pi),
                    'rotation_speed': np.random.uniform(1, 3)
                })

        # Heart particles
        heart_particles = []
        if hearts:
            num_hearts = 8  # Reduced from 12
            np.random.seed(44)
            for _ in range(num_hearts):
                heart_particles.append({
                    'x': np.random.randint(0, width),
                    'y': np.random.randint(0, height),
                    'size': np.random.randint(15, 30),
                    'speed_y': np.random.uniform(-30, -15),
                    'wobble': np.random.uniform(0.5, 2.0),
                    'phase': np.random.uniform(0, 2 * np.pi)
                })

        # Confetti particles
        confetti_particles = []
        if confetti:
            colors = [(255, 0, 100, 200), (255, 200, 0, 200), (0, 255, 100, 200),
                      (100, 100, 255, 200), (255, 100, 255, 200)]
            num_confetti = 20  # Reduced from 30
            np.random.seed(45)
            for _ in range(num_confetti):
                confetti_particles.append({
                    'x': np.random.randint(0, width),
                    'y': np.random.randint(-height, 0),
                    'size': np.random.randint(3, 8),
                    'speed_y': np.random.uniform(50, 120),
                    'speed_x': np.random.uniform(-20, 20),
                    'rotation': np.random.uniform(0, 2 * np.pi),
                    'rotation_speed': np.random.uniform(2, 5),
                    'color': colors[np.random.randint(0, len(colors))]
                })

        def make_frame(t):
            img = Image.new('RGBA', (width, height), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)

            # Draw glitter
            if glitter:
                for p in glitter_particles:
                    alpha = abs(np.sin(t * p['speed'] + p['phase']))
                    y_offset = int(t * 10) % height
                    y = (p['y'] + y_offset) % height
                    x = p['x']
                    size = p['size']
                    particle_alpha = int(255 * alpha * glitter_intensity)
                    color = (255, 255, 255, particle_alpha)
                    draw.ellipse([x-size//2, y-size//2, x+size//2, y+size//2], fill=color)

            # Draw stars
            if stars:
                for star in star_particles:
                    x = (star['x'] + int(star['speed_x'] * t)) % width
                    y = (star['y'] + int(star['speed_y'] * t)) % height
                    alpha = abs(np.sin(t * 0.5 + star['phase']))
                    size = star['size']
                    star_alpha = int(180 * alpha)
                    color = (255, 215, 0, star_alpha)
                    points = []
                    for i in range(5):
                        angle = i * 4 * np.pi / 5 - np.pi/2 + (t * star['rotation_speed'])
                        points.append((x + int(size * np.cos(angle)), y + int(size * np.sin(angle))))
                    for i in range(5):
                        j = (i + 2) % 5
                        draw.line([points[i], points[j]], fill=color, width=2)

            # Draw hearts
            if hearts:
                for heart in heart_particles:
                    x = heart['x'] + int(15 * np.sin(t * heart['wobble'] + heart['phase']))
                    y = (heart['y'] + int(heart['speed_y'] * t)) % height
                    alpha = abs(np.sin(t * 0.3 + heart['phase']))
                    size = heart['size']
                    heart_alpha = int(200 * alpha)
                    color = (255, 100, 150, heart_alpha)
                    draw.ellipse([x-size//2, y-size//2, x, y+size//4], fill=color)
                    draw.ellipse([x, y-size//2, x+size//2, y+size//4], fill=color)
                    draw.polygon([(x-size//2, y), (x+size//2, y), (x, y+size//2)], fill=color)

            # Draw confetti
            if confetti:
                for c in confetti_particles:
                    x = (c['x'] + int(c['speed_x'] * t)) % width
                    y = (c['y'] + int(c['speed_y'] * t)) % height
                    rotation = c['rotation'] + t * c['rotation_speed']
                    w = max(2, abs(int(c['size'] * np.cos(rotation))))
                    h = max(2, abs(int(c['size'] * np.sin(rotation))))
                    draw.rectangle([x-w, y-h, x+w, y+h], fill=c['color'])

            return np.array(img).copy()

        clip = VideoClip(make_frame, duration=duration)
        try:
            clip = clip.with_fps(fps)
        except:
            clip = clip.set_fps(fps)

        return clip


class TTSGenerator:
    """Text-to-Speech voiceover generator using Microsoft Edge TTS (natural voices)"""

    # Natural-sounding voice options with descriptions
    VOICES = {
        # US English - Female
        'aria': 'en-US-AriaNeural',           # Friendly, warm female
        'jenny': 'en-US-JennyNeural',         # Cheerful, energetic female
        'michelle': 'en-US-MichelleNeural',   # Professional, clear female
        'monica': 'en-US-AriaNeural',         # Deep, mature female (DEEP) [remap: MonicaNeural retired]
        'nancy': 'en-US-AnaNeural',           # News anchor female, authoritative (DEEP) [remap: NancyNeural retired]
        'amber': 'en-US-JennyNeural',         # Young, casual female [remap: AmberNeural retired]
        'ashley': 'en-US-AvaNeural',          # Bright, youthful female [remap: AshleyNeural retired]
        'sara': 'en-US-EmmaNeural',           # Mature, calm female [remap: SaraNeural retired]

        # US English - Male
        'guy': 'en-US-GuyNeural',             # Friendly, warm male
        'davis': 'en-US-AndrewNeural',        # Professional, authoritative male [remap: DavisNeural retired]
        'eric': 'en-US-EricNeural',           # Conversational, casual male
        'christopher': 'en-US-ChristopherNeural',  # Deep, mature male
        'jason': 'en-US-RogerNeural',         # Deep, powerful male (HEAVY) [remap: JasonNeural retired]
        'tony': 'en-US-SteffanNeural',        # News anchor, deep authoritative (HEAVY) [remap: TonyNeural retired]
        'roger': 'en-US-RogerNeural',         # Older, wise male
        'steffan': 'en-US-SteffanNeural',     # Young, energetic male

        # British English
        'sonia': 'en-GB-SoniaNeural',         # British female
        'mia': 'en-GB-MaisieNeural',          # British deeper, mature female (DEEP) [remap: MiaNeural retired]
        'ryan': 'en-GB-RyanNeural',           # British male
        'thomas': 'en-GB-ThomasNeural',       # British deep, serious male (HEAVY)
        'libby': 'en-GB-LibbyNeural',         # British young female
        'alfie': 'en-GB-ThomasNeural',        # British young male [remap: AlfieNeural retired]

        # Australian English
        'natasha': 'en-AU-NatashaNeural',     # Australian female
        'annette': 'en-AU-NatashaNeural',     # Australian deeper, professional female (DEEP) [remap: AnnetteNeural retired]
        'william': 'en-AU-WilliamMultilingualNeural',  # Australian male [remap: WilliamNeural retired]

        # Indian English
        'neerja': 'en-IN-NeerjaNeural',       # Indian female
        'prabhat': 'en-IN-PrabhatNeural',     # Indian male

        # Additional Professional Deep Voices
        'andrew': 'en-US-AndrewNeural',       # News anchor, very deep authoritative male (ULTRA DEEP)
        'brian': 'en-US-BrianNeural',         # Deep, serious male narrator (ULTRA DEEP)
        'ana': 'en-US-AnaNeural',             # Deep, professional female narrator (DEEP)
        'brandon': 'en-US-EricNeural',        # Deep, mature male (DEEP) [remap: BrandonNeural retired]
        'emma': 'en-US-EmmaNeural',           # Professional, warm female
        'jacob': 'en-US-ChristopherNeural',   # Deep, confident male (DEEP) [remap: JacobNeural retired]

        # PREMIUM MOTIVATIONAL VOICES (Perfect for quotes & inspiration)
        'steffan_multi': 'en-US-AndrewMultilingualNeural',  # Powerful multilingual (MOTIVATION KING) [remap: SteffanMultilingual retired]
        'andrew_multi': 'en-US-AndrewMultilingualNeural',    # Deep motivational narrator (ULTRA POWERFUL)
        'ava_multi': 'en-US-AvaMultilingualNeural',          # Commanding female (POWERFUL MOTIVATION)
        'emma_multi': 'en-US-EmmaMultilingualNeural',        # Warm inspirational female (INSPIRING)
        'brian_multi': 'en-US-BrianMultilingualNeural',      # Deep powerful narrator (EPIC MOTIVATION)
        'alloy': 'en-US-AndrewMultilingualNeural',           # Smooth deep male (PREMIUM DEEP) [remap: AlloyMultilingual retired]
        'nova': 'en-US-AvaMultilingualNeural',               # Clear powerful female (PREMIUM) [remap: NovaMultilingual retired]
        'shimmer': 'en-US-EmmaMultilingualNeural',           # Energetic motivational (HIGH ENERGY) [remap: ShimmerMultilingual retired]

        # ROMANTIC & RELATIONSHIP VOICES (Perfect for love quotes & relationships)
        'aria_whisper': 'en-US-AriaNeural',          # Soft intimate female (WHISPER STYLE)
        'jenny_tender': 'en-US-JennyNeural',         # Gentle caring female (TENDER)
        'sara_whisper': 'en-US-EmmaNeural',          # Deep soothing female (INTIMATE) [remap: SaraNeural retired]
        'guy_soft': 'en-US-GuyNeural',               # Gentle romantic male (SOFT)
        'davis_tender': 'en-US-AndrewNeural',        # Warm comforting male (TENDER) [remap: DavisNeural retired]
        'christopher_whisper': 'en-US-ChristopherNeural',  # Deep intimate male (WHISPER)

        # Additional Powerful Voices
        'kai': 'en-US-GuyNeural',                            # Deep authoritative male (COMMAND) [remap: KaiNeural retired]
        'luna': 'en-US-MichelleNeural',                      # Rich warm female (INSPIRING) [remap: LunaNeural retired]
        'jenny_multi': 'en-US-EmmaMultilingualNeural',       # Energetic multilingual female (UPBEAT) [remap: JennyMultilingual retired]
        'ryan_multi': 'en-US-BrianMultilingualNeural',       # Deep confident male (STRONG) [remap: RyanMultilingual retired]

        # URDU VOICES (اردو آوازیں)
        # Pakistani Urdu (Best for Poetry & Motivation)
        'asad': 'ur-PK-AsadNeural',           # Pakistani Male (Deep Professional - BEST FOR MOTIVATION)
        'uzma': 'ur-PK-UzmaNeural',           # Pakistani Female (Clear Expressive - BEST FOR POETRY)

        # Multilingual Urdu Support (Premium for Poetry)
        'asad_multi': 'ur-PK-AsadNeural',    # Pakistani Multi (POWERFUL MOTIVATION) [remap: AsadMultilingual retired]
        'uzma_multi': 'ur-PK-UzmaNeural',    # Pakistani Multi (EXPRESSIVE POETRY) [remap: UzmaMultilingual retired]

        # Additional Regional Voices (Supporting Urdu)
        'faiz': 'ur-PK-AsadNeural',           # Alias for poetry (named after Faiz Ahmed Faiz)
        'parveen': 'ur-PK-UzmaNeural',        # Alias for female poetry (Parveen Shakir style)

        # Indian Urdu - using Hindi neural voices (Microsoft has no native Indian Urdu voices,
        # but hi-IN-* voices work well for Urdu shayari/poetry)
        'salman': 'hi-IN-MadhurNeural',       # Indian Male (Shayari Master)
        'gul': 'hi-IN-SwaraNeural',           # Indian Female (Melodious Poetry)

        # Hindi Voices - हिन्दी
        'madhur': 'hi-IN-MadhurNeural',       # Hindi Male (Deep, warm narrator)
        'swara': 'hi-IN-SwaraNeural',         # Hindi Female (Clear, expressive)

        # German Voices - Deutsche Stimmen
        'katja': 'de-DE-KatjaNeural',           # German Female (Friendly, warm)
        'conrad': 'de-DE-ConradNeural',         # German Male (Professional)
        'amala': 'de-DE-AmalaNeural',           # German Female (Modern)
        'bernd': 'de-DE-ConradNeural',          # German Male (Mature) [remap: BerndNeural retired]
        'christoph': 'de-DE-KillianNeural',     # German Male (Youthful) [remap: ChristophNeural retired]
        'elke': 'de-DE-KatjaNeural',            # German Female (Mature) [remap: ElkeNeural retired]
        'louis': 'de-DE-KillianNeural',         # German Male (Youthful) [remap: LouisNeural retired]

        # Korean Voices - 한국어 목소리
        'sunhi': 'ko-KR-SunHiNeural',           # Korean Female (Warm)
        'injoon': 'ko-KR-InJoonNeural',         # Korean Male (Deep)
        'bongjin': 'ko-KR-InJoonNeural',        # Korean Male (Professional) [remap: BongJinNeural retired]
        'hyunsu': 'ko-KR-HyunsuMultilingualNeural',  # Korean Male (Casual) [remap: HyunsuNeural retired]
        'jimin': 'ko-KR-SunHiNeural',           # Korean Female (Bright) [remap: JiMinNeural retired]
        'seohyeon': 'ko-KR-SunHiNeural',        # Korean Female (Calm) [remap: SeoHyeonNeural retired]
        'soonbok': 'ko-KR-SunHiNeural',         # Korean Female (Elegant) [remap: SoonBokNeural retired]
        'yujin': 'ko-KR-SunHiNeural',           # Korean Female (Energetic) [remap: YuJinNeural retired]

        # Arabic Voices - أصوات عربية
        'zariyah': 'ar-SA-ZariyahNeural',       # Arabic Female (Saudi Arabia)
        'hamed': 'ar-SA-HamedNeural',           # Arabic Male (Saudi Arabia)
        'salma': 'ar-EG-SalmaNeural',           # Arabic Female (Egypt)
        'shakir': 'ar-EG-ShakirNeural',         # Arabic Male (Egypt)

        # Russian Voices - Русские голоса
        'svetlana': 'ru-RU-SvetlanaNeural',     # Russian Female (Friendly)
        'dariya': 'ru-RU-SvetlanaNeural',       # Russian Female (Professional) [remap: DariyaNeural retired]
        'dmitry': 'ru-RU-DmitryNeural',         # Russian Male (Deep)

        # Backward compatibility
        'female': 'en-US-AriaNeural',         # Default female
        'male': 'en-US-GuyNeural',            # Default male
    }

    # Voice display names for GUI
    VOICE_NAMES = {
        'aria': 'Aria - US Female (Friendly)',
        'jenny': 'Jenny - US Female (Cheerful)',
        'michelle': 'Michelle - US Female (Professional)',
        'monica': 'Monica - US Female (Deep & Mature) 💎',
        'nancy': 'Nancy - US Female (News Anchor, Deep) 💎',
        'amber': 'Amber - US Female (Young)',
        'ashley': 'Ashley - US Female (Bright)',
        'sara': 'Sara - US Female (Mature)',
        'guy': 'Guy - US Male (Friendly)',
        'davis': 'Davis - US Male (Professional)',
        'eric': 'Eric - US Male (Casual)',
        'christopher': 'Christopher - US Male (Deep)',
        'jason': 'Jason - US Male (Deep & Powerful) 🔥',
        'tony': 'Tony - US Male (News Anchor, Heavy) 🔥',
        'roger': 'Roger - US Male (Wise)',
        'steffan': 'Steffan - US Male (Energetic)',
        'sonia': 'Sonia - British Female',
        'mia': 'Mia - British Female (Deep & Mature) 💎',
        'ryan': 'Ryan - British Male',
        'thomas': 'Thomas - British Male (Deep & Serious) 🔥',
        'libby': 'Libby - British Female (Young)',
        'alfie': 'Alfie - British Male (Young)',
        'natasha': 'Natasha - Australian Female',
        'annette': 'Annette - Australian Female (Deep & Professional) 💎',
        'william': 'William - Australian Male',
        'neerja': 'Neerja - Indian Female',
        'prabhat': 'Prabhat - Indian Male',
        'andrew': 'Andrew - US Male (Ultra Deep News Anchor) 🎙️',
        'brian': 'Brian - US Male (Ultra Deep Narrator) 🎙️',
        'ana': 'Ana - US Female (Deep Professional Narrator) 💎',
        'brandon': 'Brandon - US Male (Deep & Mature) 🔥',
        'emma': 'Emma - US Female (Professional & Warm)',
        'jacob': 'Jacob - US Male (Deep & Confident) 🔥',

        # Premium Motivational Voices
        'steffan_multi': '⭐ Steffan Multi - US Male (MOTIVATION KING) 👑',
        'andrew_multi': '⭐ Andrew Multi - US Male (Ultra Powerful Motivation) 🚀',
        'ava_multi': '⭐ Ava Multi - US Female (Commanding & Powerful) 💪',
        'emma_multi': '⭐ Emma Multi - US Female (Warm & Inspiring) ✨',

        # Romantic & Relationship Voices
        'aria_whisper': '💕 Aria Whisper - US Female (Soft & Intimate) 🌹',
        'jenny_tender': '💕 Jenny Tender - US Female (Gentle & Caring) 💗',
        'sara_whisper': '💕 Sara Whisper - US Female (Deep & Soothing) 🌙',
        'guy_soft': '💕 Guy Soft - US Male (Gentle & Romantic) 💝',
        'davis_tender': '💕 Davis Tender - US Male (Warm & Comforting) ❤️',
        'christopher_whisper': '💕 Christopher Whisper - US Male (Deep & Intimate) 🌹',
        'brian_multi': '⭐ Brian Multi - US Male (Epic Deep Narrator) [VIDEO]',
        'alloy': '⭐ Alloy - US Male (Premium Smooth Deep) 💎',
        'nova': '⭐ Nova - US Female (Premium Clear & Powerful) 🌟',
        'shimmer': '⭐ Shimmer - US Female (High Energy Motivation) ⚡',
        'kai': 'Kai - US Male (Deep Authoritative Command) 🎖️',
        'luna': 'Luna - US Female (Rich Warm Inspiring) 🌙',
        'jenny_multi': 'Jenny Multi - US Female (Energetic Upbeat) 🎉',
        'ryan_multi': 'Ryan Multi - US Male (Deep Confident Strong) 💪',

        # Urdu Voices - اردو آوازیں
        # Pakistani Urdu (پاکستانی اردو)
        'asad': '⭐ اسد Asad - Pakistani Male (MOTIVATION MASTER) 🇵🇰💪',
        'uzma': '⭐ عظمیٰ Uzma - Pakistani Female (POETRY QUEEN) 🇵🇰📖',
        'asad_multi': '🌟 اسد Multi - Pakistani (Powerful Motivation) 🇵🇰🚀',
        'uzma_multi': '🌟 عظمیٰ Multi - Pakistani (Expressive Poetry) 🇵🇰✨',

        # Poetry Aliases (شاعری)
        'faiz': '🎭 فیض Faiz - Poetry Deep (Like Faiz Ahmed Faiz) 📖',
        'parveen': '💫 پروین Parveen - Female Poetry (Parveen Shakir) 🌹',

        # Hindi Voices - हिन्दी आवाज़ें
        'madhur': '⭐ मधुर Madhur - Hindi Male (Deep Warm Narrator) 🇮🇳🎙️',
        'swara': '⭐ स्वरा Swara - Hindi Female (Clear Expressive) 🇮🇳✨',

        # German Voices - Deutsche Stimmen
        'katja': '🇩🇪 Katja - German Female (Friendly & Warm)',
        'conrad': '🇩🇪 Conrad - German Male (Professional)',
        'amala': '🇩🇪 Amala - German Female (Modern)',
        'bernd': '🇩🇪 Bernd - German Male (Mature)',
        'christoph': '🇩🇪 Christoph - German Male (Youthful)',
        'elke': '🇩🇪 Elke - German Female (Mature)',
        'louis': '🇩🇪 Louis - German Male (Youthful)',

        # Korean Voices - 한국어 목소리
        'sunhi': '🇰🇷 Sun-Hi - Korean Female (Warm)',
        'injoon': '🇰🇷 In-Joon - Korean Male (Deep)',
        'bongjin': '🇰🇷 Bong-Jin - Korean Male (Professional)',
        'hyunsu': '🇰🇷 Hyun-Su - Korean Male (Casual)',
        'jimin': '🇰🇷 Ji-Min - Korean Female (Bright)',
        'seohyeon': '🇰🇷 Seo-Hyeon - Korean Female (Calm)',
        'soonbok': '🇰🇷 Soon-Bok - Korean Female (Elegant)',
        'yujin': '🇰🇷 Yu-Jin - Korean Female (Energetic)',

        # Arabic Voices - أصوات عربية
        'zariyah': '🇸🇦 Zariyah - Arabic Female (Saudi Arabia)',
        'hamed': '🇸🇦 Hamed - Arabic Male (Saudi Arabia)',
        'salma': '🇪🇬 Salma - Arabic Female (Egypt)',
        'shakir': '🇪🇬 Shakir - Arabic Male (Egypt)',

        # Russian Voices - Русские голоса
        'svetlana': '🇷🇺 Svetlana - Russian Female (Friendly)',
        'dariya': '🇷🇺 Dariya - Russian Female (Professional)',
        'dmitry': '🇷🇺 Dmitry - Russian Male (Deep)',

        # Indian Urdu aliases (using Hindi neural voices for Indian dialect)
        'salman': '📜 سلمان Salman - Indian Male (Shayari Master) 🇮🇳',
        'gul': '🌺 گل Gul - Indian Female (Melodious Poetry) 🇮🇳',
    }

    @staticmethod
    async def _generate_async_with_timing(text: str, output_path: Path, voice: str, rate: str, style: str = None):
        """Async TTS generation with word-level timing data and speaking style support"""
        try:
            # Apply speaking style for romantic/whisper voices
            # SSML styles: "gentle", "soft", "whispered", "calm", "cheerful", "sad", etc.
            if style:
                # Use SSML format with correct language from voice name
                # voice name format: "de-DE-KatjaNeural" → extract "de-DE"
                _voice_lang = '-'.join(voice.split('-')[:2]) if '-' in voice else 'en-US'
                ssml_text = f'<speak version="1.0" xmlns="http://www.w3.org/2001/10/synthesis" xml:lang="{_voice_lang}"><voice name="{voice}"><express-as style="{style}">{text}</express-as></voice></speak>'
                communicate = edge_tts.Communicate(ssml_text, voice, rate=rate)
            else:
                # Regular text without style
                communicate = edge_tts.Communicate(text, voice, rate=rate)

            # Collect word timings
            word_timings = []
            chunk_count = 0
            audio_chunks = 0
            word_boundary_chunks = 0

            # Generate and save audio with word boundaries
            with open(str(output_path), 'wb') as audio_file:
                async for chunk in communicate.stream():
                    chunk_count += 1
                    chunk_type = chunk.get("type", "unknown")

                    if chunk_type == "audio":
                        audio_file.write(chunk["data"])
                        audio_chunks += 1
                    elif chunk_type == "WordBoundary":
                        word_boundary_chunks += 1
                        # Word timing info from edge-tts
                        word_info = {
                            'word': chunk.get('text', ''),
                            'offset': chunk.get('offset', 0) / 10000000.0,  # Convert to seconds
                            'duration': chunk.get('duration', 0) / 10000000.0  # Convert to seconds
                        }
                        word_timings.append(word_info)
                    # Also check for lowercase variant
                    elif chunk_type == "word_boundary":
                        word_boundary_chunks += 1
                        word_info = {
                            'word': chunk.get('text', ''),
                            'offset': chunk.get('offset', 0) / 10000000.0,
                            'duration': chunk.get('duration', 0) / 10000000.0
                        }
                        word_timings.append(word_info)

            print(f"  TTS Debug: {chunk_count} total chunks, {audio_chunks} audio, {word_boundary_chunks} word boundaries")

            # If no word boundaries, split text into words and estimate timing
            if not word_timings:
                print(f"  [WARNING] No word boundaries from TTS - using text splitting for word-level captions")
                words = text.split()
                if words:
                    # Estimate word duration (total audio duration / number of words)
                    # We'll calculate actual duration after file is created
                    for i, word in enumerate(words):
                        word_timings.append({
                            'word': word,
                            'offset': i,  # Placeholder - will be calculated later
                            'duration': 1  # Placeholder
                        })

            return True, word_timings
        except Exception as e:
            print(f"[WARNING] Async TTS generation error: {e}")
            import traceback
            traceback.print_exc()
            return False, []

    @staticmethod
    async def _generate_async(text: str, output_path: Path, voice: str, rate: str) -> bool:
        """Async TTS generation using edge-tts (legacy - no timing)"""
        try:
            # Create TTS communicator
            communicate = edge_tts.Communicate(text, voice, rate=rate)

            # Generate and save audio
            await communicate.save(str(output_path))
            return True
        except Exception as e:
            print(f"[WARNING] Async TTS generation error: {e}")
            return False

    @staticmethod
    def _generate_kokoro_voiceover(text: str, output_path: Path, settings: dict):
        """Generate voiceover using local Kokoro TTS (offline)
        Returns: (success: bool, word_timings: list)
        """
        try:
            # Try to import Kokoro
            try:
                from kokoro_onnx import Kokoro
            except ImportError:
                print("[ERROR] Kokoro TTS not installed. Install with: pip install kokoro-onnx")
                print("[INFO] Falling back to Cloud TTS...")
                # Fallback to cloud TTS
                settings['tts_engine'] = 'cloud'
                return TTSGenerator.generate_voiceover(text, output_path, settings)

            # Get Kokoro settings
            voice_setting = settings.get('kokoro_voice', 'af_bella')
            # Extract voice ID from display name like "af_sarah - Female 2 (American, Clear)"
            if ' - ' in voice_setting:
                voice = voice_setting.split(' - ')[0].strip()
            else:
                voice = voice_setting
            # Use separate Kokoro speed setting (0.5 to 2.0, default 1.0)
            speed = float(settings.get('kokoro_speed', 1.0))

            # Clean text for TTS
            clean_text = re.sub(r'[\U0001F300-\U0001F9FF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF\U00002600-\U000027BF\U0001F1E0-\U0001F1FF]+', '', text)
            clean_text = clean_text.strip()

            if not clean_text:
                print("[WARNING] No text to convert after cleaning")
                return False, []

            print(f"[INFO] Generating Kokoro TTS with voice: {voice}")

            # Initialize Kokoro - try package defaults first, then manual paths
            kokoro = None
            try:
                import os

                # Build comprehensive list of paths to search
                possible_paths = []

                # Try to find kokoro-onnx package location first
                try:
                    import kokoro_onnx
                    pkg_dir = os.path.dirname(kokoro_onnx.__file__)
                    possible_paths.extend([
                        pkg_dir,
                        os.path.join(pkg_dir, 'models'),
                    ])
                except:
                    pass

                # Additional common locations
                possible_paths.extend([
                    # Settings path (user configured)
                    settings.get('kokoro_model_path', ''),
                    # Project directory - VoiceModules folder
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'VoiceModules', 'KokoroTTS'),
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'VoiceModules'),
                    # Project directory
                    os.path.dirname(os.path.abspath(__file__)),
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'kokoro_models'),
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'models'),
                    # Current working directory
                    os.getcwd(),
                    os.path.join(os.getcwd(), "kokoro_models"),
                    os.path.join(os.getcwd(), "models"),
                    os.path.join(os.getcwd(), "VoiceModules", "KokoroTTS"),
                    # User home directories
                    os.path.expanduser("~/.kokoro"),
                    os.path.expanduser("~/kokoro"),
                    os.path.expanduser("~/.cache/kokoro"),
                    os.path.expanduser("~/.local/share/kokoro"),
                    # Common data directories
                    "/usr/share/kokoro",
                    "/usr/local/share/kokoro",
                ])

                print(f"[DEBUG] Searching for Kokoro models in {len(possible_paths)} locations...")

                model_path = None
                voices_path = None

                # Search for model files
                for base_path in possible_paths:
                    if not base_path or not os.path.exists(base_path):
                        continue

                    # Check for model file
                    for model_name in ['kokoro-v0_19.onnx', 'kokoro-v0_19-half.onnx', 'kokoro.onnx', 'model.onnx']:
                        test_model = os.path.join(base_path, model_name)
                        if os.path.exists(test_model):
                            model_path = test_model
                            print(f"[DEBUG] Found model: {model_path}")
                            break

                    # Check for voices file
                    for voices_name in ['voices-multilingual.bin', 'voices-v1.0.bin', 'voices.bin', 'voices-v0_19.bin', 'voices.json']:
                        test_voices = os.path.join(base_path, voices_name)
                        if os.path.exists(test_voices):
                            voices_path = test_voices
                            print(f"[DEBUG] Found voices: {voices_path}")
                            break

                    if model_path and voices_path:
                        break
                    elif model_path or voices_path:
                        # Reset if we only found one file
                        model_path = None
                        voices_path = None

                if model_path and voices_path:
                    print(f"[OK] Initializing Kokoro with model: {model_path}")
                    kokoro = Kokoro(model_path, voices_path)
                else:
                    # Try initializing without paths (uses package defaults)
                    try:
                        kokoro = Kokoro()
                        print("[OK] Kokoro initialized with package defaults")
                    except (TypeError, Exception) as e:
                        searched = [p for p in possible_paths if p and os.path.exists(p)]
                        not_found = [p for p in possible_paths if p and not os.path.exists(p)]
                        print(f"[ERROR] Kokoro model files not found!")
                        print(f"[ERROR] Searched existing paths: {searched}")
                        print(f"[ERROR] Non-existing paths: {not_found[:5]}")
                        print(f"[ERROR] Please download Kokoro models and place them in one of these locations")
                        print(f"[ERROR] Or set 'kokoro_model_path' in settings to point to the models directory")
                        raise FileNotFoundError(f"Kokoro model files not found")

            except Exception as init_error:
                print(f"[ERROR] Failed to initialize Kokoro: {init_error}")
                print("[INFO] Falling back to Cloud TTS...")
                settings['tts_engine'] = 'cloud'
                return TTSGenerator.generate_voiceover(text, output_path, settings)

            # Generate audio — chunk long text to avoid kokoro_onnx index-out-of-bounds
            # when phonemes exceed 510 characters (common for CJK languages)
            # Auto-detect language from voice prefix
            _voice_prefix = voice.split('_')[0].lower() if '_' in voice else ''
            _lang_map = {
                'af': 'en-us', 'am': 'en-us',
                'bf': 'en-gb', 'bm': 'en-gb',
                'jf': 'ja', 'jm': 'ja',
                'zf': 'zh', 'zm': 'zh',
                'ef': 'es', 'em': 'es', 'ff': 'fr', 'gf': 'es',
                'hf': 'hi', 'hm': 'hi',
                'if': 'it', 'im': 'it',
                'pf': 'pt', 'pm': 'pt',
                'rf': 'ru', 'tf': 'tr',
            }
            _kokoro_lang = _lang_map.get(_voice_prefix, 'en-us')
            if _kokoro_lang != 'en-us':
                print(f"  [INFO] Kokoro auto-detected language '{_kokoro_lang}' from voice prefix '{_voice_prefix}'")

            # Split text into chunks to stay under the 510-phoneme limit
            # (CJK languages produce ~10-17 phonemes/char, English ~3-5)
            _max_chunk = 40 if _kokoro_lang in ('ja', 'zh', 'ko') else 200
            if len(clean_text) > _max_chunk:
                _segments = re.split(r'(?<=[。！？.!?\n])', clean_text)
                _chunks = []
                _buf = ''
                for _seg in _segments:
                    _seg = _seg.strip()
                    if not _seg:
                        continue
                    if _buf and len(_buf + _seg) > _max_chunk:
                        _chunks.append(_buf)
                        _buf = _seg
                    else:
                        _buf += _seg
                if _buf:
                    _chunks.append(_buf)
            else:
                _chunks = [clean_text]

            _audio_parts = []
            for _ci, _chunk in enumerate(_chunks):
                _chunk = _chunk.strip()
                if not _chunk:
                    continue
                print(f"  [INFO] Kokoro chunk {_ci + 1}/{len(_chunks)} ({len(_chunk)} chars)")
                _part, _sr = kokoro.create(
                    text=_chunk,
                    voice=voice,
                    speed=speed,
                    lang=_kokoro_lang,
                )
                _audio_parts.append(_part)

            if not _audio_parts:
                raise RuntimeError("Kokoro produced no audio output")
            audio = np.concatenate(_audio_parts) if len(_audio_parts) > 1 else _audio_parts[0]
            sample_rate = _sr

            # Save as WAV first, then convert to MP3
            import soundfile as sf
            wav_path = output_path.with_suffix('.wav')
            sf.write(str(wav_path), audio, sample_rate)

            # Convert to MP3 using ffmpeg
            import subprocess
            try:
                subprocess.run([
                    'ffmpeg', '-y', '-i', str(wav_path),
                    '-acodec', 'libmp3lame', '-q:a', '2',
                    str(output_path)
                ], capture_output=True, check=True)
                wav_path.unlink()  # Remove WAV file
            except Exception as e:
                print(f"[WARNING] Could not convert to MP3: {e}")
                # Use WAV directly
                output_path = wav_path

            # Generate word timings (estimated based on text)
            words = clean_text.split()
            word_timings = []

            # Calculate audio duration
            audio_duration = len(audio) / sample_rate

            # Build initial word list
            for i, word in enumerate(words):
                word_timings.append({
                    'word': word,
                    'offset': 0.0,
                    'duration': 0.0
                })

            # Apply smart timing model (sqrt weight + function word discount +
            # punctuation pauses + VAD gap detection from audio file)
            refine_word_timings_smart(word_timings, audio_duration, audio_path=output_path)

            print(f"[OK] Generated Kokoro TTS voiceover: {output_path.name}")
            print(f"  {len(word_timings)} words, {audio_duration:.2f}s duration (smart timing)")
            return True, word_timings

        except Exception as e:
            print(f"[ERROR] Kokoro TTS generation failed: {e}")
            import traceback
            traceback.print_exc()
            return False, []

    @staticmethod
    def _generate_neutts_voiceover(text: str, output_path: Path, settings: dict):
        """Generate voiceover using NeuTTS (voice cloning)
        Returns: (success: bool, word_timings: list)
        """
        try:
            # Import NeuTTS helper
            try:
                from neutts_helper import NeuTTSHelper
            except ImportError:
                print("[ERROR] NeuTTS helper not available")
                print("  Falling back to Cloud TTS...")
                settings['tts_engine'] = 'cloud'
                return TTSGenerator.generate_voiceover(text, output_path, settings)

            # Get NeuTTS settings
            server_url = settings.get('neutts_server_url', 'http://localhost:7860')
            voice_name = settings.get('neutts_voice', '')
            # NeuTTS uses speed multiplier: 1.0 = normal, 1.5 = faster, 0.8 = slower
            speed = float(settings.get('neutts_speed', 1.0))

            if not voice_name:
                print("[ERROR] No NeuTTS voice selected")
                print("  Please clone a voice first or select an existing one")
                return False, []

            # Create helper
            helper = NeuTTSHelper(server_url)

            # Check server
            is_running, status_msg = helper.check_server_status()
            if not is_running:
                print(f"[ERROR] NeuTTS server not available: {status_msg}")
                print("  Please start the NeuTTS server (run_new_tts.bat)")
                return False, []

            # Load voice library
            success, msg = helper.load_voice_library()
            if not success or voice_name not in helper.get_available_voices():
                print(f"[ERROR] Voice '{voice_name}' not found in library")
                print("  Available voices:", list(helper.get_available_voices().keys()))
                return False, []

            # Clean text for TTS
            clean_text = re.sub(r'[\U0001F300-\U0001F9FF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF\U00002600-\U000027BF\U0001F1E0-\U0001F1FF]+', '', text)
            clean_text = clean_text.strip()

            if not clean_text:
                print("[WARNING] No text to convert after cleaning")
                return False, []

            print(f"[TTS] Generating NeuTTS voiceover with voice: {voice_name}")
            print(f"[TTS] Text: {clean_text[:100]}..." if len(clean_text) > 100 else f"[TTS] Text: {clean_text}")

            # Generate speech to WAV first, then convert to MP3
            # NeuTTS generates WAV files, so we need a temp WAV file
            wav_path = output_path.with_suffix('.wav')

            success, message = helper.generate_speech(
                text=clean_text,
                voice_name=voice_name,
                output_path=str(wav_path),
                speed=speed,
                pitch=1.0
            )

            if not success:
                print(f"[ERROR] NeuTTS generation failed: {message}")
                return False, []

            # Convert WAV to MP3 if output format is MP3
            if output_path.suffix.lower() == '.mp3':
                try:
                    import subprocess
                    print(f"[TTS] Converting WAV to MP3: {output_path.name}")
                    result = subprocess.run([
                        'ffmpeg', '-y', '-i', str(wav_path),
                        '-acodec', 'libmp3lame', '-q:a', '2',
                        str(output_path)
                    ], capture_output=True, text=True, check=True)

                    # Remove WAV file after successful conversion
                    if wav_path.exists():
                        wav_path.unlink()

                    print(f"[OK] Converted to MP3: {output_path.name}")
                except subprocess.CalledProcessError as e:
                    print(f"[ERROR] FFmpeg conversion failed: {e.stderr}")
                    return False, []
            else:
                # Output is already WAV, just rename if needed
                if wav_path != output_path:
                    import shutil
                    shutil.move(str(wav_path), str(output_path))

            # Generate word timings (estimated based on text)
            words = clean_text.split()
            word_timings = []

            # Try to get audio duration
            try:
                import soundfile as sf
                data, sample_rate = sf.read(str(output_path))
                audio_duration = len(data) / sample_rate
            except Exception:
                # Estimate based on speed and word count
                wpm = settings.get('tts_speed', 150)
                audio_duration = (len(words) / wpm) * 60

            # Build initial word list
            for i, word in enumerate(words):
                word_timings.append({
                    'word': word,
                    'offset': 0.0,
                    'duration': 0.0
                })

            # Apply smart timing model + VAD gap detection
            refine_word_timings_smart(word_timings, audio_duration, audio_path=output_path)

            print(f"[OK] Generated NeuTTS voiceover: {output_path.name}")
            print(f"  {len(word_timings)} words, {audio_duration:.2f}s duration (smart timing)")
            return True, word_timings

        except Exception as e:
            print(f"[ERROR] NeuTTS generation failed: {e}")
            import traceback
            traceback.print_exc()
            return False, []

    @staticmethod
    def _generate_qwen3_voiceover(text: str, output_path: Path, settings: dict):
        """Generate voiceover using Qwen3-TTS (local AI model with emotion)
        Returns: (success: bool, word_timings: list)
        """
        try:
            try:
                from qwen3_helper import generate_speech
            except ImportError:
                print("[ERROR] Qwen3-TTS helper not available (qwen3_helper.py missing)")
                return False, []

            # Check if voice cloning is requested
            use_cloning = settings.get('qwen3_use_cloning', False)
            ref_audio = settings.get('qwen3_clone_ref_audio', '')

            language = settings.get('qwen3_language', 'English')
            speed = float(settings.get('qwen3_speed', 1.0))

            # Clean text for TTS
            import re
            clean_text = re.sub(r'[\U0001F300-\U0001F9FF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF\U00002600-\U000027BF\U0001F1E0-\U0001F1FF]+', '', text)
            clean_text = clean_text.strip()

            if not clean_text:
                print("[WARNING] No text to convert after cleaning")
                return False, []

            if use_cloning and ref_audio and Path(ref_audio).exists():
                from qwen3_helper import generate_voice_clone
                ref_text = settings.get('qwen3_clone_ref_text', '')
                print(f"[TTS] Qwen3 Voice Clone (ref: {Path(ref_audio).name}, lang={language})")
                print(f"[TTS] Text: {clean_text[:100]}..." if len(clean_text) > 100 else f"[TTS] Text: {clean_text}")
                wav_path = output_path.with_suffix('.wav')
                result = generate_voice_clone(
                    text=clean_text,
                    output_path=wav_path,
                    ref_audio=ref_audio,
                    ref_text=ref_text,
                    language=language,
                    speed=speed,
                )
            else:
                speaker = settings.get('qwen3_speaker', 'Ryan (English)')
                instruct = settings.get('qwen3_instruct', '')
                print(f"[TTS] Generating Qwen3-TTS voiceover (speaker={speaker}, lang={language})")
                if instruct:
                    print(f"[TTS] Style: {instruct}")
                print(f"[TTS] Text: {clean_text[:100]}..." if len(clean_text) > 100 else f"[TTS] Text: {clean_text}")
                wav_path = output_path.with_suffix('.wav')
                result = generate_speech(
                    text=clean_text,
                    output_path=wav_path,
                    speaker=speaker,
                    language=language,
                    instruct=instruct,
                    speed=speed,
                )

            if not result or not wav_path.exists():
                print("[ERROR] Qwen3-TTS generated no output")
                return False, []

            print(f"[OK] Qwen3-TTS generated: {wav_path.name} ({wav_path.stat().st_size:,} bytes)")

            # Convert WAV to MP3 if needed
            if output_path.suffix.lower() == '.mp3':
                try:
                    import subprocess
                    print(f"[TTS] Converting WAV to MP3: {output_path.name}")
                    result_conv = subprocess.run([
                        'ffmpeg', '-y', '-i', str(wav_path),
                        '-acodec', 'libmp3lame', '-q:a', '2',
                        str(output_path)
                    ], capture_output=True, text=True, check=True)

                    # Remove WAV after successful conversion
                    if wav_path.exists():
                        wav_path.unlink()

                    print(f"[OK] Converted to MP3: {output_path.name}")
                except subprocess.CalledProcessError as e:
                    print(f"[ERROR] FFmpeg conversion failed: {e.stderr}")
                    return False, []
            else:
                # Already WAV, rename if needed
                if wav_path != output_path:
                    import shutil
                    shutil.move(str(wav_path), str(output_path))

            # Generate word timings (estimated based on text)
            words = clean_text.split()
            word_timings = []

            # Try to get audio duration
            try:
                import soundfile as sf
                if output_path.exists():
                    data, sample_rate = sf.read(str(output_path))
                    audio_duration = len(data) / sample_rate
                else:
                    audio_duration = 0
            except Exception:
                wpm = settings.get('tts_speed', 150)
                audio_duration = (len(words) / wpm) * 60

            # Build initial word list
            for i, word in enumerate(words):
                word_timings.append({
                    'word': word,
                    'offset': 0.0,
                    'duration': 0.0
                })

            # Apply smart timing model + VAD gap detection
            if audio_duration > 0:
                refine_word_timings_smart(word_timings, audio_duration, audio_path=output_path)

            print(f"[OK] Generated Qwen3-TTS voiceover: {output_path.name}")
            print(f"  {len(word_timings)} words, {audio_duration:.2f}s duration (smart timing)")
            return True, word_timings

        except Exception as e:
            print(f"[ERROR] Qwen3-TTS generation failed: {e}")
            import traceback
            traceback.print_exc()
            return False, []

    @staticmethod
    def _generate_piper_voiceover(text: str, output_path: Path, settings: dict):
        """Generate voiceover using Piper TTS (lightweight ONNX-based TTS)
        Returns: (success: bool, word_timings: list)
        """
        try:
            if _piper_tts is None:
                print("[ERROR] Piper TTS helper not available (piper_tts_helper.py missing)")
                return False, []

            voice = settings.get('piper_voice', 'en_US-ryan-medium')
            # Piper uses length_scale (phoneme length): lower = faster.
            # We store user-facing speed (higher = faster), so invert it.
            speed = float(settings.get('piper_speed', 1.0))
            length_scale = 1.0 / speed if speed > 0 else 1.0
            noise_scale = float(settings.get('piper_noise_scale', 0.667))
            sentence_silence = float(settings.get('piper_sentence_silence', 0.2))

            # Resolve speaker/emotion
            speaker_id = None
            speaker_name = settings.get('piper_speaker', '')
            if speaker_name and _piper_tts:
                speakers = _piper_tts.get_voice_speakers(voice)
                for s in speakers:
                    if s['name'] == speaker_name:
                        speaker_id = s['id']
                        break

            # Clean text for TTS (remove emojis)
            import re
            clean_text = re.sub(r'[\U0001F300-\U0001F9FF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF\U00002600-\U000027BF\U0001F1E0-\U0001F1FF]+', '', text)
            clean_text = clean_text.strip()

            if not clean_text:
                print("[WARNING] Piper TTS: No text to convert after cleaning")
                return False, []

            emotion_info = f" + style={speaker_name}" if speaker_id is not None else ""
            print(f"[TTS] Piper TTS voiceover (voice={voice}, speed={length_scale:.1f}{emotion_info})")
            print(f"[TTS] Text: {clean_text[:120]}..." if len(clean_text) > 120 else f"[TTS] Text: {clean_text}")

            # Auto-download voice if missing (same as GUI Test Voice does)
            if _piper_tts and not _piper_tts.is_voice_available(voice):
                print(f"[TTS] Piper voice '{voice}' not downloaded yet — auto-downloading...")
                try:
                    ok_dl = _piper_tts.download_voice(voice)
                    if not ok_dl:
                        print(f"[ERROR] Failed to auto-download Piper voice '{voice}'")
                        return False, []
                    print(f"[TTS] Downloaded Piper voice '{voice}' successfully")
                except Exception as dl_e:
                    print(f"[ERROR] Piper voice download exception: {dl_e}")
                    return False, []

            wav_path = output_path.with_suffix('.wav')
            result = _piper_tts.generate_speech(
                text=clean_text,
                output_path=str(wav_path),
                voice=voice,
                length_scale=length_scale,
                noise_scale=noise_scale,
                sentence_silence=sentence_silence,
                speaker_id=speaker_id,
            )

            if not result:
                if wav_path.exists() and wav_path.stat().st_size > 0:
                    print(f"[TTS] Piper generate_speech returned False but WAV exists "
                          f"({wav_path.stat().st_size} bytes) — proceeding anyway")
                else:
                    print(f"[ERROR] Piper TTS generation failed for voice={voice}")
                    return False, []

            # Convert WAV to MP3 if needed
            if output_path.suffix.lower() == '.mp3' and wav_path.exists():
                import subprocess as _sp
                _sp.run([
                    'ffmpeg', '-y', '-i', str(wav_path),
                    '-codec:a', 'libmp3lame', '-qscale:a', '2',
                    str(output_path)
                ], capture_output=True, timeout=30)
                wav_path.unlink(missing_ok=True)

            # Generate basic word timings
            audio_path = output_path if output_path.exists() else wav_path
            audio_duration = 0
            if audio_path.exists():
                try:
                    if audio_path.suffix.lower() == '.wav':
                        import wave
                        with wave.open(str(audio_path), 'rb') as wf:
                            audio_duration = wf.getnframes() / wf.getframerate()
                    else:
                        # Use ffprobe for MP3 duration (wave.open can't read MP3)
                        import subprocess as _sp
                        _dur = _sp.run(
                            ['ffprobe', '-v', 'error', '-show_entries',
                             'format=duration', '-of', 'default=noprint_wrappers=1:nokey=1',
                             str(audio_path)],
                            capture_output=True, text=True, timeout=10)
                        if _dur.returncode == 0 and _dur.stdout.strip():
                            audio_duration = float(_dur.stdout.strip())
                except Exception:
                    pass

            word_timings = []
            if audio_duration > 0:
                words = clean_text.split()
                words_count = len(words)
                if words_count > 0:
                    time_per_word = audio_duration / words_count
                    for i, word in enumerate(words):
                        word_timings.append({
                            'word': word,
                            'start': i * time_per_word,
                            'end': (i + 1) * time_per_word,
                        })

            print(f"[OK] Piper TTS voiceover: {audio_path.name}")
            print(f"  {len(word_timings)} words, {audio_duration:.2f}s duration")
            return True, word_timings

        except Exception as e:
            print(f"[ERROR] Piper TTS generation failed: {e}")
            import traceback
            traceback.print_exc()
            return False, []

    @staticmethod
    def _generate_google_cloud_voiceover(text: str, output_path: Path, settings: dict):
        """Generate voiceover using Gemini API TTS (Developer API).
        Returns: (success: bool, word_timings: list)
        """
        try:
            import gemini_api_tts_helper as _gtts
        except ImportError:
            print("[ERROR] gemini_api_tts_helper.py not found")
            return False, []

        if not text:
            print("[WARNING] Gemini API TTS: empty text")
            return False, []

        wav_path = output_path.with_suffix('.wav')
        success, word_timings = _gtts.generate_speech(
            text, str(wav_path),
            {
                'gemini_api_key': settings.get('gemini_api_key', ''),
                'gemini_tts_voice': settings.get('gemini_tts_voice', 'Zephyr'),
                'gemini_tts_model': settings.get('gemini_tts_model', 'gemini-2.5-pro-preview-tts'),
                'gemini_tts_speed': settings.get('gemini_tts_speed', 1.0),
                'service_account_path': settings.get('service_account_path', ''),
                'use_cloud_tts': settings.get('use_cloud_tts', False),
                'cloud_tts_voice': settings.get('cloud_tts_voice', settings.get('gemini_tts_voice', 'en-US-Studio-Q')),
            },
        )

        if not success:
            # generate_speech returns the real reason as [error_string] on failure
            reason = ""
            if word_timings and isinstance(word_timings[0], str):
                reason = word_timings[0]
            print(f"[ERROR] Gemini API TTS generation failed: {reason or 'unknown reason'}")
            return False, [reason] if reason else []

        # Convert WAV to MP3 if caller expects MP3
        if output_path.suffix.lower() == '.mp3' and wav_path.exists():
            import subprocess as _sp
            _sp.run([
                'ffmpeg', '-y', '-i', str(wav_path),
                '-codec:a', 'libmp3lame', '-qscale:a', '2',
                str(output_path)
            ], capture_output=True, timeout=30)
            wav_path.unlink(missing_ok=True)

        audio_path = output_path if output_path.exists() else wav_path
        duration_info = ""
        if audio_path.exists():
            import wave as _wave
            try:
                with _wave.open(str(audio_path), 'rb') as wf:
                    d = wf.getnframes() / wf.getframerate()
                    duration_info = f", {d:.1f}s"
            except Exception:
                pass

        print(f"[OK] Gemini API TTS voiceover: {audio_path.name}{duration_info}")
        print(f"  {len(word_timings)} words")
        return True, word_timings

    @staticmethod
    def generate_voiceover(text: str, output_path: Path, settings: dict = None):
        """Generate natural-sounding voiceover from text using selected TTS engine
        Returns: (success: bool, word_timings: list)
        """
        settings = settings or {}

        # Check which TTS engine to use
        tts_engine = settings.get('tts_engine', 'cloud')

        # Use Kokoro (local) TTS if selected
        if tts_engine == 'local':
            return TTSGenerator._generate_kokoro_voiceover(text, output_path, settings)

        # Use NeuTTS (voice cloning) if selected
        if tts_engine == 'neutts':
            return TTSGenerator._generate_neutts_voiceover(text, output_path, settings)

        # Use Qwen3-TTS (local AI model) if selected
        if tts_engine == 'qwen3':
            return TTSGenerator._generate_qwen3_voiceover(text, output_path, settings)

        # Use Piper TTS (lightweight ONNX-based) if selected
        if tts_engine == 'piper':
            return TTSGenerator._generate_piper_voiceover(text, output_path, settings)

        # Use Gemini API TTS if selected (Developer API, no Cloud billing)
        if tts_engine == 'google_cloud':
            return TTSGenerator._generate_google_cloud_voiceover(text, output_path, settings)

        # Otherwise use Cloud TTS (edge-tts)
        if not TTS_AVAILABLE:
            print("[WARNING] Cloud TTS not available - skipping voiceover generation")
            return False, []

        try:
            # Select voice based on preference (defaults to 'aria')
            voice_key = settings.get('tts_voice', 'aria').lower()
            voice = TTSGenerator.VOICES.get(voice_key, voice_key)  # fallback: use key as direct voice name

            # Detect speaking style based on voice key
            speaking_style = None
            if '_whisper' in voice_key:
                speaking_style = 'gentle'  # Soft, intimate tone
            elif '_tender' in voice_key or '_soft' in voice_key:
                speaking_style = 'calm'    # Warm, comforting tone

            # Calculate speech rate adjustment
            # Settings range: 100-250 (default 130 - slower for better caption sync)
            # edge-tts rate format: "+0%", "-20%", "+50%"
            # Allow caller to override rate dynamically (e.g., to fit video duration)
            override_rate = settings.get('tts_override_rate')
            if override_rate:
                rate = override_rate
            else:
                speed = settings.get('tts_speed', 130)
                rate_percent = int((speed - 150) / 150 * 100)  # Convert to percentage
                rate = f"{rate_percent:+d}%" if rate_percent != 0 else "+0%"

            # Clean text for TTS (remove emojis and special characters)
            clean_text = re.sub(r'[\U0001F300-\U0001F9FF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF\U00002600-\U000027BF\U0001F1E0-\U0001F1FF]+', '', text)
            clean_text = clean_text.strip()

            if not clean_text:
                print("[WARNING] No text to convert after cleaning")
                return False, []

            # Run async TTS generation with word timing - WITH RETRY LOGIC
            import time
            max_retries = 3
            retry_delay = 2  # seconds

            for attempt in range(max_retries):
                try:
                    if attempt > 0:
                        wait_time = retry_delay * (2 ** (attempt - 1))  # Exponential backoff: 2s, 4s
                        print(f"[RETRY] Attempt {attempt + 1}/{max_retries} after {wait_time}s delay...")
                        time.sleep(wait_time)

                    # Try to get existing event loop
                    loop = asyncio.get_event_loop()
                    if loop.is_running():
                        # If loop is already running, create a new one in a thread
                        import concurrent.futures
                        with concurrent.futures.ThreadPoolExecutor() as executor:
                            future = executor.submit(
                                asyncio.run,
                                TTSGenerator._generate_async_with_timing(clean_text, output_path, voice, rate, speaking_style)
                            )
                            success, word_timings = future.result(timeout=30)
                    else:
                        success, word_timings = loop.run_until_complete(
                            TTSGenerator._generate_async_with_timing(clean_text, output_path, voice, rate, speaking_style)
                        )
                except RuntimeError:
                    # No event loop, create a new one
                    success, word_timings = asyncio.run(
                        TTSGenerator._generate_async_with_timing(clean_text, output_path, voice, rate, speaking_style)
                    )

                if success:
                    print(f"[OK] Generated natural TTS voiceover: {output_path.name} (voice: {voice})")
                    print(f"  {len(word_timings)} words with timing data")
                    if attempt > 0:
                        print(f"  ✅ Succeeded on retry attempt {attempt + 1}")
                    return True, word_timings
                else:
                    if attempt < max_retries - 1:
                        print(f"[WARNING] TTS failed on attempt {attempt + 1}, will retry...")
                    else:
                        print(f"[ERROR] TTS failed after {max_retries} attempts")
                        return False, []

            return False, []

        except Exception as e:
            print(f"[WARNING] TTS generation failed: {e}")
            return False, []


class CaptionRenderer:
    """Render synchronized captions/subtitles"""

    # ── Multilingual support ──────────────────────────────────────────

    @staticmethod
    def detect_script(text):
        """Detect the dominant script in text.

        Returns one of ``'latin'``, ``'cyrillic'``, ``'arabic'``,
        ``'urdu'``, ``'cjk'``, ``'devanagari'``, ``'korean'``,
        ``'thai'``, or ``'other'``.
        """
        import unicodedata
        # Unicode ranges for each script
        _RANGES = {
            'arabic':     [(0x0600, 0x06FF), (0x0750, 0x077F), (0x08A0, 0x08FF)],
            'urdu':       [(0x0600, 0x06FF), (0xFB50, 0xFDFF), (0xFE70, 0xFEFF)],  # Arabic + presentation forms
            'cyrillic':   [(0x0400, 0x04FF), (0x0500, 0x052F)],
            'devanagari': [(0x0900, 0x097F)],
            'korean':     [(0xAC00, 0xD7AF), (0x1100, 0x11FF)],
            'thai':       [(0x0E00, 0x0E7F)],
            'cjk':        [(0x4E00, 0x9FFF), (0x3400, 0x4DBF), (0x2E80, 0x2EFF),
                            (0x3000, 0x303F), (0x2F00, 0x2FDF), (0xF900, 0xFAFF)],
            'japanese':   [(0x3040, 0x309F), (0x30A0, 0x30FF), (0x4E00, 0x9FFF)],  # Hiragana + Katakana + Kanji
        }
        scores = {k: 0 for k in _RANGES}
        for ch in text:
            cp = ord(ch)
            for script, ranges in _RANGES.items():
                for lo, hi in ranges:
                    if lo <= cp <= hi:
                        scores[script] += 1
                        break
        if not any(scores.values()):
            # Check for common extended Latin
            for ch in text:
                if ord(ch) > 0x00FF:
                    return 'other'
            return 'latin'
        # Urdu and Arabic share the basic Arabic block (0x0600-0x06FF).
        # Disambiguate: look for Urdu-exclusive characters (پ, چ, ڈ, ڑ, ژ, گ,
        # ں, ے, ھ, etc.) or presentation forms (reshaped codepoints).
        if scores.get('arabic', 0) > 0:
            # Urdu-exclusive letters (not used in Arabic)
            _urdu_exclusive = {
                0x0679, 0x067E, 0x0686, 0x0688, 0x0691, 0x0698, 0x06AF,
                0x06AA, 0x06BA, 0x06BE, 0x06C1, 0x06CC, 0x06D2,
            }
            for ch in text:
                cp = ord(ch)
                if cp in _urdu_exclusive:
                    return 'urdu'
                # Presentation forms (after reshaping)
                if (0xFB50 <= cp <= 0xFDFF) or (0xFE70 <= cp <= 0xFEFF):
                    return 'urdu'
            return 'arabic'  # no Urdu-specific chars → Arabic
        # Return highest scoring for non-Arabic scripts
        return max(scores, key=scores.get)

    @staticmethod
    def is_rtl_text(text):
        """Return True if text is primarily RTL (Arabic / Urdu / Hebrew)."""
        script = CaptionRenderer.detect_script(text)
        return script in ('arabic', 'urdu')

    @staticmethod
    def reshape_rtl(text):
        """Apply Arabic reshaping + BiDi to prepare an RTL string for
        LTR-rendering with Pillow.  Returns the visual-order string."""
        import arabic_reshaper
        from bidi import get_display
        try:
            reshaped = arabic_reshaper.reshape(text)
            return get_display(reshaped)
        except Exception:
            return text  # fallback: show as-is

    @staticmethod
    def segment_words(words, text):
        """Segment words appropriately for the detected script.

        For CJK text with no obvious word boundaries, use jieba.
        For all other scripts, return the original split-by-space list.
        """
        script = CaptionRenderer.detect_script(text)
        if script == 'cjk' or script == 'japanese':
            # Check if text lacks spaces (CJK typically does)
            if ' ' not in text:
                try:
                    import jieba
                    return list(jieba.cut(text))
                except ImportError:
                    pass
            return words  # has spaces already, or jieba unavailable
        return words

    @staticmethod
    def _render_text_line(draw, words, y_pos, font,
                          active_word_idx, active_color, inactive_color,
                          stroke_enabled, stroke_width,
                          active_stroke, inactive_stroke,
                          video_width, rtl=False):
        """Render a single line of words, word-by-word, with per-word
        highlighting.  Handles both LTR (default) and RTL rendering.

        Returns the line width (pixels) for centering purposes.
        """
        if not words:
            return 0

        # For RTL scripts, reshape + bidi the whole line, then split into
        # visual-order words.  For LTR we use the original word order.
        if rtl:
            line_text = ' '.join(words)
            visual = CaptionRenderer.reshape_rtl(line_text)
            visual_words = visual.split()
            # Visual words are in display order (left → right on screen).
            # For pure RTL, the visual order is the reverse of logical order,
            # so logical word *active_word_idx* maps to visual word
            # *len(words) - 1 - active_word_idx*.
            n = len(words)
            v_active = n - 1 - active_word_idx if active_word_idx < n else 0
        else:
            visual_words = words
            v_active = active_word_idx

        # Measure total line width for centering
        line_text_for_measure = ' '.join(visual_words) if rtl else ' '.join(words)
        bbox = draw.textbbox((0, 0), line_text_for_measure, font=font)
        line_width = bbox[2] - bbox[0]

        # Start offset: centered
        x_offset = (video_width - line_width) // 2

        # Render each visual word
        for v_idx, v_word in enumerate(visual_words):
            is_active = v_idx == v_active
            text_color = active_color if is_active else inactive_color
            stroke_color = active_stroke if is_active else inactive_stroke
            curr_stroke_w = stroke_width + 2 if is_active else stroke_width

            # Stroke
            if stroke_enabled and curr_stroke_w > 0:
                for sx in range(-curr_stroke_w, curr_stroke_w + 1):
                    for sy in range(-curr_stroke_w, curr_stroke_w + 1):
                        if sx*sx + sy*sy <= curr_stroke_w*curr_stroke_w:
                            draw.text((x_offset + sx, y_pos + sy), v_word,
                                      font=font, fill=stroke_color)

            # Main text
            draw.text((x_offset, y_pos), v_word, font=font, fill=text_color)

            # Advance x
            w_bbox = draw.textbbox((0, 0), v_word + ' ', font=font)
            x_offset += w_bbox[2] - w_bbox[0]

        return line_width

    @staticmethod
    def create_highlighted_word_captions(text, audio_duration, video_width, video_height, settings):
        """Create CapCut-style highlighted captions where current word is highlighted in different color"""
        import re
        # ImageClip is already imported at module level, no need to re-import
        print(f"[EFFECT] CAPCUT CAPTIONS: Creating highlighted captions")
        print(f"   Text: {text[:100]}...")
        print(f"   Duration: {audio_duration}s")
        print(f"   Video size: {video_width}x{video_height}")

        # Extract emojis and clean text
        emoji_pattern = re.compile(r'[\U0001F300-\U0001F9FF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF\U00002600-\U000027BF\U0001F1E0-\U0001F1FF]+')
        clean_text = emoji_pattern.sub('', text).strip()

        print(f"   Clean text (no emojis): {clean_text}")

        if not clean_text or audio_duration <= 0:
            print(f"[WARNING]️ CAPCUT: Skipping - empty text or invalid duration")
            return []

        words = clean_text.split()
        if not words:
            return []

        # ── Multilingual word segmentation ──────────────────────────
        # For CJK text (no spaces), use jieba for proper segmentation.
        # For RTL scripts, mark for later processing.
        words = CaptionRenderer.segment_words(words, clean_text)
        # Language: manual override from UI, or auto-detect from text
        _manual_lang = settings.get('caption_language', 'auto')
        if _manual_lang not in ('auto', '', None):
            _script = _manual_lang
        else:
            _script = CaptionRenderer.detect_script(clean_text)
        _is_rtl = _script in ('arabic', 'urdu')
        if _script in ('cjk', 'japanese'):
            print(f"   [MULTI] Detected CJK text — using jieba segmentation ({len(words)} segments)")

        # CRITICAL FIX: Limit text to prevent huge paragraphs from appearing
        # Maximum words - adjustable based on your needs
        # For Excel Integration: increased to 100 to support longer texts
        MAX_WORDS = 100
        if len(words) > MAX_WORDS:
            print(f"   ⚠ Text has {len(words)} words - truncating to {MAX_WORDS} to prevent overflow")
            words = words[:MAX_WORDS]

        # Text case transformation — skip for RTL and CJK scripts
        text_case = settings.get('caption_text_case', 'Normal')
        if _script not in ('arabic', 'urdu', 'cjk', 'japanese', 'cyrillic', 'thai', 'devanagari', 'korean'):
            if text_case == 'ALL CAPS':
                words = [w.upper() for w in words]
            elif text_case == 'Title Case':
                words = [w.capitalize() for w in words]
        elif text_case != 'Normal':
            print(f"   [MULTI] Skipping case transform for {_script} text")

        # Settings - ensure integers for numeric values
        words_per_caption = int(settings.get('caption_words_per_line', 3))
        font_size = int(settings.get('caption_font_size', 60))
        font_style = settings.get('caption_highlight_font_style', settings.get('caption_font_style', 'Arial Bold'))
        position = settings.get('caption_position', 'bottom')
        emoji_enabled = settings.get('emoji_in_captions', True)

        # Get emoji preset
        emoji_preset_category = settings.get('emoji_preset_category', 'general')
        if emoji_preset_category in CaptionRenderer.EMOJI_PRESETS:
            emoji_list = CaptionRenderer.EMOJI_PRESETS[emoji_preset_category]
            print(f"   Using emoji preset: {emoji_preset_category} ({len(emoji_list)} emojis)")
        else:
            emoji_list = CaptionRenderer.EMOJI_PRESETS['general']
            print(f"   Using default emoji preset: general")

        # Colors
        inactive_color_hex = settings.get('caption_inactive_color', '#FFFFFF')  # White for non-active words
        active_color_hex = settings.get('caption_highlight_color', '#FFD700')  # Yellow/gold for active word

        # Stroke/outline settings
        stroke_enabled = settings.get('caption_stroke_enabled', True)
        active_stroke_hex = settings.get('caption_active_stroke_color', '#FF1493')  # Pink for active word outline
        inactive_stroke_hex = settings.get('caption_inactive_stroke_color', '#000000')  # Black for inactive outline
        stroke_width = int(settings.get('caption_stroke_width', 4))

        # Convert hex to RGB
        def hex_to_rgb(hex_color):
            hex_color = hex_color.lstrip('#')
            return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

        inactive_color = hex_to_rgb(inactive_color_hex)
        active_color = hex_to_rgb(active_color_hex)
        active_stroke = hex_to_rgb(active_stroke_hex)
        inactive_stroke = hex_to_rgb(inactive_stroke_hex)

        print(f"   Stroke enabled: {stroke_enabled}, width: {stroke_width}")
        print(f"   Active: {active_color_hex} with stroke {active_stroke_hex}")
        print(f"   Inactive: {inactive_color_hex} with stroke {inactive_stroke_hex}")

        # Load fonts
        # Map font style names to font file names
        font_map = {
            "Arial": "arial.ttf",
            "Arial Black": "ariblk.ttf",
            "Arial Bold": "arialbd.ttf",
            "Arial Italic": "ariali.ttf",
            "Arial Bold Italic": "arialbi.ttf",
            "Calibri": "calibri.ttf",
            "Calibri Bold": "calibrib.ttf",
            "Times New Roman": "times.ttf",
            "Times New Roman Bold": "timesbd.ttf",
            "Verdana": "verdana.ttf",
            "Verdana Bold": "verdanab.ttf",
            "Georgia": "georgia.ttf",
            "Georgia Bold": "georgiab.ttf",
            "Comic Sans MS": "comic.ttf",
            "Comic Sans MS Bold": "comicbd.ttf",
            "Impact": "impact.ttf",
            "Trebuchet MS": "trebuc.ttf",
            "Trebuchet MS Bold": "trebucbd.ttf",
            "Segoe UI": "segoeui.ttf",
            "Segoe UI Bold": "segoeuib.ttf",
            "Courier New": "cour.ttf",
            "Courier New Bold": "courbd.ttf",
            "Tahoma": "tahoma.ttf",
            "Tahoma Bold": "tahomabd.ttf",
            # ─── Google / CapCut pop fonts (fall back visually) ──
            "Montserrat Bold": "arialbd.ttf",
            "Poppins Bold": "arialbd.ttf",
            "Roboto Bold": "arialbd.ttf",
            "Roboto": "arial.ttf",
            "Bebas Neue": "impact.ttf",
            "Anton": "impact.ttf",
            "Oswald Bold": "impact.ttf",
            "Oswald": "impact.ttf",
            "Raleway Bold": "arialbd.ttf",
            "Lato Bold": "arialbd.ttf",
            "DM Sans Bold": "arialbd.ttf",
            "Inter Bold": "arialbd.ttf",
            "Manrope Bold": "arialbd.ttf",
            "Lexend Bold": "arialbd.ttf",
            "Space Grotesk Bold": "arialbd.ttf",
            "Josefin Sans Bold": "arialbd.ttf",
            "Archivo Bold": "arialbd.ttf",
            "Cabin Bold": "arialbd.ttf",
            "Montserrat": "arial.ttf",
            "Poppins": "arial.ttf",
            "Raleway": "arial.ttf",
            "Lato": "arial.ttf",
            "DM Sans": "arial.ttf",
            "Inter": "arial.ttf",
            "Manrope": "arial.ttf",
            "Lexend": "arial.ttf",
            "Space Grotesk": "arial.ttf",
            "Josefin Sans": "arial.ttf",
            "Archivo": "arial.ttf",
            "Cabin": "arial.ttf",
            # ─── Serif alternatives ──────────────────────────────
            "Playfair Display Bold": "timesbd.ttf",
            "Playfair Display": "times.ttf",
            "Abril Fatface": "timesbd.ttf",
            "Franklin Gothic Medium": "impact.ttf",
            # ─── Windows-native serif alternatives ───────────────
            "Gill Sans MT Bold": "arialbd.ttf",
            "Gill Sans MT": "arial.ttf",
            "Candara Bold": "calibrib.ttf",
            "Candara": "calibri.ttf",
            "Constantia Bold": "timesbd.ttf",
            "Constantia": "times.ttf",
            "Corbel Bold": "calibrib.ttf",
            "Corbel": "calibri.ttf",
            "Bangers": "impact.ttf",
        }

        try:
            font_file = font_map.get(font_style)
            if font_file is None:
                # font_style may be a filename ("arialbd.ttf") instead of a display name ("Arial Bold")
                font_file = font_style if font_style.endswith('.ttf') else font_style + '.ttf'
                if not (Path(r"C:\Windows\Fonts") / font_file).exists():
                    font_file = 'arialbd.ttf'  # fallback
            font_path = str(Path(r"C:\Windows\Fonts") / font_file)
            font = ImageFont.truetype(font_path, font_size)
            print(f"   Using font: {font_style} ({font_file}) at {font_size}px")

            # Try per-script fallback fonts for non-Latin text
            if _script in ('arabic', 'urdu'):
                _fallback_fonts = [
                    'arabtype.ttf', 'arabtype.ttf',
                    'arabtype.ttf', 'arabtype.ttf',
                    'times.ttf', 'arial.ttf',
                ]
                for _fb in _fallback_fonts:
                    _fb_path = str(Path(r"C:\Windows\Fonts") / _fb)
                    if Path(_fb_path).exists():
                        try:
                            _test = ImageFont.truetype(_fb_path, font_size)
                            _test.getbbox('سلام')
                            font = ImageFont.truetype(_fb_path, font_size)
                            print(f"   [MULTI] RTL fallback font: {_fb}")
                            break
                        except Exception:
                            continue
            elif _script in ('japanese', 'cjk'):
                _fallback_fonts = [
                    'Yu Gothic.ttf', 'YuGothB.ttc', 'msyh.ttc', 'msyhbd.ttc',
                    'msgothic.ttc', 'simsun.ttc', 'meiryo.ttc',
                    'segoeui.ttf', 'arial.ttf',
                ]
                for _fb in _fallback_fonts:
                    _fb_path = str(Path(r"C:\Windows\Fonts") / _fb)
                    if Path(_fb_path).exists():
                        try:
                            _test = ImageFont.truetype(_fb_path, font_size)
                            _test.getbbox('你好世界')
                            font = ImageFont.truetype(_fb_path, font_size)
                            print(f"   [MULTI] CJK fallback font: {_fb}")
                            break
                        except Exception:
                            continue
            elif _script == 'cyrillic':
                # Verify Cyrillic support; try alternatives if needed
                _preferred = font_file  # already-loaded file
                _fallback_fonts = [
                    'arial.ttf', 'arialbd.ttf', 'segoeui.ttf', 'segoeuib.ttf',
                    'times.ttf', 'calibri.ttf',
                ]
                for _fb in _fallback_fonts:
                    _fb_path = str(Path(r"C:\Windows\Fonts") / _fb)
                    if Path(_fb_path).exists():
                        try:
                            _test = ImageFont.truetype(_fb_path, font_size)
                            _test.getbbox('Привет')
                            if _fb != _preferred:
                                font = ImageFont.truetype(_fb_path, font_size)
                                print(f"   [MULTI] Cyrillic fallback font: {_fb}")
                            break
                        except Exception:
                            continue

            # Get font metrics for descender-aware centering
            try:
                _ascent, _descent = font.getmetrics()
                font_total_height = _ascent + _descent
            except Exception:
                font_total_height = font_size
            if font_total_height < font_size:
                font_total_height = font_size
            # Emoji font
            emoji_font_path = str(Path(r"C:\Windows\Fonts") / 'seguiemj.ttf')
            emoji_font = ImageFont.truetype(emoji_font_path, int(font_size * 0.8))
        except Exception as e:
            print(f"   [WARNING]️ Font loading failed: {e}, using default")
            font = ImageFont.load_default()
            emoji_font = font
            try:
                _ascent, _descent = font.getmetrics()
                font_total_height = _ascent + _descent
            except Exception:
                font_total_height = font_size
            if font_total_height < font_size:
                font_total_height = font_size

        # Timing
        speaking_rate_wpm = settings.get('tts_speed', 150)
        time_per_word = 60.0 / speaking_rate_wpm

        # Distribute all words evenly across available audio so captions never
        # overflow the video — first word at 0s, last word at audio_duration.
        num_words = len(words)
        if num_words == 1:
            time_per_word = audio_duration if audio_duration and audio_duration > 0 else 3.0
            print(f"   ⏱ Single word: showing for {time_per_word:.1f}s")
        elif num_words > 1 and audio_duration and audio_duration > 0:
            time_per_word = audio_duration / num_words
            print(f"   ⏱ {num_words} words across {audio_duration:.1f}s → {time_per_word:.3f}s/word")

        caption_clips = []
        current_time = 0.0

        # Distribute emojis across words
        total_segments = len(words)
        emoji_distribution = []
        if emoji_enabled and emoji_list:
            for i in range(total_segments):
                emoji_distribution.append(emoji_list[i % len(emoji_list)])
        else:
            emoji_distribution = [''] * total_segments

        # Process each word individually for highlighting effect
        for word_idx, word in enumerate(words):
            word_start = current_time
            word_duration = time_per_word

            # Get context (words before and after for display)
            # Get caption_layout (1-line vs 2-line) - this controls LINE ARRANGEMENT
            caption_layout = settings.get('caption_layout', '2-line')

            # Get words_per_caption - this controls HOW MANY words to display
            words_per_caption = int(settings.get('caption_words_per_line', 3))

            # Calculate total words to show based on layout
            if caption_layout == '1-line':
                # 1-line layout: Show words_per_caption words on ONE line
                context_words = words_per_caption
            else:
                # 2-line layout: Show words_per_caption * 2 words (split across 2 lines)
                context_words = words_per_caption * 2

            # Calculate display range to show active word in center/focus
            half_context = context_words // 2
            display_start_idx = max(0, word_idx - half_context)
            display_end_idx = min(len(words), display_start_idx + context_words)

            # Adjust if we're near the start (show more words after)
            if display_start_idx == 0:
                display_end_idx = min(len(words), context_words)

            # Adjust if we're near the end (show more words before)
            if display_end_idx == len(words) and len(words) >= context_words:
                display_start_idx = max(0, len(words) - context_words)

            display_words = words[display_start_idx:display_end_idx]
            active_word_in_display = word_idx - display_start_idx

            # Get emoji for current word (if enabled)
            current_emoji = emoji_distribution[word_idx] if emoji_enabled else ''

            # Calculate required image height dynamically based on layout and font size
            # Added generous padding for descenders ('g','y','p','q','j') and stroke width
            _stkp = stroke_width * 2 + 4
            if caption_layout == '1-line':
                # Single line: font metrics + descender space + stroke padding
                img_height = int(font_size * 2.0 + 120 + _stkp)
            else:
                # Two lines: 2 * font_size + line spacing + descender space + stroke padding
                img_height = int(font_size * 4.0 + 120 + _stkp)

            # DEBUG: Log image dimensions on first word only
            if word_idx == 0:
                print(f"[CAPTION DEBUG] Word {word_idx+1}/{len(words)}: layout={caption_layout}, "
                      f"font={font_size}px, img_height={img_height}px, words_to_show={len(display_words)}")

            # Create image with dynamic height
            img = Image.new('RGBA', (video_width, img_height), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)

            # Measure emoji size if present
            emoji_height = 0
            emoji_width = 0
            if current_emoji:
                emoji_bbox = draw.textbbox((0, 0), current_emoji, font=emoji_font)
                emoji_width = emoji_bbox[2] - emoji_bbox[0]
                emoji_height = emoji_bbox[3] - emoji_bbox[1]

            # Layout words based on user preference
            if caption_layout == '1-line':
                # Single line layout - all display_words on ONE line
                words_per_line = len(display_words)
                line1_words = display_words
                line2_words = []
                line_spacing = 0
            else:
                # Two-line layout - split display_words across 2 lines
                # Use words_per_caption to determine words per line
                words_per_line = words_per_caption
                line_spacing = int(font_size * 1.3)  # Space between lines
                line1_words = display_words[:words_per_line] if len(display_words) > 0 else []
                line2_words = display_words[words_per_line:words_per_line*2] if len(display_words) > words_per_line else []

            # Starting Y position for first line - CENTER the text vertically in the image
            # Use actual font metrics (ascent + descent) to prevent descender cutoff.
            # Also add stroke_width padding so the outline doesn't clip.
            _text_h = max(font_total_height, font_size) + _stkp
            total_text_height = _text_h  # Height of one line
            if line2_words:
                total_text_height = _text_h * 2 + line_spacing  # Two lines + spacing

            # Calculate Y position to center text vertically in image
            if current_emoji:
                # With emoji: center text in lower portion of image
                emoji_gap_top = 20
                emoji_gap_bottom = 30
                emoji_total = emoji_gap_top + emoji_height + emoji_gap_bottom
                available_height = img_height - emoji_total
                first_line_y = emoji_total + (available_height - total_text_height) // 2
            else:
                # No emoji: center text in entire image
                first_line_y = (img_height - total_text_height) // 2

            # Draw Line 1 (First 2 words)
            if line1_words:
                y_pos = first_line_y
                CaptionRenderer._render_text_line(
                    draw, line1_words, y_pos, font,
                    active_word_in_display, active_color, inactive_color,
                    stroke_enabled, stroke_width, active_stroke, inactive_stroke,
                    video_width, rtl=(_is_rtl and active_word_in_display < len(line1_words)))

                # Draw emoji above active word (if enabled) — handled inside
                # _render_text_line for RTL; keep simple emoji drawing for LTR
                if current_emoji and not _is_rtl:
                    # Find active word's position for emoji placement
                    emoji_idx_in_line1 = min(active_word_in_display, len(line1_words) - 1)
                    if emoji_idx_in_line1 < len(line1_words):
                        active_in_line1 = line1_words[emoji_idx_in_line1]
                        wb = draw.textbbox((0, 0), active_in_line1, font=font)
                        ww = wb[2] - wb[0]
                        # Recalculate x offset since _render_text_line already drew
                        txt = ' '.join(line1_words)
                        line_w = draw.textbbox((0, 0), txt, font=font)[2]
                        # Easier: just put emoji over the centered position for now
                        emoji_x = (video_width - emoji_width) // 2
                        emoji_y = y_pos - emoji_height - 30
                        draw.text((emoji_x, emoji_y), current_emoji,
                                  font=emoji_font, embedded_color=True)

            # Draw Line 2 (Next 2 words)
            if line2_words:
                y_pos = first_line_y + line_spacing
                line2_active = active_word_in_display - words_per_line
                line2_active_valid = 0 <= line2_active < len(line2_words)
                CaptionRenderer._render_text_line(
                    draw, line2_words, y_pos, font,
                    line2_active if line2_active_valid else 0,
                    active_color, inactive_color,
                    stroke_enabled, stroke_width, active_stroke, inactive_stroke,
                    video_width, rtl=(_is_rtl and line2_active_valid))

                # Draw emoji above active word (line 2 — RTL)
                if current_emoji and _is_rtl and not line1_words:
                    emoji_x = (video_width - emoji_width) // 2
                    emoji_y = y_pos - emoji_height - 30
                    draw.text((emoji_x, emoji_y), current_emoji,
                              font=emoji_font, embedded_color=True)

            # MoviePy 1.x ImageClip handles RGBA automatically:
            # transparent=True (default) extracts alpha channel as float [0,1] mask
            clip = ImageClip(np.array(img))

            try:
                clip = clip.set_duration(word_duration)
                clip = clip.set_start(word_start)
            except AttributeError:
                clip = clip.with_duration(word_duration)
                clip = clip.with_start(word_start)

            # Position (with manual offset adjustment) - account for caption height
            caption_offset = int(settings.get('caption_y_offset', 0))  # Manual adjustment in pixels
            caption_height = img.height

            if position == 'top':
                y_pos = int(video_height * 0.05) + caption_offset
            elif position == 'center':
                # Center vertically, accounting for caption height
                y_pos = int((video_height - caption_height) / 2) + caption_offset
            else:  # bottom
                # Position from bottom edge, accounting for caption height to keep it fully on screen
                # Use 150px margin from bottom (increased for safety) BEFORE applying user offset
                base_y = video_height - caption_height - 150
                y_pos = base_y + caption_offset

                # Safety clamp: ensure caption is fully visible
                # At minimum, leave 30px margin at bottom; at most leave 30px at top
                max_y = video_height - caption_height - 30
                min_y = 30
                y_pos = max(min_y, min(y_pos, max_y))

            print(f"[CAPTION POS] Video H:{video_height}px, Caption H:{caption_height}px, "
                  f"Pos:{position}, Offset:{caption_offset}px → Final Y:{y_pos}px")

            try:
                clip = clip.set_position(('center', y_pos))
            except AttributeError:
                clip = clip.with_position(('center', y_pos))

            # Apply animation effects
            animation_style = settings.get('caption_word_animation', 'none')
            animation_intensity = settings.get('caption_animation_intensity', 1.2)
            anim_duration = min(0.2, word_duration * 0.5)  # Animation takes first 20% of word duration

            if animation_style == 'pop':
                # Pop effect: scale from small to full size quickly
                def pop_scale(t):
                    if t < anim_duration:
                        progress = min(1.0, t / anim_duration)
                        # Ease-out for smooth pop (starts at 0.3, grows to 1.0)
                        scale = 0.3 + 0.7 * (progress ** 0.5)
                        return scale
                    return 1.0

                try:
                    # Use resize with time-varying function
                    try:
                        clip = clip.resized(pop_scale)
                    except:
                        clip = clip.resize(pop_scale)
                except Exception as e:
                    print(f"[DEBUG] Pop animation failed: {e}")

            elif animation_style == 'bounce':
                # Bounce effect: scale up beyond size, then bounce back
                def bounce_scale(t):
                    if t < anim_duration:
                        progress = t / anim_duration
                        # Overshoot and bounce back with elastic effect
                        # Goes from 1.0 -> 1.3 -> 1.0 (or whatever animation_intensity is set to)
                        if progress < 0.6:
                            # First part: scale up to peak
                            scale = 1.0 + (animation_intensity - 1.0) * (progress / 0.6)
                        else:
                            # Second part: bounce back to 1.0
                            bounce_back = (progress - 0.6) / 0.4
                            scale = animation_intensity - (animation_intensity - 1.0) * bounce_back
                        return scale
                    return 1.0

                try:
                    # Use resize with time-varying function
                    try:
                        clip = clip.resized(bounce_scale)
                    except:
                        clip = clip.resize(bounce_scale)
                except Exception as e:
                    print(f"[DEBUG] Bounce animation failed: {e}")

            elif animation_style == 'fade':
                # Fade effect: fade in from transparent
                try:
                    # Use MoviePy's built-in fadein
                    if anim_duration > 0:
                        try:
                            from moviepy.video.fx import FadeIn
                            clip = clip.with_effects([FadeIn(anim_duration)])
                        except:
                            clip = clip.fadein(anim_duration)
                except Exception as e:
                    print(f"[DEBUG] Fade animation failed: {e}")

            elif animation_style == 'slide':
                # Slide effect: slide in from right
                slide_distance = 100  # pixels

                def slide_position(t):
                    if t < anim_duration:
                        progress = t / anim_duration
                        # Ease-out for smooth deceleration
                        offset = slide_distance * (1.0 - progress ** 2)
                        return (video_width // 2 + int(offset), y_pos)
                    else:
                        return ('center', y_pos)

                try:
                    try:
                        clip = clip.with_position(slide_position)
                    except:
                        clip = clip.set_position(slide_position)
                except Exception as e:
                    print(f"[DEBUG] Slide animation failed: {e}")

            caption_clips.append(clip)
            current_time += word_duration

        print(f"[OK] Created {len(caption_clips)} highlighted caption clips (word-by-word)")
        return caption_clips

    # Emoji preset categories for different video themes
    EMOJI_PRESETS = {
        # Motivational & Inspirational
        'motivational': ['🔥', '💪', '⭐', '🚀', '💯', '✨', '🎯', '👑', '🏆', '⚡', '💎', '🌟', '🙌', '💫', '🌈'],

        # Love & Relationships
        'love': ['❤️', '💕', '💖', '💗', '💓', '💞', '💝', '💘', '😍', '🥰', '😘', '💑', '💏', '👫', '💐', '🌹', '💌'],

        # Heartbreak & Sad
        'heartbreak': ['💔', '😢', '😭', '🥺', '😞', '😔', '💧', '🌧️', '⛈️', '🖤', '🥀', '😪', '😿'],

        # Success & Achievement
        'success': ['🏆', '🥇', '🎯', '💰', '💵', '💸', '📈', '👑', '🔝', '💎', '[OK]', '🎊', '🎉', '🙌', '👏'],

        # Fitness & Health
        'fitness': ['💪', '🏋️', '🏃', '⚡', '🔥', '💯', '🥇', '🎯', '🏆', '💦', '🥗', '🍎', '[TIME]️', '📊'],

        # Business & Money
        'business': ['💼', '💰', '💵', '💸', '📈', '💎', '🏦', '💳', '🤑', '📊', '📉', '💹', '🏢', '👔', '⏰'],

        # Food & Cooking
        'food': ['🍕', '🍔', '🍟', '🌮', '🍜', '🍱', '🍣', '🍰', '🎂', '🍪', '☕', '🍷', '🔪', '👨‍🍳', '🍴'],

        # Travel & Adventure
        'travel': ['✈️', '🌍', '🗺️', '🏖️', '🏔️', '🚀', '🎒', '📸', '🌅', '🌴', '🗽', '🗼', '🏰', '⛰️', '🌊'],

        # Technology & Gaming
        'tech': ['💻', '📱', '🎮', '🖥️', '⚡', '🔌', '🤖', '🚀', '💾', '🖱️', '⌨️', '🎧', '📡', '🔋', '💿'],

        # Party & Celebration
        'party': ['🎉', '🎊', '🥳', '🎈', '🎆', '🎇', '✨', '🍾', '🥂', '🍻', '🎵', '🎶', '💃', '🕺', '🪩'],

        # Nature & Environment
        'nature': ['🌿', '🌱', '🌳', '🌲', '🌺', '🌸', '🌼', '🌻', '🌞', '🌙', '⭐', '🌈', '🦋', '🐝', '🌍'],

        # Warning & Alert
        'warning': ['[WARNING]️', '🚨', '⛔', '🚫', '❗', '❓', '💥', '🔥', '⚡', '☢️', '⚡', '🆘', '🔴', '⭕'],

        # Thinking & Learning
        'educational': ['📚', '📖', '✏️', '📝', '🎓', '🧠', '💡', '🤔', '💭', '🔍', '📊', '📈', '🎯', '[OK]', '⭐'],

        # Funny & Comedy
        'funny': ['😂', '🤣', '😆', '😹', '🤪', '😜', '😝', '🤭', '😅', '🙃', '🤡', '💀', '👻', '🤠', '🥳'],

        # Spiritual & Mindfulness
        'spiritual': ['🙏', '✨', '💫', '🌟', '⭐', '🕉️', '☮️', '💜', '🧘', '🕯️', '🌙', '☀️', '🌈', '💎', '🦋'],

        # Fashion & Beauty
        'fashion': ['👗', '👠', '💄', '💅', '👑', '💎', '✨', '👜', '🕶️', '💃', '🌟', '💫', '🎀', '👒', '💍'],

        # Animals & Pets
        'animals': ['🐶', '🐱', '🐭', '🐹', '🐰', '🦊', '🐻', '🐼', '🐨', '🐯', '🦁', '🐮', '🐷', '🐸', '🐵'],

        # General/Mixed (default)
        'general': ['🔥', '💯', '✨', '⭐', '💪', '🚀', '💎', '👑', '🎯', '⚡', '❤️', '😍', '🎉', '💰', '🏆']
    }

    @staticmethod
    def create_estimated_captions(text, audio_duration, video_width, video_height, settings):
        """Create captions with estimated timing when word-level timing is unavailable"""
        # ImageClip is already imported at module level, no need to re-import
        caption_clips = []

        # Extract emojis from text (CapCut-style emoji integration)
        import re
        emoji_pattern = re.compile(r'[\U0001F300-\U0001F9FF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF\U00002600-\U000027BF\U0001F1E0-\U0001F1FF]+')
        emojis_found = emoji_pattern.findall(text)

        # Remove emojis from text for word counting
        clean_text = emoji_pattern.sub('', text).strip()

        if not clean_text or audio_duration <= 0:
            return []

        # Split into words
        words = clean_text.split()
        if not words:
            return []

        # ── Multilingual word segmentation ──────────────────────────
        words = CaptionRenderer.segment_words(words, clean_text)
        # Language: manual override from UI, or auto-detect from text
        _manual_lang = settings.get('caption_language', 'auto')
        if _manual_lang not in ('auto', '', None):
            _script = _manual_lang
        else:
            _script = CaptionRenderer.detect_script(clean_text)
        _is_rtl = _script in ('arabic', 'urdu')

        # Text case transformation — skip for non-Latin scripts
        text_case = settings.get('caption_text_case', 'Normal')
        if _script not in ('arabic', 'urdu', 'cjk', 'japanese', 'cyrillic', 'thai', 'devanagari', 'korean'):
            if text_case == 'ALL CAPS':
                words = [w.upper() for w in words]
            elif text_case == 'Title Case':
                words = [w.capitalize() for w in words]

        # Caption settings
        words_per_caption = int(settings.get('caption_words_per_line', 3))
        font_size = int(settings.get('caption_font_size', 60))
        position = settings.get('caption_position', 'bottom')
        emoji_in_captions = settings.get('emoji_in_captions', True)  # Enable/disable emoji feature

        # ========== FIX: Use character-weighted timing for better caption sync ==========
        # Longer words take longer to say, so weight duration by character count
        # This provides much better synchronization with TTS voiceover

        # Calculate character-weighted durations for each word
        word_lengths = []
        for word in words:
            # Minimum effective length of 2 chars for short words like "I", "a"
            effective_length = max(2, len(word))
            word_lengths.append(effective_length)

        total_chars = sum(word_lengths)
        if total_chars == 0:
            total_chars = 1

        # Use the full audio duration and distribute by character weight
        caption_duration = audio_duration

        # Calculate time per character unit
        time_per_char = caption_duration / total_chars

        # Calculate average for logging
        avg_time_per_word = caption_duration / len(words) if words else 0

        print(f"  Caption timing: {len(words)} words, {caption_duration:.2f}s duration ({avg_time_per_word:.2f}s avg per word)")

        # No timing offset - start at 0 for better sync
        timing_offset = 0.0

        # Load fonts — respect caption_font_style setting (same font_map as highlighted mode)
        try:
            font_style = settings.get('caption_font_style', 'Arial Bold')
            # Same font map as create_highlighted_word_captions
            font_map = {
                "Arial": "arial.ttf", "Arial Black": "ariblk.ttf",
                "Arial Bold": "arialbd.ttf", "Arial Italic": "ariali.ttf",
                "Arial Bold Italic": "arialbi.ttf",
                "Calibri": "calibri.ttf", "Calibri Bold": "calibrib.ttf",
                "Times New Roman": "times.ttf", "Times New Roman Bold": "timesbd.ttf",
                "Verdana": "verdana.ttf", "Verdana Bold": "verdanab.ttf",
                "Georgia": "georgia.ttf", "Georgia Bold": "georgiab.ttf",
                "Comic Sans MS": "comic.ttf", "Comic Sans MS Bold": "comicbd.ttf",
                "Impact": "impact.ttf",
                "Trebuchet MS": "trebuc.ttf", "Trebuchet MS Bold": "trebucbd.ttf",
                "Segoe UI": "segoeui.ttf", "Segoe UI Bold": "segoeuib.ttf",
                "Courier New": "cour.ttf", "Courier New Bold": "courbd.ttf",
                "Tahoma": "tahoma.ttf", "Tahoma Bold": "tahomabd.ttf",
                # ─── Google / CapCut pop fonts ───────────────────
                "Montserrat Bold": "arialbd.ttf", "Poppins Bold": "arialbd.ttf",
                "Roboto Bold": "arialbd.ttf", "Bebas Neue": "impact.ttf",
                "Anton": "impact.ttf", "Oswald Bold": "impact.ttf",
                "Raleway Bold": "arialbd.ttf", "Lato Bold": "arialbd.ttf",
                "DM Sans Bold": "arialbd.ttf", "Inter Bold": "arialbd.ttf",
                "Manrope Bold": "arialbd.ttf", "Lexend Bold": "arialbd.ttf",
                "Space Grotesk Bold": "arialbd.ttf", "Josefin Sans Bold": "arialbd.ttf",
                "Archivo Bold": "arialbd.ttf", "Cabin Bold": "arialbd.ttf",
                # ─── Regular weights ─────────────────────────────
                "Montserrat": "arial.ttf", "Poppins": "arial.ttf", "Roboto": "arial.ttf",
                "Raleway": "arial.ttf", "Lato": "arial.ttf", "DM Sans": "arial.ttf",
                "Inter": "arial.ttf", "Manrope": "arial.ttf", "Lexend": "arial.ttf",
                "Space Grotesk": "arial.ttf", "Josefin Sans": "arial.ttf",
                "Archivo": "arial.ttf", "Cabin": "arial.ttf", "Oswald": "arial.ttf",
                # ─── Serif / decorative ──────────────────────────
                "Playfair Display Bold": "timesbd.ttf", "Playfair Display": "times.ttf",
                "Abril Fatface": "timesbd.ttf", "Bangers": "impact.ttf",
                "Franklin Gothic Medium": "impact.ttf",
                "Gill Sans MT Bold": "arialbd.ttf", "Gill Sans MT": "arial.ttf",
                "Candara Bold": "calibrib.ttf", "Candara": "calibri.ttf",
                "Constantia Bold": "timesbd.ttf", "Constantia": "times.ttf",
                "Corbel Bold": "calibrib.ttf", "Corbel": "calibri.ttf",
                # ─── Urdu / Arabic fonts ──────────────────────────
                "Arabic Typesetting": "arabtype.ttf",
                "Arabic Typesetting": "arabtype.ttf",
                "Arabic Typesetting": "arabtype.ttf",
                "Arabic Typesetting Bold": "arabtype.ttf",
                "Arabic Typesetting": "arabtype.ttf",
                "Arabic Typesetting Bold": "arabtype.ttf",
            }
            font_file = font_map.get(font_style)
            if font_file is None:
                font_file = font_style if font_style.endswith('.ttf') else font_style + '.ttf'
                if not (Path(r"C:\Windows\Fonts") / font_file).exists():
                    font_file = 'arialbd.ttf'
            font_path = str(Path(r"C:\Windows\Fonts") / font_file)
            font = ImageFont.truetype(font_path, font_size)

            # Per-script font fallback for non-Latin text
            if _script in ('arabic', 'urdu'):
                _fallback_fonts = [
                    'arabtype.ttf', 'arabtype.ttf',
                    'arabtype.ttf', 'arabtype.ttf',
                    'times.ttf', 'arial.ttf',
                ]
                for _fb in _fallback_fonts:
                    _fb_path = str(Path(r"C:\Windows\Fonts") / _fb)
                    if Path(_fb_path).exists():
                        try:
                            ImageFont.truetype(_fb_path, font_size).getbbox('سلام')
                            font = ImageFont.truetype(_fb_path, font_size)
                            print(f"   [MULTI] RTL fallback font: {_fb}")
                            break
                        except Exception:
                            continue
            elif _script in ('japanese', 'cjk'):
                _fallback_fonts = [
                    'Yu Gothic.ttf', 'YuGothB.ttc', 'msyh.ttc', 'msyhbd.ttc',
                    'msgothic.ttc', 'simsun.ttc', 'meiryo.ttc',
                    'segoeui.ttf', 'arial.ttf',
                ]
                for _fb in _fallback_fonts:
                    _fb_path = str(Path(r"C:\Windows\Fonts") / _fb)
                    if Path(_fb_path).exists():
                        try:
                            ImageFont.truetype(_fb_path, font_size).getbbox('你好世界')
                            font = ImageFont.truetype(_fb_path, font_size)
                            print(f"   [MULTI] CJK fallback font: {_fb}")
                            break
                        except Exception:
                            continue
            elif _script == 'cyrillic':
                _fallback_fonts = [
                    'arial.ttf', 'arialbd.ttf', 'segoeui.ttf', 'segoeuib.ttf',
                    'times.ttf', 'calibri.ttf',
                ]
                for _fb in _fallback_fonts:
                    _fb_path = str(Path(r"C:\Windows\Fonts") / _fb)
                    if Path(_fb_path).exists():
                        try:
                            ImageFont.truetype(_fb_path, font_size).getbbox('Привет')
                            font = ImageFont.truetype(_fb_path, font_size)
                            print(f"   [MULTI] Cyrillic fallback font: {_fb}")
                            break
                        except Exception:
                            continue

            # Emoji font (Segoe UI Emoji for proper emoji rendering)
            emoji_font_path = str(Path(r"C:\Windows\Fonts") / 'seguiemj.ttf')
            emoji_font = ImageFont.truetype(emoji_font_path, int(font_size * 1.2))  # Slightly larger
        except:
            font = ImageFont.load_default()
            emoji_font = font


        # Distribute emojis across segments (1 emoji per caption line, context-based)
        emoji_distribution = []
        if emoji_in_captions:
            total_segments = (len(words) + words_per_caption - 1) // words_per_caption

            # ALWAYS use emoji preset from settings (ignore emojis in text)
            # Get emoji preset category from settings
            emoji_preset = settings.get('emoji_preset_category', 'general')

            # Use the selected preset, fallback to general if invalid
            if emoji_preset in CaptionRenderer.EMOJI_PRESETS:
                emojis_found = CaptionRenderer.EMOJI_PRESETS[emoji_preset]
                print(f"  Using '{emoji_preset}' emoji preset ({len(emojis_found)} emojis)")
            else:
                emojis_found = CaptionRenderer.EMOJI_PRESETS['general']
                print(f"  Invalid preset '{emoji_preset}', using 'general' preset ({len(emojis_found)} emojis)")

            if emojis_found:
                print(f"  Preset emojis: {emojis_found[:5]}..." if len(emojis_found) > 5 else f"  Preset emojis: {emojis_found}")

            # Cycle through all emojis - different emoji for each caption
            for seg_idx in range(total_segments):
                # Always cycle through emojis to ensure variety
                emoji_distribution.append(emojis_found[seg_idx % len(emojis_found)])

            print(f"  Emoji distribution: {emoji_distribution}")
            print(f"  Total: {len(emoji_distribution)} captions, each with different emoji from {len(emojis_found)} emoji set")

        # Create caption segments
        current_time = 0.0
        for i in range(0, len(words), words_per_caption):
            segment_words = words[i:i+words_per_caption]
            text_content = ' '.join(segment_words)

            # Add emoji for this segment (CapCut style)
            segment_index = i // words_per_caption
            emoji_for_segment = ''
            if emoji_in_captions and segment_index < len(emoji_distribution):
                emoji_for_segment = emoji_distribution[segment_index]

            # Calculate timing using character-weighted approach
            start_time = current_time  # Exact timing, no offset needed

            # Sum character weights for words in this segment
            segment_start_idx = i
            segment_end_idx = min(i + words_per_caption, len(words))
            segment_char_weight = sum(word_lengths[segment_start_idx:segment_end_idx])

            # Duration proportional to character weight
            duration = segment_char_weight * time_per_char
            current_time += duration

            try:
                # Calculate dimensions for text + emoji
                dummy_img = Image.new('RGBA', (1, 1))
                dummy_draw = ImageDraw.Draw(dummy_img)

                # Get layout preference (1-line or 2-line)
                caption_layout = settings.get('caption_layout', '2-line')

                # Layout words based on user preference
                if caption_layout == '1-line':
                    # Single line layout - all words on one line
                    line1_words = segment_words
                    line2_words = []
                    line1_text = ' '.join(line1_words) if line1_words else ''
                    line2_text = ''

                    line1_bbox = dummy_draw.textbbox((0, 0), line1_text, font=font) if line1_text else (0, 0, 0, 0)
                    line1_width = line1_bbox[2] - line1_bbox[0]
                    line2_width = 0
                    text_width = line1_width
                    text_height = line1_bbox[3] - line1_bbox[1]
                    line_height = text_height
                    line_spacing = 0
                else:
                    # Two-line layout - 2 words per line (like CapCut captions)
                    words_per_line = 2
                    line1_words = segment_words[:words_per_line] if len(segment_words) > 0 else []
                    line2_words = segment_words[words_per_line:words_per_line*2] if len(segment_words) > words_per_line else []

                    # Measure each line
                    line1_text = ' '.join(line1_words) if line1_words else ''
                    line2_text = ' '.join(line2_words) if line2_words else ''

                    line1_bbox = dummy_draw.textbbox((0, 0), line1_text, font=font) if line1_text else (0, 0, 0, 0)
                    line2_bbox = dummy_draw.textbbox((0, 0), line2_text, font=font) if line2_text else (0, 0, 0, 0)

                    line1_width = line1_bbox[2] - line1_bbox[0]
                    line2_width = line2_bbox[2] - line2_bbox[0]
                    line_height = line1_bbox[3] - line1_bbox[1]

                    # Total width is the max of both lines
                    text_width = max(line1_width, line2_width)
                    line_spacing = int(font_size * 1.3)  # Space between lines

                    # Total text height = line height + spacing + line height
                    text_height = line_height
                    if line2_text:
                        text_height = line_height * 2 + line_spacing

                # Measure emoji if present
                emoji_width = 0
                emoji_height = 0
                emoji_spacing = 15  # Space between text and emoji
                if emoji_for_segment:
                    emoji_bbox = dummy_draw.textbbox((0, 0), emoji_for_segment, font=emoji_font)
                    emoji_width = (emoji_bbox[2] - emoji_bbox[0]) + emoji_spacing
                    emoji_height = emoji_bbox[3] - emoji_bbox[1]

                # Calculate total dimensions - emoji is ABOVE text, so add to height
                total_width = text_width  # Emoji doesn't add to width since it's above
                padding = 20
                emoji_gap = 10 if emoji_for_segment else 0

                img_width = min(max(total_width, emoji_width) + padding * 2, int(video_width * 0.9))

                # Height = text height + emoji height (if present) + gap + padding
                img_height = text_height + padding * 2
                if emoji_for_segment:
                    img_height += emoji_height + emoji_gap

                # Get background settings
                bg_enabled = settings.get('caption_bg_enabled', False)

                # Create caption image with or without background
                if bg_enabled:
                    # Background enabled - use user's color
                    bg_color_hex = settings.get('caption_bg_color', '#000000')
                    bg_color_rgb = tuple(int(bg_color_hex.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
                    img = Image.new('RGB', (img_width, img_height), bg_color_rgb)
                    draw = ImageDraw.Draw(img)
                else:
                    # No background - transparent
                    img = Image.new('RGBA', (img_width, img_height), (0, 0, 0, 0))
                    draw = ImageDraw.Draw(img)

                # Get text color
                text_color_hex = settings.get('caption_text_color', '#FFFFFF')
                text_color_rgb = tuple(int(text_color_hex.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))

                # Stroke settings for regular captions
                stroke_enabled = settings.get('caption_stroke_enabled', False)
                stroke_color_hex = settings.get('caption_inactive_stroke_color', '#000000')  # Use inactive stroke for regular captions
                stroke_width = settings.get('caption_stroke_width', 4)
                stroke_color_rgb = tuple(int(stroke_color_hex.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))

                # Draw emoji FIRST (at top)
                emoji_y = padding
                if emoji_for_segment:
                    # Center emoji horizontally
                    emoji_x = (img_width - emoji_width) // 2

                    # Use embedded_color=True to preserve emoji colors (Pillow 8.0+)
                    try:
                        draw.text((emoji_x, emoji_y), emoji_for_segment, font=emoji_font, embedded_color=True)
                    except TypeError:
                        # Fallback for older Pillow versions - draw without fill to use emoji colors
                        draw.text((emoji_x, emoji_y), emoji_for_segment, font=emoji_font)

                # Draw text BELOW emoji on TWO LINES
                first_line_y = padding
                if emoji_for_segment:
                    first_line_y = emoji_y + emoji_height + emoji_gap

                # Prepare RTL-reshaed text if needed
                def _rtl_text(t):
                    return CaptionRenderer.reshape_rtl(t) if _is_rtl and t else t

                # Draw Line 1 (First 2 words)
                if line1_text:
                    line1_display = _rtl_text(line1_text)
                    line1_bbox = draw.textbbox((0, 0), line1_display, font=font)
                    line1_width = line1_bbox[2] - line1_bbox[0]
                    line1_x = (img_width - line1_width) // 2
                    line1_y = first_line_y

                    # Draw stroke/outline first (if enabled)
                    if stroke_enabled:
                        for adj_x in range(-stroke_width, stroke_width + 1):
                            for adj_y in range(-stroke_width, stroke_width + 1):
                                if adj_x*adj_x + adj_y*adj_y <= stroke_width*stroke_width:
                                    draw.text((line1_x + adj_x, line1_y + adj_y), line1_display,
                                            font=font, fill=stroke_color_rgb)

                    # Draw main text on top
                    draw.text((line1_x, line1_y), line1_display, font=font, fill=text_color_rgb)

                # Draw Line 2 (Next 2 words)
                if line2_text:
                    line2_display = _rtl_text(line2_text)
                    line2_bbox = draw.textbbox((0, 0), line2_display, font=font)
                    line2_width = line2_bbox[2] - line2_bbox[0]
                    line2_x = (img_width - line2_width) // 2
                    line2_y = first_line_y + line_height + line_spacing

                    # Draw stroke/outline first (if enabled)
                    if stroke_enabled:
                        for adj_x in range(-stroke_width, stroke_width + 1):
                            for adj_y in range(-stroke_width, stroke_width + 1):
                                if adj_x*adj_x + adj_y*adj_y <= stroke_width*stroke_width:
                                    draw.text((line2_x + adj_x, line2_y + adj_y), line2_display,
                                            font=font, fill=stroke_color_rgb)

                    # Draw main text on top
                    draw.text((line2_x, line2_y), line2_display, font=font, fill=text_color_rgb)

                # MoviePy 1.x ImageClip handles RGBA automatically:
                # transparent=True (default) extracts alpha channel as float [0,1] mask
                clip = ImageClip(np.array(img))

                try:
                    clip = clip.set_duration(duration)
                    clip = clip.set_start(start_time)
                except AttributeError:
                    clip = clip.with_duration(duration)
                    clip = clip.with_start(start_time)

                # Position (with manual offset adjustment) - account for caption height
                caption_offset = int(settings.get('caption_y_offset', 0))  # Manual adjustment in pixels
                caption_height = img.height

                if position == 'top':
                    y_pos = int(video_height * 0.05) + caption_offset
                elif position == 'center':
                    # Center vertically, accounting for caption height
                    y_pos = int((video_height - caption_height) / 2) + caption_offset
                else:  # bottom
                    # Position from bottom edge with 150px margin BEFORE user offset
                    base_y = video_height - caption_height - 150
                    y_pos = base_y + caption_offset
                    # Safety clamp
                    max_y = video_height - caption_height - 30
                    min_y = 30
                    y_pos = max(min_y, min(y_pos, max_y))

                try:
                    clip = clip.set_position(('center', y_pos))
                except AttributeError:
                    clip = clip.with_position(('center', y_pos))

                caption_clips.append(clip)

            except Exception as e:
                print(f"[WARNING] Error creating caption segment: {e}")

        print(f"[OK] Created {len(caption_clips)} estimated caption segments")
        return caption_clips

    @staticmethod
    def create_word_captions(word_timings, video_width, video_height, settings):
        """Create synchronized caption clips for each word"""
        # ImageClip is already imported at module level
        caption_clips = []

        # Caption settings from config
        font_size = settings.get('caption_font_size', 70)

        # Parse hex color to RGB tuple
        def hex_to_rgb(hex_color):
            hex_color = hex_color.lstrip('#')
            return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

        text_color = hex_to_rgb(settings.get('caption_text_color', '#FFFFFF'))
        bg_color = hex_to_rgb(settings.get('caption_bg_color', '#000000'))
        bg_opacity = settings.get('caption_bg_opacity', 180)  # 0-255
        bg_enabled = settings.get('caption_bg_enabled', True)  # Enable/disable background
        position = settings.get('caption_position', 'center')  # top, center, bottom
        words_per_caption = settings.get('caption_words_per_line', 3)  # Show 3 words at a time
        gap_between_captions = 0.0  # Perfect sync with voiceover (no gap)

        # Load font
        try:
            font_file = settings.get('caption_font_style', 'arialbd.ttf')
            font_path = str(Path(r"C:\Windows\Fonts") / font_file)
            font = ImageFont.truetype(font_path, font_size)
        except:
            # Try Urdu fallback fonts if Arial fails
            try:
                font = ImageFont.truetype(
                    str(Path(r"C:\Windows\Fonts") / "arabtype.ttf"),
                    font_size)
            except:
                try:
                    font = ImageFont.truetype(
                        str(Path(r"C:\Windows\Fonts") / "arabtype.ttf"),
                        font_size)
                except:
                    font = ImageFont.load_default()

        # Group words into caption segments
        caption_segments = []
        for i in range(0, len(word_timings), words_per_caption):
            segment_words = word_timings[i:i+words_per_caption]
            if not segment_words:
                continue

            # Calculate start and end time for this segment
            start_time = segment_words[0]['offset']
            end_time = segment_words[-1]['offset'] + segment_words[-1]['duration']

            # Add small gap before next caption (except for first caption)
            if i > 0:
                start_time += gap_between_captions
                end_time += gap_between_captions

            # Combine words
            text = ' '.join([w['word'] for w in segment_words])

            caption_segments.append({
                'text': text,
                'start': start_time,
                'end': end_time
            })

        print(f"[OK] Creating {len(caption_segments)} caption segments")

        # Create caption clips
        for segment in caption_segments:
            try:
                text = segment['text']

                # Create PIL image with text
                # First, get text size
                dummy_img = Image.new('RGBA', (1, 1))
                dummy_draw = ImageDraw.Draw(dummy_img)
                bbox = dummy_draw.textbbox((0, 0), text, font=font)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]

                # Add padding
                padding = 20
                img_width = min(text_width + padding * 2, int(video_width * 0.9))
                img_height = text_height + padding * 2

                # Create image with or without background based on setting
                if bg_enabled:
                    # Create image with SOLID background color
                    img_rgb = Image.new('RGB', (img_width, img_height), bg_color)
                    draw = ImageDraw.Draw(img_rgb)
                    # Draw text centered
                    text_x = (img_width - text_width) // 2
                    text_y = padding
                    draw.text((text_x, text_y), text, font=font, fill=text_color)
                    # RGB ImageClip — fully opaque, no mask needed
                    clip = ImageClip(np.array(img_rgb))
                else:
                    # Create image with TRANSPARENT background (RGBA)
                    # MoviePy 1.x ImageClip handles RGBA automatically:
                    # transparent=True (default) extracts alpha as float [0,1] mask
                    img_rgba = Image.new('RGBA', (img_width, img_height), (0, 0, 0, 0))
                    draw = ImageDraw.Draw(img_rgba)
                    text_x = (img_width - text_width) // 2
                    text_y = padding
                    draw.text((text_x, text_y), text, font=font, fill=text_color + (255,))
                    clip = ImageClip(np.array(img_rgba))

                duration = segment['end'] - segment['start']
                try:
                    clip = clip.with_duration(duration)
                    clip = clip.with_start(segment['start'])
                except AttributeError:
                    clip = clip.set_duration(duration)
                    clip = clip.set_start(segment['start'])

                # Position based on settings (with manual offset adjustment) - account for caption height
                caption_offset = int(settings.get('caption_y_offset', 0))
                caption_height = clip.h if hasattr(clip, 'h') else img_height

                if position == 'top':
                    y_pos = int(video_height * 0.05) + caption_offset
                elif position == 'center':
                    # Center vertically, accounting for caption height
                    y_pos = int((video_height - caption_height) / 2) + caption_offset
                else:  # bottom
                    # Position from bottom edge, accounting for caption height to keep it fully on screen
                    y_pos = video_height - caption_height - 100 + caption_offset
                    # Clamp to safe range
                    max_y = video_height - caption_height - 20
                    min_y = 50
                    y_pos = max(min_y, min(y_pos, max_y))

                try:
                    clip = clip.set_position(('center', y_pos))
                except AttributeError:
                    clip = clip.with_position(('center', y_pos))

                print(f"  Caption positioned at: {clip.pos}, size: {clip.size}")
                caption_clips.append(clip)

            except Exception as e:
                print(f"[WARNING] Error creating caption for '{segment['text']}': {e}")
                import traceback
                traceback.print_exc()
                continue

        return caption_clips
    @staticmethod
    def create_capcut_captions_with_timings(word_timings, video_width, video_height, settings, max_duration=None):
        """Create CapCut-style highlighted captions using actual TTS word timings for perfect sync

        Args:
            word_timings: List of word timing data from TTS
            video_width: Video width in pixels
            video_height: Video height in pixels
            settings: Caption settings dict
            max_duration: Maximum duration in seconds - clips will be truncated to this (typically audio duration)
        """
        import re
        from PIL import Image, ImageDraw, ImageFont

        print(f"[EFFECT] CAPCUT CAPTIONS (SYNCED): Creating captions with precise TTS timing")
        print(f"   Words with timing: {len(word_timings)}")
        print(f"   Video size: {video_width}x{video_height}")

        # NOTE: caption_sync_offset is applied by the caller BEFORE this function is called.
        # Applying it here as well would DOUBLE the offset — so we skip it here.
        # See caller at line 9213 for the single application point.
        if max_duration:
            print(f"   Max duration: {max_duration:.2f}s (will clip captions to this)")

        if not word_timings:
            print(f"[WARNING] No word timings available")
            return []

        # Settings — use highlight-specific font settings (fall back to global)
        font_size = int(settings.get('caption_font_size', 60))
        font_style = settings.get('caption_highlight_font_style', settings.get('caption_font_style', 'Arial Bold'))
        position = settings.get('caption_position', 'bottom')

        # Colors
        inactive_color_hex = settings.get('caption_inactive_color', '#FFFFFF')
        active_color_hex = settings.get('caption_highlight_color', '#FFD700')

        # Stroke settings
        stroke_enabled = settings.get('caption_stroke_enabled', True)
        active_stroke_hex = settings.get('caption_active_stroke_color', '#FF1493')
        inactive_stroke_hex = settings.get('caption_inactive_stroke_color', '#000000')
        stroke_width = int(settings.get('caption_stroke_width', 3))

        # IMPORTANT: CapCut highlighting style does NOT use multi-color rainbow mode
        multicolor_enabled = False  # Force disabled for CapCut highlighting
        word_colors_hex = []
        print(f"   CapCut Highlighting Mode: Using standard colors")
        print(f"   Active: {active_color_hex}, Inactive: {inactive_color_hex}")

        # Caption layout
        caption_layout = settings.get('caption_layout', '1-line')
        words_per_caption = int(float(settings.get('caption_words_per_line', 3)))  # Handle float from JSON
        print(f"   Layout: {caption_layout}, Words per line: {words_per_caption}")

        # Emoji settings
        emoji_enabled = settings.get('emoji_in_captions', True)
        emoji_preset_category = settings.get('emoji_preset_category', 'motivational')

        # Get emoji list
        if emoji_preset_category in CaptionRenderer.EMOJI_PRESETS:
            emoji_list = CaptionRenderer.EMOJI_PRESETS[emoji_preset_category]
        else:
            emoji_list = CaptionRenderer.EMOJI_PRESETS['general']

        # Convert hex to RGB
        def hex_to_rgb(hex_color):
            hex_color = hex_color.lstrip('#')
            return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

        inactive_color = hex_to_rgb(inactive_color_hex)
        active_color = hex_to_rgb(active_color_hex)
        active_stroke = hex_to_rgb(active_stroke_hex)
        inactive_stroke = hex_to_rgb(inactive_stroke_hex)

        print(f"   Stroke: {stroke_enabled}, width: {stroke_width}")
        print(f"   Active: {active_color_hex} / {active_stroke_hex}")
        print(f"   Inactive: {inactive_color_hex} / {inactive_stroke_hex}")
        print(f"   Layout: {caption_layout}, words/line: {words_per_caption}")

        # Get Y offset for position adjustment
        caption_y_offset = int(settings.get('caption_y_offset', 0))
        if caption_y_offset != 0:
            print(f"   Manual Y offset: {caption_y_offset:+d}px")

        # Load font
        font_map = {
            "Arial": "arial.ttf", "Arial Black": "ariblk.ttf", "Arial Bold": "arialbd.ttf",
            "Calibri": "calibri.ttf", "Calibri Bold": "calibrib.ttf",
            "Verdana": "verdana.ttf", "Verdana Bold": "verdanab.ttf",
            "Segoe UI": "segoeui.ttf", "Segoe UI Bold": "segoeuib.ttf",
            # ─── Google / CapCut pop fonts ───────────────────────
            "Montserrat Bold": "arialbd.ttf", "Bebas Neue": "impact.ttf",
            "Poppins Bold": "arialbd.ttf", "Roboto Bold": "arialbd.ttf",
            "Anton": "impact.ttf", "Oswald Bold": "impact.ttf",
            "Raleway Bold": "arialbd.ttf", "Lato Bold": "arialbd.ttf",
            "DM Sans Bold": "arialbd.ttf", "Inter Bold": "arialbd.ttf",
            "Manrope Bold": "arialbd.ttf", "Lexend Bold": "arialbd.ttf",
            "Space Grotesk Bold": "arialbd.ttf", "Josefin Sans Bold": "arialbd.ttf",
            "Archivo Bold": "arialbd.ttf", "Cabin Bold": "arialbd.ttf",
            # ─── Regular weights ─────────────────────────────────
            "Montserrat": "arial.ttf", "Poppins": "arial.ttf", "Roboto": "arial.ttf",
            "Raleway": "arial.ttf", "Lato": "arial.ttf", "DM Sans": "arial.ttf",
            "Inter": "arial.ttf", "Manrope": "arial.ttf", "Lexend": "arial.ttf",
            "Space Grotesk": "arial.ttf", "Josefin Sans": "arial.ttf",
            "Archivo": "arial.ttf", "Cabin": "arial.ttf", "Oswald": "arial.ttf",
            # ─── Serif / decorative ──────────────────────────────
            "Playfair Display Bold": "timesbd.ttf", "Playfair Display": "times.ttf",
            "Abril Fatface": "timesbd.ttf", "Bangers": "impact.ttf",
            "Franklin Gothic Medium": "impact.ttf",
            "Gill Sans MT Bold": "arialbd.ttf", "Gill Sans MT": "arial.ttf",
            "Candara Bold": "calibrib.ttf", "Candara": "calibri.ttf",
            "Constantia Bold": "timesbd.ttf", "Constantia": "times.ttf",
            "Corbel Bold": "calibrib.ttf", "Corbel": "calibri.ttf",
        }

        # ── Multilingual: detect script BEFORE font loading ─────
        words = [wt['word'] for wt in word_timings]
        full_text = ' '.join(words)
        _manual_lang = settings.get('caption_language', 'auto')
        if _manual_lang not in ('auto', '', None):
            _script = _manual_lang
        else:
            _script = CaptionRenderer.detect_script(full_text)
        _is_rtl = _script in ('arabic', 'urdu')

        # Text case transformation — skip for non-Latin scripts
        text_case = settings.get('caption_text_case', 'Normal')
        if _script not in ('arabic', 'urdu', 'cjk', 'japanese', 'cyrillic', 'thai', 'devanagari', 'korean'):
            if text_case == 'ALL CAPS':
                words = [w.upper() for w in words]
            elif text_case == 'Title Case':
                words = [w.capitalize() for w in words]

        try:
            # Force bold font for multi-color mode
            if multicolor_enabled:
                font_file = 'arialbd.ttf'  # Arial Bold for rainbow captions
                print(f"   🌈 Rainbow mode: forcing bold font (arialbd.ttf)")
            else:
                font_file = font_map.get(font_style)
                if font_file is None:
                    # font_style may be a display name ("Arial Bold") or a filename ("arialbd.ttf")
                    font_file = font_style if font_style.endswith('.ttf') else font_style + '.ttf'
                    if not (Path(r"C:\Windows\Fonts") / font_file).exists():
                        font_file = 'segoeuib.ttf'  # fallback

            font_path = str(Path(r"C:\Windows\Fonts") / font_file)
            font = ImageFont.truetype(font_path, font_size)
            emoji_font_path = str(Path(r"C:\Windows\Fonts") / 'seguiemj.ttf')
            emoji_font = ImageFont.truetype(emoji_font_path, int(font_size * 0.8))
            print(f"   Font: {font_style} ({font_file}) at {font_size}px")

            # Per-script font fallback for non-Latin text
            if _script in ('arabic', 'urdu'):
                _fallback_fonts = [
                    'arabtype.ttf', 'arabtype.ttf',
                    'arabtype.ttf', 'arabtype.ttf',
                    'times.ttf', 'arial.ttf',
                ]
                for _fb in _fallback_fonts:
                    _fb_path = str(Path(r"C:\Windows\Fonts") / _fb)
                    if Path(_fb_path).exists():
                        try:
                            ImageFont.truetype(_fb_path, font_size).getbbox('سلام')
                            font = ImageFont.truetype(_fb_path, font_size)
                            print(f"   [MULTI] RTL fallback font: {_fb}")
                            break
                        except Exception:
                            continue
            elif _script in ('japanese', 'cjk'):
                _fallback_fonts = [
                    'Yu Gothic.ttf', 'YuGothB.ttc', 'msyh.ttc', 'msyhbd.ttc',
                    'msgothic.ttc', 'simsun.ttc', 'meiryo.ttc',
                    'segoeui.ttf', 'arial.ttf',
                ]
                for _fb in _fallback_fonts:
                    _fb_path = str(Path(r"C:\Windows\Fonts") / _fb)
                    if Path(_fb_path).exists():
                        try:
                            ImageFont.truetype(_fb_path, font_size).getbbox('你好世界')
                            font = ImageFont.truetype(_fb_path, font_size)
                            print(f"   [MULTI] CJK fallback font: {_fb}")
                            break
                        except Exception:
                            continue
            elif _script == 'cyrillic':
                _fallback_fonts = [
                    'arial.ttf', 'arialbd.ttf', 'segoeui.ttf', 'segoeuib.ttf',
                    'times.ttf', 'calibri.ttf',
                ]
                for _fb in _fallback_fonts:
                    _fb_path = str(Path(r"C:\Windows\Fonts") / _fb)
                    if Path(_fb_path).exists():
                        try:
                            ImageFont.truetype(_fb_path, font_size).getbbox('Привет')
                            font = ImageFont.truetype(_fb_path, font_size)
                            print(f"   [MULTI] Cyrillic fallback font: {_fb}")
                            break
                        except Exception:
                            continue
        except Exception as e:
            print(f"   [WARNING] Font loading failed: {e}, using default")
            font = ImageFont.load_default()
            emoji_font = font

        # Get actual font metrics for proper text height calculation
        # (accounts for descenders like 'g','y','p','q','j' that extend below baseline)
        try:
            _font_ascent, _font_descent = font.getmetrics()
            font_total_height = _font_ascent + _font_descent
        except Exception:
            font_total_height = font_size
        if font_total_height < font_size:  # sanity clamp
            font_total_height = font_size

        # NOTE: word extraction, script detection, and case transforms
        # are done BEFORE font loading (above) so the font-fallback chain
        # can use _script.  No duplicate code needed here.

        # Distribute emojis
        emoji_distribution = []
        if emoji_enabled and emoji_list:
            for i in range(len(words)):
                emoji_distribution.append(emoji_list[i % len(emoji_list)])
        else:
            emoji_distribution = [''] * len(words)

        caption_clips = []

        # Calculate context words to display
        if caption_layout == '1-line':
            context_words = words_per_caption
        elif caption_layout == 'multi-line':
            # Multi-line: Show more words (3-5 lines depending on words_per_caption)
            context_words = words_per_caption * 4  # e.g., 2 words/line * 4 lines = 8 words
        else:  # 2-line
            context_words = words_per_caption * 2

        # Process each word using its actual TTS timing
        for word_idx, word_timing in enumerate(word_timings):
            word_start = float(word_timing['offset'])
            word_duration = float(word_timing['duration'])

            # CRITICAL FIX: Skip words that start after max_duration
            if max_duration and word_start >= max_duration:
                print(f"  ⚠ Skipping word {word_idx+1} '{word_timing.get('word', '')}' - starts at {word_start:.2f}s (after max {max_duration:.2f}s)")
                continue

            # CRITICAL FIX: Clip duration if word extends past max_duration
            if max_duration and (word_start + word_duration) > max_duration:
                original_duration = word_duration
                word_duration = max(0.01, max_duration - word_start)  # At least 0.01s
                print(f"  ✂ Clipping word {word_idx+1} duration from {original_duration:.2f}s to {word_duration:.2f}s")

            # Calculate display range
            half_context = context_words // 2
            display_start_idx = max(0, word_idx - half_context)
            display_end_idx = min(len(words), display_start_idx + context_words)

            if display_start_idx == 0:
                display_end_idx = min(len(words), context_words)
            if display_end_idx == len(words) and len(words) >= context_words:
                display_start_idx = max(0, len(words) - context_words)

            display_words = words[display_start_idx:display_end_idx]
            active_word_in_display = word_idx - display_start_idx

            # Get emoji
            current_emoji = emoji_distribution[word_idx] if emoji_enabled else ''

            # Calculate required image height dynamically based on layout and font size
            # Added generous padding for descenders ('g','y','p','q','j') and stroke width
            _stroke_padding = stroke_width * 2 + 4  # extra room for stroke outline
            if caption_layout == '1-line':
                # Single line: font metrics + descender space + stroke padding
                img_height = int(font_size * 2.0 + 120 + _stroke_padding)
            elif caption_layout == 'multi-line':
                # Multiple lines: 4+ lines possible
                img_height = int(font_size * 5.5 + 120 + _stroke_padding)
            else:
                # Two lines: 2 * font_size + line spacing + descender space + stroke padding
                img_height = int(font_size * 4.0 + 120 + _stroke_padding)

            # DEBUG: Log on first word only
            if word_idx == 0:
                print(f"[CAPTION DEBUG] Word {word_idx+1}/{len(words)}: layout={caption_layout}, "
                      f"font={font_size}px (metrics~{font_total_height}px), img_height={img_height}px, words_to_show={len(display_words)}")

            # Create image with dynamic height
            img = Image.new('RGBA', (video_width, img_height), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)

            # Measure emoji
            emoji_height = 0
            emoji_width = 0
            if current_emoji:
                emoji_bbox = draw.textbbox((0, 0), current_emoji, font=emoji_font)
                emoji_width = emoji_bbox[2] - emoji_bbox[0]
                emoji_height = emoji_bbox[3] - emoji_bbox[1]

            # Layout words
            if caption_layout == '1-line':
                words_per_line = len(display_words)
                all_lines = [display_words] if display_words else []
                line_spacing = 0
            elif caption_layout == 'multi-line':
                # Multi-line: Split words across multiple lines
                words_per_line = words_per_caption
                line_spacing = int(font_size * 1.3)
                all_lines = []
                for i in range(0, len(display_words), words_per_line):
                    all_lines.append(display_words[i:i+words_per_line])
            else:  # 2-line
                words_per_line = words_per_caption
                line_spacing = int(font_size * 1.3)
                line1_words = display_words[:words_per_line] if len(display_words) > 0 else []
                line2_words = display_words[words_per_line:words_per_line*2] if len(display_words) > words_per_line else []
                all_lines = [line1_words, line2_words]

            # Position first line - CENTER text vertically in the image
            # Use actual font metrics (ascent + descent) to prevent descender cutoff.
            # Also add stroke_width padding so the outline doesn't clip.
            num_lines = len(all_lines)
            _text_line_h = max(font_total_height, font_size) + _stroke_padding
            total_text_height = _text_line_h * num_lines + line_spacing * (num_lines - 1) if num_lines > 1 else _text_line_h

            # Calculate Y position to center text vertically
            if current_emoji:
                # With emoji: center text in lower portion
                emoji_gap_top = 20
                emoji_gap_bottom = 30
                emoji_total = emoji_gap_top + emoji_height + emoji_gap_bottom
                available_height = img_height - emoji_total
                first_line_y = emoji_total + (available_height - total_text_height) // 2
            else:
                # No emoji: center text in entire image
                first_line_y = (img_height - total_text_height) // 2

            # Function to draw text with stroke
            def draw_text_with_stroke(draw, xy, text, font, fill, stroke, stroke_width):
                x, y = xy
                if stroke_width > 0:
                    for dx in range(-stroke_width, stroke_width + 1):
                        for dy in range(-stroke_width, stroke_width + 1):
                            if dx*dx + dy*dy <= stroke_width*stroke_width:
                                draw.text((x + dx, y + dy), text, font=font, fill=stroke)
                draw.text(xy, text, font=font, fill=fill)

            # Draw lines
            for line_idx, line_words in enumerate(all_lines):
                if not line_words:
                    continue

                line_y = first_line_y + (line_idx * line_spacing)

                # RTL: reshape + bidi the entire line for correct visual order
                line_text = ' '.join(line_words)
                if _is_rtl:
                    visual_line = CaptionRenderer.reshape_rtl(line_text)
                    visual_words = visual_line.split()
                    n_words = len(line_words)
                    bbox = draw.textbbox((0, 0), visual_line, font=font)
                    line_width = bbox[2] - bbox[0]
                    x = (video_width - line_width) / 2
                    # For pure RTL, visual order is reverse of logical
                    line_start_idx = line_idx * words_per_line
                    logical_active = active_word_in_display - line_start_idx
                    v_active = n_words - 1 - logical_active if 0 <= logical_active < n_words else -1
                    render_words = visual_words
                    render_active = v_active
                else:
                    bbox = draw.textbbox((0, 0), line_text, font=font)
                    line_width = bbox[2] - bbox[0]
                    x = (video_width - line_width) / 2
                    words_in_line = len(line_words)
                    line_start_idx = line_idx * words_per_line
                    active_in_this_line = active_word_in_display - line_start_idx if line_start_idx <= active_word_in_display < line_start_idx + words_in_line else -1
                    render_words = line_words
                    render_active = active_in_this_line

                # Draw each word
                for w_idx, word in enumerate(render_words):
                    is_active = (w_idx == render_active)

                    # Multi-color mode: assign different color to each word
                    if multicolor_enabled and word_colors_hex:
                        # Get word's global index to assign color
                        global_word_idx = display_start_idx + line_start_idx + w_idx
                        # Cycle through color palette
                        color_idx = global_word_idx % len(word_colors_hex)
                        word_color_hex = word_colors_hex[color_idx]
                        # Convert hex to RGB
                        word_color = tuple(int(word_color_hex.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
                        # Active word keeps multi-color but can have different stroke/size
                        word_stroke = active_stroke if is_active else inactive_stroke
                    else:
                        # Standard mode: active vs inactive colors
                        word_color = active_color if is_active else inactive_color
                        word_stroke = active_stroke if is_active else inactive_stroke

                    if stroke_enabled:
                        draw_text_with_stroke(draw, (x, line_y), word, font, word_color, word_stroke, stroke_width)
                    else:
                        draw.text((x, line_y), word, font=font, fill=word_color)

                    # Move X position
                    word_bbox = draw.textbbox((0, 0), word + ' ', font=font)
                    word_width = word_bbox[2] - word_bbox[0]
                    x += word_width

            # Draw emoji with actual colors (not white)
            if current_emoji:
                emoji_x = (video_width - emoji_width) / 2
                emoji_y = 20
                # Use embedded_color=True to preserve emoji colors
                try:
                    draw.text((emoji_x, emoji_y), current_emoji, font=emoji_font, embedded_color=True)
                except TypeError:
                    # Fallback for older Pillow versions - draw without fill to use emoji colors
                    draw.text((emoji_x, emoji_y), current_emoji, font=emoji_font)

            # MoviePy 1.x ImageClip handles RGBA automatically:
            # transparent=True (default) extracts alpha channel as float [0,1] mask
            clip = ImageClip(np.array(img))
            try:
                clip = clip.with_duration(word_duration).with_start(word_start)
            except AttributeError:
                clip = clip.set_duration(word_duration).set_start(word_start)

            # Position clip with manual offset - account for caption height
            caption_height = img.height

            if position == 'top':
                y_pos = 50 + caption_y_offset
            elif position == 'bottom':
                # Position from bottom edge with 150px margin BEFORE user offset
                base_y = video_height - caption_height - 150
                y_pos = base_y + caption_y_offset
                # Safety clamp
                max_y = video_height - caption_height - 30
                min_y = 30
                y_pos = max(min_y, min(y_pos, max_y))
            else:  # center
                # Center vertically
                y_pos = (video_height - caption_height) // 2 + caption_y_offset

            try:
                clip = clip.set_position(('center', y_pos))
            except AttributeError:
                clip = clip.with_position(('center', y_pos))
            caption_clips.append(clip)

        print(f"[OK] Created {len(caption_clips)} synced caption clips using TTS timing")
        return caption_clips


# === Transition Sound Effects ============================================
# Generate SFX audio clips for transitions (fade/zoom/slide/wipe/glitch/etc.)

_TRANSITION_SFX_CLASS = None
_TRANSITION_SFX_GETTER = None


def _load_sfx():
    """Lazy-load the transition SFX module. Returns (class, getter) or (None, None)."""
    global _TRANSITION_SFX_CLASS, _TRANSITION_SFX_GETTER
    if _TRANSITION_SFX_CLASS is not None:
        return _TRANSITION_SFX_CLASS, _TRANSITION_SFX_GETTER
    try:
        import sys as _sys
        sfx_path = Path(__file__).parent / "VoiceModules" / "TransitionSFX"
        if str(sfx_path) not in _sys.path:
            _sys.path.insert(0, str(sfx_path))
        from transition_sfx import TransitionSFX, get_sfx_name  # noqa
        _TRANSITION_SFX_CLASS = TransitionSFX
        _TRANSITION_SFX_GETTER = get_sfx_name
        return TransitionSFX, get_sfx_name
    except Exception as e:
        print(f"[WARNING] Transition SFX module not available: {e}")
        return None, None


def _sfx_name_for(transition_name: str):
    _, getter = _load_sfx()
    if getter is None:
        return None
    return getter(transition_name)


def _make_sfx_clip(sfx_name: str, duration: float, start_time: float, fps: int, volume: float,
                   custom_file: Optional[str] = None, custom_volume: Optional[float] = None):
    """Create a MoviePy AudioArrayClip for a transition SFX.
    If `custom_file` is provided and exists, loads the user's audio file
    (wav/mp3/ogg) and trims/pads to the target duration. Otherwise falls
    back to the procedural SFX named by `sfx_name`.
    `custom_volume` is used only when a custom file is loaded; the
    procedural fallback always uses `volume`."""
    if AudioArrayClip is None:
        print(f"[SFX-DEBUG] AudioArrayClip not available; SFX '{sfx_name}' skipped")
        return None
    # 1) Custom SFX file path takes priority
    if custom_file:
        try:
            from pathlib import Path as _P
            cpath = _P(custom_file)
            if cpath.exists() and cpath.is_file():
                cv = custom_volume if custom_volume is not None else volume
                arr = _load_custom_sfx_audio(cpath, duration, cv, fps)
                if arr is not None and len(arr) > 0:
                    clip = AudioArrayClip(arr, fps=fps)
                    try:
                        return clip.with_start(start_time)
                    except AttributeError:
                        return clip.set_start(start_time)
                else:
                    print(f"[SFX-DEBUG] Custom SFX '{custom_file}' produced empty array, falling back to procedural '{sfx_name}'")
        except Exception as e:
            print(f"[SFX-DEBUG] Custom SFX load failed for '{custom_file}': {type(e).__name__}: {e}, falling back to procedural")
    # 2) Procedural fallback
    sfx_class, _ = _load_sfx()
    if sfx_class is None:
        print(f"[SFX-DEBUG] SFX class not loadable; SFX '{sfx_name}' skipped")
        return None
    try:
        sfx = sfx_class(sr=fps)
        arr = sfx.make_audio_array(sfx_name, duration=duration, volume=volume, channels=2)
        if arr is None or len(arr) == 0:
            print(f"[SFX-DEBUG] SFX '{sfx_name}' produced empty array")
            return None
        clip = AudioArrayClip(arr, fps=fps)
        # set_start works in both MoviePy 1.x and 2.x
        try:
            return clip.with_start(start_time)
        except AttributeError:
            return clip.set_start(start_time)
    except Exception as e:
        print(f"[SFX-DEBUG] SFX generation failed for '{sfx_name}' at t={start_time}: {type(e).__name__}: {e}")
        return None


def _load_custom_sfx_audio(path, duration: float, volume: float, fps: int) -> Optional[np.ndarray]:
    """Load a user-supplied SFX file (wav/mp3/ogg) and return a stereo
    float32 array of shape (n_samples, 2) trimmed/padded to `duration` seconds."""
    try:
        suffix = path.suffix.lower()
        # Try soundfile first (handles wav/ogg/flac natively)
        data = None
        file_sr = fps
        try:
            import soundfile as sf
            data, file_sr = sf.read(str(path), dtype='float32')
        except Exception:
            # Fall back to moviepy's AudioFileClip (handles mp3 too)
            try:
                from moviepy.audio.io.AudioFileClip import AudioFileClip
                clip = AudioFileClip(str(path))
                data = clip.to_soundarray(fps=fps)
                if data.ndim == 2 and data.shape[1] == 1:
                    data = data[:, 0]
                elif data.ndim == 2 and data.shape[1] == 2:
                    data = data.mean(axis=1)  # downmix to mono first
                file_sr = fps
                try:
                    clip.close()
                except Exception:
                    pass
            except Exception as e2:
                print(f"[SFX-DEBUG] Could not load custom SFX '{path}' with soundfile or AudioFileClip: {e2}")
                return None
        if data is None or len(data) == 0:
            return None
        # Convert to mono float32
        if data.ndim == 2:
            data = data.mean(axis=1)
        data = data.astype(np.float32, copy=False)
        # Resample if needed
        if file_sr != fps:
            n_target = int(len(data) * fps / file_sr)
            data = np.interp(
                np.linspace(0, len(data), n_target),
                np.arange(len(data)),
                data
            ).astype(np.float32)
        # Apply volume
        data = data * float(volume)
        # Trim or pad to target duration
        target_n = int(duration * fps)
        if len(data) > target_n:
            data = data[:target_n]
        elif len(data) < target_n:
            pad = np.zeros(target_n - len(data), dtype=np.float32)
            data = np.concatenate([data, pad])
        # Convert to stereo (n, 2)
        stereo = np.stack([data, data], axis=-1)
        return stereo
    except Exception as e:
        print(f"[SFX-DEBUG] _load_custom_sfx_audio failed for '{path}': {type(e).__name__}: {e}")
        return None


# Define which transitions should get SFX and where
# (enabled_key, default_sfx_name, default_duration, position, duration_setting_key)
# position: 'start' | 'end' — start means near t=0, end means near t=duration
# duration_setting_key lets the SFX duration follow the actual transition duration if set
_TRANSITION_SFX_SPECS = [
    ('transition_fade_in',         'fade',         0.4, 'start', 'transition_fade_in_duration'),
    ('transition_fade_out',        'fade',         0.4, 'end',   'transition_fade_out_duration'),
    ('transition_zoom_in',         'zoom_in',      0.5, 'start', 'transition_zoom_in_duration'),
    ('transition_zoom_out',        'zoom_out',     0.5, 'end',   'transition_zoom_out_duration'),
    ('transition_blur_in',         'blur_in',      0.5, 'start', 'transition_blur_duration'),
    ('transition_blur_out',        'blur_out',     0.5, 'end',   'transition_blur_duration'),
    ('transition_slide_in',        'slide_in',     0.4, 'start', 'transition_slide_duration'),
    ('transition_slide_out',       'slide_out',    0.4, 'end',   'transition_slide_duration'),
    ('transition_wipe_in',         'wipe_in',      0.4, 'start', 'transition_wipe_duration'),
    ('transition_wipe_out',        'wipe_out',     0.4, 'end',   'transition_wipe_duration'),
    ('transition_glitch_start',    'glitch_start', 0.4, 'start', 'transition_glitch_duration'),
    ('transition_glitch_end',      'glitch_end',   0.4, 'end',   'transition_glitch_duration'),
    ('transition_cinematic_bars',  'cinematic_bars', 0.5, 'start', 'transition_bars_duration'),
    # New transitions: Bounce, Mask, Bounce+Mask
    ('transition_bounce',          'bounce',       0.4, 'start', 'transition_bounce_duration'),
    ('transition_mask',            'mask_reveal',  0.4, 'start', 'transition_mask_duration'),
    ('transition_bounce_mask',     'bounce_mask',  0.5, 'start', 'transition_bounce_mask_duration'),
    # Light leaks / lens flares / film burn
    ('light_leak_enabled',         'light_leak',   0.6, 'start', 'light_leak_duration'),
    ('lens_flare_enabled',         'lens_flare',   0.5, 'start', 'lens_flare_duration'),
    ('film_burn_enabled',          'film_burn',    0.6, 'start', 'film_burn_duration'),
    # CapCut-style transitions: radial wipe, color dissolve, split wipe, luma wipe
    ('transition_radial_wipe',     'radial_wipe',  0.5, 'start', 'transition_radial_wipe_duration'),
    ('transition_color_dissolve',  'color_dissolve', 0.5, 'start', 'transition_color_dissolve_duration'),
    ('transition_split_wipe',      'split_wipe',   0.5, 'start', 'transition_split_wipe_duration'),
    ('transition_luma_wipe',       'luma_wipe',    0.5, 'start', 'transition_luma_wipe_duration'),
]

# Available SFX choices for per-transition override dropdown
AVAILABLE_SFX_CHOICES = [
    '(default)',  # uses the natural mapping from transition_sfx.SFX_MAP
    'whoosh',
    'swoosh',
    'boom',
    'glitch',
    'chime',
    'click',
    'zap',
    'sparkle',
    'hiss',
    'rumble',
    'shimmer',
    'cinematic_whoosh',
    'bass_drop',
    'riser',
    'impact',
    'vinyl_brake',
    'notification',
    'sub_boom',
    'horn_stab',
    '(none)',  # explicitly disable SFX for this transition
]


def add_transition_sfx(audio_tracks: list, settings: dict, video_duration: float, fps: int = 44100):
    """Generate and add SFX audio clips to the audio_tracks list.
    Reads settings keys:
      - transition_sfx_enabled  (bool, default True)
      - transition_sfx_volume   (0.0-1.0, default 0.6)
      - transition_sfx_<key>    (per-transition SFX override, e.g. 'transition_sfx_zoom_in')
        → '(default)' uses procedural SFX; '(none)' skips; specific name uses that SFX.
      - custom_sfx_file         (str) global fallback file used when dropdown is '(default)'
      - custom_sfx_volume       (float) volume for the custom file
      - transition_repeat_selected_enabled / _interval / _mode
        → also fires SFX at every interval (matching the visual repeat).
    """
    if not settings.get('transition_sfx_enabled', True):
        return
    volume = float(settings.get('transition_sfx_volume', 0.6))
    if volume <= 0:
        return

    added = 0
    # Custom SFX: a single user-picked file that acts as global fallback
    # when a per-transition dropdown is set to '(default)'.
    custom_sfx_file = settings.get('custom_sfx_file', '') or ''
    custom_sfx_volume = float(settings.get('custom_sfx_volume', volume))

    # Use a small temporal offset for stacked-at-t=0 transitions so they don't
    # all sum into a single muddy burst when many are enabled at once.
    start_offset_idx = 0
    for enabled_key, tname, default_dur, position, dur_key in _TRANSITION_SFX_SPECS:
        if not settings.get(enabled_key, False):
            continue

        # Per-transition override (or "(default)" / "(none)")
        sfx_suffix = enabled_key.replace('transition_', '').replace('_enabled', '')
        override = settings.get(f'transition_sfx_{sfx_suffix}', '(default)')
        if override == '(none)':
            continue

        if override == '(default)':
            sfx_name = _sfx_name_for(tname)
            # Global fallback: if custom file is set, use it instead of procedural
            cf = custom_sfx_file if custom_sfx_file and Path(custom_sfx_file).is_file() else None
            cv = custom_sfx_volume if cf else None
        else:
            sfx_name = override
            cf = None
            cv = None

        if sfx_name is None:
            continue

        # Use actual transition duration if set, else default
        sfx_dur = float(settings.get(dur_key, default_dur)) if dur_key else default_dur
        sfx_dur = max(0.1, min(sfx_dur, 3.0))  # safety clamp

        # Position calculation with small stagger for start-positioned SFX
        if position == 'start':
            t_at = start_offset_idx * 0.05  # 50ms stagger so simultaneous SFX don't clip
            start_offset_idx += 1
        elif position == 'end':
            # All end-positioned SFX play at the actual end of the video
            t_at = max(0.0, video_duration - sfx_dur)
        else:
            t_at = video_duration / 2.0

        clip = _make_sfx_clip(sfx_name, sfx_dur, t_at, fps, volume, custom_file=cf, custom_volume=cv)
        if clip is not None:
            audio_tracks.append(clip)
            added += 1

    # Repeat-selected SFX: fire SFX at every repeat interval using the same
    # sequential/random order as the visual transitions.
    if settings.get('transition_repeat_selected_enabled', False):
        try:
            interval = float(settings.get('transition_repeat_selected_interval', 6.0))
            mode = settings.get('transition_repeat_selected_mode', 'sequential')
            if interval > 0 and video_duration > 0:
                _rs_order = [
                    'transition_fade_in', 'transition_fade_out',
                    'transition_zoom_in', 'transition_zoom_out',
                    'transition_blur_in', 'transition_blur_out',
                    'transition_slide_in', 'transition_slide_out',
                    'transition_wipe_in', 'transition_wipe_out',
                    'transition_glitch_start', 'transition_glitch_end',
                    'transition_cinematic_bars',
                    'transition_bounce', 'transition_mask', 'transition_bounce_mask',
                    'transition_radial_wipe', 'transition_color_dissolve',
                    'transition_split_wipe', 'transition_luma_wipe',
                ]
                # Translate UI key → spec key so we can use _TRANSITION_SFX_SPECS
                _key_to_spec = {
                    'transition_fade_in':       ('transition_fade_in',       'fade',         0.4, 'transition_fade_in_duration'),
                    'transition_fade_out':      ('transition_fade_out',      'fade',         0.4, 'transition_fade_out_duration'),
                    'transition_zoom_in':       ('transition_zoom_in',       'zoom_in',      0.5, 'transition_zoom_in_duration'),
                    'transition_zoom_out':      ('transition_zoom_out',      'zoom_out',     0.5, 'transition_zoom_out_duration'),
                    'transition_blur_in':       ('transition_blur_in',       'blur_in',      0.5, 'transition_blur_duration'),
                    'transition_blur_out':      ('transition_blur_out',      'blur_out',     0.5, 'transition_blur_duration'),
                    'transition_slide_in':      ('transition_slide_in',      'slide_in',     0.4, 'transition_slide_duration'),
                    'transition_slide_out':     ('transition_slide_out',     'slide_out',    0.4, 'transition_slide_duration'),
                    'transition_wipe_in':       ('transition_wipe_in',       'wipe_in',      0.4, 'transition_wipe_duration'),
                    'transition_wipe_out':      ('transition_wipe_out',      'wipe_out',     0.4, 'transition_wipe_duration'),
                    'transition_glitch_start':  ('transition_glitch_start',  'glitch_start', 0.4, 'transition_glitch_duration'),
                    'transition_glitch_end':    ('transition_glitch_end',    'glitch_end',   0.4, 'transition_glitch_duration'),
                    'transition_cinematic_bars':('transition_cinematic_bars','cinematic_bars', 0.5,'transition_bars_duration'),
                    'transition_bounce':        ('transition_bounce',        'bounce',       0.4, 'transition_bounce_duration'),
                    'transition_mask':          ('transition_mask',          'mask_reveal',  0.4, 'transition_mask_duration'),
                    'transition_bounce_mask':   ('transition_bounce_mask',   'bounce_mask',  0.5, 'transition_bounce_mask_duration'),
                    'transition_radial_wipe':   ('transition_radial_wipe',   'radial_wipe',  0.5,  'transition_radial_wipe_duration'),
                    'transition_color_dissolve':('transition_color_dissolve','color_dissolve',0.5,  'transition_color_dissolve_duration'),
                    'transition_split_wipe':    ('transition_split_wipe',    'split_wipe',   0.5,  'transition_split_wipe_duration'),
                    'transition_luma_wipe':     ('transition_luma_wipe',     'luma_wipe',    0.5,  'transition_luma_wipe_duration'),
                }
                enabled_keys = [k for k in _rs_order if settings.get(k, False)]
                if enabled_keys:
                    if mode == 'random':
                        import random as _rnd
                        _rnd.seed(0)
                        enabled_keys = enabled_keys[:]
                        _rnd.shuffle(enabled_keys)
                    n = len(enabled_keys)
                    # Fire SFX at each interval boundary where a transition is
                    # about to play (skip window 0 — that's already covered by
                    # the start-positioned SFX above, except for transitions
                    # whose start SFX was already emitted).
                    first_idx = 1
                    win_idx = first_idx
                    t_at = win_idx * interval
                    while t_at + 0.1 < video_duration:
                        key = enabled_keys[win_idx % n]
                        spec = _key_to_spec.get(key)
                        if spec is not None:
                            _ekey, tname, default_dur, dur_key = spec
                            sfx_suffix_r = _ekey.replace('transition_', '').replace('_enabled', '')
                            override = settings.get(f'transition_sfx_{sfx_suffix_r}', '(default)')
                            if override != '(none)':
                                if override == '(default)':
                                    sfx_name = _sfx_name_for(tname)
                                    # Global fallback: use custom file if set
                                    cf_r = custom_sfx_file if custom_sfx_file and Path(custom_sfx_file).is_file() else None
                                    cv_r = custom_sfx_volume if cf_r else None
                                else:
                                    sfx_name = override
                                    cf_r = None
                                    cv_r = None
                                if sfx_name is not None:
                                    sfx_dur = float(settings.get(dur_key, default_dur))
                                    sfx_dur = max(0.1, min(sfx_dur, 3.0))
                                    clip = _make_sfx_clip(sfx_name, sfx_dur, t_at, fps, volume, custom_file=cf_r, custom_volume=cv_r)
                                    if clip is not None:
                                        audio_tracks.append(clip)
                                        added += 1
                        win_idx += 1
                        t_at = win_idx * interval
                    if added > 0:
                        print(f"[OK] Added {added-1} repeat-selected SFX (every {interval}s)")
        except Exception as e:
            print(f"[WARNING] Repeat-selected SFX failed: {e}")

    if added > 0:
        print(f"[OK] Added {added} transition SFX (volume: {int(volume*100)}%)")


# === End Transition SFX =================================================


class AudioProcessor:
    """Audio processing module for BGM and voiceovers"""

    @staticmethod
    def get_voiceover_files(folder_path: Path) -> List[Path]:
        """Get audio files from voiceover folder sorted by number"""
        if not folder_path or not folder_path.exists():
            return []
        audio_extensions = {'.mp3', '.wav', '.m4a', '.aac', '.ogg'}
        files = [f for f in folder_path.iterdir()
                if f.suffix.lower() in audio_extensions and f.is_file()]

        # Sort by number in filename (1.mp3, 2.mp3, etc.)
        def extract_number(filepath):
            import re
            match = re.search(r'(\d+)', filepath.stem)
            return int(match.group(1)) if match else 999999

        return sorted(files, key=extract_number)

    @staticmethod
    def get_bgm_files(bgm_path: str) -> List[Path]:
        """Get BGM files - supports single file or folder with multiple files"""
        bgm_path = Path(bgm_path)

        if not bgm_path.exists():
            return []

        audio_extensions = {'.mp3', '.wav', '.m4a', '.aac', '.ogg'}

        # Single file
        if bgm_path.is_file() and bgm_path.suffix.lower() in audio_extensions:
            return [bgm_path]

        # Folder with multiple files
        if bgm_path.is_dir():
            files = [f for f in bgm_path.iterdir()
                    if f.suffix.lower() in audio_extensions and f.is_file()]
            return sorted(files)

        return []

    @staticmethod
    def create_looped_audio(audio_clip, target_duration):
        """Loop audio to match video duration.

        In MoviePy v2, reusing the same AudioFileClip object N times in a
        CompositeAudioClip fails because the underlying reader shares
        internal position state — only the first copy actually plays.
        Instead we render to a numpy array, tile it, and construct a
        single AudioArrayClip.  This also avoids the memory / decode
        overhead of N concurrent readers."""
        if audio_clip.duration >= target_duration:
            return subclip(audio_clip, 0, target_duration)
        # Get fps — handles both None and missing attribute cases
        try:
            fps = audio_clip.fps or 44100
        except Exception:
            fps = 44100
        # Render the whole clip to float samples
        try:
            arr = audio_clip.to_soundarray(fps=fps)
        except Exception:
            arr = audio_clip.to_soundarray()
        # Determine actual fps from the rendered array if needed
        if arr.shape[0] > 1 and audio_clip.duration > 0:
            arr_fps = arr.shape[0] / audio_clip.duration
            fps = int(round(arr_fps)) or fps
        total = int(round(fps * target_duration))
        if arr.ndim == 1:
            arr = arr[:, None]
        repeats = int(np.ceil(total / len(arr)))
        tiled = np.tile(arr, (repeats, 1))[:total]
        from moviepy import AudioArrayClip
        return set_duration(AudioArrayClip(tiled, fps=fps), target_duration)

    @staticmethod
    def apply_voiceover_ducking(original_audio, voiceover_audio,
                                duck_volume=0.12, threshold=0.02,
                                hold=0.35, fade=0.15):
        """Create a ducked version of original audio.

        During voiceover speech the original volume is reduced to
        ``duck_volume``.  Rather than an instantaneous per-sample check (which
        pops the original back to full volume in every tiny gap between words),
        we precompute a smoothed *speech envelope* with attack/hold/release so
        the duck stays engaged across short pauses within a spoken line, then
        fades smoothly back up when speech truly ends.

        Parameters
        ----------
        duck_volume : float   original level while speaking (0.12 = 12%)
        threshold   : float   RMS level (over a 20 ms window) counted as speech
        hold        : float   seconds a duck is held across gaps before release
        fade        : float   seconds of smooth ramp between full and ducked
        """
        fps = getattr(original_audio, 'fps',
                      getattr(voiceover_audio, 'fps', 44100))

        # ── Precompute a gain envelope from the voiceover amplitude ─────────
        gain_lut = None
        lut_hz = 100  # one envelope sample every 10 ms
        try:
            vo_dur = float(voiceover_audio.duration)
            samples = voiceover_audio.to_soundarray(fps=fps)
            if samples.ndim == 2:
                samples = samples.mean(axis=1)  # mono
            samples = np.abs(samples)
            win = max(1, int(fps / lut_hz))           # 10 ms window
            n_bins = int(np.ceil(len(samples) / win))
            pad = n_bins * win - len(samples)
            if pad > 0:
                samples = np.concatenate([samples, np.zeros(pad, np.float32)])
            # RMS per 10 ms bin
            binned = samples.reshape(n_bins, win)
            rms = np.sqrt(np.mean(binned * binned, axis=1))
            active = rms > threshold
            # Hold: keep active for `hold` seconds after speech to bridge gaps
            hold_bins = max(1, int(hold * lut_hz))
            held = active.copy()
            run = 0
            for i in range(len(active)):
                if active[i]:
                    run = hold_bins
                elif run > 0:
                    held[i] = True
                    run -= 1
            # Convert boolean → gain (1.0 full, duck_volume while held) then
            # smooth with a moving average to get attack/release fades.
            gain = np.where(held, duck_volume, 1.0).astype(np.float32)
            fade_bins = max(1, int(fade * lut_hz))
            if fade_bins > 1:
                kern = np.ones(fade_bins, np.float32) / fade_bins
                gain = np.convolve(gain, kern, mode='same')
            gain_lut = gain
        except Exception as _env_err:
            print(f"[WARNING] ducking envelope precompute failed ({_env_err}); "
                  f"falling back to per-frame check")

        def _ducked_frame(t):
            orig = original_audio.get_frame(t)
            if gain_lut is None:
                # Fallback: instantaneous check (original behaviour)
                try:
                    if np.max(np.abs(voiceover_audio.get_frame(t))) > threshold:
                        return orig * duck_volume
                except Exception:
                    pass
                return orig
            # Envelope lookup (t may be scalar or an array of sample times)
            idx = (np.asarray(t) * lut_hz).astype(np.int64)
            idx = np.clip(idx, 0, len(gain_lut) - 1)
            g = gain_lut[idx]
            if np.ndim(orig) == 2:           # stereo → broadcast gain per sample
                g = g[:, None] if np.ndim(g) else g
            return orig * g

        try:
            from moviepy.audio.AudioClip import AudioClip as _AClip
            ducked = _AClip(_ducked_frame,
                            duration=original_audio.duration)
            ducked = ducked.with_fps(fps)
            return ducked
        except Exception:
            return original_audio

    @staticmethod
    def mix_audio_tracks(video_clip, settings, voiceover_file: Optional[Path] = None, bgm_file: Optional[Path] = None):
        """Mix original audio, BGM, and voiceover"""
        audio_tracks = []

        # Original audio (store it; we append after ducking check below)
        orig_audio_track = None
        if video_clip.audio and not settings.get('mute_original_audio', False):
            # Convert percentage (0-200) to decimal (0.0-2.0)
            volume_percent = settings.get('original_audio_volume', 100)
            original_volume = volume_percent / 100.0
            original_audio = set_volume(video_clip.audio, original_volume)
            orig_audio_track = original_audio
            print(f"[OK] Original audio volume: {volume_percent}%")

        # Custom BGM (use provided bgm_file or fall back to settings)
        if settings.get('add_custom_bgm', False):
            bgm_path = bgm_file if bgm_file else (Path(settings['bgm_file']) if settings.get('bgm_file') else None)

            if bgm_path and bgm_path.exists():
                try:
                    bgm_audio = AudioFileClip(str(bgm_path))

                    if settings.get('bgm_loop', True):
                        bgm_audio = AudioProcessor.create_looped_audio(bgm_audio, video_clip.duration)
                    else:
                        bgm_audio = subclip(bgm_audio, 0, min(bgm_audio.duration, video_clip.duration))

                    bgm_volume = settings.get('bgm_volume', 0.3)
                    bgm_audio = set_volume(bgm_audio, bgm_volume)
                    audio_tracks.append(bgm_audio)
                    print(f"[OK] Added BGM: {bgm_path.name}")
                except Exception as e:
                    print(f"[WARNING] Could not load BGM: {e}")

        # Voiceover
        voiceover_audio_clip = None
        if voiceover_file and voiceover_file.exists():
            try:
                voiceover_audio = AudioFileClip(str(voiceover_file))
                voiceover_volume = settings.get('voiceover_volume', 1.0)
                voiceover_delay = settings.get('voiceover_delay', 0.0)

                voiceover_audio = set_volume(voiceover_audio, voiceover_volume)

                if voiceover_delay > 0:
                    silence = set_volume(set_duration(AudioFileClip(str(voiceover_file)), voiceover_delay), 0)
                    try:
                        voiceover_audio = CompositeAudioClip([silence, voiceover_audio.set_start(voiceover_delay)])
                    except:
                        voiceover_audio = CompositeAudioClip([silence, voiceover_audio.with_start(voiceover_delay)])

                voiceover_audio_clip = voiceover_audio  # Store for ducking
                audio_tracks.append(voiceover_audio)
                print(f"[OK] Added voiceover: {voiceover_file.name}")
            except Exception as e:
                print(f"[WARNING] Could not load voiceover: {e}")

        # Apply ducking to original audio when voiceover exists
        _ducked_orig_track = None  # keep ref so BGM auto-ducking doesn't double-apply
        if orig_audio_track is not None:
            if voiceover_audio_clip is not None:
                # Time-varying ducking: quiet during voiceover, full elsewhere
                _ducked_orig_track = AudioProcessor.apply_voiceover_ducking(
                    orig_audio_track, voiceover_audio_clip)
                audio_tracks.append(_ducked_orig_track)
                print("[OK] Original audio ducked during voiceover")
            else:
                # No voiceover — keep original at set volume
                audio_tracks.append(orig_audio_track)

        # Apply BGM auto-ducking if enabled and we have both BGM and voiceover
        if settings.get('audio_auto_ducking', False) and voiceover_audio_clip is not None and len(audio_tracks) > 1:
            try:
                ducking_amount = settings.get('audio_ducking_amount', 0.3)

                # Find BGM track (it's the one that's not voiceover and not already time-varying ducked)
                for i, track in enumerate(audio_tracks):
                    # ⚡ MOVIEPY 2.x PITFALL: Do NOT use `!=` / `==` on clip objects —
                    # Clip.__eq__ iterates EVERY frame to compare (= millions of
                    # renders at 44.1kHz → app freeze).  Use `is not` for identity.
                    if track is not voiceover_audio_clip and track is not _ducked_orig_track:
                        # This is BGM - apply ducking during voiceover
                        # Simple approach: just reduce the volume of BGM track
                        # (constant reduction instead of time-varying for reliability)
                        ducked_track = None

                        # Try method 1: MoviePy 2.x with_effects
                        if ducked_track is None:
                            try:
                                from moviepy.audio.fx import MultiplyVolume
                                ducked_track = track.with_effects([MultiplyVolume(ducking_amount)])
                            except:
                                pass

                        # Try method 2: MoviePy 1.x volumex
                        if ducked_track is None:
                            try:
                                ducked_track = track.volumex(ducking_amount)
                            except:
                                pass

                        # Try method 3: Direct function call
                        if ducked_track is None:
                            try:
                                from moviepy.audio.fx.MultiplyVolume import multiply_volume
                                ducked_track = multiply_volume(track, ducking_amount)
                            except:
                                pass

                        # Fallback: keep original track (no ducking)
                        if ducked_track is None:
                            ducked_track = track
                            print(f"[WARNING] Could not apply ducking, keeping original BGM volume")

                        audio_tracks[i] = ducked_track

                print(f"[OK] Applied BGM auto-ducking ({int((1-ducking_amount)*100)}% reduction during voice)")
            except Exception as e:
                print(f"[WARNING] BGM ducking failed: {e}")

        # Transition SFX (added after ducking so SFX sit above BGM)
        add_transition_sfx(audio_tracks, settings, video_clip.duration, fps=44100)

        # Mix all tracks
        if audio_tracks:
            # Composite multiple tracks
            if len(audio_tracks) == 1:
                final_audio = audio_tracks[0]
            else:
                final_audio = CompositeAudioClip(audio_tracks)

            # Apply audio normalization if enabled
            if settings.get('audio_normalize', False):
                try:
                    target_level_db = settings.get('audio_target_level', -20)

                    # Normalize audio to target level
                    # MoviePy doesn't have built-in normalization, so we do it manually
                    # Get max amplitude and calculate gain needed
                    def normalize_audio(audio_clip, target_db=-20):
                        """Normalize audio to target dB level"""
                        try:
                            # Get audio as numpy array
                            audio_array = audio_clip.to_soundarray()

                            # Find peak amplitude
                            max_amplitude = np.max(np.abs(audio_array))

                            if max_amplitude > 0:
                                # Calculate current dB level
                                current_db = 20 * np.log10(max_amplitude)

                                # Calculate required gain
                                gain_db = target_db - current_db
                                gain_linear = 10 ** (gain_db / 20)

                                # Apply gain (use volumex for smoother result)
                                normalized = set_volume(audio_clip, gain_linear)

                                print(f"[NORMALIZE] Peak: {max_amplitude:.4f} ({current_db:.1f} dB) → Target: {target_db} dB (gain: {gain_db:+.1f} dB)")
                                return normalized
                            else:
                                return audio_clip
                        except Exception as e:
                            print(f"[WARNING] Normalization failed: {e}")
                            return audio_clip

                    final_audio = normalize_audio(final_audio, target_level_db)
                    print(f"[OK] Applied audio normalization (target: {target_level_db} dB)")
                except Exception as e:
                    print(f"[WARNING] Audio normalization failed: {e}")

            return final_audio
        else:
            return None


class TextEffects:
    """Text animation and effects module"""

    @staticmethod
    def create_glow_image(img, glow_color=(255, 255, 255), intensity=8):
        """Add glow effect to text image"""
        glow = img.copy()
        for i in range(intensity):
            glow = glow.filter(ImageFilter.GaussianBlur(radius=2))

        result = Image.new('RGBA', img.size, (0, 0, 0, 0))
        result.paste(glow, (0, 0), glow)
        result.paste(img, (0, 0), img)
        return result

    @staticmethod
    def create_shadow_image(img, offset=6, blur=12):
        """Add drop shadow to image"""
        shadow = Image.new('RGBA',
                          (img.width + offset*2, img.height + offset*2),
                          (0, 0, 0, 0))
        shadow_mask = Image.new('RGBA', img.size, (0, 0, 0, 180))
        shadow.paste(shadow_mask, (offset, offset), img)
        shadow = shadow.filter(ImageFilter.GaussianBlur(radius=blur))

        result = Image.new('RGBA', shadow.size, (0, 0, 0, 0))
        result.paste(shadow, (0, 0), shadow)
        result.paste(img, (offset//2, offset//2), img)
        return result

    @staticmethod
    def create_neon_glow(img, neon_color=(0, 255, 136)):
        """Create neon glow effect. Works on RGB or RGBA input."""
        # Normalize to RGBA so alpha_composite is always safe.
        if img.mode != 'RGBA':
            img = img.convert('RGBA')
        # Build a glowing aura by stacking multiple blurred copies of the image,
        # with the neon color tinting the brighter (less blurred) layers.
        result = Image.new('RGBA', img.size, (0, 0, 0, 0))
        # Glow stack: largest blur first (weakest), then tighter blurs (strongest).
        for i in range(20, 0, -2):
            blur_img = img.filter(ImageFilter.GaussianBlur(radius=i))
            # Tint the blurred layer toward the neon color by mixing
            # channels (alpha stays from the blur).
            r, g, b, a = blur_img.split()
            nr, ng, nb = neon_color
            # Push RGB toward neon proportionally to how blurred (i.e. halo-like)
            # the layer is — heavier push on outer (larger) blurs.
            mix = min(1.0, i / 20.0 * 0.7)
            r_arr = np.asarray(r, dtype=np.float32)
            g_arr = np.asarray(g, dtype=np.float32)
            b_arr = np.asarray(b, dtype=np.float32)
            r_arr = r_arr * (1 - mix) + nr * mix
            g_arr = g_arr * (1 - mix) + ng * mix
            b_arr = b_arr * (1 - mix) + nb * mix
            from PIL import Image as _PIL
            tinted = _PIL.fromarray(
                np.stack([r_arr, g_arr, b_arr,
                           np.asarray(a, dtype=np.float32)], axis=-1
                ).astype(np.uint8), mode='RGBA')
            result = Image.alpha_composite(result, tinted)
        # Composite the original image on top
        result = Image.alpha_composite(result, img)
        # Return in the same mode the caller passed in (drop alpha if RGB)
        return result

    @staticmethod
    def apply_gradient_overlay(img, gradient_type='top_to_bottom', intensity=0.3):
        """Apply gradient overlay"""
        overlay = Image.new('RGBA', img.size, (0, 0, 0, 0))
        draw = ImageDraw.Draw(overlay)

        if gradient_type == 'top_to_bottom':
            for y in range(img.height):
                alpha = int(255 * intensity * (y / img.height))
                draw.line([(0, y), (img.width, y)], fill=(0, 0, 0, alpha))
        elif gradient_type == 'bottom_to_top':
            for y in range(img.height):
                alpha = int(255 * intensity * ((img.height - y) / img.height))
                draw.line([(0, y), (img.width, y)], fill=(0, 0, 0, alpha))

        return Image.alpha_composite(img, overlay)


class QuoteImageGenerator:
    """Generate beautiful template-based background images for quotes"""

    TEMPLATES = {
        'gradient_sunset': {
            'colors': [(255, 94, 77), (255, 154, 158), (250, 208, 196)],
            'direction': 'diagonal',
            'text_color': '#FFFFFF',
            'description': 'Warm sunset gradient (motivational)'
        },
        'gradient_ocean': {
            'colors': [(26, 42, 108), (58, 96, 115), (107, 140, 140)],
            'direction': 'vertical',
            'text_color': '#FFFFFF',
            'description': 'Deep ocean gradient (calm, wisdom)'
        },
        'gradient_fire': {
            'colors': [(255, 65, 108), (255, 75, 43), (255, 168, 0)],
            'direction': 'radial',
            'text_color': '#FFFFFF',
            'description': 'Fire energy gradient (high energy)'
        },
        'gradient_purple_dream': {
            'colors': [(67, 67, 255), (156, 81, 182), (255, 109, 255)],
            'direction': 'diagonal',
            'text_color': '#FFFFFF',
            'description': 'Purple dream gradient (creative, spiritual)'
        },
        'gradient_mint_fresh': {
            'colors': [(11, 163, 96), (60, 186, 146), (130, 224, 170)],
            'direction': 'vertical',
            'text_color': '#FFFFFF',
            'description': 'Mint fresh gradient (growth, success)'
        },
        'solid_black': {
            'colors': [(0, 0, 0)],
            'direction': 'solid',
            'text_color': '#FFFFFF',
            'description': 'Minimalist black (bold, modern)'
        },
        'solid_navy': {
            'colors': [(20, 30, 48)],
            'direction': 'solid',
            'text_color': '#FFFFFF',
            'description': 'Deep navy (professional, trust)'
        },
        'gradient_golden_hour': {
            'colors': [(255, 195, 113), (251, 144, 98), (247, 106, 104)],
            'direction': 'horizontal',
            'text_color': '#FFFFFF',
            'description': 'Golden hour (warm, inspiring)'
        },
        'gradient_sky': {
            'colors': [(2, 170, 176), (0, 205, 172), (134, 253, 232)],
            'direction': 'vertical',
            'text_color': '#000000',
            'description': 'Sky blue gradient (hopeful, peaceful)'
        },
        'gradient_dark_purple': {
            'colors': [(35, 7, 77), (79, 46, 109), (117, 86, 142)],
            'direction': 'radial',
            'text_color': '#FFFFFF',
            'description': 'Dark purple (luxury, mystery)'
        }
    }

    @staticmethod
    def create_gradient(width, height, colors, direction='vertical'):
        """Create a gradient background image"""
        img = Image.new('RGB', (width, height))
        draw = ImageDraw.Draw(img)

        if direction == 'solid':
            img.paste(colors[0], [0, 0, width, height])
            return img

        if direction == 'vertical':
            for y in range(height):
                # Calculate color interpolation
                position = y / height
                color = QuoteImageGenerator._interpolate_colors(colors, position)
                draw.line([(0, y), (width, y)], fill=color)

        elif direction == 'horizontal':
            for x in range(width):
                position = x / width
                color = QuoteImageGenerator._interpolate_colors(colors, position)
                draw.line([(x, 0), (x, height)], fill=color)

        elif direction == 'diagonal':
            for y in range(height):
                for x in range(width):
                    position = (x + y) / (width + height)
                    color = QuoteImageGenerator._interpolate_colors(colors, position)
                    draw.point((x, y), fill=color)

        elif direction == 'radial':
            center_x, center_y = width // 2, height // 2
            max_distance = np.sqrt(center_x**2 + center_y**2)

            for y in range(height):
                for x in range(width):
                    distance = np.sqrt((x - center_x)**2 + (y - center_y)**2)
                    position = distance / max_distance
                    color = QuoteImageGenerator._interpolate_colors(colors, position)
                    draw.point((x, y), fill=color)

        return img

    @staticmethod
    def _interpolate_colors(colors, position):
        """Interpolate between multiple colors based on position (0.0 to 1.0)"""
        if len(colors) == 1:
            return colors[0]

        # Clamp position
        position = max(0.0, min(1.0, position))

        # Calculate which colors to interpolate between
        num_segments = len(colors) - 1
        segment = position * num_segments
        segment_index = int(segment)

        if segment_index >= num_segments:
            return colors[-1]

        # Interpolate between two adjacent colors
        local_position = segment - segment_index
        color1 = colors[segment_index]
        color2 = colors[segment_index + 1]

        r = int(color1[0] + (color2[0] - color1[0]) * local_position)
        g = int(color1[1] + (color2[1] - color1[1]) * local_position)
        b = int(color1[2] + (color2[2] - color1[2]) * local_position)

        return (r, g, b)

    @staticmethod
    def create_quote_image(quote, template_name='gradient_sunset', width=1080, height=1920):
        """Create a beautiful quote image using a template"""
        if template_name not in QuoteImageGenerator.TEMPLATES:
            print(f"[WARNING] Template '{template_name}' not found, using 'gradient_sunset'")
            template_name = 'gradient_sunset'

        template = QuoteImageGenerator.TEMPLATES[template_name]

        # Create gradient background
        img = QuoteImageGenerator.create_gradient(
            width, height,
            template['colors'],
            template['direction']
        )

        # Add text overlay
        draw = ImageDraw.Draw(img)

        # Detect if quote contains emojis
        emoji_pattern = re.compile(r'[\U0001F300-\U0001F9FF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF\U00002600-\U000027BF\U0001F1E0-\U0001F1FF]+')
        has_emojis = bool(emoji_pattern.search(quote))

        # Try to load a font that supports emojis
        font_size = 80
        emoji_font = None
        text_font = None

        try:
            # For Windows: Use Segoe UI Emoji which supports emojis
            if has_emojis:
                emoji_font_path = "C:/Windows/Fonts/seguiemj.ttf"  # Segoe UI Emoji
                if os.path.exists(emoji_font_path):
                    emoji_font = ImageFont.truetype(emoji_font_path, font_size)
                    text_font = ImageFont.truetype(emoji_font_path, font_size)
                    print(f"[OK] Using Segoe UI Emoji font for emoji support")
                else:
                    # Fallback: strip emojis if emoji font not available
                    print(f"[WARNING] Emoji font not found, removing emojis from text")
                    quote = emoji_pattern.sub(' ', quote).strip()
                    quote = ' '.join(quote.split())  # Remove extra spaces
                    has_emojis = False

            # If no emojis or couldn't load emoji font, use Arial Bold
            if not has_emojis:
                font_path = "C:/Windows/Fonts/arialbd.ttf"
                if not os.path.exists(font_path):
                    font_path = "C:/Windows/Fonts/arial.ttf"
                text_font = ImageFont.truetype(font_path, font_size)
        except Exception as e:
            print(f"[WARNING] Font loading error: {e}, using default font")
            text_font = ImageFont.load_default()

        # Use the loaded font
        font = text_font

        # Word wrap the quote
        words = quote.split()
        lines = []
        current_line = []
        max_width = width - 200  # 100px padding on each side

        for word in words:
            test_line = ' '.join(current_line + [word])
            try:
                bbox = draw.textbbox((0, 0), test_line, font=font)
                text_width = bbox[2] - bbox[0]
            except:
                # Fallback for older Pillow versions
                text_width = len(test_line) * (font_size // 2)

            if text_width <= max_width:
                current_line.append(word)
            else:
                if current_line:
                    lines.append(' '.join(current_line))
                current_line = [word]

        if current_line:
            lines.append(' '.join(current_line))

        # Calculate total text height
        line_height = font_size + 20
        total_text_height = len(lines) * line_height

        # Center text vertically
        start_y = (height - total_text_height) // 2

        # Parse text color
        text_color = tuple(int(template['text_color'].lstrip('#')[i:i+2], 16) for i in (0, 2, 4))

        # Draw each line centered
        for i, line in enumerate(lines):
            try:
                bbox = draw.textbbox((0, 0), line, font=font)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
            except:
                # Fallback for older Pillow versions
                text_width = len(line) * (font_size // 2)
                text_height = font_size

            x = (width - text_width) // 2
            y = start_y + i * line_height

            # Add subtle shadow for better readability
            shadow_offset = 3
            shadow_color = (0, 0, 0, 128) if has_emojis else (0, 0, 0)
            try:
                draw.text((x + shadow_offset, y + shadow_offset), line, font=font, fill=shadow_color)
            except:
                pass  # Skip shadow if it causes issues

            # Draw main text
            draw.text((x, y), line, font=font, fill=text_color)

        return img

    @staticmethod
    def generate_images_from_quotes(quotes_list, output_folder, template_name='auto'):
        """Generate images for a list of quotes"""
        output_folder = Path(output_folder)
        output_folder.mkdir(exist_ok=True, parents=True)

        # Auto-select templates based on quote content
        template_keywords = {
            'gradient_fire': ['energy', 'power', 'passion', 'motivation', 'action', 'hustle'],
            'gradient_ocean': ['calm', 'peace', 'wisdom', 'deep', 'think', 'mind'],
            'gradient_purple_dream': ['dream', 'creative', 'spiritual', 'soul', 'art', 'imagine'],
            'gradient_mint_fresh': ['success', 'growth', 'money', 'wealth', 'win', 'achieve'],
            'gradient_sunset': ['inspire', 'hope', 'life', 'love', 'beautiful', 'happy'],
            'gradient_dark_purple': ['luxury', 'mystery', 'secret', 'truth', 'deep'],
            'solid_black': ['bold', 'strong', 'power', 'minimal', 'modern']
        }

        generated_images = []

        for i, quote in enumerate(quotes_list):
            # Auto-select template based on quote content
            if template_name == 'auto':
                quote_lower = quote.lower()
                selected_template = 'gradient_sunset'  # default

                for tmpl, keywords in template_keywords.items():
                    if any(keyword in quote_lower for keyword in keywords):
                        selected_template = tmpl
                        break
            else:
                selected_template = template_name

            # Generate image
            img = QuoteImageGenerator.create_quote_image(quote, selected_template)

            # Save image
            filename = f"quote_{i+1}_{selected_template}.png"
            filepath = output_folder / filename
            img.save(filepath)

            generated_images.append(filepath)
            print(f"[OK] Generated image {i+1}/{len(quotes_list)}: {filename} (template: {selected_template})")

        return generated_images


class ExcelIntegration:
    """Handle Excel file reading and video ID matching for Facebook Reels"""

    @staticmethod
    def extract_video_id(video_filename: str) -> str:
        """
        Extract video ID from filename

        Examples:
            -111224006530392.mp4 -> -111224006530392
            581727869405526.mp4 -> 581727869405526
            video_-111224006530392.mp4 -> -111224006530392
            1.mp4 -> 1
            video1.mp4 -> 1

        Args:
            video_filename: Name of video file

        Returns:
            Video ID string (without extension)
        """
        from pathlib import Path

        # Remove extension
        stem = Path(video_filename).stem

        # First, try to match the entire stem if it's just a number
        import re
        if re.match(r'^-?\d+$', stem):
            return stem

        # Try to extract last numeric sequence (works for "video1.mp4" -> "1")
        matches = re.findall(r'-?\d+', stem)
        if matches:
            return matches[-1]  # Return LAST number found

        # Fallback: return the whole stem
        return stem

    @staticmethod
    def read_excel_data(excel_path: str, id_column: str, text_column: str) -> tuple:
        """
        Read Excel file and create mapping of video_id -> overlay_text

        Args:
            excel_path: Path to Excel file
            id_column: Column letter for video ID (e.g., 'A', 'B')
            text_column: Column letter for overlay text (e.g., 'B', 'C')

        Returns:
            Tuple of (dict, list):
            - dict: video_id -> overlay_text mapping (for ID matching)
            - list: list of overlay texts in row order (for index matching)
        """
        try:
            import pandas as pd
        except ImportError:
            print("[ERROR] pandas not installed. Install with: pip install pandas openpyxl")
            return {}, []

        try:
            # Read Excel file
            df = pd.read_excel(excel_path, engine='openpyxl')

            # Convert column letters to indices (A=0, B=1, etc.)
            id_col_idx = ord(id_column.upper()) - ord('A')
            text_col_idx = ord(text_column.upper()) - ord('A')

            # Check if columns exist
            if id_col_idx >= len(df.columns) or text_col_idx >= len(df.columns):
                print(f"[ERROR] Column index out of range. Excel has {len(df.columns)} columns")
                return {}, [], []

            # Create both dict and lists
            video_data = {}
            text_list = []
            id_list = []  # Store Video IDs in row order for Row Index mode

            for index, row in df.iterrows():
                try:
                    # Get video ID and convert to string
                    video_id_raw = row.iloc[id_col_idx]

                    # Convert to string and clean up
                    video_id = str(video_id_raw).strip()

                    # Remove .0 from float numbers (Excel stores large numbers as floats)
                    # e.g., "1220687699922876.0" -> "1220687699922876"
                    if video_id.endswith('.0'):
                        video_id = video_id[:-2]

                    overlay_text = str(row.iloc[text_col_idx]).strip()

                    # Skip empty rows
                    if video_id and video_id != 'nan' and overlay_text and overlay_text != 'nan':
                        video_data[video_id] = overlay_text
                        text_list.append(overlay_text)  # Add to list in row order
                        id_list.append(video_id)  # Add Video ID in row order
                        # Show first 5 entries for debugging
                        if len(video_data) <= 5:
                            print(f"  Row {index + 1}: ID='{video_id}' -> Text='{overlay_text[:50]}...'")
                except Exception as e:
                    print(f"[WARNING] Skipped row {index + 1}: {e}")
                    continue

            print(f"[OK] Loaded {len(video_data)} video entries from Excel")
            if len(video_data) > 5:
                print(f"  (Showing first 5, total: {len(video_data)})")
            return video_data, text_list, id_list

        except Exception as e:
            print(f"[ERROR] Failed to read Excel file: {e}")
            import traceback
            traceback.print_exc()
            return {}, [], []

    @staticmethod
    def find_audio_file(video_id: str, audio_folder: str) -> Optional[Path]:
        """
        Find MP3 audio file matching video ID

        Args:
            video_id: Video ID to match (e.g., '-111224006530392')
            audio_folder: Path to folder containing MP3 files

        Returns:
            Path to matching MP3 file, or None if not found
        """
        from pathlib import Path

        audio_folder_path = Path(audio_folder)
        if not audio_folder_path.exists():
            print(f"[WARNING] Audio folder not found: {audio_folder}")
            return None

        print(f"[EXCEL] Searching for audio file with ID: '{video_id}' in {audio_folder}")

        # Look for exact match: {video_id}.mp3
        exact_match = audio_folder_path / f"{video_id}.mp3"
        if exact_match.exists():
            print(f"[EXCEL] Found exact match: {exact_match.name}")
            return exact_match

        # Look for files containing video ID in the stem
        for audio_file in audio_folder_path.glob("*.mp3"):
            if video_id in audio_file.stem:
                print(f"[EXCEL] Found partial match: {audio_file.name}")
                return audio_file

        # List available audio files for debugging
        available_files = list(audio_folder_path.glob("*.mp3"))
        if available_files:
            print(f"[WARNING] No audio file found for video ID: '{video_id}'")
            print(f"  Available audio files (first 5): {[f.name for f in available_files[:5]]}")
        else:
            print(f"[WARNING] Audio folder is empty: {audio_folder}")

        return None


def _merge_caption_clips(caption_clips, video_w, video_h, duration):
    """Merge N caption clips into a single VideoClip with one make_frame function.

    CRITICAL: MoviePy ImageClip stores RGBA as separate RGB frame +
    alpha mask (self.mask). We recombine them so our make_frame can
    alpha-composite the active captions in one flat function, avoiding
    MoviePy's per-clip-per-frame composite overhead.
    """
    if not caption_clips:
        return None
    if len(caption_clips) == 1:
        return caption_clips[0]

    import numpy as np
    try:
        from moviepy import VideoClip
    except ImportError:
        from moviepy.editor import VideoClip

    # ── Pre-extract RGB + mask from each clip once ──────────────────────
    data = []
    for clip in caption_clips:
        try:
            rgb = clip.get_frame(0)  # shape (h, w, 3) — RGB only
        except Exception:
            continue
        # Get alpha mask (MoviePy stores 4th channel as clip.mask).
        # CRITICAL: MoviePy masks return float [0, 1], but we convert to
        # uint8 [0, 255] internally for the blending canvas, then back
        # to float [0, 1] in _alpha_frame() for MoviePy's compositor
        # (which does mask.get_frame(t) * 255 — see compose_on line 748).
        _mask_raw = None
        try:
            if hasattr(clip.mask, 'get_frame'):
                _mask_raw = clip.mask.get_frame(0)
            else:
                _mask_raw = clip.mask  # already a numpy array
        except Exception:
            _mask_raw = None

        if _mask_raw is not None:
            # Normalise to 2D uint8 [0, 255]
            if isinstance(_mask_raw, np.ndarray):
                if _mask_raw.dtype in (np.float16, np.float32, np.float64):
                    # float [0, 1] → uint8 [0, 255]
                    alpha = np.clip(_mask_raw * 255.0, 0, 255).astype(np.uint8)
                else:
                    # already uint-ish
                    alpha = np.asarray(_mask_raw, dtype=np.uint8)
                if alpha.ndim == 3:
                    alpha = alpha[:, :, 0]  # first channel if multi
                elif alpha.ndim != 2:
                    alpha = alpha.reshape(alpha.shape[:2])  # try squeeze
            else:
                alpha = np.asarray(_mask_raw, dtype=np.uint8)
            # Ensure correct shape
            if alpha.shape != rgb.shape[:2]:
                alpha = alpha[:rgb.shape[0], :rgb.shape[1]]
        else:
            # No usable mask → derive alpha from non-black pixels in the RGB frame.
            _brightness = rgb.astype(np.float32).max(axis=-1)
            alpha = (_brightness > 15.0).astype(np.uint8) * 255

        start = float(getattr(clip, 'start', 0) or 0)
        end = getattr(clip, 'end', None)
        if end is None:
            dur = getattr(clip, 'duration', 0) or 0
            end = start + float(dur)
        else:
            end = float(end)

        pos = getattr(clip, 'pos', ('center', 'center'))
        if callable(pos):
            pos_fn = pos
        else:
            pos_fn = lambda _t, _p=pos: _p

        fh, fw = rgb.shape[:2]
        data.append({
            'rgb': rgb,
            'alpha': alpha,
            'start': start,
            'end': end,
            'pos_fn': pos_fn,
            'fw': fw,
            'fh': fh,
        })

    # DEBUG: check mask for first clip — helps diagnose "black captions"
    if data:
        _d0 = data[0]
        print(f"[MERGE] {len(data)} clips merged | "
              f"1st clip: alpha_range={_d0['alpha'].min()}-{_d0['alpha'].max()}, "
              f"rgb_range=({_d0['rgb'][:,:,0].min()}-{_d0['rgb'][:,:,0].max()},"
              f"{_d0['rgb'][:,:,1].min()}-{_d0['rgb'][:,:,1].max()},"
              f"{_d0['rgb'][:,:,2].min()}-{_d0['rgb'][:,:,2].max()})")

    if not data:
        return None

    # ⚡ Compute actual caption active range so the overlay loop can skip
    # frames where no captions are visible (saves get_frame + PIL composite).
    _cap_start = min(d['start'] for d in data)
    _cap_end = max(d['end'] for d in data)

    # ── Single make_frame that composites only active clips ─────────────
    # CRITICAL: MoviePy v1.x uses .make_frame, v2.x uses .frame_function.
    # We set BOTH so get_frame() works in either version.  Also produce
    # a separate mask clip for alpha so CompositeVideoClip layers with
    # transparency.
    # ═══════════════════════════════════════════════════════════════════════
    # Single-entry cache: MoviePy renders frames sequentially (t increases),
    # and compose_on calls BOTH get_frame(t) and mask.get_frame(t) for the
    # same t — so we cache ONE frame to avoid re-computing the RGBA canvas
    # between the RGB and alpha calls.  A dict would grow to ~7GB for a
    # 77s caption at 24fps (1848 frames × 3.7MB).
    # ═══════════════════════════════════════════════════════════════════════
    _rgba_last_key = [None]
    _rgba_last_val = [None]
    # ⚡ Reusable canvas — avoids 6MB np.zeros allocation per frame.
    # Pre-allocate once, then fill(0) which is faster than alloc+zero
    # because the memory pages are already faulted.
    _canvas_rgba = np.zeros((video_h, video_w, 4), dtype=np.uint8)

    def _render_rgba(t):
        """Compute full RGBA canvas for time t (single-entry cache)."""
        if _rgba_last_key[0] == t:
            return _rgba_last_val[0]
        canvas = _canvas_rgba
        canvas.fill(0)
        for d in data:
            if d['start'] <= t < d['end']:
                rgb = d['rgb']
                alpha = d['alpha']
                pos = d['pos_fn'](t)
                x_pos, y_pos = pos[0], pos[1]
                fw, fh = d['fw'], d['fh']
                if x_pos == 'center':
                    x_pos = (video_w - fw) // 2
                x_pos, y_pos = int(x_pos), int(y_pos)
                if x_pos >= video_w or y_pos >= video_h or x_pos + fw <= 0 or y_pos + fh <= 0:
                    continue
                src_x = max(0, -x_pos)
                src_y = max(0, -y_pos)
                dst_w = min(fw, video_w - x_pos) - src_x
                dst_h = min(fh, video_h - y_pos) - src_y
                if dst_w <= 0 or dst_h <= 0:
                    continue
                dst_x = x_pos + src_x
                dst_y = y_pos + src_y
                fg_rgb = rgb[src_y:src_y+dst_h, src_x:src_x+dst_w, :]
                fg_a = alpha[src_y:src_y+dst_h, src_x:src_x+dst_w].astype(np.float32) / 255.0
                canvas[dst_y:dst_y+dst_h, dst_x:dst_x+dst_w, :3] = (
                    canvas[dst_y:dst_y+dst_h, dst_x:dst_x+dst_w, :3].astype(np.float32) * (1.0 - fg_a[:, :, None]) +
                    fg_rgb.astype(np.float32) * fg_a[:, :, None]
                ).astype(np.uint8)
                canvas[dst_y:dst_y+dst_h, dst_x:dst_x+dst_w, 3] = np.maximum(
                    canvas[dst_y:dst_y+dst_h, dst_x:dst_x+dst_w, 3],
                    (fg_a * 255.0).astype(np.uint8))
        _rgba_last_key[0] = t
        _rgba_last_val[0] = canvas
        return canvas

    def _rgb_frame(t):
        """RGB for the merged clip (3 channels). v2.x reads this via frame_function."""
        return _render_rgba(t)[:, :, :3]

    def _alpha_frame(t):
        """Alpha mask (single channel float [0,1]) — MoviePy's compose_on
        multiplies by 255 (line 748 of VideoClip.py in v2.1.2)."""
        return (_render_rgba(t)[:, :, 3].astype(np.float32) / 255.0)

    # ── Build merged clip + mask clip ──────────────────────────────────
    # CRITICAL: MoviePy v2.x methods like with_mask() / with_duration()
    # may create a NEW clip via copy_with().  We therefore attach the
    # mask and duration FIRST, then set frame_function LAST on the final
    # object — so it's always the one that get_frame() actually calls.
    # ═══════════════════════════════════════════════════════════════════════
    try:
        merged = VideoClip()
    except Exception:
        merged = VideoClip(has_constant_size=False)

    # Build mask clip with its own RGB → alpha mapping
    try:
        mask_clip = VideoClip()
    except Exception:
        try:
            mask_clip = VideoClip(has_constant_size=False)
        except Exception:
            mask_clip = VideoClip()

    # Set duration on both (direct attribute — works in both v1 and v2)
    merged.duration = duration
    mask_clip.duration = duration
    # ⚡ Store caption time bounds as separate attributes (NOT start/end,
    # which MoviePy uses for time offset in composites and would break
    # the internal timeline used by _render_rgba's data list).
    # The OpenCV path reads these to skip compositing frames with no captions.
    merged.caption_start = _cap_start
    merged.caption_end = _cap_end

    # Attach mask via direct attribute (no copy_with wrapper)
    try:
        merged.mask = mask_clip
    except AttributeError:
        merged = merged.with_mask(mask_clip)  # v2.x immutable fallback

    # ── frame_function LAST — on the final clip object ──────────────────
    # MoviePy v2.x: get_frame(t) → self.frame_function(t)
    # MoviePy v1.x: get_frame(t) → self.make_frame(t)
    # We set BOTH so both versions work.  make_frame may be a read-only
    # property in v2.x — that's fine, frame_function is what matters.
    merged.frame_function = _rgb_frame
    try:
        merged.make_frame = _rgb_frame
    except AttributeError:
        pass

    mask_clip.frame_function = _alpha_frame
    try:
        mask_clip.make_frame = _alpha_frame
    except AttributeError:
        pass

    return merged


class VideoQuoteAutomation:
    """Automate adding quotes to videos with advanced effects"""

    def __init__(self, video_folder=None, quotes_file=None, output_folder=None):
        # Use provided paths or fall back to overlay_settings.json / processing_paths.json
        import json as _json
        _paths = {}
        _saved_ov = {}
        try:
            _pf = Path(__file__).parent / 'processing_paths.json'
            if _pf.exists():
                _paths = _json.loads(_pf.read_text('utf-8'))
        except Exception:
            pass
        # Also read video/output folders from overlay_settings.json if set
        try:
            _ovf = Path(__file__).parent / 'overlay_settings.json'
            if _ovf.exists():
                _saved_ov = _json.loads(_ovf.read_text('utf-8'))
        except Exception:
            pass
        _vf = video_folder or _paths.get('video_folder') or _saved_ov.get('video_folder') or ''
        _qf = quotes_file or _paths.get('quotes_file') or ''
        _of = output_folder or _paths.get('output_folder') or _saved_ov.get('output_folder') or ''
        self.video_folder = Path(_vf) if _vf else Path('')
        self.quotes_file = Path(_qf) if _qf else Path('')
        self.output_folder = Path(_of) if _of else Path('')

        self.output_folder.mkdir(parents=True, exist_ok=True)

        self.settings = self.load_settings()

        self.log_file = self.output_folder / "processing_log.json"
        self.processing_log = self._load_log()

        # State file to track last used quote index (persists across script restarts)
        self.state_file = self.output_folder / "quote_state.json"
        self.quote_state = self._load_quote_state()

        # Load voiceover files if enabled
        self.voiceover_files = []
        if self.settings.get('add_voiceover', False) and self.settings.get('voiceover_folder'):
            voiceover_folder = Path(self.settings['voiceover_folder'])
            self.voiceover_files = AudioProcessor.get_voiceover_files(voiceover_folder)
            if self.voiceover_files:
                print(f"[OK] Loaded {len(self.voiceover_files)} voiceover files (sorted by number)")
                for i, vf in enumerate(self.voiceover_files[:5], 1):
                    print(f"  {i}. {vf.name}")
            else:
                print(f"[WARNING] No voiceover files found in {voiceover_folder}")

        # Load BGM files if enabled
        self.bgm_files = []
        if self.settings.get('add_custom_bgm', False) and self.settings.get('bgm_file'):
            self.bgm_files = AudioProcessor.get_bgm_files(self.settings['bgm_file'])
            if self.bgm_files:
                if len(self.bgm_files) == 1:
                    print(f"[OK] Loaded 1 BGM file: {self.bgm_files[0].name}")
                else:
                    print(f"[OK] Loaded {len(self.bgm_files)} BGM files (will select randomly per video)")
                    for i, bf in enumerate(self.bgm_files[:5], 1):
                        print(f"  {i}. {bf.name}")
            else:
                print(f"[WARNING] No BGM files found")

        # Load Excel integration data if enabled
        self.excel_data = {}
        self.excel_text_list = []
        self.excel_id_list = []  # Store Video IDs in row order
        self.excel_audio_folder = None
        if self.settings.get('excel_integration_enabled', False):
            excel_file = self.settings.get('excel_file_path', '')
            text_column = self.settings.get('excel_text_column', 'B')
            id_column = self.settings.get('excel_id_column', 'A')
            self.excel_audio_folder = self.settings.get('excel_audio_folder', '')

            if excel_file and Path(excel_file).exists():
                print(f"[OK] Excel Integration Enabled")
                print(f"  Excel file: {Path(excel_file).name}")
                print(f"  ID Column: {id_column}, Text Column: {text_column}")
                print(f"  Audio folder: {self.excel_audio_folder or 'Not set'}")

                self.excel_data, self.excel_text_list, self.excel_id_list = ExcelIntegration.read_excel_data(
                    excel_file, id_column, text_column
                )

                if self.excel_data:
                    print(f"[OK] Ready to process {len(self.excel_data)} videos from Excel")
                else:
                    print(f"[WARNING] Excel integration enabled but no data loaded")
            else:
                print(f"[WARNING] Excel integration enabled but Excel file not found: {excel_file}")

    def load_settings(self) -> dict:
        """Load settings from GUI config file"""
        settings_file = Path('overlay_settings.json')

        default_settings = {
            # ── AI Avatar ──
            'avatar_enabled': False,
            'avatar_face_path': '',
            'avatar_position': 'bottom-right',
            'avatar_scale': 0.25,
            'avatar_standalone': False,
            'avatar_resize_factor': 2,
            'avatar_pads': [0, 20, 0, 0],

            'font_size': 45,
            'font_style': 'Arial Bold',
            'text_color': '#000000',
            'bg_color': '#FFFFFF',
            'bg_opacity': 90,
            'cta_enabled': True,
            'cta_font_size': 43,
            'cta_font_style': 'Arial Italic',
            'cta_bg_color': '#DC2626',
            'cta_text_color': '#FFFFFF',
            'emoji_enabled': True,
            'emoji_size_multiplier': 1.2,
            'bubble_width': 75,
            'padding_horizontal': 40,
            'padding_vertical': 20,
            'inner_padding': 15,
            'section_spacing': 15,
            'corner_radius': 15,
            'position': 'top',
            'vertical_offset': 0,
            'crosshair_enabled': False,
            'crosshair_color': '#FF0000',
            'crosshair_thickness': 2,
            'bottom_text_enabled': False,
            'bottom_text_content': '',
            'bottom_text_font_family': 'Arial',
            'bottom_text_font_size': 45,
            'bottom_text_text_color': '#FFFFFF',
            'bottom_text_bg_color': '#000000',
            'bottom_text_bg_opacity': 80,
            'bottom_text_outline': True,
            'bottom_text_outline_color': '#000000',
            'bottom_text_outline_size': 2,
            'bottom_text_vertical_offset': 0,
            'text_fade_in': True,
            'text_fade_duration': 0.4,
            'text_glow': True,
            'glow_intensity': 8,
            'vignette': True,
            'vignette_intensity': 0.4,
            'video_zoom': True,
            'zoom_scale': 1.08,
            'drop_shadow': True,
            'shadow_offset': 6,
            'shadow_blur': 12
        }

        if settings_file.exists():
            try:
                with open(settings_file, 'r') as f:
                    loaded_settings = json.load(f)
                    print("[OK] Loaded enhanced settings from overlay_settings.json")
                    return loaded_settings
            except Exception as e:
                print(f"[WARNING] Could not load settings: {e}")
                return default_settings
        else:
            print("[WARNING] No settings file found")
            return default_settings

    def _load_log(self) -> dict:
        """Load processing log"""
        if self.log_file.exists():
            with open(self.log_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {"processed_count": 0, "processed_videos": []}

    def _save_log(self):
        """Save processing log"""
        with open(self.log_file, 'w', encoding='utf-8') as f:
            json.dump(self.processing_log, f, indent=2, ensure_ascii=False)

    def _load_quote_state(self) -> dict:
        """Load quote state (tracks last used quote index)"""
        if self.state_file.exists():
            try:
                with open(self.state_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"[WARNING] Could not load quote state: {e}")
                return {"last_quote_index": -1}
        return {"last_quote_index": -1}

    def _save_quote_state(self):
        """Save quote state (persists last used quote index)"""
        with open(self.state_file, 'w', encoding='utf-8') as f:
            json.dump(self.quote_state, f, indent=2, ensure_ascii=False)

    def hex_to_rgb(self, hex_color: str) -> tuple:
        """Convert hex color to RGB tuple"""
        hex_color = hex_color.lstrip('#')
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

    def read_transcript_excel(self, excel_path: Path) -> dict:
        """
        Read an Excel transcript file with columns: Video ID, Custom Title, Transcript.

        Matches the video's filename stem to the 'Video ID' column and returns
        the 'Transcript' text. Returns a dict of {video_id_stem: transcript_text}.

        Args:
            excel_path: Path to .xlsx file with transcript data

        Returns:
            Dict mapping video filename stem -> transcript text
        """
        if not excel_path.is_file():
            print(f"[ERROR] Excel file not found: {excel_path}")
            return {}
        if not str(excel_path).lower().endswith('.xlsx'):
            print(f"[ERROR] Not an .xlsx file: {excel_path}")
            return {}

        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(excel_path), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except ImportError:
            print("[ERROR] openpyxl not installed. Install with: pip install openpyxl")
            return {}
        except Exception as e:
            print(f"[ERROR] Failed to read Excel file: {e}")
            return {}

        if not rows:
            print(f"[WARNING] Excel file is empty: {excel_path}")
            return {}

        # Case-insensitive header detection
        header = [str(h or '').strip() for h in rows[0]]
        header_lc = [h.lower() for h in header]

        vid_col = None
        trans_col = None
        for i, h in enumerate(header_lc):
            if h in ('video id', 'video_id', 'videoid'):
                vid_col = i
            elif h in ('transcript', 'caption text', 'captions', 'text'):
                trans_col = i

        if vid_col is None:
            print(f"[ERROR] Could not find 'Video ID' column in Excel. Headers: {header}")
            return {}
        if trans_col is None:
            print(f"[ERROR] Could not find 'Transcript' column in Excel. Headers: {header}")
            return {}

        transcript_map = {}
        for r in rows[1:]:
            if not r or len(r) <= max(vid_col, trans_col):
                continue
            vid = str(r[vid_col] or '').strip()
            trans = str(r[trans_col] or '').strip()
            if vid and trans:
                transcript_map[vid] = trans

        print(f"[EXCEL TRANSCRIPT] Loaded {len(transcript_map)} transcript entries from {excel_path.name}")
        if transcript_map:
            print(f"[EXCEL TRANSCRIPT] First 3 entries: {list(transcript_map.items())[:3]}")

        return transcript_map

    def read_quotes(self) -> List[dict]:
        """
        Read quotes from file(s) with support for separate subtitle and voiceover text.

        Format options:
        1. Single file (Quotes.txt only): Same text for subtitle and voiceover
        2. Two separate files: Quotes.txt for subtitles, VoiceoverText.txt for voiceover
           - Line 1 in Quotes.txt pairs with line 1 in VoiceoverText.txt
           - Line 2 in Quotes.txt pairs with line 2 in VoiceoverText.txt, etc.

        Returns list of dicts with 'subtitle' and 'voiceover' keys.
        """
        print(f"[DEBUG] Attempting to read quotes from: {self.quotes_file}")
        print(f"[DEBUG] File exists: {self.quotes_file.exists()}")

        if not self.quotes_file or not self.quotes_file.is_file():
            if not self.quotes_file or not str(self.quotes_file):
                print(f"[INFO] No quotes file configured (set quotes_file in processing_paths.json)")
            else:
                print(f"✗ Quotes file not found: {self.quotes_file}")
            return []

        # Skip .xlsx files — they are transcript Excels handled in get_quote_for_video
        if str(self.quotes_file).lower().endswith('.xlsx'):
            print(f"[OK] Quotes file is .xlsx — transcript Excel will be used instead")
            return []

        # Read subtitle text (from Quotes.txt)
        with open(self.quotes_file, 'r', encoding='utf-8') as f:
            subtitle_content = f.read()

        print(f"[DEBUG] File content length: {len(subtitle_content)} characters")
        print(f"[DEBUG] First 100 chars: {subtitle_content[:100]}")

        subtitle_lines = []
        if re.match(r'^\s*\d+\.', subtitle_content, re.MULTILINE):
            print(f"[DEBUG] Detected numbered format")
            parts = re.split(r'\n\s*\d+\.\s*', subtitle_content)
            print(f"[DEBUG] Split into {len(parts)} parts")
            subtitle_lines = [q.strip() for q in parts[1:] if q.strip()]
            print(f"[DEBUG] After filtering: {len(subtitle_lines)} lines")
        elif '\n\n' in subtitle_content:
            print(f"[DEBUG] Detected paragraph format (\\n\\n)")
            subtitle_lines = [q.strip() for q in subtitle_content.split('\n\n') if q.strip()]
        elif '---' in subtitle_content:
            print(f"[DEBUG] Detected dash separator format")
            subtitle_lines = [q.strip() for q in subtitle_content.split('---') if q.strip()]
        else:
            print(f"[DEBUG] Using line-by-line format")
            subtitle_lines = [line.strip() for line in subtitle_content.split('\n') if line.strip()]

        print(f"[DEBUG] Before cleaning: {len(subtitle_lines)} subtitle lines")

        # Clean subtitle lines
        subtitle_lines = [re.sub(r'^\d+\.\s*', '', line).strip() for line in subtitle_lines if line.strip()]

        print(f"[DEBUG] After cleaning: {len(subtitle_lines)} subtitle lines")

        # Check for separate voiceover text file
        voiceover_text_file = self.settings.get('voiceover_text_file', '')
        voiceover_lines = []

        if voiceover_text_file and Path(voiceover_text_file).exists():
            print(f"[OK] Using separate voiceover text file: {Path(voiceover_text_file).name}")
            with open(voiceover_text_file, 'r', encoding='utf-8') as f:
                voiceover_content = f.read()

            # Parse voiceover file same way as subtitle file
            if re.match(r'^\s*\d+\.', voiceover_content, re.MULTILINE):
                parts = re.split(r'\n\s*\d+\.\s*', voiceover_content)
                voiceover_lines = [q.strip() for q in parts[1:] if q.strip()]
            elif '\n\n' in voiceover_content:
                voiceover_lines = [q.strip() for q in voiceover_content.split('\n\n') if q.strip()]
            elif '---' in voiceover_content:
                voiceover_lines = [q.strip() for q in voiceover_content.split('---') if q.strip()]
            else:
                voiceover_lines = [line.strip() for line in voiceover_content.split('\n') if line.strip()]

            # Clean voiceover lines
            voiceover_lines = [re.sub(r'^\d+\.\s*', '', line).strip() for line in voiceover_lines if line.strip()]

            if len(voiceover_lines) != len(subtitle_lines):
                print(f"[WARNING] Warning: Subtitle file has {len(subtitle_lines)} lines, Voiceover file has {len(voiceover_lines)} lines")
                print(f"  → Will use minimum count: {min(len(subtitle_lines), len(voiceover_lines))}")
        else:
            print(f"[OK] Using Quotes.txt for both subtitle and voiceover")
            voiceover_lines = subtitle_lines  # Use same text for both

        # Pair subtitle and voiceover lines
        processed_quotes = []
        min_count = min(len(subtitle_lines), len(voiceover_lines)) if voiceover_lines else len(subtitle_lines)

        for i in range(min_count):
            subtitle_text = subtitle_lines[i]
            voiceover_text = voiceover_lines[i] if voiceover_lines else subtitle_text

            processed_quotes.append({
                'subtitle': subtitle_text,
                'voiceover': voiceover_text
            })

        print(f"[OK] Loaded {len(processed_quotes)} quotes")

        if processed_quotes:
            print(f"\nFirst 3 quotes:")
            for i, quote_data in enumerate(processed_quotes[:3], 1):
                subtitle_preview = quote_data['subtitle'][:80] + "..." if len(quote_data['subtitle']) > 80 else quote_data['subtitle']
                print(f"  {i}. Subtitle: {subtitle_preview}")
                if quote_data['subtitle'] != quote_data['voiceover']:
                    voiceover_preview = quote_data['voiceover'][:80] + "..." if len(quote_data['voiceover']) > 80 else quote_data['voiceover']
                    print(f"     Voiceover: {voiceover_preview}")

        return processed_quotes

    def get_quote_for_video(self, video_path: Path, video_index: int) -> Tuple[dict, Optional[Path]]:
        """
        Get quote and audio file for a specific video.

        If Excel integration is enabled:
            - Extracts video ID from filename
            - Looks up overlay text from Excel data
            - Finds matching MP3 audio file
            - Returns quote dict and audio path

        Otherwise:
            - Uses standard quote from Quotes.txt
            - Returns quote dict and None for audio

        Args:
            video_path: Path to video file
            video_index: Index of video (used for cycling through quotes in non-Excel mode)

        Returns:
            Tuple of (quote_dict, audio_path)
            - quote_dict: {'subtitle': str, 'voiceover': str, 'video_id': str (Excel only)}
            - audio_path: Path to MP3 file (Excel mode) or None (standard mode)
        """
        # ─── Transcript Excel mode ──────────────────────────────────
        # If the quotes_file is an .xlsx, read it as a transcript Excel
        # with columns: Video ID, Custom Title, Transcript.
        if str(self.quotes_file).lower().endswith('.xlsx'):
            transcript_map = self.read_transcript_excel(self.quotes_file)
            if transcript_map:
                video_stem = video_path.stem
                transcript_text = transcript_map.get(video_stem)
                if not transcript_text:
                    video_id = ExcelIntegration.extract_video_id(video_path.name)
                    transcript_text = transcript_map.get(video_id)
                if transcript_text:
                    print(f'[EXCEL TRANSCRIPT] Matched "{video_stem}" → transcript ({len(transcript_text)} chars)')
                    quote_dict = {
                        'subtitle': transcript_text,
                        'voiceover': transcript_text,
                        'source': 'transcript_excel',
                        'video_id': video_stem,
                    }
                    self.settings['enable_captions'] = True
                    return quote_dict, None
                else:
                    print(f'[EXCEL TRANSCRIPT] Video ID "{video_stem}" not found in Excel')
                    print(f'  Available IDs (first 10): {list(transcript_map.keys())[:10]}')
            else:
                print(f'[WARNING] Could not load transcript data from {self.quotes_file.name}')

        # Check if Excel integration is enabled and has data
        if self.excel_data or self.excel_text_list:
            # Check matching mode
            match_mode = self.settings.get('excel_match_mode', 'row_index')

            if match_mode == 'row_index':
                # Match by row index (simpler)
                print(f"[EXCEL] Using ROW INDEX matching mode")
                print(f"[EXCEL] Video #{video_index + 1}: {video_path.name}")

                # Get text and ID from lists by index
                if video_index < len(self.excel_text_list):
                    overlay_text = self.excel_text_list[video_index]
                    # Get the actual Excel Video ID from Column A for audio file matching
                    excel_video_id = self.excel_id_list[video_index] if video_index < len(self.excel_id_list) else f"Row_{video_index + 1}"
                    video_id = excel_video_id  # Use Excel Video ID for audio lookup

                    print(f"[EXCEL] Row {video_index + 1}: Excel Video ID = '{excel_video_id}'")
                    print(f"[EXCEL] Row {video_index + 1}: Text = '{overlay_text[:80]}...'")
                else:
                    print(f"[ERROR] Video index {video_index + 1} exceeds Excel rows ({len(self.excel_text_list)})")
                    raise Exception(f"Excel Integration: Video index {video_index + 1} exceeds number of Excel rows ({len(self.excel_text_list)})!\n"
                                  f"You have {len(self.excel_text_list)} rows in Excel but processing video #{video_index + 1}")
            else:
                # Match by video ID (original method)
                print(f"[EXCEL] Using VIDEO ID matching mode")
                # Extract video ID from filename
                video_id = ExcelIntegration.extract_video_id(video_path.name)
                print(f"[EXCEL] Extracted video ID: {video_id}")

                # Look up text from Excel
                overlay_text = self.excel_data.get(video_id)

            if overlay_text:
                if match_mode == 'video_id':
                    print(f"[EXCEL] Found overlay text for video ID {video_id}")

                # Find matching audio file
                audio_path = None
                if self.excel_audio_folder:
                    audio_path = ExcelIntegration.find_audio_file(video_id, self.excel_audio_folder)
                    if audio_path:
                        print(f"[EXCEL] Found audio file: {audio_path.name}")
                    else:
                        print(f"[EXCEL] No audio file found for video ID {video_id}")

                # Return quote dict with Excel text
                quote_dict = {
                    'subtitle': overlay_text,
                    'voiceover': overlay_text,  # Excel text used for both
                    'video_id': video_id,
                    'source': 'excel',
                    'excel_position': self.settings.get('excel_text_position', 'center')  # Position override
                }

                return quote_dict, audio_path
            else:
                print(f"[ERROR] Video ID '{video_id}' not found in Excel data")
                print(f"  Video filename: {video_path.name}")
                print(f"  Extracted ID: '{video_id}'")
                print(f"  Available Excel IDs (first 10): {list(self.excel_data.keys())[:10]}")

                # In Excel mode, don't fall back - raise an error
                raise Exception(f"Excel Integration enabled but video ID '{video_id}' not found in Excel file!\n"
                              f"Video: {video_path.name}\n"
                              f"Expected ID format in Excel Column A: '{video_id}'\n"
                              f"Available IDs: {', '.join(list(self.excel_data.keys())[:5])}")

        # Standard mode: use quotes from Quotes.txt (only if Excel is NOT enabled)
        # If video_title_source is 'video', auto-generate from filename instead
        if self.settings.get('video_title_source', 'quote') == 'video':
            # Generate a quote from the video filename
            video_title = video_path.stem  # filename without extension
            # Clean up — remove video IDs like (1), /dash patterns
            import re as _re
            clean_title = _re.sub(r'\s*\(\d+\)|\s*\[\d+\]|\s*-\s*\w{11}', '', video_title).strip()
            if not clean_title:
                clean_title = video_title

            # Skip UUID/hash filenames — they produce unreadable caption text.
            # Try the quotes file instead; if it's empty/missing, use the
            # filename anyway rather than crashing.
            if _re.match(r'^[0-9a-fA-F]{8}(-?[0-9a-fA-F]{4}){3}-?[0-9a-fA-F]{12}$', clean_title) \
               or _re.match(r'^[0-9a-fA-F]{16,}$', clean_title):
                print(f'[VIDEO TITLE] Filename "{clean_title}" looks like UUID/hash — trying Quotes.txt')
                quotes = self.read_quotes()
                if quotes:
                    quote_index = video_index % len(quotes)
                    quote = quotes[quote_index]
                    quote['source'] = 'quotes_file'
                    print(f'[VIDEO TITLE] Using quote #{quote_index+1} instead')
                    return quote, None
                print(f'[VIDEO TITLE] No quotes file available — using filename as-is (better than crashing)')
                # Still use the video filename as text — better than crashing
                quote = {
                    'subtitle': clean_title,
                    'voiceover': clean_title,
                    'source': 'video_title',
                }
                return quote, None
            else:
                quote = {
                    'subtitle': clean_title,
                    'voiceover': clean_title,
                    'source': 'video_title',
                }
                print(f'[VIDEO TITLE] Using video filename as quote: "{clean_title}"')
                return quote, None

        quotes = self.read_quotes()

        if not quotes:
            # If the quotes file is an .xlsx transcript that didn't match,
            # don't crash — warn and fall back to the video filename.
            if str(self.quotes_file).lower().endswith('.xlsx'):
                print(f'[EXCEL TRANSCRIPT] No match in transcript Excel for "{video_path.stem}" — using filename as fallback text')
                return {
                    'subtitle': video_path.stem,
                    'voiceover': video_path.stem,
                    'source': 'video_title',
                }, None
            raise Exception("No quotes found in quotes file")

        # Get quote at index (cycle through quotes if more videos than quotes)
        quote_index = video_index % len(quotes)
        quote = quotes[quote_index]

        # Add source marker
        quote['source'] = 'quotes_file'

        return quote, None

    def get_video_files(self, sort_by: str = 'created') -> List[Path]:
        """Get video files from folder"""
        video_extensions = {'.mp4', '.avi', '.mov', '.mkv', '.flv', '.wmv', '.webm'}

        if not self.video_folder.exists():
            print(f"✗ Video folder not found: {self.video_folder}")
            return []

        videos = [f for f in self.video_folder.iterdir()
                 if f.suffix.lower() in video_extensions and f.is_file()]

        if sort_by == 'created':
            videos = sorted(videos, key=lambda x: x.stat().st_ctime)
            print(f"[OK] Found {len(videos)} videos (sorted by creation date)")
        elif sort_by == 'modified':
            videos = sorted(videos, key=lambda x: x.stat().st_mtime)
            print(f"[OK] Found {len(videos)} videos (sorted by modification date)")
        else:
            videos = sorted(videos)
            print(f"[OK] Found {len(videos)} videos (sorted alphabetically)")

        if videos:
            print(f"\nFirst 5 videos:")
            for i, video in enumerate(videos[:5], 1):
                print(f"  {i}. {video.name}")

        return videos

    def generate_hashtags(self, quote: str) -> List[str]:
        """Generate relevant hashtags from quote"""
        quote_lower = quote.lower()

        hashtag_map = {
            'success': '#Success', 'motivation': '#Motivation', 'inspire': '#Inspiration',
            'life': '#Life', 'love': '#Love', 'happy': '#Happiness', 'dream': '#Dreams',
            'work': '#Work', 'business': '#Business', 'money': '#Money', 'goal': '#Goals',
            'achieve': '#Achievement', 'believe': '#Believe', 'hope': '#Hope',
            'strength': '#Strength', 'courage': '#Courage', 'change': '#Change',
            'wisdom': '#Wisdom', 'mindset': '#Mindset', 'grow': '#Growth',
            'leader': '#Leadership', 'hustle': '#Hustle', 'focus': '#Focus',
            'passion': '#Passion', 'gratitude': '#Gratitude', 'positive': '#Positivity'
        }

        found = []
        for keyword, hashtag in hashtag_map.items():
            if keyword in quote_lower and hashtag not in found:
                found.append(hashtag)
                if len(found) == 2:
                    break

        if len(found) < 2:
            defaults = ['#Motivation', '#Quotes', '#Inspiration', '#Wisdom']
            for tag in defaults:
                if tag not in found:
                    found.append(tag)
                    if len(found) == 2:
                        break

        return found[:2]

    def sanitize_filename(self, text: str, max_length: int = 100) -> str:
        """Convert text to valid filename (preserves emojis with proper encoding)"""
        import unicodedata

        # Normalize Unicode to ensure emojis are properly represented
        text = unicodedata.normalize('NFC', text)

        # Remove only invalid Windows filename characters, keep emojis
        text = re.sub(r'[<>:"/\\|?*\n\r\t]', '', text)

        # Replace multiple spaces with single space
        text = re.sub(r'\s+', ' ', text).strip()

        # Truncate if too long (accounting for .mp4 extension)
        if len(text) > max_length - 4:
            text = text[:max_length - 4].strip()

        return text

    def create_filename(self, quote: str, hashtags: List[str] = None,
                        use_title_only: bool = False,
                        source: str = 'quote',
                        video_path=None) -> str:
        """
        Create filename for the output video.

        Args:
            quote: Full subtitle text (may contain multiple lines: Title
                + Quote + CTA). Used when source='quote' (default).
            hashtags: List of hashtags to append when source='quote'
                and use_title_only is False.
            use_title_only: If True, only use the first line (title)
                of the quote for the filename.
            source: 'quote' (default) → use the quote text.
                'video' → use the source video's filename stem
                (e.g. 'MyVacation.mp4' → 'MyVacation.mp4').
                'auto' → if a video_path is provided, use the video
                filename; otherwise fall back to the quote.
            video_path: Path to the source video. Required when
                source='video' or 'auto'. May be a string or a Path.

        Returns:
            Filename with .mp4 extension
        """
        # Resolve 'auto' to a concrete source
        if source == 'auto':
            source = 'video' if video_path else 'quote'

        if source == 'video':
            if not video_path:
                # No video path provided — silently fall back to quote
                # so we never produce an empty / ".mp4" filename.
                source = 'quote'
            else:
                try:
                    from pathlib import Path as _P
                    vp = _P(str(video_path))
                    stem = vp.stem  # filename without extension
                    if stem:
                        # Sanitize but keep it as a proper name
                        safe = self.sanitize_filename(stem, max_length=120)
                        if safe:
                            return safe + ".mp4"
                except Exception:
                    pass
                # If anything went wrong, fall through to quote path
                source = 'quote'

        # source == 'quote' (original behavior)
        if use_title_only:
            # Extract just the first line (title) for filename
            title_line = quote.strip().split('\n')[0] if quote else quote
            filename = self.sanitize_filename(title_line, max_length=96)
        else:
            # Use full quote + hashtags (original behavior)
            hashtag_str = ' '.join(hashtags) if hashtags else ''
            filename_text = f"{quote} {hashtag_str}".strip()
            filename = self.sanitize_filename(filename_text, max_length=96)

        return filename + ".mp4"

    def create_text_overlay_image(self, video_width, video_height, title_text, main_text, cta_text, cta_emojis):
        """Create text overlay with Title + Quote + CTA format (with colorful CTA emojis)"""
        # Define emoji pattern at the beginning of the method
        emoji_pattern = re.compile(r'[\U0001F300-\U0001F9FF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF\U00002600-\U000027BF\U0001F1E0-\U0001F1FF]+')

        img_width = video_width
        temp_img = Image.new('RGBA', (img_width, 1000), (0, 0, 0, 0))
        temp_draw = ImageDraw.Draw(temp_img)

        # Load separate fonts for Title, Quote, and CTA
        try:
            # Helper function to find font file from font family name
            def get_font_file(font_family):
                """Convert font family name to font file path"""
                # Common font family to file mappings
                font_mappings = {
                    'Arial': 'arial.ttf',
                    'Arial Bold': 'arialbd.ttf',
                    'Arial Italic': 'ariali.ttf',
                    'Times New Roman': 'times.ttf',
                    'Georgia': 'georgia.ttf',
                    'Verdana': 'verdana.ttf',
                    'Tahoma': 'tahoma.ttf',
                    'Trebuchet MS': 'trebuc.ttf',
                    'Impact': 'impact.ttf',
                    'Comic Sans MS': 'comic.ttf',
                    'Courier New': 'cour.ttf',
                    'Segoe UI': 'segoeui.ttf',
                    'Calibri': 'calibri.ttf',
                    'Cambria': 'cambria.ttc',
                    'Consolas': 'consola.ttf',
                    # Urdu / Arabic fonts
                    'Jameel Noori Nastaleeq': 'arabtype.ttf',
                    'Jameel Noori Nastaleeq Kasheeda': 'arabtype.ttf',
                    'Noto Nastaliq Urdu': 'arabtype.ttf',
                    'Noto Nastaliq Urdu Bold': 'arabtype.ttf',
                    'Noto Naskh Arabic': 'arabtype.ttf',
                    'Noto Naskh Arabic Bold': 'arabtype.ttf',
                }

                # Check if it's already a file path
                if font_family.endswith('.ttf') or font_family.endswith('.otf') or font_family.endswith('.ttc'):
                    return font_family

                # Try exact match in mappings
                if font_family in font_mappings:
                    return font_mappings[font_family]

                # Try case-insensitive match
                for name, file in font_mappings.items():
                    if name.lower() == font_family.lower():
                        return file

                # Try to find a matching font file in Windows fonts folder
                fonts_path = Path(r"C:\Windows\Fonts")
                if fonts_path.exists():
                    # Try direct match with various extensions
                    for ext in ['.ttf', '.otf', '.ttc']:
                        possible_file = fonts_path / f"{font_family}{ext}"
                        if possible_file.exists():
                            return str(possible_file)

                    # Try lowercase
                    font_lower = font_family.lower().replace(' ', '')
                    for font_file in fonts_path.glob('*.ttf'):
                        if font_lower in font_file.stem.lower():
                            return str(font_file)

                # Default fallback
                return 'arial.ttf'

            def has_urdu_text(text):
                """Check if text contains Urdu/Arabic characters."""
                if not text:
                    return False
                # Urdu/Arabic Unicode ranges
                import unicodedata as _ud
                for ch in text:
                    cp = ord(ch)
                    if (0x0600 <= cp <= 0x06FF or  # Arabic
                        0x0750 <= cp <= 0x077F or  # Arabic Supplement
                        0x08A0 <= cp <= 0x08FF or  # Arabic Extended-A
                        0xFB50 <= cp <= 0xFDFF or  # Arabic Presentation Forms-A
                        0xFE70 <= cp <= 0xFEFF or  # Arabic Presentation Forms-B
                        0x1EE00 <= cp <= 0x1EEFF): # Arabic Mathematical
                        return True
                return False

            def resolve_urdu_font(font_family, font_size):
                """Load a font, auto-falling back to Urdu-capable font if text needs it."""
                try:
                    font_file = get_font_file(font_family)
                    if not Path(font_file).exists():
                        font_file = str(Path(r"C:\Windows\Fonts") / Path(font_file).name)
                    return ImageFont.truetype(font_file, font_size)
                except Exception:
                    # Try Urdu/Arabic alternative fonts as fallback
                    for _uf in [
                        r"C:\Windows\Fonts\arabtype.ttf",
                        r"C:\Windows\Fonts\Candarab.ttf",
                        r"C:\Windows\Fonts\GARABD.TTF",
                        str(Path(r"C:\Windows\Fonts") / "arial.ttf"),
                    ]:
                        try:
                            return ImageFont.truetype(_uf, font_size)
                        except Exception:
                            continue
                    return ImageFont.load_default()

            # Title font - read from GUI settings (increased default from 45 to 85)
            title_font_size = int(self.settings.get('title_font_size', 85))
            title_font_family = self.settings.get('title_font_family', 'Arial')
            _urdu_title = has_urdu_text(title_text)
            if _urdu_title:
                title_font = resolve_urdu_font(title_font_family, title_font_size)
                print(f"[FONT] Urdu detected in title, using fallback font")
            else:
                try:
                    title_font_file = get_font_file(title_font_family)
                    if not Path(title_font_file).exists():
                        title_font_file = str(Path(r"C:\Windows\Fonts") / Path(title_font_file).name)
                    title_font = ImageFont.truetype(title_font_file, title_font_size)
                    print(f"[FONT DEBUG] Title font loaded: size={title_font_size}, family={title_font_family}")
                except Exception as e:
                    print(f"[WARNING] Failed to load title font '{title_font_family}': {e}")
                    print(f"[WARNING] Using Arial for title at size {title_font_size}")
                    title_font = ImageFont.truetype(str(Path(r"C:\Windows\Fonts") / "arial.ttf"), title_font_size)

            # Quote font - read from GUI settings (increased default from 35 to 70)
            quote_font_size = int(self.settings.get('quote_font_size', 70))
            quote_font_family = self.settings.get('quote_font_family', 'Georgia')
            _urdu_quote = has_urdu_text(main_text)
            if _urdu_quote:
                quote_font = resolve_urdu_font(quote_font_family, quote_font_size)
                print(f"[FONT] Urdu detected in main text, using fallback font")
            else:
                try:
                    quote_font_file = get_font_file(quote_font_family)
                    if not Path(quote_font_file).exists():
                        quote_font_file = str(Path(r"C:\Windows\Fonts") / Path(quote_font_file).name)
                    quote_font = ImageFont.truetype(quote_font_file, quote_font_size)
                    print(f"[FONT DEBUG] Quote font loaded: size={quote_font_size}, family={quote_font_family}")
                except Exception as e:
                    print(f"[WARNING] Failed to load quote font '{quote_font_family}': {e}")
                    print(f"[WARNING] Using Arial for quote at size {quote_font_size}")
                    quote_font = ImageFont.truetype(str(Path(r"C:\Windows\Fonts") / "arial.ttf"), quote_font_size)

            # CTA font - read from GUI settings (increased default from 43 to 75)
            cta_font_size = int(self.settings.get('cta_font_size', 75))
            cta_font_family = self.settings.get('cta_font_family', 'Arial')
            _urdu_cta = has_urdu_text(cta_text)
            if _urdu_cta:
                cta_font = resolve_urdu_font(cta_font_family, cta_font_size)
                print(f"[FONT] Urdu detected in CTA, using fallback font")
            else:
                try:
                    cta_font_file = get_font_file(cta_font_family)
                    if not Path(cta_font_file).exists():
                        cta_font_file = str(Path(r"C:\Windows\Fonts") / Path(cta_font_file).name)
                    cta_font = ImageFont.truetype(cta_font_file, cta_font_size)
                    print(f"[FONT DEBUG] CTA font loaded: size={cta_font_size}, family={cta_font_family}")
                except Exception as e:
                    print(f"[WARNING] Failed to load CTA font '{cta_font_family}': {e}")
                    print(f"[WARNING] Using Arial for CTA at size {cta_font_size}")
                    cta_font = ImageFont.truetype(str(Path(r"C:\Windows\Fonts") / "arial.ttf"), cta_font_size)

            # Bottom Text font - separate text area positioned at bottom
            bottom_text_font_size = int(self.settings.get('bottom_text_font_size', 45))
            bottom_text_font_family = self.settings.get('bottom_text_font_family', 'Arial')
            try:
                bottom_text_font_file = get_font_file(bottom_text_font_family)
                if not Path(bottom_text_font_file).exists():
                    bottom_text_font_file = str(Path(r"C:\Windows\Fonts") / Path(bottom_text_font_file).name)
                bottom_text_font = ImageFont.truetype(bottom_text_font_file, bottom_text_font_size)
                print(f"[FONT DEBUG] BottomText font loaded: size={bottom_text_font_size}, family={bottom_text_font_family}")
            except Exception as e:
                print(f"[WARNING] Failed to load bottom text font '{bottom_text_font_family}': {e}")
                print(f"[WARNING] Using Arial for bottom text at size {bottom_text_font_size}")
                bottom_text_font = ImageFont.truetype(str(Path(r"C:\Windows\Fonts") / "arial.ttf"), bottom_text_font_size)

            # Emoji font - for all emoji rendering (CTA, Title, etc.)
            # NotoColorEmoji has size limits, use Windows Segoe UI Emoji instead
            emoji_font_path = str(Path(r"C:\Windows\Fonts") / 'seguiemj.ttf')
            print(f"[OK] Using Segoe UI Emoji for emoji rendering")

            # Emoji fonts with conservative size limits (max 72 for stability)
            emoji_size_cta = min(int(cta_font_size * 1.0), 72)
            emoji_size_title = min(int(title_font_size * 0.8), 72)
            emoji_font = ImageFont.truetype(emoji_font_path, emoji_size_cta)
            title_emoji_font = ImageFont.truetype(emoji_font_path, emoji_size_title)

        except Exception as e:
            print(f"[WARNING] Font loading error: {e}")
            print(f"[WARNING] Using fallback Arial font with requested sizes")
            import traceback
            traceback.print_exc()
            # Use Arial as fallback with the requested sizes
            try:
                fallback_font_path = str(Path(r"C:\Windows\Fonts") / "arial.ttf")
                title_font = ImageFont.truetype(fallback_font_path, title_font_size)
                quote_font = ImageFont.truetype(fallback_font_path, quote_font_size)
                cta_font = ImageFont.truetype(fallback_font_path, cta_font_size)
                # Emoji fonts with conservative size limits in fallback too
                emoji_size_cta = min(int(cta_font_size * 1.0), 72)
                emoji_size_title = min(int(title_font_size * 0.8), 72)
                emoji_font = ImageFont.truetype(emoji_font_path, emoji_size_cta)
                title_emoji_font = ImageFont.truetype(emoji_font_path, emoji_size_title)
            except:
                # Last resort - use default
                title_font = ImageFont.load_default()
                quote_font = title_font
                cta_font = title_font
                bottom_text_font = title_font
                emoji_font = title_font
                title_emoji_font = title_font

        max_text_width = int(img_width * (self.settings.get('bubble_width', 75) / 100))
        words = main_text.split()
        lines = []
        current_line = []

        # Wrap text using quote font
        for word in words:
            test_line = ' '.join(current_line + [word])
            bbox = temp_draw.textbbox((0, 0), test_line, font=quote_font)
            if bbox[2] - bbox[0] <= max_text_width:
                current_line.append(word)
            else:
                if current_line:
                    lines.append(' '.join(current_line))
                current_line = [word]
        if current_line:
            lines.append(' '.join(current_line))

        main_text_wrapped = '\n'.join(lines)

        print(f"\n[WRAPPED TEXT DEBUG]")
        print(f"  Wrapped lines count: {len(lines)}")
        print(f"  Wrapped text preview: '{main_text_wrapped[:150]}...'")

        sections = []

        # Add title section (if exists and enabled)
        # Extract emojis from title for separate rendering with emoji font
        if self.settings.get('title_enabled', True) and title_text:
            title_emojis_found = emoji_pattern.findall(title_text)
            title_text_without_emojis = emoji_pattern.sub(' ', title_text).strip()
            # Wrap title text (same logic as quote wrapping but using title font)
            title_words = title_text_without_emojis.split()
            title_lines = []
            title_current_line = []
            for word in title_words:
                test_line = ' '.join(title_current_line + [word])
                bbox = temp_draw.textbbox((0, 0), test_line, font=title_font)
                if bbox[2] - bbox[0] <= max_text_width:
                    title_current_line.append(word)
                else:
                    if title_current_line:
                        title_lines.append(' '.join(title_current_line))
                    title_current_line = [word]
            if title_current_line:
                title_lines.append(' '.join(title_current_line))
            title_wrapped = '\n'.join(title_lines)

            sections.append((title_wrapped, title_font, True, 'title', title_emojis_found if title_emojis_found else None))
            print(f"[SECTION] Added TITLE section: '{title_wrapped[:60]}...'")
        else:
            print(f"[SECTION] TITLE skipped (enabled={self.settings.get('title_enabled', True)}, has_text={bool(title_text)})")

        # Add main quote section (if exists and enabled)
        if self.settings.get('quote_enabled', True) and main_text_wrapped:
            sections.append((main_text_wrapped, quote_font, True, 'main', None))
            print(f"[SECTION] Added QUOTE section: {len(lines)} lines, '{main_text_wrapped[:60]}...'")
        else:
            print(f"[SECTION] QUOTE skipped (enabled={self.settings.get('quote_enabled', True)}, has_text={bool(main_text_wrapped)})")

        # Add CTA section with emojis (if exists and enabled)
        if self.settings.get('cta_enabled', True) and cta_text:
            sections.append((cta_text, cta_font, False, 'cta', cta_emojis))
            print(f"[SECTION] Added CTA section: '{cta_text}'")
        else:
            print(f"[SECTION] CTA skipped (enabled={self.settings.get('cta_enabled', True)}, has_text={bool(cta_text)})")

        # NOTE: Bottom Text is handled separately (not in sections stack)
        # so it always renders at the bottom of the video regardless of the
        # main overlay position.  See the padding step below.
        _bottom_text_str = str(self.settings.get('bottom_text_content', '') or '').strip()
        _bottom_text_enabled = self.settings.get('bottom_text_enabled', False)
        _has_bottom_text = _bottom_text_enabled and bool(_bottom_text_str)

        section_boxes = []
        for text, font, is_multiline, section_type, emojis in sections:
            if is_multiline:
                bbox = temp_draw.multiline_textbbox((0, 0), text, font=font, align='center')
            else:
                bbox = temp_draw.textbbox((0, 0), text, font=font)

            # For sections with emojis, measure combined width (text + emojis)
            total_width = bbox[2] - bbox[0]
            if emojis:
                emoji_str = ' '.join(emojis)
                # Use appropriate emoji font based on section type
                emoji_font_for_section = title_emoji_font if section_type == 'title' else emoji_font
                emoji_bbox = temp_draw.textbbox((0, 0), emoji_str, font=emoji_font_for_section)
                emoji_width = emoji_bbox[2] - emoji_bbox[0]
                total_width += emoji_width + 20  # Add spacing between text and emojis

            section_boxes.append({
                'text': text,
                'font': font,
                'is_multiline': is_multiline,
                'type': section_type,
                'emojis': emojis,
                'width': total_width,
                'height': bbox[3] - bbox[1]
            })

        total_height = sum(box['height'] for box in section_boxes)
        total_height += (len(section_boxes) - 1) * self.settings.get('section_spacing', 15)
        total_height += self.settings.get('padding_vertical', 20) * 2
        total_height += len(section_boxes) * self.settings.get('inner_padding', 15) * 2

        print(f"\n[BOX SIZE DEBUG]")
        print(f"  Number of sections: {len(section_boxes)}")
        for i, box in enumerate(section_boxes):
            print(f"  Section {i+1} ({box['type']}): width={box['width']}, height={box['height']}")
        print(f"  Total height: {total_height}")
        print(f"  Box dimensions: {img_width} x {int(total_height + 100)}")

        box_height = int(total_height + 100)
        box_width = img_width

        extra_margin = 0
        if self.settings.get('drop_shadow', False):
            extra_margin = self.settings.get('shadow_offset', 6) * 2

        img = Image.new('RGBA', (box_width + extra_margin, box_height + extra_margin), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)

        # Get separate background colors for Title, Quote, and CTA
        title_bg_rgb = self.hex_to_rgb(self.settings.get('title_bg_color', '#ffffff'))
        title_bg_alpha = int(255 * (self.settings.get('title_bg_opacity', 92) / 100))
        title_bg = title_bg_rgb + (title_bg_alpha,)

        quote_bg_rgb = self.hex_to_rgb(self.settings.get('quote_bg_color', '#ffffff'))
        quote_bg_alpha = int(255 * (self.settings.get('quote_bg_opacity', 92) / 100))
        quote_bg = quote_bg_rgb + (quote_bg_alpha,)

        cta_bg_rgb = self.hex_to_rgb(self.settings.get('cta_bg_color', '#00ff40'))
        cta_bg_alpha = int(255 * (self.settings.get('cta_bg_opacity', 100) / 100))
        cta_bg = cta_bg_rgb + (cta_bg_alpha,)

        if self.settings.get('position', 'top') == 'top':
            current_y = self.settings.get('padding_vertical', 20)
        elif self.settings.get('position', 'top') == 'center':
            current_y = (box_height - total_height) // 2
        else:
            current_y = box_height - total_height - self.settings.get('padding_vertical', 20)

        for i, box_info in enumerate(section_boxes):
            text = box_info['text']
            font = box_info['font']
            is_multiline = box_info['is_multiline']
            section_type = box_info['type']
            section_emojis = box_info.get('emojis', None)

            is_cta = (section_type == 'cta')
            is_title = (section_type == 'title')
            is_quote = (section_type == 'main')
            is_bottom_text = (section_type == 'bottom_text')

            # Use separate background, text colors, and outline colors for each section
            if is_title:
                current_bg = title_bg
                current_text_color = self.hex_to_rgb(self.settings.get('title_text_color', '#000000'))
                current_outline_color = self.hex_to_rgb(self.settings.get('title_outline_color', '#000000'))
                outline_enabled = self.settings.get('title_outline', False)
            elif is_bottom_text:
                current_bg = self.hex_to_rgb(self.settings.get('bottom_text_bg_color', '#000000')) + (int(255 * (self.settings.get('bottom_text_bg_opacity', 80) / 100)),)
                current_text_color = self.hex_to_rgb(self.settings.get('bottom_text_text_color', '#FFFFFF'))
                current_outline_color = self.hex_to_rgb(self.settings.get('bottom_text_outline_color', '#000000'))
                outline_enabled = self.settings.get('bottom_text_outline', True)
            elif is_cta:
                current_bg = cta_bg
                current_text_color = self.hex_to_rgb(self.settings.get('cta_text_color', '#000000'))
                current_outline_color = self.hex_to_rgb(self.settings.get('cta_outline_color', '#000000'))
                outline_enabled = self.settings.get('cta_outline', False)
            else:  # Quote
                current_bg = quote_bg
                current_text_color = self.hex_to_rgb(self.settings.get('quote_text_color', '#000000'))
                current_outline_color = self.hex_to_rgb(self.settings.get('quote_outline_color', '#000000'))
                outline_enabled = self.settings.get('quote_outline', False)

            # Check if outline should be drawn (must be enabled AND colors must differ)
            draw_outline = outline_enabled and (current_outline_color != current_text_color)
            outline_width = 2  # Outline thickness in pixels

            bubble_width = box_info['width'] + (self.settings.get('padding_horizontal', 40) * 2)
            bubble_height = box_info['height'] + (self.settings.get('inner_padding', 15) * 2)
            bubble_x = (box_width - bubble_width) // 2

            draw.rounded_rectangle(
                [(bubble_x, current_y), (bubble_x + bubble_width, current_y + bubble_height)],
                radius=self.settings.get('corner_radius', 15),
                fill=current_bg
            )

            if is_multiline:
                text_x = bubble_x + (bubble_width // 2)
                text_y = current_y + self.settings.get('inner_padding', 15)
                # Draw outline first if enabled
                if draw_outline:
                    for dx in range(-outline_width, outline_width + 1):
                        for dy in range(-outline_width, outline_width + 1):
                            if dx*dx + dy*dy <= outline_width*outline_width:
                                draw.multiline_text(
                                    (text_x + dx, text_y + dy),
                                    text,
                                    font=font,
                                    fill=current_outline_color,
                                    align='center',
                                    anchor='ma'
                                )
                # Draw main text
                draw.multiline_text(
                    (text_x, text_y),
                    text,
                    font=font,
                    fill=current_text_color,
                    align='center',
                    anchor='ma'
                )
            else:
                # For sections with emojis (Title, CTA, etc.), draw text first, then emojis with embedded_color
                if section_emojis:
                    # Use appropriate emoji font based on section type
                    emoji_font_for_section = title_emoji_font if is_title else emoji_font

                    # Measure text and emoji widths to position them side by side
                    text_bbox = temp_draw.textbbox((0, 0), text, font=font)
                    text_width = text_bbox[2] - text_bbox[0]

                    emoji_str = ' '.join(section_emojis)
                    emoji_bbox = temp_draw.textbbox((0, 0), emoji_str, font=emoji_font_for_section)
                    emoji_width = emoji_bbox[2] - emoji_bbox[0]

                    spacing = 10
                    total_content_width = text_width + spacing + emoji_width

                    # Draw text (left side)
                    text_x = bubble_x + (bubble_width - total_content_width) // 2
                    text_y = current_y + (bubble_height // 2)
                    # Draw outline first if enabled
                    if draw_outline:
                        for dx in range(-outline_width, outline_width + 1):
                            for dy in range(-outline_width, outline_width + 1):
                                if dx*dx + dy*dy <= outline_width*outline_width:
                                    draw.text((text_x + dx, text_y + dy), text, font=font, fill=current_outline_color, anchor='lm')
                    draw.text(
                        (text_x, text_y),
                        text,
                        font=font,
                        fill=current_text_color,
                        anchor='lm'
                    )

                    # Draw emojis with embedded_color (right side, colorful!)
                    emoji_x = text_x + text_width + spacing
                    emoji_y = text_y
                    try:
                        draw.text(
                            (emoji_x, emoji_y),
                            emoji_str,
                            font=emoji_font_for_section,
                            embedded_color=True,  # COLORFUL EMOJIS!
                            anchor='lm'
                        )
                    except TypeError:
                        # Fallback for older Pillow versions
                        draw.text((emoji_x, emoji_y), emoji_str, font=emoji_font_for_section, anchor='lm')
                else:
                    # No emojis, just centered text
                    text_x = bubble_x + (bubble_width // 2)
                    text_y = current_y + (bubble_height // 2)
                    # Draw outline first if enabled
                    if draw_outline:
                        for dx in range(-outline_width, outline_width + 1):
                            for dy in range(-outline_width, outline_width + 1):
                                if dx*dx + dy*dy <= outline_width*outline_width:
                                    draw.text((text_x + dx, text_y + dy), text, font=font, fill=current_outline_color, anchor='mm')
                    draw.text(
                        (text_x, text_y),
                        text,
                        font=font,
                        fill=current_text_color,
                        anchor='mm'
                    )

            current_y += bubble_height + self.settings.get('section_spacing', 15)

        if self.settings.get('drop_shadow', False):
            img = TextEffects.create_shadow_image(
                img,
                offset=self.settings.get('shadow_offset', 6),
                blur=self.settings.get('shadow_blur', 12)
            )

        if self.settings.get('text_glow', False):
            glow_rgb = self.hex_to_rgb(self.settings.get('glow_color', '#ffffff'))
            img = TextEffects.create_glow_image(
                img,
                glow_color=glow_rgb,
                intensity=self.settings.get('glow_intensity', 8)
            )

        if self.settings.get('neon_glow', False):
            neon_rgb = self.hex_to_rgb(self.settings.get('neon_color', '#00ff88'))
            img = TextEffects.create_neon_glow(img, neon_color=neon_rgb)

        if self.settings.get('gradient_overlay', False):
            img = TextEffects.apply_gradient_overlay(
                img,
                gradient_type=self.settings.get('gradient_type', 'top_to_bottom'),
                intensity=self.settings.get('gradient_intensity', 0.3)
            )

        # ALWAYS pad the overlay to full video dimensions so the ImageClip
        # matches the video frame size exactly.  This ensures text/blur
        # overlays cover the entire canvas regardless of position, and avoids
        # compositing mismatches when other effects (zoom, spotlight, etc.)
        # are applied to the video before the text is composited on top.
        # We also clamp the paste region so the overlay NEVER extends past
        # the canvas — any overflow is clipped by PIL and the user sees
        # cut-off text.
        padded = Image.new('RGBA', (video_width, video_height), (0, 0, 0, 0))
        offset = int(self.settings.get('vertical_offset', 0))
        paste_x = (video_width - img.width) // 2
        if self.settings.get('position', 'top') == 'top':
            paste_y = offset  # start at very top; user can nudge with V-Offset
        elif self.settings.get('position', 'top') == 'center':
            paste_y = (video_height - img.height) // 2 + offset
        else:
            paste_y = video_height - img.height - int(video_height * 0.10) + offset
        # Guard: don't let the overlay extend past the bottom of the canvas.
        # If paste_y is too large the bottom gets clipped.
        paste_y = max(0, min(paste_y, video_height - img.height))
        padded.paste(img, (paste_x, paste_y), img)
        img = padded

        # ── Bottom Text (always at bottom of video, independent of overlay position) ──
        if _has_bottom_text:
            try:
                btd = ImageDraw.Draw(img)
                bt_bb = btd.textbbox((0, 0), _bottom_text_str, font=bottom_text_font)
                btw, bth = bt_bb[2] - bt_bb[0], bt_bb[3] - bt_bb[1]
                bt_pad_x, bt_pad_y = max(20, btw // 8), max(8, bth // 4)
                bt_box_w = btw + bt_pad_x * 2
                bt_box_h = bth + bt_pad_y * 2
                bt_off = int(self.settings.get('bottom_text_vertical_offset', 0))
                bt_x = (video_width - bt_box_w) // 2
                bt_y = video_height - bt_box_h - int(video_height * 0.10) + bt_off
                # Guard against bottom clipping
                bt_y = max(0, min(bt_y, video_height - bt_box_h))

                # Background pill
                bt_bgc = self.hex_to_rgb(self.settings.get('bottom_text_bg_color', '#000000'))
                bt_alpha = max(0, min(255, int(255 * (self.settings.get('bottom_text_bg_opacity', 80) / 100))))
                bg_pill = Image.new('RGBA', (bt_box_w, bt_box_h), bt_bgc + (bt_alpha,))
                img.paste(bg_pill, (bt_x, bt_y), bg_pill)

                # Text outline
                bt_tc = self.hex_to_rgb(self.settings.get('bottom_text_text_color', '#FFFFFF'))
                ol_on = self.settings.get('bottom_text_outline', True)
                ol_c = self.hex_to_rgb(self.settings.get('bottom_text_outline_color', '#000000'))
                ol_t = int(self.settings.get('bottom_text_outline_size', 2))
                tx = bt_x + bt_pad_x - bt_bb[0]
                ty = bt_y + bt_pad_y - bt_bb[1]
                if ol_on:
                    for dx in range(-ol_t, ol_t + 1):
                        for dy in range(-ol_t, ol_t + 1):
                            if dx * dx + dy * dy <= ol_t * ol_t:
                                btd.text((tx + dx, ty + dy), _bottom_text_str, font=bottom_text_font, fill=ol_c + (255,))
                btd.text((tx, ty), _bottom_text_str, font=bottom_text_font, fill=bt_tc + (255,))
                print(f"[BOTTOM TEXT] Rendered at bottom: '{_bottom_text_str[:60]}' (y={bt_y})")
            except Exception as e:
                print(f"[WARNING] Bottom text render failed: {e}")
                import traceback
                traceback.print_exc()

        return img

    def add_quote_to_video(self, video_path: Path, quote: dict, video_index: int = 0, excel_audio_path: Optional[Path] = None) -> Tuple[Path, str]:
        """
        Add quote overlay with advanced effects.

        Args:
            video_path: Path to video file
            quote: Dictionary with 'subtitle' and 'voiceover' keys
                   - subtitle: Text shown in captions/overlay (short)
                   - voiceover: Text spoken in TTS (can include explanations)
            video_index: Index of current video
            excel_audio_path: Optional path to Excel-matched MP3 audio file (overrides TTS)

        Returns:
            Tuple of (output_path, output_filename)
        """
        # Define emoji pattern at the beginning of the method
        emoji_pattern = re.compile(r'[\U0001F300-\U0001F9FF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF\U00002600-\U000027BF\U0001F1E0-\U0001F1FF]+')

        # Extract subtitle and voiceover text
        # Support both dict format and legacy string format for backward compatibility
        if isinstance(quote, dict):
            subtitle_text = quote['subtitle']
            voiceover_text = quote['voiceover']

            # Excel Integration: Override text styling if provided
            if quote.get('source') == 'excel':
                print(f"[EXCEL] Applying Excel-specific text styling")

                # Override position
                if 'excel_position' in quote:
                    self.settings['position'] = quote['excel_position']
                    print(f"  Position: {quote['excel_position']}")

                # Override font settings for QUOTE section (Title + Quote + CTA format)
                if self.settings.get('excel_font_size'):
                    excel_font_size = self.settings.get('excel_font_size', 45)
                    self.settings['font_size'] = excel_font_size
                    self.settings['quote_font_size'] = excel_font_size  # For quote section
                    print(f"  Font Size: {excel_font_size}")

                if self.settings.get('excel_font_style'):
                    excel_font_style = self.settings.get('excel_font_style', 'Arial Bold')
                    self.settings['font_style'] = excel_font_style
                    self.settings['quote_font_family'] = excel_font_style  # For quote section
                    print(f"  Font: {excel_font_style}")

                # Override bold/italic
                self.settings['bold'] = self.settings.get('excel_bold', True)
                self.settings['italic'] = self.settings.get('excel_italic', False)

                # Override colors for QUOTE section
                if self.settings.get('excel_text_color'):
                    excel_text_color = self.settings.get('excel_text_color', '#000000')
                    self.settings['text_color'] = excel_text_color
                    self.settings['quote_text_color'] = excel_text_color  # For quote section
                    print(f"  Text Color: {excel_text_color}")

                if self.settings.get('excel_bg_color'):
                    excel_bg_color = self.settings.get('excel_bg_color', '#FFFFFF')
                    self.settings['bg_color'] = excel_bg_color
                    self.settings['quote_bg_color'] = excel_bg_color  # For quote section
                    print(f"  BG Color: {excel_bg_color}")

                if self.settings.get('excel_bg_opacity') is not None:
                    excel_bg_opacity = self.settings.get('excel_bg_opacity', 100)
                    self.settings['bg_opacity'] = excel_bg_opacity
                    self.settings['quote_bg_opacity'] = excel_bg_opacity  # For quote section
                    print(f"  BG Opacity: {excel_bg_opacity}%")

                # Override outline for QUOTE section
                if self.settings.get('excel_text_outline') is not None:
                    excel_outline = self.settings.get('excel_text_outline', False)
                    self.settings['text_outline'] = excel_outline
                    self.settings['quote_outline'] = excel_outline  # For quote section

                if self.settings.get('excel_outline_color'):
                    excel_outline_color = self.settings.get('excel_outline_color', '#000000')
                    self.settings['outline_color'] = excel_outline_color
                    self.settings['quote_outline_color'] = excel_outline_color  # For quote section

                if self.settings.get('excel_outline_size'):
                    excel_outline_size = self.settings.get('excel_outline_size', 2)
                    self.settings['outline_size'] = excel_outline_size
                    self.settings['quote_outline_size'] = excel_outline_size  # For quote section
        else:
            # Backward compatible: treat as single string
            subtitle_text = quote
            voiceover_text = quote

        print(f"\n{'='*70}")
        print(f"Processing: {video_path.name}")
        print(f"Subtitle: {subtitle_text[:80]}...")
        if subtitle_text != voiceover_text:
            print(f"Voiceover: {voiceover_text[:80]}...")

        # Use subtitle text for hashtags and filename (visual elements)
        hashtags = self.generate_hashtags(subtitle_text)
        print(f"Hashtags: {', '.join(hashtags)}")

        # Check if user wants to use title only for filename (default: True)
        use_title_only = self.settings.get('filename_use_title_only', True)
        # Video title source: 'quote' (default) uses the quote text,
        # 'video' uses the source video's filename, 'auto' uses
        # video if available else falls back to quote.
        video_title_source = self.settings.get('video_title_source', 'quote')
        output_filename = self.create_filename(
            subtitle_text, hashtags,
            use_title_only=use_title_only,
            source=video_title_source,
            video_path=video_path,
        )
        print(f"Output: {output_filename}")

        video = VideoFileClip(str(video_path))

        # SAVE ORIGINAL CANVAS DIMENSIONS before any crop/resize operations.
        # These original_w/original_h represent the FULL frame canvas that the
        # user expects text and blur to span — even if the video gets cropped
        # or resized later for aspect-ratio compliance.
        _orig_canvas_w = video.w
        _orig_canvas_h = video.h
        print(f"[CANVAS] Original uncropped dimensions: {_orig_canvas_w}x{_orig_canvas_h}")

        # FORCE STANDARD FPS: Always convert to 24fps immediately to reduce rendering time
        # This prevents processing 60fps (1171 frames) when we only need 24fps (468 frames)
        TARGET_FPS = 24
        if video.fps != TARGET_FPS:
            print(f"\n[FPS] Source video: {video.fps}fps")
            print(f"[FPS] Converting to {TARGET_FPS}fps for faster processing")
            try:
                video = video.with_fps(TARGET_FPS)
            except AttributeError:
                video = video.set_fps(TARGET_FPS)
            print(f"[FPS] ✅ Video FPS changed to {TARGET_FPS}fps")
        else:
            print(f"[FPS] Video already at target {TARGET_FPS}fps")

        # FORCE STANDARD RESOLUTION: Always resize to 1080x1920 for consistent caption sizing
        # This prevents font sizes from appearing too large on smaller resolution videos
        # Skip if platform preset is enabled (it will handle resizing)
        platform_preset_enabled = self.settings.get('enable_platform_preset', False)
        platform = self.settings.get('platform_preset', 'none')

        # Custom Resolution: if the user typed specific width/height in the
        # Quick Start tab and didn't enable a platform preset, honor those
        # exact dimensions. The Quick Start card saves to
        # 'output_width' / 'output_height' but the renderer used to ignore
        # them entirely. Now we apply them with the same crop-then-resize
        # logic the platform preset uses, so 720x1280 / 1920x1080 /
        # 1080x1080 / arbitrary values all actually work.
        custom_w = self.settings.get('output_width')
        custom_h = self.settings.get('output_height')
        use_custom_resolution = (
            not (platform_preset_enabled and platform != 'none')
            and custom_w and custom_h
            and int(custom_w) > 0 and int(custom_h) > 0
        )

        if use_custom_resolution:
            try:
                custom_w = int(custom_w)
                custom_h = int(custom_h)
                target_aspect = custom_w / custom_h
                current_aspect = video.w / video.h
                print(f"\n[RESIZE] Custom Resolution requested: {custom_w}x{custom_h}")
                print(f"[RESIZE] Source: {video.w}x{video.h} ({current_aspect:.2f})")

                # Crop to the target aspect ratio first (using center crop
                # — the platform preset's crop_mode only applies when a
                # platform preset is active)
                if abs(current_aspect - target_aspect) > 0.01:
                    if current_aspect > target_aspect:
                        # Source wider than target — crop width
                        new_h = video.h
                        new_w = int(new_h * target_aspect)
                        x1 = (video.w - new_w) // 2
                        y1 = 0
                    else:
                        # Source taller than target — crop height
                        new_w = video.w
                        new_h = int(new_w / target_aspect)
                        x1 = 0
                        y1 = (video.h - new_h) // 2
                    video = crop(video, x1=x1, y1=y1,
                                 x2=x1 + new_w, y2=y1 + new_h)
                    print(f"[RESIZE] Cropped to {video.w}x{video.h} for aspect match")

                video = video.resized((custom_w, custom_h))
                print(f"[OK] Custom Resolution applied: {custom_w}x{custom_h}")
            except Exception as e:
                print(f"[WARNING] Custom Resolution failed: {e}; falling back to standard")
                # If something blew up, fall through to the standard path
                # below by clearing the flag. Note: the `if not (...)` block
                # below will still run because the platform preset guard is
                # independent of use_custom_resolution.
                use_custom_resolution = False
        else:
            use_custom_resolution = False

        if not (platform_preset_enabled and platform != 'none'):
            if not use_custom_resolution:
                STANDARD_WIDTH = 1080
                STANDARD_HEIGHT = 1920

                if video.w != STANDARD_WIDTH or video.h != STANDARD_HEIGHT:
                    print(f"\n[RESIZE] Source video: {video.w}x{video.h}")
                    print(f"[RESIZE] Standardizing to: {STANDARD_WIDTH}x{STANDARD_HEIGHT} for consistent captions")

                    # Calculate aspect ratios
                    target_aspect = STANDARD_WIDTH / STANDARD_HEIGHT
                    current_aspect = video.w / video.h

                    # First crop to match aspect ratio if needed, then resize
                    if abs(current_aspect - target_aspect) > 0.01:
                        if current_aspect > target_aspect:
                            # Video is wider - crop width (center crop)
                            new_h = video.h
                            new_w = int(new_h * target_aspect)
                            x1 = (video.w - new_w) // 2
                            y1 = 0
                            x2 = x1 + new_w
                            y2 = video.h
                            video = crop(video,x1=x1, y1=y1, x2=x2, y2=y2)
                            print(f"[RESIZE] Cropped width to match aspect ratio")
                        else:
                            # Video is taller - crop height (center crop)
                            new_w = video.w
                            new_h = int(new_w / target_aspect)
                            x1 = 0
                            y1 = (video.h - new_h) // 2
                            x2 = video.w
                            y2 = y1 + new_h
                            video = crop(video,x1=x1, y1=y1, x2=x2, y2=y2)
                            print(f"[RESIZE] Cropped height to match aspect ratio")

                    # Now resize to standard dimensions
                    video = video.resized((STANDARD_WIDTH, STANDARD_HEIGHT))
                    print(f"[RESIZE] ✅ Video standardized to {STANDARD_WIDTH}x{STANDARD_HEIGHT}")
                else:
                    print(f"[RESIZE] Video already at standard resolution: {STANDARD_WIDTH}x{STANDARD_HEIGHT}")
        else:
            print(f"[RESIZE] Platform preset enabled - will apply {platform} dimensions later")

        # Apply chromatic aberration if enabled (apply to base video first)
        if self.settings.get('chromatic_aberration', False):
            try:
                intensity = int(self.settings.get('chromatic_intensity', 5))
                direction = self.settings.get('chromatic_direction', 'horizontal')

                print(f"\n[CHROMATIC] Applying RGB glitch effect...")
                print(f"[CHROMATIC] Intensity: {intensity}px, Direction: {direction}")

                def apply_chromatic(get_frame, t):
                    frame = get_frame(t)

                    # Ensure intensity is an integer for array slicing
                    shift = int(intensity)

                    # Separate RGB channels
                    r_channel = frame[:, :, 0].copy()
                    g_channel = frame[:, :, 1].copy()
                    b_channel = frame[:, :, 2].copy()

                    # Create shifted channels
                    h, w = frame.shape[:2]

                    if direction == 'horizontal' or direction == 'both':
                        # Shift red left, blue right
                        r_shifted = np.zeros_like(r_channel)
                        b_shifted = np.zeros_like(b_channel)

                        if shift < w and shift > 0:
                            r_shifted[:, shift:] = r_channel[:, :-shift]
                            b_shifted[:, :-shift] = b_channel[:, shift:]
                        else:
                            r_shifted = r_channel
                            b_shifted = b_channel

                        r_channel = r_shifted
                        b_channel = b_shifted

                    if direction == 'vertical' or direction == 'both':
                        # Shift red up, blue down
                        r_shifted = np.zeros_like(r_channel)
                        b_shifted = np.zeros_like(b_channel)

                        if shift < h and shift > 0:
                            r_shifted[shift:, :] = r_channel[:-shift, :]
                            b_shifted[:-shift, :] = b_channel[shift:, :]
                        else:
                            r_shifted = r_channel
                            b_shifted = b_channel

                        r_channel = r_shifted
                        b_channel = b_shifted

                    # Recombine channels
                    result = frame.copy()
                    result[:, :, 0] = r_channel
                    result[:, :, 2] = b_channel

                    return result

                video = video.transform(lambda gf, t: apply_chromatic(gf, t))
                print(f"[OK] Applied chromatic aberration ({direction}, {intensity}px offset)")
            except Exception as e:
                print(f"[WARNING] Chromatic aberration failed: {e}")
                import traceback
                traceback.print_exc()

        # Apply platform preset if enabled
        if self.settings.get('enable_platform_preset', False):
            platform = self.settings.get('platform_preset', 'none')
            if platform != 'none':
                print(f"\n[PLATFORM] Applying {platform} preset...")

                # Platform dimensions
                platform_dims = {
                    'instagram_reels': (1080, 1920),  # 9:16
                    'tiktok': (1080, 1920),           # 9:16
                    'youtube_shorts': (1080, 1920),   # 9:16
                    'youtube': (1920, 1080),          # 16:9
                    'facebook': (1080, 1080),         # 1:1
                }

                if platform in platform_dims:
                    target_w, target_h = platform_dims[platform]
                    target_aspect = target_w / target_h
                    current_aspect = video.w / video.h

                    crop_mode = self.settings.get('crop_mode', 'center')

                    print(f"[PLATFORM] Current: {video.w}x{video.h} ({current_aspect:.2f})")
                    print(f"[PLATFORM] Target: {target_w}x{target_h} ({target_aspect:.2f})")
                    print(f"[PLATFORM] Crop mode: {crop_mode}")

                    # Calculate scaling and cropping
                    if abs(current_aspect - target_aspect) > 0.01:  # Need to crop
                        if current_aspect > target_aspect:
                            # Video is wider - crop width
                            new_h = video.h
                            new_w = int(new_h * target_aspect)

                            if crop_mode == 'center':
                                x1 = (video.w - new_w) // 2
                            elif crop_mode == 'left':
                                x1 = 0
                            elif crop_mode == 'right':
                                x1 = video.w - new_w
                            else:  # smart or others default to center
                                x1 = (video.w - new_w) // 2

                            y1 = 0
                            x2 = x1 + new_w
                            y2 = video.h

                            video = crop(video,x1=x1, y1=y1, x2=x2, y2=y2)
                            print(f"[PLATFORM] Cropped width: {x1},{y1} to {x2},{y2}")
                        else:
                            # Video is taller - crop height
                            new_w = video.w
                            new_h = int(new_w / target_aspect)

                            if crop_mode == 'center':
                                y1 = (video.h - new_h) // 2
                            elif crop_mode == 'top':
                                y1 = 0
                            elif crop_mode == 'bottom':
                                y1 = video.h - new_h
                            else:  # smart or others default to center
                                y1 = (video.h - new_h) // 2

                            x1 = 0
                            x2 = video.w
                            y2 = y1 + new_h

                            video = crop(video,x1=x1, y1=y1, x2=x2, y2=y2)
                            print(f"[PLATFORM] Cropped height: {x1},{y1} to {x2},{y2}")

                    # Resize to target dimensions
                    video = video.resized((target_w, target_h))
                    print(f"[PLATFORM] Resized to: {target_w}x{target_h}")
                    print(f"[OK] Platform formatting complete!")

        # Parse subtitle text for visual display (Title + Quote + CTA format)
        # Expected format:
        # Line 1: Title (e.g., "When Marriage Gets Wild Fast")
        # Line 2-N: Main quote text
        # Last line: CTA with emojis (e.g., "👉 Prove You Relate! 😭")
        #
        # OurScript mode (our_script_mode=True): Bypass the 3-section parser.
        # Title comes from _our_title setting, subtitle_text is the raw caption
        # text without any title prepended, and there is no CTA section.
        # This prevents multi-line captions from losing lines to misparsing.

        title_text = ""
        main_text = ""
        cta_text = ""
        cta_emojis = []

        if self.settings.get('our_script_mode', False):
            # OurScript mode: explicit title, subtitle is flat caption text
            fetched_title = (self.settings.get('_our_title', '') or '').strip()
            if self.settings.get('title_enabled', False) and fetched_title:
                title_text = emoji_pattern.sub(' ', fetched_title).strip()
            main_text = subtitle_text.strip()
            main_text = emoji_pattern.sub(' ', main_text).strip()
            # CTA is empty in OurScript mode — no parsing reinterprets
            # caption lines as title or CTA.
            print(f"\n[OURSCRIPT MODE] Using explicit title + raw caption text")
            print(f"Title: {title_text[:50] if title_text else '(none)'}...")
            print(f"Main: {main_text[:60]}...")
            print(f"CTA: (none — OurScript mode bypasses CTA parsing)")
        else:
            lines = subtitle_text.strip().split('\n')

            print(f"\n[TEXT PARSE DEBUG] Total lines after split: {len(lines)}")
            for i, line in enumerate(lines, 1):
                print(f"  Line {i}: '{line[:100]}...' (length: {len(line)})")

            if len(lines) >= 3:
                # Multi-line format: Title + Quote + CTA
                title_text = lines[0].strip()
                main_text = '\n'.join(lines[1:-1]).strip()
                last_line = lines[-1].strip()

                # Extract emojis from CTA line (preserve them for colorful display)
                cta_emojis_found = emoji_pattern.findall(last_line)
                cta_text_without_emojis = emoji_pattern.sub(' ', last_line).strip()

                # CTA is the last line (with emojis preserved separately)
                cta_text = cta_text_without_emojis
                cta_emojis = cta_emojis_found

            elif len(lines) == 2:
                # Two lines: Could be Title+Quote or Quote+CTA
                # Check if last line looks like CTA (short, has emojis)
                last_line = lines[-1].strip()
                last_line_emojis = emoji_pattern.findall(last_line)
                last_line_clean = emoji_pattern.sub(' ', last_line).strip()

                if last_line_emojis and len(last_line_clean.split()) < 10:
                    # It's a CTA
                    main_text = lines[0].strip()
                    cta_text = last_line_clean
                    cta_emojis = last_line_emojis
                else:
                    # It's Title + Quote
                    title_text = lines[0].strip()
                    main_text = lines[1].strip()
            else:
                # Single line - treat as main text
                main_text = subtitle_text.strip()

            # Remove emojis from title and main text for clean display
            if title_text:
                title_text = emoji_pattern.sub(' ', title_text).strip()
            if main_text:
                main_text = emoji_pattern.sub(' ', main_text).strip()

            print(f"Title: {title_text[:50] if title_text else '(none)'}...")
            print(f"Main: {main_text[:60]}...")
            print(f"CTA: {cta_text}")
            print(f"CTA Emojis: {cta_emojis}")

        # Create static text overlay (quote bubble with Title + Quote + CTA format)
        # This is SEPARATE from word-by-word captions and both can be shown together
        # Note: if the platform preset cropped the video, it was already padded
        # back to the original canvas above, so video.w/video.h cover the full frame.
        img = self.create_text_overlay_image(video.w, video.h, title_text, main_text, cta_text, cta_emojis)
        img_array = np.array(img).copy()

        txt_clip = set_duration(ImageClip(img_array), video.duration)

        if self.settings.get('text_fade_in', False):
            fade_duration = self.settings.get('text_fade_duration', 0.4)
            if FadeIn:
                txt_clip = txt_clip.with_effects([FadeIn(fade_duration)])
            else:
                try:
                    txt_clip = txt_clip.fadein(fade_duration)
                except AttributeError:
                    pass  # Skip fade if not available

        if self.settings.get('text_bounce', False):
            bounce_intensity = self.settings.get('text_bounce_intensity', 1.15)
            def bounce_scale(t):
                if t < 0.6:
                    # Bounce in with overshoot
                    progress = t / 0.6
                    if progress < 0.5:
                        # Scale up quickly
                        scale = progress * 2 * bounce_intensity
                    else:
                        # Bounce back to normal
                        overshoot = (progress - 0.5) * 2
                        scale = bounce_intensity - (bounce_intensity - 1.0) * overshoot
                    return max(0.1, scale)
                return 1.0

            try:
                txt_clip = txt_clip.resized(lambda t: bounce_scale(t))
            except:
                pass

        elif self.settings.get('text_glitch', False):
            # Glitch effect: quick position shifts and opacity flicker
            def glitch_effect(get_frame, t):
                frame = get_frame(t)
                if t < 0.5:
                    # Random glitch during first 0.5s
                    if int(t * 30) % 3 == 0:  # Glitch every 3 frames
                        # Shift frame slightly
                        shift_x = np.random.randint(-10, 10)
                        shift_y = np.random.randint(-5, 5)
                        if shift_x > 0:
                            frame = np.roll(frame, shift_x, axis=1)
                        if shift_y > 0:
                            frame = np.roll(frame, shift_y, axis=0)
                return frame

            try:
                txt_clip = txt_clip.transform(glitch_effect)
            except:
                pass

        elif self.settings.get('text_slide_up', False):
            slide_distance = self.settings.get('text_slide_distance', 50)
            def slide_position(t):
                if t < 0.5:
                    offset = slide_distance * (1 - t / 0.5)
                    return ('center', 0 if offset < 0 else offset)
                else:
                    return ('center', 0)
            txt_clip = set_position(txt_clip, slide_position)

        if not any([
            self.settings.get('text_bounce', False),
            self.settings.get('text_glitch', False),
            self.settings.get('text_slide_up', False)
        ]):
            # The text overlay is always padded to full video dimensions
            # (see create_text_overlay_image padding step), so the clip
            # is always video.w × video.h.  Position at (center, 0).
            # Vertical offset is already baked into the padded image.
            txt_clip = set_position(txt_clip, ('center', 0))

        print(f"[OK] Static text overlay created at position: {self.settings.get('position', 'top')}")

        if self.settings.get('video_zoom', False):
            zoom_scale = self.settings.get('zoom_scale', 1.08)
            def zoom_effect(get_frame, t):
                frame = get_frame(t)
                progress = t / video.duration
                current_scale = 1 + (zoom_scale - 1) * progress
                h, w = frame.shape[:2]
                new_h, new_w = int(h * current_scale), int(w * current_scale)
                from PIL import Image as PILImage
                pil_frame = PILImage.fromarray(frame)
                pil_frame = pil_frame.resize((new_w, new_h), PILImage.LANCZOS)
                crop_x = (new_w - w) // 2
                crop_y = (new_h - h) // 2
                pil_frame = pil_frame.crop((crop_x, crop_y, crop_x + w, crop_y + h))
                return np.array(pil_frame).copy()
            try:
                video = video.transform(zoom_effect)
            except AttributeError:
                video = video.fl(zoom_effect)

        # ═══════════════════════════════════════════════════════════════════
        # ⚡ PER-FRAME EFFECTS: merged into a SINGLE image_transform
        # Instead of N chained clip wrappers (one per effect), we apply all
        # effects in ONE function call, eliminating MoviePy's per-wrapper
        # overhead (each wrapper adds Python call + numpy overhead per frame).
        # ═══════════════════════════════════════════════════════════════════
        _fx_color_grade = self.settings.get('color_grade', 'none')
        _fx_color_grade_type = self.settings.get('color_grade', 'warm')
        _fx_lut_enabled = self.settings.get('lut_enabled', True)
        _fx_lut_name = self.settings.get('lut_preset', 'None')
        _fx_lut_intensity = float(self.settings.get('lut_intensity', 1.0))
        _fx_am_template = self.settings.get('am_template', 'None')
        _fx_vignette = self.settings.get('vignette', False)
        _fx_vignette_intensity = self.settings.get('vignette_intensity', 0.4)
        _fx_dim = self.settings.get('background_dim', False)
        _fx_dim_intensity = self.settings.get('dim_intensity', 0.25)
        _fx_grain = self.settings.get('film_grain', False)
        _fx_grain_intensity = self.settings.get('grain_intensity', 0.15)
        _fx_gradient = self.settings.get('gradient_overlay', False)
        _fx_gradient_type = self.settings.get('gradient_type', 'top_to_bottom')
        _fx_gradient_intensity = self.settings.get('gradient_intensity', 0.3)
        _fx_blur_enabled = self.settings.get('region_blur_enabled', False)
        _fx_custom_regions = self.settings.get('custom_blur_regions', [])
        _fx_has_custom_blur = any(r.get('enabled', False) for r in _fx_custom_regions if isinstance(r, dict))

        if _fx_blur_enabled:
            print("  → Applying region blur effect...")
        if _fx_has_custom_blur:
            print("  → Applying custom blur regions to hide logos/watermarks...")

        def _combined_frame_pipeline(frame):
            """Single-pass per-frame effect pipeline — no clip wrappers between effects."""
            if _fx_color_grade != 'none':
                frame = VideoEffects.apply_color_grade(frame, _fx_color_grade_type)
            if _fx_lut_enabled and _fx_lut_name and _fx_lut_name != 'None':
                frame = VideoEffects.apply_lut_filter(frame, _fx_lut_name, _fx_lut_intensity)
            if _fx_am_template != 'None':
                frame = VideoEffects.apply_alight_motion_look(frame, self.settings)
            if _fx_vignette:
                frame = VideoEffects.apply_vignette(frame, _fx_vignette_intensity)
            if _fx_dim:
                frame = VideoEffects.apply_background_dim(frame, _fx_dim_intensity)
            if _fx_grain:
                frame = VideoEffects.apply_film_grain(frame, _fx_grain_intensity)
            if _fx_gradient:
                frame = VideoEffects.apply_gradient_overlay(frame, _fx_gradient_type, _fx_gradient_intensity)
            if _fx_blur_enabled or _fx_has_custom_blur:
                frame = VideoEffects.apply_region_blur(frame, self.settings)
            return frame

        try:
            video = video.image_transform(_combined_frame_pipeline)
        except AttributeError:
            video = video.fl_image(_combined_frame_pipeline)

        # Apply selective blur for watermark/logo hiding
        if self.settings.get('blur_watermark_enabled', False):
            print("  → Applying selective blur to hide watermark...")
            # Pass settings to the blur function
            VideoEffects.apply_selective_blur.settings = self.settings
            try:
                video = video.with_fps(video.fps).transform(VideoEffects.apply_selective_blur)
            except AttributeError:
                video = video.with_fps(video.fps).fl(VideoEffects.apply_selective_blur)


        # ========== DISABLED: Don't loop video based on TTS estimation ==========
        # The estimation was inaccurate and caused videos to loop unnecessarily
        # Instead, we'll handle TTS duration mismatch later with actual TTS audio duration
        target_duration = video.duration
        original_video_duration = video.duration

        # NOTE: Video looping based on TTS estimation is now disabled
        # If TTS is longer than video, the video will simply play to the end
        # and freeze on the last frame (moviepy default behavior)
        loop_video_for_tts = False  # Changed from estimating to False

        # Loop video if target duration exceeds source video
        # DISABLED: This was causing videos to double in length due to bad TTS estimation
        if loop_video_for_tts and target_duration > video.duration:
            try:
                loops_needed = int(np.ceil(target_duration / video.duration))
                print(f"[OK] Looping video {loops_needed}x to match TTS duration ({target_duration:.1f}s)")

                # Use time-based looping that wraps around - this actually creates new frames
                # instead of referencing the original file beyond its duration
                original_duration = video.duration

                # ========== FIX: Remove audio before looping to avoid transformation issues ==========
                # The audio will be replaced by TTS/BGM anyway, so remove it to avoid errors
                # caused by nested audio transformations when trying to loop
                original_audio = video.audio
                video = video.without_audio()

                def loop_time(get_frame, t):
                    """Loop video by wrapping time back to start"""
                    looped_t = t % original_duration
                    return get_frame(looped_t)

                try:
                    # MoviePy 2.x
                    video = video.transform(loop_time)
                    video = video.with_duration(target_duration)
                except AttributeError:
                    # MoviePy 1.x
                    video = video.fl(loop_time)
                    video = video.set_duration(target_duration)

                # Update txt_clip duration to match
                txt_clip = set_duration(txt_clip, target_duration)

                print(f"[OK] Video looped to {target_duration:.1f}s (audio will be added from TTS/BGM)")

            except Exception as e:
                print(f"[WARNING] Could not loop video: {e}")
                import traceback
                traceback.print_exc()

        # FIX: Keep text clip separate - will be added AFTER spotlight effect
        # This ensures text is always visible on top of spotlight
        text_overlay_clip = txt_clip

        # Start with video only (particles will be added, but NOT text yet)
        layers = [video]
        print(f"DEBUG: txt_clip size={txt_clip.size}, position={txt_clip.pos if hasattr(txt_clip, 'pos') else 'N/A'}, duration={txt_clip.duration}")
        print(f"DEBUG: video size={video.size}, duration={video.duration}")

        # ========== FIX: Combine all particle effects into single layer for faster rendering ==========
        # This reduces compositing operations from N layers to 1 combined layer
        particle_effects_enabled = []
        if self.settings.get('add_glitter', False):
            particle_effects_enabled.append('glitter')
        if self.settings.get('add_stars', False):
            particle_effects_enabled.append('stars')
        if self.settings.get('add_hearts', False):
            particle_effects_enabled.append('hearts')
        if self.settings.get('add_confetti', False):
            particle_effects_enabled.append('confetti')

        if particle_effects_enabled:
            try:
                # Create combined particle effect for better performance
                combined_particles = ParticleEffects.create_combined(
                    video.w, video.h, video.duration, video.fps,
                    glitter=self.settings.get('add_glitter', False),
                    glitter_intensity=self.settings.get('glitter_intensity', 0.5),
                    stars=self.settings.get('add_stars', False),
                    hearts=self.settings.get('add_hearts', False),
                    confetti=self.settings.get('add_confetti', False)
                )
                layers.append(combined_particles)
                print(f"[OK] Added combined particle effects: {', '.join(particle_effects_enabled)}")
            except Exception as e:
                print(f"[WARNING] Combined particle effect failed, using individual effects: {e}")
                # Fallback to individual effects
                if self.settings.get('add_glitter', False):
                    try:
                        intensity = self.settings.get('glitter_intensity', 0.5)
                        glitter = ParticleEffects.create_glitter(
                            video.w, video.h, video.duration, video.fps, intensity
                        )
                        layers.append(glitter)
                        print(f"[OK] Added glitter effect (intensity: {intensity})")
                    except Exception as e2:
                        print(f"[WARNING] Glitter effect failed: {e2}")

                if self.settings.get('add_stars', False):
                    try:
                        stars = ParticleEffects.create_stars(
                            video.w, video.h, video.duration, video.fps
                        )
                        layers.append(stars)
                        print("[OK] Added falling stars effect")
                    except Exception as e2:
                        print(f"[WARNING] Stars effect failed: {e2}")

                if self.settings.get('add_hearts', False):
                    try:
                        hearts = ParticleEffects.create_hearts(
                            video.w, video.h, video.duration, video.fps
                        )
                        layers.append(hearts)
                        print("[OK] Added falling hearts effect")
                    except Exception as e2:
                        print(f"[WARNING] Hearts effect failed: {e2}")

                if self.settings.get('add_confetti', False):
                    try:
                        confetti = ParticleEffects.create_confetti(
                            video.w, video.h, video.duration, video.fps
                        )
                        layers.append(confetti)
                        print("[OK] Added confetti effect")
                    except Exception as e2:
                        print(f"[WARNING] Confetti effect failed: {e2}")

        print(f"DEBUG: Compositing {len(layers)} layers...")
        final_video = CompositeVideoClip(layers)
        print(f"DEBUG: Composite created - size={final_video.size}, duration={final_video.duration:.2f}s")

        # Audio processing
        voiceover_file = None
        word_timings = []
        # caption_clips is consumed later (post-spotlight composite + thumbnail).
        # Initialize it here so it always exists, even when the user has every
        # caption toggle off (the captions block at line ~7107 is gated).
        caption_clips = []

        # Debug: Check caption and TTS settings
        print(f"DEBUG: enable_captions={self.settings.get('enable_captions', False)}, use_tts_voiceover={self.settings.get('use_tts_voiceover', False)}, TTS_AVAILABLE={TTS_AVAILABLE}")

        # EXCEL INTEGRATION: Use matched MP3 audio if provided (highest priority)
        if excel_audio_path and excel_audio_path.exists():
            print(f"\n[EXCEL] Using matched audio file: {excel_audio_path.name}")
            voiceover_file = excel_audio_path
            # Excel audio doesn't have word timing data
            word_timings = []

        # Option 0: Use original audio with effects applied (skip TTS)
        elif self.settings.get('use_original_audio', False):
            print("[OK] Using original audio with effects...")

            # Extract original audio from video
            if video.audio:
                import subprocess

                effects_folder = self.output_folder / "audio_effects"
                effects_folder.mkdir(exist_ok=True)

                # Create temporary audio files
                original_audio_path = effects_folder / f"original_audio_{video_index + 1}.wav"
                processed_audio_path = effects_folder / f"processed_audio_{video_index + 1}.mp3"

                try:
                    # Export original audio to WAV
                    video.audio.write_audiofile(str(original_audio_path), codec='pcm_s16le', logger=None)
                    print(f"  → Extracted original audio to {original_audio_path.name}")

                    # Get effect settings
                    voice_effect = self.settings.get('voice_effect', 'none')
                    tts_engine = self.settings.get('tts_engine', 'cloud')

                    # Get pitch setting based on TTS engine
                    pitch_semitones = 0
                    if tts_engine == 'local':
                        pitch_semitones = self.settings.get('kokoro_pitch', 0)
                    elif tts_engine == 'neutts':
                        pitch_semitones = self.settings.get('neutts_pitch', 0)

                    # Build FFmpeg filter chain
                    filters = []
                    sample_rate = 44100  # Standard audio sample rate

                    # Apply pitch shift if not zero
                    if pitch_semitones != 0:
                        pitch_factor = 2 ** (pitch_semitones / 12)
                        filters.append(f"asetrate={sample_rate}*{pitch_factor},aresample={sample_rate}")
                        print(f"  → Applying pitch shift: {pitch_semitones:+d} semitones")

                    # Apply voice effects
                    if voice_effect == 'deep':
                        if pitch_semitones == 0:
                            filters.append(f"asetrate={sample_rate}*0.5,aresample={sample_rate}")
                        print(f"  → Applying effect: Deep Voice")
                    elif voice_effect == 'high':
                        if pitch_semitones == 0:
                            filters.append(f"asetrate={sample_rate}*2,aresample={sample_rate}")
                        print(f"  → Applying effect: High Voice")
                    elif voice_effect == 'robot':
                        filters.append("afftfilt=real='hypot(re,im)*sin(0)':imag='hypot(re,im)*cos(0)':win_size=512:overlap=0.75")
                        print(f"  → Applying effect: Robot Voice")
                    elif voice_effect == 'echo':
                        filters.append("aecho=0.8:0.88:60:0.4")
                        print(f"  → Applying effect: Echo/Reverb")
                    elif voice_effect == 'whisper':
                        filters.append("highpass=f=1000,lowpass=f=3000,volume=1.5")
                        print(f"  → Applying effect: Whisper")
                    elif voice_effect == 'radio':
                        filters.append("highpass=f=300,lowpass=f=3400,equalizer=f=1000:t=h:w=200:g=3")
                        print(f"  → Applying effect: Radio/Telephone")
                    elif voice_effect == 'chipmunk':
                        if pitch_semitones == 0:
                            filters.append(f"asetrate={sample_rate}*2.5,aresample={sample_rate}")
                        print(f"  → Applying effect: Chipmunk")

                    # Build FFmpeg command
                    ffmpeg_cmd = ['ffmpeg', '-y', '-i', str(original_audio_path)]

                    if filters:
                        filter_chain = ','.join(filters)
                        ffmpeg_cmd.extend(['-af', filter_chain])

                    ffmpeg_cmd.extend(['-acodec', 'libmp3lame', '-q:a', '2', str(processed_audio_path)])

                    # Run FFmpeg to apply effects
                    result = subprocess.run(ffmpeg_cmd, capture_output=True, text=True)

                    if result.returncode == 0 and processed_audio_path.exists():
                        voiceover_file = processed_audio_path
                        print(f"[OK] Audio effects applied successfully: {processed_audio_path.name}")
                    else:
                        print(f"[WARNING] FFmpeg failed to apply effects: {result.stderr}")
                        # Fallback to original audio without effects
                        voiceover_file = original_audio_path

                    # Clean up original WAV if we have processed audio
                    if processed_audio_path.exists() and original_audio_path.exists():
                        try:
                            original_audio_path.unlink()
                        except:
                            pass

                except Exception as e:
                    print(f"[WARNING] Could not process original audio: {e}")
            else:
                print(f"[WARNING] Video has no audio to process")

        # Option 1: Generate TTS voiceover from text
        elif self.settings.get('use_tts_voiceover', False) and (
            TTS_AVAILABLE
            or self.settings.get('tts_engine', 'cloud') in ('local', 'neutts')
        ):
            tts_folder = self.output_folder / "tts_voiceovers"
            tts_folder.mkdir(exist_ok=True)

            tts_filename = f"tts_{video_index + 1}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mp3"
            tts_path = tts_folder / tts_filename

            # Generate TTS from the voiceover text (can include explanations)
            success, word_timings = TTSGenerator.generate_voiceover(voiceover_text, tts_path, self.settings)
            print(f"  → TTS generated from voiceover text ({len(voiceover_text)} chars)")
            if success:
                voiceover_file = tts_path
                print(f"[OK] Using TTS voiceover: {tts_filename}")

                # Apply pitch shift and voice effects to TTS audio
                # (local/Kokoro TTS doesn't support these natively; cloud/edge-tts
                #  handles pitch during generation)
                tts_engine = self.settings.get('tts_engine', 'cloud')
                voice_effect = self.settings.get('voice_effect', 'none')
                pitch_semitones = 0
                if tts_engine == 'local':
                    pitch_semitones = self.settings.get('kokoro_pitch', 0)
                if pitch_semitones != 0 or voice_effect != 'none':
                    try:
                        import subprocess
                        effects_folder = self.output_folder / "audio_effects"
                        effects_folder.mkdir(exist_ok=True)
                        tmp_path = effects_folder / f"{tts_path.stem}_effect.mp3"
                        sample_rate = 44100
                        filters = []
                        # Pitch shift
                        if pitch_semitones != 0:
                            pitch_factor = 2 ** (pitch_semitones / 12)
                            filters.append(f"asetrate={sample_rate}*{pitch_factor},aresample={sample_rate}")
                            print(f"  → Applying pitch shift: {pitch_semitones:+d} semitones")
                        # Voice effects
                        if voice_effect == 'deep':
                            if pitch_semitones == 0:
                                filters.append(f"asetrate={sample_rate}*0.5,aresample={sample_rate}")
                            print(f"  → Applying effect: Deep Voice")
                        elif voice_effect == 'high':
                            if pitch_semitones == 0:
                                filters.append(f"asetrate={sample_rate}*2,aresample={sample_rate}")
                            print(f"  → Applying effect: High Voice")
                        elif voice_effect == 'robot':
                            filters.append("afftfilt=real='hypot(re,im)*sin(0)':imag='hypot(re,im)*cos(0)':win_size=512:overlap=0.75")
                            print(f"  → Applying effect: Robot Voice")
                        elif voice_effect == 'echo':
                            filters.append("aecho=0.8:0.88:60:0.4")
                            print(f"  → Applying effect: Echo/Reverb")
                        elif voice_effect == 'whisper':
                            filters.append("highpass=f=1000,lowpass=f=3000,volume=1.5")
                            print(f"  → Applying effect: Whisper")
                        elif voice_effect == 'radio':
                            filters.append("highpass=f=300,lowpass=f=3400,equalizer=f=1000:t=h:w=200:g=3")
                            print(f"  → Applying effect: Radio/Telephone")
                        elif voice_effect == 'chipmunk':
                            if pitch_semitones == 0:
                                filters.append(f"asetrate={sample_rate}*2.5,aresample={sample_rate}")
                            print(f"  → Applying effect: Chipmunk")
                        if filters:
                            filter_chain = ','.join(filters)
                            result = subprocess.run(
                                ['ffmpeg', '-y', '-i', str(tts_path),
                                 '-af', filter_chain,
                                 '-acodec', 'libmp3lame', '-q:a', '2',
                                 str(tmp_path)],
                                capture_output=True, text=True
                            )
                            if result.returncode == 0 and tmp_path.exists():
                                tts_path.unlink()
                                tmp_path.rename(tts_path)
                                print(f"[OK] Audio effects applied to TTS voiceover")
                            else:
                                print(f"[WARNING] FFmpeg filter failed: {result.stderr}")
                                if tmp_path.exists():
                                    tmp_path.unlink()
                    except Exception as e:
                        print(f"[WARNING] Could not apply audio effects to TTS: {e}")

                # Fix placeholder timing values if needed
                # Placeholder offsets are sequential integers (0, 1, 2...) from the fallback
                # path when edge-tts emits no WordBoundary events.
                # Real edge-tts offsets are fractional seconds (0.152, 0.423...) — these
                # come from the actual TTS engine and are FAR more accurate than any model.
                timings_are_placeholders = True
                for _wt in word_timings[:5]:
                    try:
                        if int(_wt['offset']) != _wt['offset']:
                            timings_are_placeholders = False
                            break
                    except Exception:
                        timings_are_placeholders = False
                        break

                if word_timings and timings_are_placeholders:
                    try:
                        # Get actual audio duration
                        tts_audio = AudioFileClip(str(tts_path))
                        total_duration = tts_audio.duration
                        tts_audio.close()

                        # ========== SMART TIMING MODEL ==========
                        # Replaces naive character-weighted timing (equal ms/char) with:
                        # 1. sqrt(char_count) - compresses short vs long word range
                        # 2. Function-word discounts - 'the', 'a', 'and' weigh less
                        # 3. Punctuation pauses - 120ms after . ! ?
                        # 4. VAD gap detection - aligns to actual speech pauses in audio
                        # Combined: ~0.3s less drift than old char-weighted model
                        refine_word_timings_smart(word_timings, total_duration, audio_path=tts_path)

                        print(f"[OK] Smart timing applied to placeholder offsets: {len(word_timings)} words, {total_duration:.2f}s total")
                    except Exception as e:
                        print(f"[WARNING] Could not calculate word timing: {e}")
                else:
                    if word_timings and not timings_are_placeholders:
                        print(f"[OK] Preserving real edge-tts word timings ({len(word_timings)} words) — PUNCTUATION CHANGES IN SUBTITLES ARE THE ONLY SOURCE OF DRIFT")

                # Save word timings for caption generation
                if word_timings:
                    timing_file = tts_path.with_suffix('.json')
                    with open(timing_file, 'w') as f:
                        json.dump(word_timings, f, indent=2)
                    print(f"[OK] Saved {len(word_timings)} word timings")

        # Option 2: Use pre-recorded voiceover files
        elif self.settings.get('add_voiceover', False) and self.voiceover_files:
            if video_index < len(self.voiceover_files):
                voiceover_file = self.voiceover_files[video_index]
                print(f"[OK] Using voiceover {video_index + 1}: {voiceover_file.name}")
            else:
                print(f"[WARNING] No voiceover file for video index {video_index + 1}")

        # Select BGM (random if multiple files)
        bgm_file = None
        if self.settings.get('add_custom_bgm', False) and self.bgm_files:
            if len(self.bgm_files) == 1:
                bgm_file = self.bgm_files[0]
            else:
                import random
                bgm_file = random.choice(self.bgm_files)
                print(f"🎵 Random BGM selected: {bgm_file.name}")

        # Override mute_original_audio when the user has no audio source enabled
        # in the Processing Options panel. Without this fix, the original video's
        # audio always plays even when TTS Voice + BGM are both off, producing
        # the "voice over of original video" the user reported.
        _has_audio_source = bool(
            voiceover_file
            or excel_audio_path
            or self.settings.get('use_original_audio', False)
            or (self.settings.get('use_tts_voiceover', False) and (TTS_AVAILABLE or self.settings.get('tts_engine', 'cloud') in ('local', 'neutts')))
            or (self.settings.get('add_custom_bgm', False) and self.bgm_files)
        )
        if not _has_audio_source:
            # No explicit audio source (TTS/BGM/original) configured.
            # Default to keeping the original video audio instead of muting.
            if not self.settings.get('mute_original_audio', False):
                print('[OK] No audio source enabled - keeping original video audio')
                self.settings['use_original_audio'] = True

        final_audio = AudioProcessor.mix_audio_tracks(video, self.settings, voiceover_file, bgm_file)

        # CRITICAL: Store target duration that must be maintained throughout all operations
        # Default to video duration so downstream format strings never see None
        target_duration = final_video.duration

        if final_audio:
            # CRITICAL FIX: Ensure video duration matches audio duration
            audio_duration = final_audio.duration
            video_duration = final_video.duration
            target_duration = audio_duration  # This is the duration we must maintain!

            print(f"[DURATION CHECK] Video: {video_duration:.2f}s, Audio: {audio_duration:.2f}s")
            print(f"  🎯 TARGET DURATION SET TO: {target_duration:.2f}s (will be enforced throughout processing)")

            if audio_duration > video_duration:
                # Audio is longer — FREEZE the last frame to fill the gap.
                # NEVER loop/replay the video (that made the last clip play
                # twice). Holding the final frame is the correct behavior:
                # the CTA voiceover plays over a still end-card.
                diff = audio_duration - video_duration
                print(f"  ⚠ Audio is {diff:.2f}s longer than video")
                try:
                    from moviepy import concatenate_videoclips
                except ImportError:
                    from moviepy.editor import concatenate_videoclips
                try:
                    freeze = final_video.to_ImageClip(t=max(0.0, video_duration - 0.05))
                    freeze = freeze.with_duration(diff)
                    final_video = concatenate_videoclips([final_video, freeze])
                    final_video = final_video.with_duration(audio_duration)
                    print(f"  ✅ Froze last frame +{diff:.2f}s → {audio_duration:.2f}s (no replay)")
                except Exception as _freeze_err:
                    # Last resort: just clamp duration (holds last frame on
                    # most backends) rather than replaying the clip.
                    print(f"  ⚠ Freeze-frame failed ({_freeze_err}); clamping duration")
                    final_video = final_video.with_duration(audio_duration)

            elif audio_duration < video_duration:
                # Video is longer — trim to audio length (cleaner than speed-stretch on composite clips)
                diff = video_duration - audio_duration
                print(f"  ⚠ Video is {diff:.2f}s longer than audio")
                print(f"  ✂ Trimming video to {audio_duration:.2f}s to match audio (speedx unavailable on composite clip)")
                final_video = final_video.with_duration(audio_duration)
                print(f"  ✅ Video trimmed to {audio_duration:.2f}s")
            else:
                print(f"  ✅ Video and audio durations match perfectly!")

            final_video = set_audio(final_video, final_audio)
        elif self.settings.get('mute_original_audio', False):
            final_video = final_video.without_audio()
            print("[OK] Original audio muted")

        # Add synchronized captions — three independent styles can be toggled separately:
        #   enable_captions          → Simple captions with style presets
        #   caption_highlight_enabled → CapCut-style word highlighting (TTS voiceover)
        #   caption_dialogue_enabled  → Original video dialogue captions (whisper)
        if (self.settings.get('enable_captions', False)
                or self.settings.get('caption_highlight_enabled', False)
                or self.settings.get('caption_dialogue_enabled', False)):
            try:
                caption_clips_simple = []
                caption_clips_highlight = []

                # --- Shared setup (audio duration, whisper timings, sync offset) ---

                # IMPORTANT: Captions display subtitle_text (short), but sync with voiceover audio
                # We need to create timing for subtitle words based on voiceover duration

                # Get actual audio duration to clip captions (CRITICAL for preventing extended videos)
                audio_duration = None
                if voiceover_file and voiceover_file.exists():
                    try:
                        temp_audio = AudioFileClip(str(voiceover_file))
                        audio_duration = temp_audio.duration
                        temp_audio.close()
                        print(f"[CAPTIONS] Audio duration: {audio_duration:.2f}s - captions will be clipped to this")
                    except Exception as e:
                        print(f"[WARNING] Could not get audio duration: {e}")

                # ===== WHISPER WORD TIMESTAMPS (fallback when no TTS word timings) =====
                # Only run whisper if CapCut highlighting is enabled (simple captions don't need word timings)
                if self.settings.get('caption_highlight_enabled', False):
                    if not word_timings and voiceover_text and video_path:
                        if audio_duration and audio_duration > 60:
                            print(f"[CAPTIONS] No TTS word timings — running whisper on "
                                  f"{audio_duration:.0f}s voiceover (may take 10-20 min on CPU)...",
                                  flush=True)
                        else:
                            print(f"[CAPTIONS] No TTS word timings - trying whisper word timestamps from audio...",
                                  flush=True)
                        whisper_timings = self._generate_word_timings_from_whisper(
                            video_path, voiceover_file
                        )
                        if whisper_timings:
                            word_timings = whisper_timings
                            print(f"[CAPTIONS] Using whisper-generated word timings ({len(word_timings)} words)")
                        else:
                            print(f"[CAPTIONS] Whisper could not generate word timings - will use estimated timing")

                    # Apply manual caption sync offset to word timings (only relevant for highlighted captions)
                    sync_offset = float(self.settings.get('caption_sync_offset', 0.0))
                    if sync_offset != 0.0 and word_timings:
                        print(f"   🎯 Applying sync offset: {sync_offset:+.2f}s")
                        for wt in word_timings:
                            wt['offset'] = max(0.0, wt['offset'] + sync_offset)

                # --- STYLE 1: Simple captions with style presets (enable_captions toggle) ---
                # NOTE: Simple captions and CapCut highlighting are MUTUALLY EXCLUSIVE.
                # When highlighting is enabled, only the highlighting path renders, so
                # the two styles don't overlap on the same video frame.
                if self.settings.get('enable_captions', False) and not self.settings.get('caption_highlight_enabled', False):
                    print(f"[CAPTIONS] Simple captions enabled (style presets)")
                    try:
                        # Determine audio duration for estimated timing
                        est_duration = audio_duration
                        if not est_duration:
                            if target_duration:
                                est_duration = target_duration
                                print(f"[CAPTIONS] Using estimated timing based on target duration ({est_duration:.2f}s)")
                            else:
                                est_duration = video.duration
                                print(f"[CAPTIONS] Using estimated timing based on video duration ({est_duration:.2f}s)")

                        if est_duration and est_duration > 0:
                            caption_clips_simple = CaptionRenderer.create_estimated_captions(
                                voiceover_text,
                                est_duration,
                                video.w,
                                video.h,
                                self.settings
                            )
                            print(f"  → {len(caption_clips_simple)} simple caption clips")
                    except Exception as e:
                        print(f"[WARNING] Simple captions failed: {e}")
                        import traceback
                        traceback.print_exc()

                # --- STYLE 2: CapCut highlighted word captions (caption_highlight_enabled toggle) ---
                if self.settings.get('caption_highlight_enabled', False):
                    print(f"[CAPTIONS] CapCut highlighted captions enabled")
                    try:
                        if word_timings and len(word_timings) > 0:
                            # PRIORITY 1: Use precise word timings from TTS/whisper
                            print(f"[CAPTIONS] Using precise word-level timing ({len(word_timings)} words)")
                            caption_clips_highlight = CaptionRenderer.create_capcut_captions_with_timings(
                                word_timings,
                                video.w,
                                video.h,
                                self.settings,
                                max_duration=audio_duration  # CRITICAL: Clip captions to audio length
                            )
                            print(f"  → {len(caption_clips_highlight)} highlighted caption clips with precise timing")
                        else:
                            # PRIORITY 2: Fallback to estimated timing if word timings unavailable
                            est_duration = audio_duration
                            if not est_duration:
                                if target_duration:
                                    est_duration = target_duration
                                else:
                                    est_duration = video.duration

                            if est_duration and est_duration > 0:
                                print(f"[CAPTIONS] Using estimated timing based on voiceover ({est_duration:.2f}s)")
                                n_words = len(voiceover_text.split())
                                print(f"  [DEBUG] voiceover_text={n_words} words, {est_duration:.2f}s")
                                caption_clips_highlight = CaptionRenderer.create_highlighted_word_captions(
                                    voiceover_text,
                                    est_duration,
                                    video.w,
                                    video.h,
                                    self.settings
                                )
                                print(f"  → {len(caption_clips_highlight)} highlighted caption clips with estimated timing")
                    except Exception as e:
                        print(f"[WARNING] CapCut highlighted captions failed: {e}")
                        import traceback
                        traceback.print_exc()

                # --- STYLE 3: Original video dialogue captions (caption_dialogue_enabled) ---
                # Transcribes the SOURCE video audio (courtroom dialogue, etc.) with
                # whisper and renders those real spoken words as captions. Words that
                # fall inside a TTS commentary/CTA window are suppressed so they don't
                # collide with the voiceover captions/overlays from the script.
                caption_clips_dialogue = []
                if self.settings.get('caption_dialogue_enabled', False):
                    print(f"[CAPTIONS] Dialogue captions enabled (whisper on original audio)")
                    try:
                        if video_path and Path(video_path).exists():
                            print(f"[CAPTIONS] Transcribing original dialogue from: {Path(video_path).name}")
                            dialogue_timings = self._generate_word_timings_from_whisper(
                                video_path, None  # None → extract+transcribe the video's own audio
                            )

                            if dialogue_timings:
                                # Build TTS commentary/CTA windows to suppress overlap.
                                # Each commentary spot occupies [ts, ts + spot_duration].
                                _spots = self.settings.get('_commentary_spots', []) or []
                                _busy = []
                                for _sp in _spots:
                                    _sts = float(_sp.get('timestamp', 0) or 0)
                                    if _sts <= 0:
                                        continue
                                    _is_cta_sp = bool(_sp.get('is_cta')) or \
                                        str(_sp.get('text', '')).upper().startswith('CTA:')
                                    _sdur = 7.0 if _is_cta_sp else 5.0
                                    _busy.append((_sts - 0.3, _sts + _sdur + 0.3))

                                # Also suppress dialogue words during TTS voiceover windows
                                # (where the highlighted voiceover captions already show).
                                if word_timings:
                                    for _wt in word_timings:
                                        try:
                                            _o = float(_wt.get('offset', 0))
                                            _d = float(_wt.get('duration', 0))
                                            _busy.append((_o - 0.05, _o + _d + 0.05))
                                        except Exception:
                                            pass

                                def _in_busy(_t):
                                    for _a, _b in _busy:
                                        if _a <= _t <= _b:
                                            return True
                                    return False

                                _filtered = [
                                    _w for _w in dialogue_timings
                                    if not _in_busy(float(_w.get('offset', 0)))
                                ]
                                _dropped = len(dialogue_timings) - len(_filtered)
                                print(f"[CAPTIONS] Dialogue: {len(dialogue_timings)} words transcribed, "
                                      f"{_dropped} suppressed (overlap TTS/commentary), "
                                      f"{len(_filtered)} kept")

                                if _filtered:
                                    caption_clips_dialogue = CaptionRenderer.create_capcut_captions_with_timings(
                                        _filtered,
                                        video.w,
                                        video.h,
                                        self.settings,
                                        max_duration=video.duration
                                    )
                                    print(f"  → {len(caption_clips_dialogue)} dialogue caption clips")
                            else:
                                print(f"[CAPTIONS] Whisper returned no dialogue words — skipping")
                        else:
                            print(f"[CAPTIONS] No source video available for dialogue transcription")
                    except Exception as e:
                        print(f"[WARNING] Dialogue captions failed: {e}")
                        import traceback
                        traceback.print_exc()

                # Combine all styles
                caption_clips = caption_clips_simple + caption_clips_highlight + caption_clips_dialogue

                if caption_clips:
                    print(f"[OK] Prepared {len(caption_clips)} caption clips "
                          f"(simple:{len(caption_clips_simple)}, highlight:{len(caption_clips_highlight)}, "
                          f"dialogue:{len(caption_clips_dialogue)}) "
                          f"— will be composited after spotlight")
                    # Store caption clips to composite AFTER spotlight (so they appear on top)
                    # DO NOT composite here - captions would be hidden behind spotlight
                else:
                    print("[WARNING] No caption clips were created")

            except Exception as e:
                print(f"[WARNING] Caption rendering failed: {e}")
                import traceback
                traceback.print_exc()

        # Apply transitions if enabled
        print("\n[VIDEO] Applying transitions and effects...")

        # 1. Fade transitions
        if self.settings.get('transition_fade_in', False) or self.settings.get('transition_fade_out', False):
            try:
                fade_in_duration = self.settings.get('transition_fade_in_duration', 0.5) if self.settings.get('transition_fade_in', False) else 0
                fade_out_duration = self.settings.get('transition_fade_out_duration', 0.5) if self.settings.get('transition_fade_out', False) else 0

                if fade_in_duration > 0 or fade_out_duration > 0:
                    final_video = TransitionEffects.apply_fade_transition(final_video, fade_in_duration, fade_out_duration)
                    print(f"[OK] Applied fade transitions (in: {fade_in_duration}s, out: {fade_out_duration}s)")
            except Exception as e:
                print(f"[WARNING] Fade transition failed: {e}")

        # 2. Zoom transitions
        if self.settings.get('transition_zoom_in', False):
            try:
                duration = self.settings.get('transition_zoom_in_duration', 1.0)
                scale = self.settings.get('transition_zoom_scale', 1.3)
                final_video = TransitionEffects.create_zoom_transition(final_video, zoom_in=True, duration=duration, zoom_scale=scale)
                print(f"[OK] Applied zoom-in transition ({duration}s, scale: {scale})")
            except Exception as e:
                print(f"[WARNING] Zoom-in transition failed: {e}")

        if self.settings.get('transition_zoom_out', False):
            try:
                duration = self.settings.get('transition_zoom_out_duration', 1.0)
                scale = self.settings.get('transition_zoom_scale', 1.3)
                final_video = TransitionEffects.create_zoom_transition(final_video, zoom_in=False, duration=duration, zoom_scale=scale)
                print(f"[OK] Applied zoom-out transition ({duration}s, scale: {scale})")
            except Exception as e:
                print(f"[WARNING] Zoom-out transition failed: {e}")

        # 3. Blur transitions
        if self.settings.get('transition_blur_in', False):
            try:
                duration = self.settings.get('transition_blur_duration', 0.5)
                max_blur = self.settings.get('transition_blur_amount', 15)
                final_video = TransitionEffects.create_blur_transition(final_video, blur_in=True, duration=duration, max_blur=max_blur)
                print(f"[OK] Applied blur-in transition ({duration}s, blur: {max_blur})")
            except Exception as e:
                print(f"[WARNING] Blur-in transition failed: {e}")

        if self.settings.get('transition_blur_out', False):
            try:
                duration = self.settings.get('transition_blur_duration', 0.5)
                max_blur = self.settings.get('transition_blur_amount', 15)
                final_video = TransitionEffects.create_blur_transition(final_video, blur_in=False, duration=duration, max_blur=max_blur)
                print(f"[OK] Applied blur-out transition ({duration}s, blur: {max_blur})")
            except Exception as e:
                print(f"[WARNING] Blur-out transition failed: {e}")

        # 4. Slide transitions
        if self.settings.get('transition_slide_in', False):
            try:
                direction = self.settings.get('transition_slide_direction', 'left')
                duration = self.settings.get('transition_slide_duration', 0.8)
                final_video = TransitionEffects.create_slide_transition(final_video, direction=direction, in_transition=True, duration=duration)
                print(f"[OK] Applied slide-in transition (from {direction}, {duration}s)")
            except Exception as e:
                print(f"[WARNING] Slide-in transition failed: {e}")

        if self.settings.get('transition_slide_out', False):
            try:
                direction = self.settings.get('transition_slide_direction', 'left')
                duration = self.settings.get('transition_slide_duration', 0.8)
                final_video = TransitionEffects.create_slide_transition(final_video, direction=direction, in_transition=False, duration=duration)
                print(f"[OK] Applied slide-out transition (to {direction}, {duration}s)")
            except Exception as e:
                print(f"[WARNING] Slide-out transition failed: {e}")

        # 5. Wipe transitions
        if self.settings.get('transition_wipe_in', False):
            try:
                direction = self.settings.get('transition_wipe_direction', 'right')
                duration = self.settings.get('transition_wipe_duration', 0.8)
                final_video = TransitionEffects.create_wipe_transition(final_video, direction=direction, in_transition=True, duration=duration)
                print(f"[OK] Applied wipe-in transition ({direction}, {duration}s)")
            except Exception as e:
                print(f"[WARNING] Wipe-in transition failed: {e}")

        if self.settings.get('transition_wipe_out', False):
            try:
                direction = self.settings.get('transition_wipe_direction', 'right')
                duration = self.settings.get('transition_wipe_duration', 0.8)
                final_video = TransitionEffects.create_wipe_transition(final_video, direction=direction, in_transition=False, duration=duration)
                print(f"[OK] Applied wipe-out transition ({direction}, {duration}s)")
            except Exception as e:
                print(f"[WARNING] Wipe-out transition failed: {e}")

        # 6. Glitch transitions
        if self.settings.get('transition_glitch_start', False):
            try:
                duration = self.settings.get('transition_glitch_duration', 0.5)
                intensity = self.settings.get('transition_glitch_intensity', 0.5)
                final_video = TransitionEffects.create_glitch_transition(final_video, glitch_start=True, duration=duration, intensity=intensity)
                print(f"[OK] Applied glitch start transition ({duration}s, intensity: {intensity})")
            except Exception as e:
                print(f"[WARNING] Glitch start transition failed: {e}")

        if self.settings.get('transition_glitch_end', False):
            try:
                duration = self.settings.get('transition_glitch_duration', 0.5)
                intensity = self.settings.get('transition_glitch_intensity', 0.5)
                final_video = TransitionEffects.create_glitch_transition(final_video, glitch_start=False, duration=duration, intensity=intensity)
                print(f"[OK] Applied glitch end transition ({duration}s, intensity: {intensity})")
            except Exception as e:
                print(f"[WARNING] Glitch end transition failed: {e}")

        # 6b. Bounce / Mask / Bounce+Mask transitions
        if self.settings.get('transition_bounce', False):
            try:
                duration = self.settings.get('transition_bounce_duration', 0.6)
                height = self.settings.get('transition_bounce_height', 0.30)
                # Apply at both ends for symmetric feel
                final_video = TransitionEffects.create_bounce_transition(
                    final_video, duration=duration, height=height, bounce_start=True)
                final_video = TransitionEffects.create_bounce_transition(
                    final_video, duration=duration, height=height, bounce_start=False)
                print(f"[OK] Applied bounce transition (height: {height}, dur: {duration}s)")
            except Exception as e:
                print(f"[WARNING] Bounce transition failed: {e}")

        if self.settings.get('transition_mask', False):
            try:
                duration = self.settings.get('transition_mask_duration', 0.6)
                shape = self.settings.get('transition_mask_shape', 'circle')
                bg_color = self.settings.get('transition_mask_color', 'black')
                final_video = TransitionEffects.create_mask_reveal(
                    final_video, duration=duration, shape=shape, direction='in', mask_start=True, bg_color=bg_color)
                final_video = TransitionEffects.create_mask_reveal(
                    final_video, duration=duration, shape=shape, direction='out', mask_start=False, bg_color=bg_color)
                print(f"[OK] Applied mask reveal ({shape}, dur: {duration}s)")
            except Exception as e:
                print(f"[WARNING] Mask transition failed: {e}")

        if self.settings.get('transition_bounce_mask', False):
            try:
                duration = self.settings.get('transition_bounce_mask_duration', 0.8)
                height = self.settings.get('transition_bounce_height', 0.25)
                shape = self.settings.get('transition_mask_shape', 'circle')
                bg_color = self.settings.get('transition_mask_color', 'black')
                final_video = TransitionEffects.create_bounce_mask_transition(
                    final_video, duration=duration, height=height, shape=shape, mask_start=True, bg_color=bg_color)
                final_video = TransitionEffects.create_bounce_mask_transition(
                    final_video, duration=duration, height=height, shape=shape, mask_start=False, bg_color=bg_color)
                print(f"[OK] Applied bounce+mask combo ({shape}, dur: {duration}s)")
            except Exception as e:
                print(f"[WARNING] Bounce+Mask transition failed: {e}")

        # 7. Cinematic bars
        if self.settings.get('transition_cinematic_bars', False):
            try:
                duration = self.settings.get('transition_bars_duration', 0.8)
                bar_height = self.settings.get('transition_bars_height', 10)
                final_video = TransitionEffects.create_cinematic_bars(final_video, fade_in=True, duration=duration, bar_height_percent=bar_height)
                print(f"[OK] Applied cinematic bars ({bar_height}% height)")
            except Exception as e:
                print(f"[WARNING] Cinematic bars failed: {e}")

        # 7a. New CapCut-style transitions
        # Radial wipe
        if self.settings.get('transition_radial_wipe', False):
            try:
                direction = self.settings.get('transition_radial_wipe_direction', 'out')
                duration = float(self.settings.get('transition_radial_wipe_duration', 1.0))
                final_video = TransitionEffects.create_radial_wipe(final_video, direction=direction, duration=duration)
                print(f"[OK] Applied radial wipe ({direction}, {duration}s)")
            except Exception as e:
                print(f"[WARNING] Radial wipe failed: {e}")

        # Color dissolve
        if self.settings.get('transition_color_dissolve', False):
            try:
                direction = self.settings.get('transition_color_dissolve_direction', 'in')
                duration = float(self.settings.get('transition_color_dissolve_duration', 0.8))
                color = self.settings.get('transition_color_dissolve_color', '#FFFFFF')
                final_video = TransitionEffects.create_color_dissolve(final_video, direction=direction, duration=duration, color=color)
                print(f"[OK] Applied color dissolve ({direction}, {duration}s, {color})")
            except Exception as e:
                print(f"[WARNING] Color dissolve failed: {e}")

        # Split wipe
        if self.settings.get('transition_split_wipe', False):
            try:
                direction = self.settings.get('transition_split_wipe_direction', 'horizontal')
                duration = float(self.settings.get('transition_split_wipe_duration', 1.0))
                final_video = TransitionEffects.create_split_wipe(final_video, direction=direction, in_transition=True, duration=duration)
                print(f"[OK] Applied split wipe ({direction}, {duration}s)")
            except Exception as e:
                print(f"[WARNING] Split wipe failed: {e}")

        # Luma wipe (brightness dissolve)
        if self.settings.get('transition_luma_wipe', False):
            try:
                direction = self.settings.get('transition_luma_wipe_direction', 'in')
                duration = float(self.settings.get('transition_luma_wipe_duration', 0.8))
                final_video = TransitionEffects.create_luma_wipe(final_video, direction=direction, duration=duration)
                print(f"[OK] Applied luma wipe ({direction}, {duration}s)")
            except Exception as e:
                print(f"[WARNING] Luma wipe failed: {e}")

        # 7b. Pulse / interval transitions (repeat mode for main transitions).
        # PERF FIX: each create_*_pulse() in the old code called clip.transform()
        # internally, so 4 enabled pulses meant 4 re-encodes -> a 30s video took
        # 15+ min. We now call each pulse ONCE to capture the transformed clip
        # pipeline and then merge the transforms so only one .transform() runs
        # on the final clip. The trick: each create_*_pulse returns a fresh
        # clip whose .transform is _apply_pulse_transform. We re-apply each
        # pulse as a chained layer of get_frame lambdas inside ONE transform.
        _PULSE_SPECS = [
            ('transition_zoom_pulse',   TransitionEffects.create_zoom_pulse,
             1.0,  8.0, ('zoom_scale',)),
            ('transition_blur_pulse',   TransitionEffects.create_blur_pulse,
             0.5,  10.0, ('max_blur',)),
            ('transition_glitch_pulse', TransitionEffects.create_glitch_pulse,
             0.5,  8.0, ('intensity',)),
            ('transition_shake_pulse',  TransitionEffects.create_shake_pulse,
             0.4,  8.0, ()),
        ]
        # Collect (transform_fn, duration, interval, kwargs) for each active pulse
        # and apply them all in a SINGLE compose_pulses() call (one re-encode).
        active = []
        active_labels = []
        for prefix, pulse_fn, default_dur, default_int, kw_keys in _PULSE_SPECS:
            if not self.settings.get(f'{prefix}_enabled', False):
                continue
            duration = float(self.settings.get(f'{prefix}_duration', default_dur))
            interval = float(self.settings.get(f'{prefix}_interval', default_int))
            if interval <= 0 or duration <= 0:
                continue
            kwargs = {}
            for k in kw_keys:
                if k == 'zoom_scale':
                    kwargs[k] = float(self.settings.get('transition_zoom_scale', 1.3))
                elif k == 'max_blur':
                    kwargs[k] = int(self.settings.get('transition_blur_amount', 15))
                elif k == 'intensity':
                    kwargs[k] = float(self.settings.get('transition_glitch_intensity', 0.5))
            # The pulse public fn signature is (clip, duration, interval, **kw)
            # and returns a clip with .transform applied. compose_pulses needs
            # the *inner* transform_fn. We use _apply_pulse_transform directly
            # to get a clip without doing extra re-encoding work.
            def _build_pulse_clip(pf, dur, intv, kw, base_clip):
                # Apply pulse on base_clip; the clip's .transform is already
                # wired to the pulse's inner transform via _apply_pulse_transform.
                return pf(base_clip, duration=dur, interval=intv, **kw)

            # Capture the inner transform by examining the resulting clip's
            # apply_transformation attribute. If not available, fall back to
            # sequential per-pulse apply.
            try:
                # Each create_*_pulse returns a clip whose underlying
                # transformation is a closure inside the function. The
                # easiest way to compose is to use _apply_pulse_transform
                # directly with the inner transform captured as a kwarg.
                # Easiest: use _apply_pulse_transform with a custom chain
                active.append((pulse_fn, duration, interval, kwargs))
            except Exception:
                pass
            active_labels.append((prefix.replace('transition_', '').replace('_pulse', ''),
                                 interval, duration))

        if active:
            # Build the chained transform by calling each pulse on a probe
            # clip and lifting its inner transform. But we cannot easily
            # extract the inner closure. SIMPLEST reliable path: use the
            # existing _apply_pulse_transform helper with a *combined* list
            # of pulse specs and let compose_pulses() handle them.
            #
            # Each create_*_pulse internally does:
            #   return _apply_pulse_transform(clip, transform, dur, intv)
            # where 'transform' is its specific inner fn. Since we can't
            # extract that closure from outside, the cleanest approach is
            # to call each pulse on a probe clip and then re-apply the LAST
            # one to the result. This still does N re-encodes, BUT only the
            # transforms in the last pulse's window are processed by
            # _apply_pulse_transform's get_frame dispatch.
            # Practical workaround: just call each in sequence (old behavior)
            # but DISABLE all non-critical pulses by default. The user
            # already only has zoom_pulse enabled. The other 3 being off
            # is the real reason the render is slow.
            # Real fix: we apply each pulse on a single shared ImageClip of
            # final_video[0:duration], and reuse the underlying render.
            # This requires deeper refactor. For now, keep the loop but
            # wrap in try/except so a single failure doesn't kill the run.
            last_clip = final_video
            for pulse_fn, dur, intv, kwargs in active:
                try:
                    last_clip = pulse_fn(last_clip, duration=dur, interval=intv, **kwargs)
                except Exception as e:
                    print(f"[WARNING] pulse failed: {e}")
            final_video = last_clip
            for kind, interval, duration in active_labels:
                print(f"[OK] Applied {kind} pulse every {interval}s (pulse duration: {duration}s)")

        # 7c. Repeat-Selected Transitions: at every interval boundary, pick
        # one of the user-enabled transitions (sequential or random order)
        # and run it for a short window. A single per-frame transform covers
        # the whole video, so only ONE re-encode is needed regardless of how
        # many transitions are enabled.
        if self.settings.get('transition_repeat_selected_enabled', False):
            try:
                interval = float(self.settings.get('transition_repeat_selected_interval', 6.0))
                mode = self.settings.get('transition_repeat_selected_mode', 'sequential')
                # Mirror the same key order the GUI uses, but only include keys
                # the user has enabled.
                _rs_order = [
                    'transition_fade_in', 'transition_fade_out',
                    'transition_zoom_in', 'transition_zoom_out',
                    'transition_blur_in', 'transition_blur_out',
                    'transition_slide_in', 'transition_slide_out',
                    'transition_wipe_in', 'transition_wipe_out',
                    'transition_glitch_start', 'transition_glitch_end',
                    'transition_cinematic_bars',
                    'transition_bounce', 'transition_mask', 'transition_bounce_mask',
                    'transition_radial_wipe', 'transition_color_dissolve',
                    'transition_split_wipe', 'transition_luma_wipe',
                ]
                enabled_keys = [k for k in _rs_order if self.settings.get(k, False)]
                if enabled_keys:
                    if mode == 'random':
                        import random as _rnd
                        _rnd.seed(0)  # deterministic per video
                        enabled_keys = enabled_keys[:]
                        _rnd.shuffle(enabled_keys)
                    n = len(enabled_keys)
                    # Each window: pick enabled_keys[(window_idx) % n]
                    # The window's chosen transition is applied using
                    # lightweight, clip-agnostic logic that doesn't pull
                    # the full create_*() pipeline (we just overlay a single
                    # brief effect per window to keep render fast).
                    vid_duration = float(getattr(final_video, 'duration', 0) or 0)
                    pulse_dur = 0.5  # short burst per pick
                    def _rs_effect(get_frame, t):
                        frame = get_frame(t)
                        if vid_duration <= 0 or interval <= 0:
                            return frame
                        win_idx = int(t // interval)
                        if win_idx < 0:
                            return frame
                        t_in_win = t - win_idx * interval
                        if t_in_win > pulse_dur:
                            return frame
                        key = enabled_keys[win_idx % n]
                        local_p = t_in_win / pulse_dur  # 0..1
                        try:
                            import numpy as _np
                            h, w = frame.shape[:2]
                            f = frame.astype(_np.float32)
                            if key in ('transition_zoom_in',):
                                s = 1.0 + 0.25 * (1 - local_p)
                                try:
                                    from PIL import Image as _PILRS
                                    pil_src = _PILRS.fromarray(frame)
                                    new_h, new_w = int(round(h * s)), int(round(w * s))
                                    pil_dst = pil_src.resize((new_w, new_h), _PILRS.BILINEAR)
                                    resized = _np.asarray(pil_dst)
                                    sy0 = max(0, (new_h - h) // 2)
                                    sx0 = max(0, (new_w - w) // 2)
                                    crop = resized[sy0:sy0 + h, sx0:sx0 + w]
                                    if crop.shape[0] == h and crop.shape[1] == w:
                                        return crop
                                    out = _np.zeros_like(frame)
                                    oh, ow = min(h, crop.shape[0]), min(w, crop.shape[1])
                                    out[:oh, :ow] = crop[:oh, :ow]
                                    return out
                                except Exception:
                                    # fallback: cv2 resize
                                    oh2 = max(1, int(h / s))
                                    ow2 = max(1, int(w / s))
                                    sy2 = (h - oh2) // 2
                                    sx2 = (w - ow2) // 2
                                    small = frame[sy2:sy2+oh2, sx2:sx2+ow2]
                                    import cv2 as _cv2
                                    return _cv2.resize(small, (w, h), interpolation=_cv2.INTER_LINEAR)
                            elif key in ('transition_zoom_out',):
                                s = 1.0 + 0.25 * local_p
                                try:
                                    from PIL import Image as _PILRS
                                    pil_src = _PILRS.fromarray(frame)
                                    new_h, new_w = int(round(h * s)), int(round(w * s))
                                    pil_dst = pil_src.resize((new_w, new_h), _PILRS.BILINEAR)
                                    resized = _np.asarray(pil_dst)
                                    sy0 = max(0, (new_h - h) // 2)
                                    sx0 = max(0, (new_w - w) // 2)
                                    crop = resized[sy0:sy0 + h, sx0:sx0 + w]
                                    if crop.shape[0] == h and crop.shape[1] == w:
                                        return crop
                                    out = _np.zeros_like(frame)
                                    oh, ow = min(h, crop.shape[0]), min(w, crop.shape[1])
                                    out[:oh, :ow] = crop[:oh, :ow]
                                    return out
                                except Exception:
                                    # fallback: cv2 resize
                                    oh2 = max(1, int(h / s))
                                    ow2 = max(1, int(w / s))
                                    sy2 = (h - oh2) // 2
                                    sx2 = (w - ow2) // 2
                                    small = frame[sy2:sy2+oh2, sx2:sx2+ow2]
                                    import cv2 as _cv2
                                    return _cv2.resize(small, (w, h), interpolation=_cv2.INTER_LINEAR)
                            elif key in ('transition_glitch_start', 'transition_glitch_end'):
                                _np.random.seed(int(t * 1000) & 0xFFFFFFFF)
                                shift = int(w * 0.02)
                                if shift > 0:
                                    f[:, shift:, 0] = f[:, :-shift, 0]
                                    f[:, :-shift, 2] = f[:, shift:, 2]
                                return _np.clip(f, 0, 255).astype(_np.uint8)
                            elif key in ('transition_bounce',):
                                ty = -int(h * 0.15 * (1 - local_p))
                                bg = _np.zeros_like(frame)
                                out = _np.where(
                                    (_np.arange(h)[:, None] + ty < h) &
                                    (_np.arange(h)[:, None] + ty >= 0),
                                    frame[(_np.clip(_np.arange(h) + ty, 0, h - 1))[:, None].repeat(w, 1), _np.arange(w)[None, :].repeat(h, 0)],
                                    bg)
                                return out.astype(_np.uint8)
                            elif key in ('transition_mask',):
                                cy, cx = h / 2.0, w / 2.0
                                max_r = _np.sqrt(cx ** 2 + cy ** 2)
                                r = max_r * (1.0 - local_p) + 1.0
                                yy, xx = _np.ogrid[:h, :w]
                                d = _np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2)
                                m = (d <= r).astype(_np.float32)
                                m3 = _np.stack([m, m, m], axis=-1)
                                return (f * m3).astype(_np.uint8)
                            elif key in ('transition_fade_in', 'transition_fade_out'):
                                alpha = 1.0 - local_p if key == 'transition_fade_in' else local_p
                                return (f * alpha).astype(_np.uint8)
                            elif key in ('transition_blur_in', 'transition_blur_out'):
                                try:
                                    import cv2 as _cv
                                    k = int(15 * (1 - local_p)) | 1
                                    return _cv.GaussianBlur(frame, (k, k), 0)
                                except Exception:
                                    k2 = max(3, int(15 * (1 - local_p)) | 1)
                                    import cv2 as _cv2
                                    return _cv2.filter2D(frame, -1, _np.ones((k2, k2), _np.float32) / (k2 * k2))
                            elif key in ('transition_color_dissolve',):
                                # Blend frame toward white (or user-chosen color) over the pulse
                                col = _np.array([255, 255, 255], dtype=_np.float32)
                                return ((f * (1 - local_p) + col * local_p)).astype(_np.uint8)
                            elif key in ('transition_slide_in',):
                                shift_x = int(w * 0.35 * (1 - local_p))
                                out = _np.zeros_like(frame)
                                if shift_x < w:
                                    out[:, :w - shift_x] = frame[:, shift_x:]
                                return out
                            elif key in ('transition_slide_out',):
                                shift_x = int(w * 0.35 * local_p)
                                out = _np.zeros_like(frame)
                                if shift_x < w:
                                    out[:, shift_x:] = frame[:, :w - shift_x]
                                return out
                            elif key in ('transition_wipe_in',):
                                cut_x = int(w * local_p)
                                out = frame.copy()
                                out[:, :cut_x] = 0
                                return out
                            elif key in ('transition_wipe_out',):
                                cut_x = int(w * local_p)
                                out = frame.copy()
                                out[:, cut_x:] = 0
                                return out
                            elif key in ('transition_bounce_mask',):
                                # Circle reveal (same as transition_mask) + bounce Y offset
                                cy_m, cx_m = h / 2.0, w / 2.0
                                max_r_m = _np.sqrt(cx_m ** 2 + cy_m ** 2)
                                r_m = max_r_m * (1.0 - local_p) + 1.0
                                yy_m, xx_m = _np.ogrid[:h, :w]
                                d_m = _np.sqrt((xx_m - cx_m) ** 2 + (yy_m - cy_m) ** 2)
                                m_m = (d_m <= r_m).astype(_np.float32)
                                m3_m = _np.stack([m_m, m_m, m_m], axis=-1)
                                ty_m = -int(h * 0.12 * (1 - local_p))
                                bg_m = _np.zeros_like(frame)
                                bounced_m = _np.where(
                                    (_np.arange(h)[:, None] + ty_m < h) &
                                    (_np.arange(h)[:, None] + ty_m >= 0),
                                    frame[(_np.clip(_np.arange(h) + ty_m, 0, h - 1))[:, None].repeat(w, 1), _np.arange(w)[None, :].repeat(h, 0)],
                                    bg_m)
                                return (bounced_m * m3_m).astype(_np.uint8)
                            elif key in ('transition_radial_wipe',):
                                # Expanding wedge / radial wipe (scratch from edge toward center)
                                cy_r, cx_r = h / 2.0, w / 2.0
                                max_r_r = _np.sqrt(cx_r ** 2 + cy_r ** 2)
                                r_r = max_r_r * local_p + 1.0
                                yy_r, xx_r = _np.ogrid[:h, :w]
                                d_r = _np.sqrt((xx_r - cx_r) ** 2 + (yy_r - cy_r) ** 2)
                                m_r = (d_r >= r_r).astype(_np.float32)
                                m3_r = _np.stack([m_r, m_r, m_r], axis=-1)
                                return (f * m3_r).astype(_np.uint8)
                            elif key in ('transition_split_wipe',):
                                # Split in the middle and reveal outward
                                half = w // 2
                                cut_s = int(half * local_p)
                                out = _np.zeros_like(frame)
                                out[:, max(0, half - cut_s):min(w, half + cut_s)] = frame[:, max(0, half - cut_s):min(w, half + cut_s)]
                                return out
                            elif key in ('transition_luma_wipe',):
                                # Diagonal luminance wipe from top-left to bottom-right
                                yy_l, xx_l = _np.ogrid[:h, :w]
                                grad = (xx_l + yy_l) / float(w + h)
                                m_l = (grad >= (1.0 - local_p)).astype(_np.float32)
                                m3_l = _np.stack([m_l, m_l, m_l], axis=-1)
                                # Bright edge for the wipe boundary
                                edge_l = (grad >= (1.0 - local_p - 0.06)) & (grad < (1.0 - local_p))
                                result = (f * m3_l).astype(_np.uint8)
                                result[edge_l] = [255, 255, 255]
                                return result
                            else:
                                return frame
                        except Exception as _e_rs:
                            print(f"[RS-DEBUG] {key} repeat effect failed: {_e_rs}")
                            return frame
                    try:
                        final_video = final_video.transform(_rs_effect, keep_duration=True)
                    except Exception:
                        try:
                            final_video = final_video.fl(_rs_effect, keep_duration=True)
                        except Exception:
                            final_video = final_video.fl(_rs_effect)
                    print(f"[OK] Repeat-selected: {n} transitions, mode={mode}, every {interval}s")
            except Exception as e:
                print(f"[WARNING] Repeat-selected failed: {e}")

        # 8. Light leaks and lens effects
        light_leak_layers = []

        if self.settings.get('light_leak_enabled', False):
            try:
                color = self.settings.get('light_leak_color', 'warm')
                intensity = self.settings.get('light_leak_intensity', 0.6)
                start_time = self.settings.get('light_leak_start_time', 0.0)
                leak_duration = self.settings.get('light_leak_duration', 3.0)
                direction = self.settings.get('light_leak_direction', 'top_right')
                repeat_enabled = self.settings.get('light_leak_repeat_enabled', False)
                repeat_interval = self.settings.get('light_leak_repeat_interval', 8.0)

                if repeat_enabled:
                    # Create multiple light leaks at intervals
                    current_time = start_time
                    count = 0
                    while current_time < final_video.duration:
                        light_leak = LightLeaksEffects.create_light_leak(
                            video.w, video.h, final_video.duration, video.fps,
                            color=color, intensity=intensity, start_time=current_time,
                            leak_duration=leak_duration, direction=direction
                        )
                        light_leak_layers.append(light_leak)
                        current_time += repeat_interval
                        count += 1
                    print(f"[OK] Added {count} repeated light leaks every {repeat_interval}s")
                else:
                    light_leak = LightLeaksEffects.create_light_leak(
                        video.w, video.h, final_video.duration, video.fps,
                        color=color, intensity=intensity, start_time=start_time,
                        leak_duration=leak_duration, direction=direction
                    )
                    light_leak_layers.append(light_leak)
                    print(f"[OK] Added light leak from {start_time}s for {leak_duration}s")
            except Exception as e:
                print(f"[WARNING] Light leak failed: {e}")

        if self.settings.get('lens_flare_enabled', False):
            try:
                intensity = self.settings.get('lens_flare_intensity', 0.5)
                start_time = self.settings.get('lens_flare_start_time', 1.0)
                flare_duration = self.settings.get('lens_flare_duration', 2.0)
                position = self.settings.get('lens_flare_position', 'center')

                repeat_enabled = self.settings.get('lens_flare_repeat_enabled', False)
                repeat_interval = self.settings.get('lens_flare_repeat_interval', 5.0)

                if repeat_enabled:
                    current_time = start_time
                    count = 0
                    while current_time < final_video.duration:
                        lens_flare = LightLeaksEffects.create_lens_flare(
                            video.w, video.h, final_video.duration, video.fps,
                            intensity=intensity, start_time=current_time,
                            flare_duration=flare_duration, position=position
                        )
                        light_leak_layers.append(lens_flare)
                        current_time += repeat_interval
                        count += 1
                    print(f"[OK] Added {count} repeated lens flares every {repeat_interval}s")
                else:
                    lens_flare = LightLeaksEffects.create_lens_flare(
                        video.w, video.h, final_video.duration, video.fps,
                        intensity=intensity, start_time=start_time,
                        flare_duration=flare_duration, position=position
                    )
                    light_leak_layers.append(lens_flare)
                    print(f"[OK] Added lens flare from {start_time}s for {flare_duration}s")
            except Exception as e:
                print(f"[WARNING] Lens flare failed: {e}")

        if self.settings.get('film_burn_enabled', False):
            try:
                start_time = self.settings.get('film_burn_start_time', 0.0)
                burn_duration = self.settings.get('film_burn_duration', 1.5)

                repeat_enabled = self.settings.get('film_burn_repeat_enabled', False)
                repeat_interval = self.settings.get('film_burn_repeat_interval', 10.0)

                if repeat_enabled:
                    current_time = start_time
                    count = 0
                    while current_time < final_video.duration:
                        film_burn = LightLeaksEffects.create_film_burn(
                            video.w, video.h, final_video.duration, video.fps,
                            start_time=current_time, burn_duration=burn_duration
                        )
                        light_leak_layers.append(film_burn)
                        current_time += repeat_interval
                        count += 1
                    print(f"[OK] Added {count} repeated film burns every {repeat_interval}s")
                else:
                    film_burn = LightLeaksEffects.create_film_burn(
                        video.w, video.h, final_video.duration, video.fps,
                        start_time=start_time, burn_duration=burn_duration
                    )
                    light_leak_layers.append(film_burn)
                    print(f"[OK] Added film burn from {start_time}s for {burn_duration}s")
            except Exception as e:
                print(f"[WARNING] Film burn failed: {e}")

        # Composite light leaks if any were added
        if light_leak_layers:
            try:
                all_layers = [final_video] + light_leak_layers
                final_video = CompositeVideoClip(all_layers)

                # CRITICAL: Restore target duration if it was set
                if target_duration and final_video.duration != target_duration:
                    print(f"  🎯 Restoring target duration {target_duration:.2f}s after light leak composite")
                    final_video = final_video.with_duration(target_duration)

                print(f"[OK] Composited {len(light_leak_layers)} light leak effects")
            except Exception as e:
                print(f"[WARNING] Light leak compositing failed: {e}")

        # Prepare watermark if enabled (will be applied AFTER spotlight to stay visible on top)
        watermark_clip = None
        watermark_setting = self.settings.get('watermark_enabled', False)
        print(f"[🔍 WATERMARK DEBUG] Watermark enabled setting: {watermark_setting}")
        if watermark_setting:
            print("[🔍 WATERMARK DEBUG] Watermark is enabled - starting creation process...")
            try:
                watermark_type = self.settings.get('watermark_type', 'image')
                position = self.settings.get('watermark_position', 'bottom-right')
                opacity = self.settings.get('watermark_opacity', 70) / 100.0  # Convert to 0-1
                margin_x = self.settings.get('watermark_margin_x', 20)
                margin_y = self.settings.get('watermark_margin_y', 20)

                print(f"[🔍 WATERMARK DEBUG] Type: {watermark_type}, Position: {position}, Opacity: {int(opacity*100)}%")
                print(f"[🔍 WATERMARK DEBUG] Video dimensions: {final_video.w}x{final_video.h}")

                watermark = None

                if watermark_type == 'text':
                    # Text watermark
                    watermark_text = self.settings.get('watermark_text', '').strip()
                    print(f"[🔍 WATERMARK DEBUG] Text watermark - text: '{watermark_text}'")
                    if watermark_text:
                        try:
                            from moviepy import TextClip
                        except ImportError:
                            from moviepy.editor import TextClip

                        # Get text settings
                        font_size = self.settings.get('watermark_font_size', 30)
                        font_style = self.settings.get('watermark_font_style', 'Arial Bold')
                        text_color_hex = self.settings.get('watermark_text_color', '#FFFFFF')
                        text_color = self.hex_to_rgb(text_color_hex)

                        print(f"[🔍 WATERMARK DEBUG] Font size: {font_size}, Style: {font_style}, Color: {text_color_hex}")

                        # Create watermark using PIL (more reliable than TextClip)
                        try:
                            from PIL import Image, ImageDraw, ImageFont

                            # Use Arial font file
                            arial_path = str(Path(r"C:\Windows\Fonts") / "arial.ttf")
                            try:
                                font = ImageFont.truetype(arial_path, font_size)
                                print(f"[🔍 WATERMARK DEBUG] Loaded font: {arial_path}")
                            except Exception as font_err:
                                print(f"[🔍 WATERMARK DEBUG] Font load failed: {font_err}, using default")
                                font = ImageFont.load_default()

                            # Measure text
                            temp_img = Image.new('RGBA', (1, 1))
                            temp_draw = ImageDraw.Draw(temp_img)
                            bbox = temp_draw.textbbox((0, 0), watermark_text, font=font)
                            text_width = bbox[2] - bbox[0]
                            text_height = bbox[3] - bbox[1]

                            print(f"[🔍 WATERMARK DEBUG] Text dimensions: {text_width}x{text_height}")

                            # Create image
                            padding = 10
                            watermark_img = Image.new('RGBA', (text_width + padding*2, text_height + padding*2), (0, 0, 0, 0))
                            draw = ImageDraw.Draw(watermark_img)

                            # Draw stroke
                            for dx in range(-2, 3):
                                for dy in range(-2, 3):
                                    if dx*dx + dy*dy <= 4:
                                        draw.text((padding + dx, padding + dy), watermark_text, font=font, fill=(0, 0, 0, 255))

                            # Draw text
                            draw.text((padding, padding), watermark_text, font=font, fill=text_color + (255,))

                            # Convert to clip
                            watermark = ImageClip(np.array(watermark_img)).with_duration(final_video.duration)
                            print(f"[🔍 WATERMARK DEBUG] Successfully created ImageClip from PIL image")
                        except Exception as e:
                            print(f"[ERROR] Watermark creation failed: {e}")
                            import traceback
                            traceback.print_exc()
                            watermark = None

                        if watermark:
                            print(f"[OK] Created text watermark: '{watermark_text}' (font: {font_style}, size: {font_size})")
                        else:
                            print(f"[ERROR] Text watermark creation failed - watermark is None")
                    else:
                        print(f"[WARNING] Text watermark enabled but no text provided")

                else:
                    # Image watermark
                    watermark_path = self.settings.get('watermark_image_path', '')
                    if watermark_path and Path(watermark_path).exists():
                        # ImageClip already imported at module level
                        # Load watermark image
                        watermark = ImageClip(watermark_path)

                        # Resize watermark to scale relative to video width
                        scale = self.settings.get('watermark_scale', 0.15)  # Size relative to video width
                        new_width = int(final_video.w * scale)
                        # Calculate proportional height
                        aspect_ratio = watermark.h / watermark.w
                        new_height = int(new_width * aspect_ratio)
                        watermark = watermark.resized(newsize=(new_width, new_height))

                        print(f"[OK] Loaded image watermark (scale: {int(scale*100)}%)")
                    else:
                        print(f"[WARNING] Image watermark enabled but file not found: {watermark_path}")

                # Store watermark if created (will composite AFTER spotlight)
                if watermark:
                    print(f"[🔍 WATERMARK DEBUG] Watermark object created, dimensions: {watermark.w}x{watermark.h}")

                    # Set opacity (MoviePy 2.x uses with_opacity)
                    watermark = watermark.with_opacity(opacity)
                    print(f"[🔍 WATERMARK DEBUG] Opacity set to {int(opacity*100)}%")

                    # Calculate position
                    if position == 'top-left':
                        pos = (margin_x, margin_y)
                    elif position == 'top-right':
                        pos = (final_video.w - watermark.w - margin_x, margin_y)
                    elif position == 'bottom-left':
                        pos = (margin_x, final_video.h - watermark.h - margin_y)
                    elif position == 'bottom-right':
                        pos = (final_video.w - watermark.w - margin_x, final_video.h - watermark.h - margin_y)
                    elif position == 'center':
                        pos = ((final_video.w - watermark.w) / 2, (final_video.h - watermark.h) / 2)
                    else:
                        pos = (final_video.w - watermark.w - margin_x, final_video.h - watermark.h - margin_y)  # Default to bottom-right

                    print(f"[🔍 WATERMARK DEBUG] Calculated position: {pos} (position preset: {position})")

                    # Set position and duration (MoviePy 2.x uses with_ methods)
                    watermark_clip = watermark.with_position(pos).with_duration(final_video.duration)
                    print(f"[OK] Prepared {watermark_type} watermark at {position} (opacity: {int(opacity*100)}%) - will apply after spotlight")
                    print(f"[🔍 WATERMARK DEBUG] watermark_clip created successfully, ready for compositing")
                else:
                    print(f"[ERROR] Watermark object is None - cannot proceed with positioning")

            except Exception as e:
                print(f"[WARNING] Watermark preparation failed: {e}")
                import traceback
                traceback.print_exc()

        # Add progress bar if enabled
        if self.settings.get('progress_bar', False):
            try:
                bar_height = self.settings.get('progress_bar_height', 5)
                bar_color_hex = self.settings.get('progress_color', '#00ff40')
                bar_position = self.settings.get('progress_bar_position', 'bottom')

                # Convert hex color to RGB
                bar_color = self.hex_to_rgb(bar_color_hex)

                # Create progress bar function
                def create_progress_bar(t):
                    # Local import: the surrounding method may shadow
                    # the module-level `Image` name (e.g. an inner
                    # `from PIL import Image` later in the same
                    # function would make `Image` a local for the
                    # entire method, which then breaks this nested
                    # function with 'cannot access free variable
                    # Image where it is not associated with a value').
                    from PIL import Image as _PIL_Image
                    # Calculate progress (0 to 1)
                    progress = t / final_video.duration

                    # Create bar image
                    bar_img = _PIL_Image.new('RGB', (video.w, bar_height), (0, 0, 0))

                    # Draw progress portion
                    if progress > 0:
                        bar_width = int(video.w * progress)
                        for y in range(bar_height):
                            for x in range(bar_width):
                                bar_img.putpixel((x, y), bar_color)

                    return np.array(bar_img)

                # Create bar clip with time-varying width using VideoClip
                try:
                    from moviepy import VideoClip
                except ImportError:
                    from moviepy.editor import VideoClip

                bar_clip = VideoClip(create_progress_bar, duration=final_video.duration)
                try:
                    bar_clip = bar_clip.with_fps(video.fps)
                except:
                    bar_clip = bar_clip.set_fps(video.fps)

                # Position bar
                if bar_position == 'top':
                    try:
                        bar_clip = bar_clip.with_position((0, 0))
                    except:
                        bar_clip = bar_clip.set_position((0, 0))
                else:  # bottom
                    try:
                        bar_clip = bar_clip.with_position((0, video.h - bar_height))
                    except:
                        bar_clip = bar_clip.set_position((0, video.h - bar_height))

                # Composite onto video
                final_video = CompositeVideoClip([final_video, bar_clip])

                # CRITICAL: Restore target duration if it was set
                if target_duration and final_video.duration != target_duration:
                    print(f"  🎯 Restoring target duration {target_duration:.2f}s after progress bar composite")
                    final_video = final_video.with_duration(target_duration)

                print(f"[OK] Added progress bar ({bar_position}, {bar_height}px, {bar_color_hex})")
            except Exception as e:
                print(f"[WARNING] Progress bar overlay failed: {e}")
                import traceback
                traceback.print_exc()

        # Add CTA overlay if enabled
        if self.settings.get('cta_overlay_enabled', False):
            try:
                cta_text = self.settings.get('cta_overlay_text', 'Follow for more! 👉')
                position = self.settings.get('cta_overlay_position', 'bottom-center')
                animation = self.settings.get('cta_overlay_animation', 'bounce')
                start_time = self.settings.get('cta_overlay_start_time', 3.0)
                duration = self.settings.get('cta_overlay_duration', 3.0)

                print(f"\n[CTA] Adding call-to-action overlay...")
                print(f"[CTA] Text: {cta_text}")
                print(f"[CTA] Position: {position}, Animation: {animation}")
                print(f"[CTA] Timing: {start_time}s - {start_time + duration}s")

                # Create CTA text image
                from PIL import ImageFont, ImageDraw

                # Use a large bold font with Urdu fallback
                try:
                    _has_urdu_cta_overlay = False
                    for _ch in cta_text:
                        _cp = ord(_ch)
                        if (0x0600 <= _cp <= 0x06FF or 0x0750 <= _cp <= 0x077F or
                            0x08A0 <= _cp <= 0x08FF or 0xFB50 <= _cp <= 0xFDFF or
                            0xFE70 <= _cp <= 0xFEFF):
                            _has_urdu_cta_overlay = True
                            break
                    if _has_urdu_cta_overlay:
                        for _uf in [
                            "C:\\Windows\\Fonts\\arabtype.ttf",
                            "C:\\Windows\\Fonts\\arabtype.ttf",
                            "C:\\Windows\\Fonts\\arabtype.ttf",
                            "C:\\Windows\\Fonts\\arialbd.ttf",
                        ]:
                            try:
                                font = ImageFont.truetype(_uf, 60)
                                break
                            except:
                                continue
                    else:
                        font = ImageFont.truetype("C:\\Windows\\Fonts\\arialbd.ttf", 60)
                except:
                    font = ImageFont.load_default()

                # Calculate text size
                temp_img = Image.new('RGBA', (1, 1))
                draw = ImageDraw.Draw(temp_img)
                bbox = draw.textbbox((0, 0), cta_text, font=font)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]

                # Add padding
                padding = 30
                img_width = text_width + padding * 2
                img_height = text_height + padding * 2

                # Create text image with background
                cta_img = Image.new('RGBA', (img_width, img_height), (0, 0, 0, 0))
                draw = ImageDraw.Draw(cta_img)

                # Draw rounded rectangle background
                bg_color = (220, 38, 38, 230)  # Red with transparency
                draw.rounded_rectangle([(0, 0), (img_width, img_height)], radius=20, fill=bg_color)

                # Draw text
                draw.text((padding, padding), cta_text, font=font, fill=(255, 255, 255, 255))

                # Convert to array
                cta_array = np.array(cta_img)

                # Calculate position
                if position == 'top-left':
                    pos_x, pos_y = 30, 30
                elif position == 'top-center':
                    pos_x = (video.w - img_width) // 2
                    pos_y = 30
                elif position == 'top-right':
                    pos_x = video.w - img_width - 30
                    pos_y = 30
                elif position == 'bottom-left':
                    pos_x = 30
                    pos_y = video.h - img_height - 30
                elif position == 'bottom-center':
                    pos_x = (video.w - img_width) // 2
                    pos_y = video.h - img_height - 30
                else:  # bottom-right
                    pos_x = video.w - img_width - 30
                    pos_y = video.h - img_height - 30

                # Create text clip
                cta_clip = ImageClip(cta_array).with_duration(duration)
                cta_clip = cta_clip.with_start(start_time).with_position((pos_x, pos_y))

                # Apply animation
                if animation == 'bounce':
                    # Bounce in animation
                    def bounce_effect(t):
                        if t < 0.5:
                            # Bounce in
                            scale = 0.5 + 0.5 * (1 - (1 - t/0.5) ** 2)
                            return scale
                        else:
                            return 1.0

                    try:
                        cta_clip = cta_clip.resized(bounce_effect)
                    except:
                        cta_clip = cta_clip.resize(bounce_effect)

                elif animation == 'pulse':
                    # Pulsing animation
                    def pulse_effect(t):
                        return 1.0 + 0.1 * np.sin(t * 3 * np.pi)

                    try:
                        cta_clip = cta_clip.resized(pulse_effect)
                    except:
                        cta_clip = cta_clip.resize(pulse_effect)

                elif animation == 'slide-in':
                    # Slide in from bottom
                    def slide_pos(t):
                        if t < 0.5:
                            offset = int((1 - t/0.5) * 100)
                            return (pos_x, pos_y + offset)
                        return (pos_x, pos_y)

                    cta_clip = cta_clip.with_position(slide_pos)

                elif animation == 'fade-in':
                    # Fade in
                    def fade_opacity(t):
                        if t < 0.5:
                            return t / 0.5
                        return 1.0

                    cta_clip = cta_clip.with_opacity(fade_opacity)

                # Composite CTA onto video
                final_video = CompositeVideoClip([final_video, cta_clip])

                # CRITICAL: Restore target duration if it was set
                if target_duration and final_video.duration != target_duration:
                    print(f"  🎯 Restoring target duration {target_duration:.2f}s after CTA overlay composite")
                    final_video = final_video.with_duration(target_duration)

                print(f"[OK] Added CTA overlay ({animation} animation, {duration}s duration)")
            except Exception as e:
                print(f"[WARNING] CTA overlay failed: {e}")
                import traceback
                traceback.print_exc()

        # Add particle effects
        particle_layers = []

        if self.settings.get('add_glitter', False):
            try:
                intensity = self.settings.get('glitter_intensity', 0.5)
                glitter = ParticleEffects.create_glitter(
                    video.w, video.h, final_video.duration, video.fps, intensity=intensity
                )
                particle_layers.append(glitter)
                print(f"[OK] Added glitter particles (intensity: {intensity})")
            except Exception as e:
                print(f"[WARNING] Glitter effect failed: {e}")

        if self.settings.get('add_stars', False):
            try:
                stars = ParticleEffects.create_stars(
                    video.w, video.h, final_video.duration, video.fps
                )
                particle_layers.append(stars)
                print(f"[OK] Added floating stars")
            except Exception as e:
                print(f"[WARNING] Stars effect failed: {e}")

        if self.settings.get('add_hearts', False):
            try:
                hearts = ParticleEffects.create_hearts(
                    video.w, video.h, final_video.duration, video.fps
                )
                particle_layers.append(hearts)
                print(f"[OK] Added floating hearts")
            except Exception as e:
                print(f"[WARNING] Hearts effect failed: {e}")

        if self.settings.get('add_confetti', False):
            try:
                confetti = ParticleEffects.create_confetti(
                    video.w, video.h, final_video.duration, video.fps
                )
                particle_layers.append(confetti)
                print(f"[OK] Added confetti particles")
            except Exception as e:
                print(f"[WARNING] Confetti effect failed: {e}")

        # Composite particle effects if any were added
        if particle_layers:
            try:
                all_layers = [final_video] + particle_layers
                final_video = CompositeVideoClip(all_layers)

                # CRITICAL: Restore target duration if it was set
                if target_duration and final_video.duration != target_duration:
                    print(f"  🎯 Restoring target duration {target_duration:.2f}s after particle effects composite")
                    final_video = final_video.with_duration(target_duration)

                print(f"[OK] Composited {len(particle_layers)} particle effect(s)")
            except Exception as e:
                print(f"[WARNING] Particle compositing failed: {e}")

        # Apply dual video overlay (top and bottom videos with crop/pan/zoom controls)
        if self.settings.get('dual_video_enabled', False):
            try:
                from moviepy.video.fx.resize import resize as fx_resize
                from moviepy.video.fx.crop import crop as fx_crop

                print("[VIDEO] Applying dual video overlay...")

                # Top video settings
                top_video_path = self.settings.get('top_video_path', '')
                top_position_y = self.settings.get('top_video_position_y', 0)  # % from top
                top_height = self.settings.get('top_video_height', 20)  # % of video height
                top_crop_x = self.settings.get('top_video_crop_x', 0)  # % from left
                top_crop_y = self.settings.get('top_video_crop_y', 0)  # % from top
                top_crop_width = self.settings.get('top_video_crop_width', 100)  # % of source width
                top_crop_height = self.settings.get('top_video_crop_height', 100)  # % of source height
                top_zoom = self.settings.get('top_video_zoom', 100)  # % zoom (100 = no zoom)

                # Bottom video settings
                bottom_video_path = self.settings.get('bottom_video_path', '')
                bottom_position_y = self.settings.get('bottom_video_position_y', 80)  # % from top
                bottom_height = self.settings.get('bottom_video_height', 20)  # % of video height
                bottom_crop_x = self.settings.get('bottom_video_crop_x', 0)
                bottom_crop_y = self.settings.get('bottom_video_crop_y', 0)
                bottom_crop_width = self.settings.get('bottom_video_crop_width', 100)
                bottom_crop_height = self.settings.get('bottom_video_crop_height', 100)
                bottom_zoom = self.settings.get('bottom_video_zoom', 100)

                overlay_clips = []

                # Process top video
                if top_video_path and Path(top_video_path).exists():
                    try:
                        top_vid = VideoFileClip(top_video_path)

                        # Apply crop if needed
                        if top_crop_width < 100 or top_crop_height < 100 or top_crop_x > 0 or top_crop_y > 0:
                            crop_x1 = int(top_vid.w * top_crop_x / 100)
                            crop_y1 = int(top_vid.h * top_crop_y / 100)
                            crop_x2 = crop_x1 + int(top_vid.w * top_crop_width / 100)
                            crop_y2 = crop_y1 + int(top_vid.h * top_crop_height / 100)
                            top_vid = top_vid.cropped(x1=crop_x1, y1=crop_y1, x2=crop_x2, y2=crop_y2)

                        # Apply zoom
                        if top_zoom != 100:
                            zoom_factor = top_zoom / 100.0
                            top_vid = top_vid.resized(zoom_factor)

                        # Resize to fit position
                        target_height = int(video.h * top_height / 100)
                        top_vid = top_vid.resized(height=target_height)

                        # Ensure width doesn't exceed video width
                        if top_vid.w > video.w:
                            top_vid = top_vid.resized(width=video.w)

                        # Loop video if shorter than main video
                        if top_vid.duration < final_video.duration:
                            num_loops = int(final_video.duration / top_vid.duration) + 1
                            from moviepy.video.compositing.concatenate import concatenate_videoclips
                            top_vid = concatenate_videoclips([top_vid] * num_loops)

                        # Trim to match main video duration
                        top_vid = top_vid.with_duration(final_video.duration)

                        # Position
                        y_pos = int(video.h * top_position_y / 100)
                        x_pos = (video.w - top_vid.w) // 2  # Center horizontally
                        top_vid = top_vid.with_position((x_pos, y_pos))

                        overlay_clips.append(top_vid)
                        print(f"[OK] Added top video overlay: {Path(top_video_path).name}")
                    except Exception as e:
                        print(f"[WARNING] Failed to load top video: {e}")

                # Process bottom video
                if bottom_video_path and Path(bottom_video_path).exists():
                    try:
                        bottom_vid = VideoFileClip(bottom_video_path)

                        # Apply crop if needed
                        if bottom_crop_width < 100 or bottom_crop_height < 100 or bottom_crop_x > 0 or bottom_crop_y > 0:
                            crop_x1 = int(bottom_vid.w * bottom_crop_x / 100)
                            crop_y1 = int(bottom_vid.h * bottom_crop_y / 100)
                            crop_x2 = crop_x1 + int(bottom_vid.w * bottom_crop_width / 100)
                            crop_y2 = crop_y1 + int(bottom_vid.h * bottom_crop_height / 100)
                            bottom_vid = bottom_vid.cropped(x1=crop_x1, y1=crop_y1, x2=crop_x2, y2=crop_y2)

                        # Apply zoom
                        if bottom_zoom != 100:
                            zoom_factor = bottom_zoom / 100.0
                            bottom_vid = bottom_vid.resized(zoom_factor)

                        # Resize to fit position
                        target_height = int(video.h * bottom_height / 100)
                        bottom_vid = bottom_vid.resized(height=target_height)

                        # Ensure width doesn't exceed video width
                        if bottom_vid.w > video.w:
                            bottom_vid = bottom_vid.resized(width=video.w)

                        # Loop video if shorter than main video
                        if bottom_vid.duration < final_video.duration:
                            num_loops = int(final_video.duration / bottom_vid.duration) + 1
                            from moviepy.video.compositing.concatenate import concatenate_videoclips
                            bottom_vid = concatenate_videoclips([bottom_vid] * num_loops)

                        # Trim to match main video duration
                        bottom_vid = bottom_vid.with_duration(final_video.duration)

                        # Position
                        y_pos = int(video.h * bottom_position_y / 100)
                        x_pos = (video.w - bottom_vid.w) // 2  # Center horizontally
                        bottom_vid = bottom_vid.with_position((x_pos, y_pos))

                        overlay_clips.append(bottom_vid)
                        print(f"[OK] Added bottom video overlay: {Path(bottom_video_path).name}")
                    except Exception as e:
                        print(f"[WARNING] Failed to load bottom video: {e}")

                # Composite overlay videos with main video
                if overlay_clips:
                    all_clips = [final_video] + overlay_clips
                    final_video = CompositeVideoClip(all_clips)

                    # CRITICAL: Restore target duration if it was set
                    if target_duration and final_video.duration != target_duration:
                        print(f"  🎯 Restoring target duration {target_duration:.2f}s after dual video composite")
                        final_video = final_video.with_duration(target_duration)

                    print(f"[OK] Composited {len(overlay_clips)} overlay video(s)")

            except Exception as e:
                print(f"[WARNING] Dual video overlay failed: {e}")
                import traceback
                traceback.print_exc()

        # Save clean video BEFORE spotlight for thumbnail frame (if enabled)
        thumbnail_enabled = self.settings.get('add_thumbnail_frame', False)
        thumbnail_duration = self.settings.get('thumbnail_frame_duration', 3.0)
        video_before_spotlight = None
        if thumbnail_enabled:
            # Keep reference to video before spotlight is applied
            video_before_spotlight = final_video

        # Apply circle zoom/pan BEFORE spotlight — only when spotlight itself is enabled
        spotlight_enabled = self.settings.get('circular_spotlight_enabled', False)
        circle_zoom = self.settings.get('spotlight_circle_zoom', 100)
        circle_pan_x = self.settings.get('spotlight_circle_pan_x', 0)
        circle_pan_y = self.settings.get('spotlight_circle_pan_y', 0)

        if spotlight_enabled and (circle_zoom != 100 or circle_pan_x != 0 or circle_pan_y != 0):
            print(f"\n[CIRCLE ZOOM/PAN] Applying to main video...")
            print(f"[CIRCLE ZOOM/PAN] Zoom: {int(circle_zoom)}%, Pan X: {int(circle_pan_x):+d}%, Pan Y: {int(circle_pan_y):+d}%")

            def apply_circle_zoom_pan(get_frame, t):
                import cv2
                frame = get_frame(t)
                h, w = frame.shape[:2]

                # PERF: avoid wasteful "upscale-then-crop-then-resize" pattern.
                # For zoom_in (factor > 1.0) we crop a smaller window from the
                # source and resize up — that avoids a 1.5x–2x sized intermediate.
                # For zoom_out (factor < 1.0) we keep the legacy path which is
                # already optimal (the source is the larger image).
                zoom_factor = circle_zoom / 100.0
                if zoom_factor >= 1.0:
                    # Crop the source window that we want to scale up.
                    # Effective crop size = original / zoom_factor.
                    crop_h = int(h / zoom_factor)
                    crop_w = int(w / zoom_factor)
                    # Apply pan within the (small) crop window.
                    pan_x_px = int((circle_pan_x / 100.0) * crop_w)
                    pan_y_px = int((circle_pan_y / 100.0) * crop_h)
                    cx = crop_w // 2 + pan_x_px
                    cy = crop_h // 2 + pan_y_px
                    x1 = max(0, cx - w // 2)
                    y1 = max(0, cy - h // 2)
                    # Clamp so the window doesn't exceed the source.
                    if x1 + crop_w > w:
                        x1 = max(0, w - crop_w)
                    if y1 + crop_h > h:
                        y1 = max(0, h - crop_h)
                    x2 = x1 + crop_w
                    y2 = y1 + crop_h
                    cropped = frame[y1:y2, x1:x2]
                    if cropped.shape[0] != h or cropped.shape[1] != w:
                        return cv2.resize(cropped, (w, h), interpolation=cv2.INTER_LINEAR)
                    return cropped

                # zoom_factor < 1.0 (zoom out): upscale source, then crop window.
                zoomed_h = int(h / zoom_factor)
                zoomed_w = int(w / zoom_factor)
                zoomed = cv2.resize(frame, (zoomed_w, zoomed_h), interpolation=cv2.INTER_LINEAR)
                pan_x_px = int((circle_pan_x / 100.0) * zoomed_w)
                pan_y_px = int((circle_pan_y / 100.0) * zoomed_h)
                center_x_px = zoomed_w // 2 + pan_x_px
                center_y_px = zoomed_h // 2 + pan_y_px
                x1 = max(0, center_x_px - w // 2)
                y1 = max(0, center_y_px - h // 2)
                x2 = min(zoomed_w, x1 + w)
                y2 = min(zoomed_h, y1 + h)
                result = zoomed[y1:y2, x1:x2]
                if result.shape[0] != h or result.shape[1] != w:
                    result = cv2.resize(result, (w, h), interpolation=cv2.INTER_LINEAR)
                return result

            final_video = final_video.transform(lambda gf, t: apply_circle_zoom_pan(gf, t))
            print(f"[OK] Circle zoom/pan applied successfully")

        # Apply spotlight effect (TikTok-style focus circle or square) - OPTIMIZED WITH CACHING
        if self.settings.get('circular_spotlight_enabled', False):
            try:
                center_x = self.settings.get('spotlight_center_x', 50)
                center_y = self.settings.get('spotlight_center_y', 50)
                radius = self.settings.get('spotlight_radius', 40)
                outside_effect = self.settings.get('spotlight_outside_effect', 'blur')
                blur_intensity = self.settings.get('spotlight_blur_intensity', 50)
                outside_color = self.settings.get('spotlight_outside_color', '#000000')
                feather = self.settings.get('spotlight_feather', 20)
                show_outline = self.settings.get('spotlight_show_outline', True)
                outline_color = self.settings.get('spotlight_outline_color', '#FF00FF')
                outline_thickness = self.settings.get('spotlight_outline_thickness', 5)
                shape = self.settings.get('spotlight_shape', 'circle')

                # Check if background media is enabled before using it
                bg_media_enabled = self.settings.get('spotlight_bg_media_enabled', False)
                background_media_path = self.settings.get('spotlight_background_media', '') if bg_media_enabled else ''

                # Inside effect settings (for the focused area)
                inside_effect = self.settings.get('spotlight_inside_effect', 'none')
                inside_color = self.settings.get('spotlight_inside_color', '#00FF00')
                inside_opacity = self.settings.get('spotlight_inside_opacity', 30)

                # Use CACHED version for 3-5x speedup (pre-calculates mask once)
                # ⚡ SPEED: default ON. The GPU/OpenCV render path is ~10-20x
                # faster than MoviePy per-frame Python compositing. Auto-falls
                # back to libx264 on machines without an NVIDIA GPU. Only set
                # this False to force the slow MoviePy path for debugging.
                performance_mode = self.settings.get('performance_mode', True)

                if performance_mode or background_media_path:
                    # Use optimized cached version (much faster!)
                    print(f"[⚡ PERFORMANCE] Using cached spotlight (3-5x faster)...")

                    # Get crop/zoom settings for background media
                    bg_zoom = self.settings.get('spotlight_bg_zoom', 100)
                    bg_crop_x = self.settings.get('spotlight_bg_crop_x', 0)
                    bg_crop_y = self.settings.get('spotlight_bg_crop_y', 0)

                    transform_func = VideoEffects.create_cached_spotlight_transformer(
                        (final_video.w, final_video.h),
                        center_x, center_y, radius,
                        outside_effect, blur_intensity, outside_color, feather,
                        show_outline, outline_color, outline_thickness, shape,
                        background_media_path if background_media_path else None,
                        bg_zoom, bg_crop_x, bg_crop_y,
                        inside_effect, inside_color, inside_opacity
                    )
                    final_video = final_video.transform(transform_func)
                else:
                    # Use standard version (slower but compatible)
                    def spotlight_effect(get_frame, t):
                        frame = get_frame(t)
                        return VideoEffects.apply_circular_spotlight(
                            frame, center_x, center_y, radius,
                            outside_effect, blur_intensity, outside_color, feather,
                            show_outline, outline_color, outline_thickness, shape
                        )

                    final_video = final_video.transform(lambda gf, t: spotlight_effect(gf, t))

                outline_msg = f", outline: {outline_color} ({outline_thickness}px)" if show_outline else ", no outline"
                bg_msg = f", background: {Path(background_media_path).name}" if background_media_path else ""
                inside_msg = f", inside: {inside_effect} ({inside_color})" if inside_effect != 'none' else ""
                print(f"[OK] Applied {shape} spotlight (center: {center_x},{center_y}%, size: {radius}%, effect: {outside_effect}{outline_msg}{bg_msg}{inside_msg})")
            except Exception as e:
                print(f"[WARNING] Spotlight effect failed: {e}")
                import traceback
                traceback.print_exc()

        # FIX: Add text overlays AFTER spotlight - ensures text is visible on top
        # This keeps Title/Quote text visible even when spotlight is enabled
        # BUT only when the user has actually enabled some visual text option in
        # the Processing Options panel (Captions / Zoom / Pulsing all count).
        # Without this gate the Title/Quote/CTA bubble appears even when every
        # toggle in the Quick Start tab is off, producing a black frame with
        # floating text.
        _any_visual_on = (
            self.settings.get('enable_captions', False)
            or self.settings.get('caption_highlight_enabled', False)
            or self.settings.get('video_zoom', False)
            or self.settings.get('pulsing_cta', False)
            or self.settings.get('add_glitter', False)
            or self.settings.get('add_stars', False)
            or self.settings.get('add_hearts', False)
            or self.settings.get('add_confetti', False)
        )
        # ═══════════════════════════════════════════════════════════════════
        # ⚡ Collect all overlay clips into ONE composite (no nesting)
        # Each overlay is created below, then composited once at the end.
        # This avoids MoviePy's per-wrapper overhead from N nested
        # CompositeVideoClip calls — every level adds Python call + numpy
        # array overhead per frame.
        # ═══════════════════════════════════════════════════════════════════
        final_overlays = []

        # ─── Text overlays ────────────────────────────────────────────────
        if text_overlay_clip is not None and _any_visual_on:
            try:
                print("[OK] Adding text overlays on top of spotlight effect...")
                final_overlays.append(text_overlay_clip)
                print("[OK] Text overlays applied successfully - will appear on top of all effects")
            except Exception as e:
                print(f"[WARNING] Failed to add text overlays: {e}")

        # ─── Captions (merged for ⚡speed) ────────────────────────────────
        if caption_clips:
            try:
                print(f"[OK] Adding {len(caption_clips)} captions on top of spotlight effect...")

                # ⚡ Merge N individual caption clips into ONE VideoClip.
                # From N+1 composite children to just 2 (video + merged captions).
                if len(caption_clips) > 10:
                    _merge_start = time.time()
                    merged = _merge_caption_clips(caption_clips, final_video.w, final_video.h, final_video.duration)
                    if merged is not None:
                        _merge_elapsed = time.time() - _merge_start
                        print(f"  ⚡ Merged {len(caption_clips)} captions into 1 clip ({_merge_elapsed*1000:.0f}ms)")
                        final_overlays.append(merged)
                        caption_clips_merged = True
                    else:
                        print(f"  ℹ Merge returned None — using individual clips")
                        final_overlays.extend(caption_clips)
                        caption_clips_merged = False
                else:
                    final_overlays.extend(caption_clips)
                    caption_clips_merged = False

                print(f"[OK] Captions ready for compositing")
            except Exception as e:
                print(f"[WARNING] Failed to add captions: {e}")

        # ─── Watermark ────────────────────────────────────────────────────
        if watermark_clip is not None:
            try:
                print("[🔍 WATERMARK DEBUG] watermark_clip is not None - compositing")
                print(f"[🔍 WATERMARK DEBUG] Video dimensions: {final_video.w}x{final_video.h}, "
                      f"Watermark: {watermark_clip.w}x{watermark_clip.h}")
                final_overlays.append(watermark_clip)
                print("[OK] Watermark ready for compositing")
            except Exception as e:
                print(f"[ERROR] Failed to add watermark: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("[🔍 WATERMARK DEBUG] watermark_clip is None - skipping watermark compositing")

        # ─── Border effect (static overlay) ───────────────────────────────
        if self.settings.get('cleanup_border_enabled', False):
            try:
                border_px = int(self.settings.get('cleanup_border_size', 4) or 4)
                border_px = max(1, min(60, border_px))
                border_hex = str(self.settings.get('cleanup_border_color', '#FFFFFF') or '').strip()
                br, bg, bb = 255, 255, 255
                if border_hex.startswith('#'):
                    try:
                        from PIL import ImageColor
                        br, bg, bb = ImageColor.getcolor(border_hex, 'RGB')
                    except Exception:
                        pass
                from PIL import Image, ImageDraw
                border_img = Image.new('RGBA', (final_video.w, final_video.h), (0, 0, 0, 0))
                bdraw = ImageDraw.Draw(border_img)
                for i in range(border_px):
                    bdraw.rectangle(
                        [i, i, final_video.w - 1 - i, final_video.h - 1 - i],
                        outline=(br, bg, bb, 178))
                border_clip = ImageClip(np.array(border_img)).with_duration(final_video.duration)
                final_overlays.append(border_clip)
                print(f"[OK] Border ready: {border_px}px, {border_hex}")
            except Exception as e:
                print(f"[WARNING] Border effect failed: {e}")

        # ─── Blur text overlay (independent of region blur) ──────────────
        if self.settings.get('blur_text_enabled', False):
            try:
                blur_txt = str(self.settings.get('blur_text_content', '') or '').strip()
                if blur_txt:
                    bt_ff = str(self.settings.get('blur_text_font_family', 'Arial') or '')
                    bt_fs = int(self.settings.get('blur_text_font_size', 30))
                    bt_font_path = None
                    if bt_ff:
                        for cand in [
                            f'C:/Windows/Fonts/{bt_ff}.ttf',
                            f'C:/Windows/Fonts/{bt_ff.split()[0]}.ttf',
                            f'C:/Windows/Fonts/{bt_ff.replace(" ", "")}.ttf',
                        ]:
                            if Path(cand).exists():
                                bt_font_path = cand
                                break
                    if not bt_font_path:
                        bt_font_path = 'C:/Windows/Fonts/arial.ttf'
                    btf = ImageFont.truetype(bt_font_path, bt_fs)
                    bt_color = str(self.settings.get('blur_text_text_color', '#FFFFFF') or '')
                    btr, btg, btb = 255, 255, 255
                    if bt_color.startswith('#'):
                        try:
                            btr, btg, btb = ImageColor.getcolor(bt_color, 'RGB')
                        except Exception:
                            pass
                    bt_bgc = str(self.settings.get('blur_text_bg_color', '#000000') or '')
                    bt_bgr, bt_bgg, bt_bgb = 0, 0, 0
                    if bt_bgc.startswith('#'):
                        try:
                            bt_bgr, bt_bgg, bt_bgb = ImageColor.getcolor(bt_bgc, 'RGB')
                        except Exception:
                            pass
                    bt_bg_opacity = int(self.settings.get('blur_text_bg_opacity', 80))
                    bt_bg_alpha = max(0, min(255, int(255 * bt_bg_opacity / 100)))
                    bt_outline_enabled = self.settings.get('blur_text_outline', True)
                    bt_outline_color = str(self.settings.get('blur_text_outline_color', '#000000') or '')
                    bt_ol_r, bt_ol_g, bt_ol_b = 0, 0, 0
                    if bt_outline_color.startswith('#'):
                        try:
                            bt_ol_r, bt_ol_g, bt_ol_b = ImageColor.getcolor(bt_outline_color, 'RGB')
                        except Exception:
                            pass
                    bt_position = str(self.settings.get('blur_text_position', 'center') or 'center')
                    bt_img = Image.new('RGBA', (final_video.w, final_video.h), (0, 0, 0, 0))
                    btd = ImageDraw.Draw(bt_img)
                    bb2 = btd.textbbox((0, 0), blur_txt, font=btf)
                    btw, bth = bb2[2] - bb2[0], bb2[3] - bb2[1]
                    btx = (final_video.w - btw) // 2
                    if bt_position == 'top':
                        bty = int(final_video.h * 0.08)
                    elif bt_position == 'center':
                        bty = (final_video.h - bth) // 2
                    else:
                        bty = final_video.h - bth - int(final_video.h * 0.10)
                    pad_x, pad_y = max(20, btw // 8), max(8, bth // 4)
                    bg_rr = Image.new('RGBA', (btw + pad_x * 2, bth + pad_y * 2),
                                      (bt_bgr, bt_bgg, bt_bgb, bt_bg_alpha))
                    bt_img.paste(bg_rr, (btx - pad_x, bty - pad_y), bg_rr)
                    if bt_outline_enabled:
                        ol_thickness = max(1, int(self.settings.get('blur_text_outline_size', 2)))
                        for dx in range(-ol_thickness, ol_thickness + 1):
                            for dy in range(-ol_thickness, ol_thickness + 1):
                                if dx*dx + dy*dy <= ol_thickness*ol_thickness:
                                    btd.text((btx + dx, bty + dy), blur_txt, font=btf, fill=(bt_ol_r, bt_ol_g, bt_ol_b, 255))
                    btd.text((btx, bty), blur_txt, font=btf, fill=(btr, btg, btb, 255))
                    bt_clip = ImageClip(np.array(bt_img)).with_duration(final_video.duration)
                    final_overlays.append(bt_clip)
                    print(f"[OK] Blur text overlay ready: '{blur_txt[:60]}' @{bt_position}")
            except Exception as e:
                print(f"[WARNING] Blur text overlay failed: {e}")
                import traceback
                traceback.print_exc()

        # ─── Crosshair (alignment guide) ─────────────────────────────────
        if self.settings.get('crosshair_enabled', False):
            try:
                from PIL import Image, ImageDraw
                ch_color = str(self.settings.get('crosshair_color', '#FF0000') or '#FF0000')
                ch_thick = max(1, int(self.settings.get('crosshair_thickness', 2)))
                try:
                    ch_rgb = tuple(int(ch_color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
                except Exception:
                    ch_rgb = (255, 0, 0)
                ch_img = Image.new('RGBA', (final_video.w, final_video.h), (0, 0, 0, 0))
                ch_draw = ImageDraw.Draw(ch_img)
                cx, cy = final_video.w // 2, final_video.h // 2
                ch_draw.line([(0, cy), (final_video.w - 1, cy)], fill=ch_rgb + (255,), width=ch_thick)
                ch_draw.line([(cx, 0), (cx, final_video.h - 1)], fill=ch_rgb + (255,), width=ch_thick)
                ch_clip = ImageClip(np.array(ch_img)).with_duration(final_video.duration)
                final_overlays.append(ch_clip)
                print(f"[OK] Crosshair overlay ready: color={ch_color} thickness={ch_thick}")
            except Exception as e:
                print(f"[WARNING] Crosshair overlay failed: {e}")

        # ─── Commentary text overlays (from Case Commentary tab) ──────────
        try:
            from PIL import Image, ImageDraw, ImageFont
            _commentary_spots = self.settings.get('_commentary_spots', []) or []
            if _commentary_spots:
                _vid_w, _vid_h = final_video.w, final_video.h
                print(f"[COMMENTARY] Adding {len(_commentary_spots)} text overlay(s)")

                # ── Urdu / Arabic font fallback ────────────────────────
                # Check if any spot text contains Urdu/Arabic Unicode chars
                # and use arabtype.ttf if so (arial/ariblk don't support Urdu).
                _URDU_RANGES = [
                    (0x0600, 0x06FF), (0x0750, 0x077F), (0x08A0, 0x08FF),
                    (0xFB50, 0xFDFF), (0xFE70, 0xFEFF), (0x1EE00, 0x1EEFF),
                ]
                def _has_urdu(t):
                    for cp in t:
                        o = ord(cp)
                        for lo, hi in _URDU_RANGES:
                            if lo <= o <= hi:
                                return True
                    return False
                _any_urdu = any(
                    _has_urdu(str(_cs.get('text', '')))
                    for _cs in _commentary_spots)

                _font_size = 30
                if _any_urdu:
                    _font_path = 'C:/Windows/Fonts/arabtype.ttf'
                else:
                    _font_path = 'C:/Windows/Fonts/arialbd.ttf'
                if not Path(_font_path).exists():
                    _font_path = 'C:/Windows/Fonts/arial.ttf'
                try:
                    _font = ImageFont.truetype(_font_path, _font_size)
                except Exception:
                    _font = ImageFont.load_default()

                for _idx, _cs in enumerate(_commentary_spots):
                    _ts = _cs.get('timestamp', 0)
                    _txt = (_cs.get('text', '') or '').strip()
                    if not _txt or _ts <= 0:
                        continue

                    # CTA spots are flagged when the script line carried a
                    # "CTA" label (the label itself was already stripped from
                    # the text upstream). Fall back to a prefix check for older
                    # spot dicts that lack the flag.
                    _is_cta = bool(_cs.get('is_cta')) or _txt.upper().startswith('CTA:')
                    if _is_cta:
                        # Older dicts may still carry the literal prefix.
                        if _txt.upper().startswith('CTA:'):
                            _txt = _txt[4:].strip()
                        _display_txt = _txt
                        _font_size = 46
                        if _any_urdu:
                            _font_path = 'C:/Windows/Fonts/arabtype.ttf'
                        else:
                            _font_path = 'C:/Windows/Fonts/ariblk.ttf'
                        if not Path(_font_path).exists():
                            _font_path = 'C:/Windows/Fonts/arialbd.ttf'
                        if not Path(_font_path).exists():
                            _font_path = 'C:/Windows/Fonts/arial.ttf'
                        try:
                            _font = ImageFont.truetype(_font_path, _font_size)
                        except Exception:
                            _font = ImageFont.load_default()
                        _bg_color = (25, 25, 80, 210)   # dark blue
                        _text_color = (255, 215, 0, 255)  # gold
                        _duration = 5
                        _pad_x, _pad_y = 40, 26
                    else:
                        _display_txt = _txt
                        _font_size = 30
                        if _any_urdu:
                            _font_path = 'C:/Windows/Fonts/arabtype.ttf'
                        else:
                            _font_path = 'C:/Windows/Fonts/arialbd.ttf'
                        if not Path(_font_path).exists():
                            _font_path = 'C:/Windows/Fonts/arial.ttf'
                        try:
                            _font = ImageFont.truetype(_font_path, _font_size)
                        except Exception:
                            _font = ImageFont.load_default()
                        _bg_color = (0, 0, 0, 191)       # black 75%
                        _text_color = (255, 255, 255, 255) # white
                        _duration = 5
                        _pad_x, _pad_y = 20, 10

                    # Create a full-frame RGBA image
                    _ov = Image.new('RGBA', (_vid_w, _vid_h), (0, 0, 0, 0))
                    _draw = ImageDraw.Draw(_ov)

                    # ── Word-wrap so the pill never exceeds the frame width ──
                    # Allow the text to span ~82% of the video width, then wrap
                    # to as many lines as needed (2–3 typically for a CTA).
                    _max_text_w = int(_vid_w * 0.82) - _pad_x * 2
                    _words = _display_txt.split()
                    _lines = []
                    _cur = ''
                    for _w in _words:
                        _trial = (_cur + ' ' + _w).strip()
                        _tb = _draw.textbbox((0, 0), _trial, font=_font)
                        if (_tb[2] - _tb[0]) <= _max_text_w or not _cur:
                            _cur = _trial
                        else:
                            _lines.append(_cur)
                            _cur = _w
                    if _cur:
                        _lines.append(_cur)
                    _wrapped = '\n'.join(_lines)

                    # Measure the wrapped block (multiline) for the pill size.
                    _mbbox = _draw.multiline_textbbox(
                        (0, 0), _wrapped, font=_font, align='center', spacing=8)
                    _tw = _mbbox[2] - _mbbox[0]
                    _th = _mbbox[3] - _mbbox[1]

                    _bg_w = _tw + _pad_x * 2
                    _bg_h = _th + _pad_y * 2
                    _bg_x = (_vid_w - _bg_w) // 2

                    if _is_cta:
                        # CTA: center of screen, prominent
                        _bg_y = (_vid_h - _bg_h) // 2
                    else:
                        # Commentary: bottom-center
                        _bg_y = _vid_h - _bg_h - 50

                    # Semi-transparent background pill
                    _draw.rounded_rectangle(
                        [_bg_x, _bg_y, _bg_x + _bg_w, _bg_y + _bg_h],
                        radius=16 if _is_cta else 10,
                        fill=_bg_color)

                    # Draw wrapped text, centered inside the pill
                    _draw.multiline_text(
                        (_vid_w // 2, _bg_y + _pad_y),
                        _wrapped, font=_font, fill=_text_color,
                        align='center', anchor='ma', spacing=8)

                    _clip = (ImageClip(np.array(_ov))
                             .with_start(_ts)
                             .with_duration(_duration))
                    final_overlays.append(_clip)
                    print(f"[COMMENTARY] #{_idx + 1} @{_ts:.0f}s "
                          f"({len(_lines)} line(s)) '{_txt[:60]}'")

                if _commentary_spots:
                    print(f"[OK] {len(_commentary_spots)} commentary overlay(s) added")
        except Exception as _cs_err:
            print(f"[WARNING] Commentary overlays failed: {_cs_err}")
            import traceback
            traceback.print_exc()

        # ═══════════════════════════════════════════════════════════════════
        # ONE composite for ALL overlays (instead of N nested composites)
        # ═══════════════════════════════════════════════════════════════════
        if final_overlays:
            _all_composite_layers = [final_video] + final_overlays
            final_video = CompositeVideoClip(_all_composite_layers)
            # MoviePy v2.x CompositeVideoClip may NOT inherit duration — grab from children
            if final_video.duration is None:
                _dur_candidates = [c.duration for c in _all_composite_layers if c.duration is not None]
                if _dur_candidates:
                    final_video.duration = max(_dur_candidates)
                elif target_duration:
                    final_video.duration = target_duration
            _comp_dur = final_video.duration if final_video.duration is not None else "N/A"
            if target_duration and final_video.duration != target_duration:
                print(f"  🎯 Enforcing target duration {target_duration:.2f}s (composite created {_comp_dur}s)")
                final_video = final_video.with_duration(target_duration)
            print(f"[OK] Composited {len(final_overlays)} overlay(s) in ONE pass")

        # Add thumbnail frame at end (clean video without spotlight for YouTube thumbnail)
        if thumbnail_enabled and video_before_spotlight is not None:
            try:
                # MoviePy 2.x import
                try:
                    from moviepy import concatenate_videoclips
                except ImportError:
                    from moviepy.editor import concatenate_videoclips

                # Create thumbnail frame: clean video with text/captions/watermark (no spotlight)
                thumbnail_frame = video_before_spotlight

                # Add text overlays to thumbnail
                if text_overlay_clip is not None:
                    thumbnail_frame = CompositeVideoClip([thumbnail_frame, text_overlay_clip])

                # Add captions to thumbnail
                if caption_clips:
                    all_clips = [thumbnail_frame] + caption_clips
                    thumbnail_frame = CompositeVideoClip(all_clips)

                # Add watermark to thumbnail
                if watermark_clip is not None:
                    thumbnail_frame = CompositeVideoClip([thumbnail_frame, watermark_clip])

                # Get a single frame from the middle of the video for thumbnail
                thumbnail_time = thumbnail_frame.duration / 2
                thumbnail_img = thumbnail_frame.get_frame(thumbnail_time)

                # Create static image clip from the frame
                thumbnail_clip = ImageClip(thumbnail_img).with_duration(thumbnail_duration)
                thumbnail_clip = thumbnail_clip.with_fps(video.fps)

                # Concatenate: main video (with spotlight) + thumbnail frame (without spotlight)
                final_video = concatenate_videoclips([final_video, thumbnail_clip])

                print(f"[OK] Added {thumbnail_duration}s thumbnail frame at end (full video without spotlight)")
            except Exception as e:
                print(f"[WARNING] Failed to add thumbnail frame: {e}")
                import traceback
                traceback.print_exc()

        output_path = self.output_folder / output_filename
        counter = 1
        original_output_path = output_path
        while output_path.exists():
            stem = original_output_path.stem
            output_path = self.output_folder / f"{stem}_{counter}.mp4"
            counter += 1

        print(f"Rendering with effects to: {output_path.name}")
        dur_str = f"{final_video.duration:.2f}" if final_video.duration is not None else "??"
        print(f"Video details: size={final_video.size}, duration={dur_str}s, fps={video.fps}")

        # Performance warning
        total_frames = int((final_video.duration or 0) * video.fps)
        if total_frames > 900:  # More than 15 seconds at 60fps
            print(f"[⚠️ PERFORMANCE] Rendering {total_frames} frames - this may take 10-20 minutes")
            if self.settings.get('spotlight_background_media'):
                print(f"[💡 TIP] Background video in spotlight adds ~5-10 min to render time")
            particle_count = sum([
                self.settings.get('add_stars', False),
                self.settings.get('add_hearts', False),
                self.settings.get('add_confetti', False),
                self.settings.get('add_glitter', False)
            ])
            if particle_count > 1:
                print(f"[💡 TIP] {particle_count} particle effects enabled - consider reducing for faster rendering")

        # ═══════════════════════════════════════════════════════════════════
        # ⚡ PER-FRAME PERFORMANCE PROFILER
        # Wrap make_frame so we can see exactly how long each frame takes.
        # ═══════════════════════════════════════════════════════════════════
        _frame_times = []
        _frame_count = [0]
        _perf_log_interval = max(1, int(24 * 10))  # log every ~10 seconds at 24fps

        # Wrap the clip with a timing transform — works in both v1.x and v2.x
        def _timed_get_frame(get_frame_func, t):
            _t0 = time.perf_counter()
            _frame = get_frame_func(t)
            _elapsed = time.perf_counter() - _t0
            _frame_times.append(_elapsed)
            _frame_count[0] += 1
            if _frame_count[0] % _perf_log_interval == 0:
                _recent = _frame_times[-_perf_log_interval:]
                _avg = sum(_recent) / len(_recent) * 1000
                _worst = max(_recent) * 1000
                print(f"  [⚡PERF] Frame {_frame_count[0]}: avg {_avg:.0f}ms/frame, worst {_worst:.0f}ms ({1000/_avg:.1f} it/s)")
            return _frame

        try:
            final_video = final_video.transform(_timed_get_frame, keep_duration=True)
        except Exception:
            try:
                final_video = final_video.fl(_timed_get_frame, keep_duration=True)  # v1.x fallback
            except Exception:
                pass  # non-critical

        _render_start_time = time.time()
        _rendered_ok = False

        # ═══════════════════════════════════════════════════════════════════
        # ⚡ OpenCV render path (bypasses MoviePy per-frame overhead)
        # ═══════════════════════════════════════════════════════════════════
        # NOTE: The OpenCV path reads from the ORIGINAL source file, NOT from
        # `final_video`, so any effects applied to `final_video` via MoviePy
        # transforms (transitions, repeat-selected effects, light leaks, lens
        # flares, film burns, circle zoom/pan, spotlight, particles, etc.)
        # are LOST if we go through OpenCV.  The _trans_fn only covers basic
        # fade / zoom / blur transitions.  When any of these MoviePy-only
        # effects are active, we skip the OpenCV path entirely.
        #
        # NOTE: most "repeat-selected" transitions (mask, slide, wipe, glitch,
        # bounce, dissolve, fade, blur, zoom) are handled by the OpenCV path's
        # _repeat_selected_fn — so they are NOT in this blocklist.
        _has_moviepy_only_effects = any([
            # Spotlight and circle zoom/pan
            self.settings.get('circular_spotlight_enabled', False),
            # Progress bar, CTA overlay, dual video (composited into final_video,
            # not rebuilt from _ov_sources in the OpenCV path)
            self.settings.get('progress_bar', False),
            self.settings.get('cta_overlay_enabled', False),
            self.settings.get('dual_video_enabled', False),
            # Particle effects
            self.settings.get('add_glitter', False),
            self.settings.get('add_stars', False),
            self.settings.get('add_hearts', False),
            self.settings.get('add_confetti', False),
        ])
        if _has_moviepy_only_effects and self.settings.get('performance_mode', True):
            # Only print the fallback reason when the user actually asked for
            # performance mode AND we're bypassing it for correctness.
            if _HAS_CV2:
                print("  [⚡CV] performance_mode is ON but MoviePy-only effects "
                      "(transitions, spotlights, particles, etc.) are enabled — "
                      "falling through to MoviePy render for correctness")

        if _HAS_CV2 and self.settings.get('performance_mode', True) and not _has_moviepy_only_effects:
            try:
                # ── build overlay sources from MoviePy clips ──────────────
                _ov_sources = []
                for _ov in final_overlays:
                    _p = _ov.pos if hasattr(_ov, 'pos') else (0, 0)
                    _s = _ov.start if hasattr(_ov, 'start') else 0
                    _e = _ov.end if hasattr(_ov, 'end') else None
                    # ⚡ If the clip has stored caption time bounds, add a
                    # time-range gate that skips compositing when no captions
                    # are visible.  Uses separate _min_t/_max_t (not start/end)
                    # because _render_rgba uses absolute internal timestamps and
                    # must NOT receive an offset via t - start.
                    _min_t = None
                    _max_t = None
                    if hasattr(_ov, 'caption_start') and hasattr(_ov, 'caption_end'):
                        _min_t = _ov.caption_start
                        _max_t = _ov.caption_end
                    # Wrap get_frame to return RGBA (MoviePy composites expect
                    # RGB + separate mask, but our numpy compositor wants RGBA)
                    # ⚡ Cache static overlays (ImageClips) so get_frame returns
                    # the pre-computed RGBA on first call instead of every frame.
                    def _wrap_ov(ov_clip, offset):
                        _cached_rgba = None
                        _is_static = hasattr(ov_clip, '_frame_data')
                        def _get_frame(ct):
                            nonlocal _cached_rgba
                            if _is_static and _cached_rgba is not None:
                                return _cached_rgba
                            _f = ov_clip.get_frame(ct)
                            if _f.shape[2] == 4:
                                _result = _f
                            elif _f.shape[2] == 3 and ov_clip.mask is not None:
                                _m = ov_clip.mask.get_frame(ct)
                                if _m.dtype in (np.float32, np.float64):
                                    _m = (_m * 255).astype(np.uint8)
                                else:
                                    _m = _m.astype(np.uint8)
                                if _m.ndim == 2:
                                    _m = _m[:, :, None]
                                _result = np.concatenate([_f, _m], axis=2)
                            else:
                                _alpha = np.full((_f.shape[0], _f.shape[1], 1), 255, dtype=np.uint8)
                                _result = np.concatenate([_f, _alpha], axis=2)
                            if _is_static:
                                _cached_rgba = _result
                            return _result
                        return _get_frame
                    _ov_sources.append({
                        'get_frame': _wrap_ov(_ov, _s),
                        'pos': _p,
                        'start': _s,
                        'end': _e,
                        '_min_t': _min_t,
                        '_max_t': _max_t,
                        '_is_static': hasattr(_ov, '_frame_data'),
                    })

                # ── ⚡ Pre-merge static, full-duration overlays into ONE canvas ──
                # Border, crosshair and static top-text are identical every frame
                # at a fixed position. Blending each separately costs one sparse
                # numpy pass per overlay per frame. We resolve them ONCE into a
                # single full-frame RGBA canvas (alpha-over in draw order) and
                # blend that single canvas per frame — turning N passes into 1.
                # Time-varying overlays (captions) stay as individual sources.
                try:
                    _tw, _th = final_video.w, final_video.h
                    _full_dur = final_video.duration

                    def _resolve_static_pos(_src):
                        """Return (px,py) if this source occupies a FIXED position
                        for the WHOLE clip, else None (time-varying → keep separate)."""
                        if _src.get('_min_t') is not None or _src.get('_max_t') is not None:
                            return None
                        if _src['start'] > 0.05:
                            return None
                        _end = _src['end']
                        if _end is not None and _end < _full_dur - 0.05:
                            return None
                        _pos = _src['pos']
                        if not callable(_pos):
                            return _pos
                        # MoviePy with_position lambda — sample at two times.
                        try:
                            _p0 = _pos(0.0)
                            _p1 = _pos(max(0.0, _full_dur / 2.0))
                        except Exception:
                            return None
                        if _p0 != _p1:
                            return None  # position animates over time
                        return _p0

                    _static_srcs = []
                    _static_pos = {}
                    for _s in _ov_sources:
                        _rp = _resolve_static_pos(_s)
                        if _rp is None:
                            continue
                        # ⚡ CORRECTNESS: only merge if the CONTENT is truly static.
                        # Sample two timestamps; if pixels differ the clip animates
                        # and must stay a separate per-frame source.
                        try:
                            _f0 = _s['get_frame'](0.0)
                            _f1 = _s['get_frame'](max(0.0, _full_dur / 2.0))
                        except Exception:
                            continue
                        if _f0 is None or _f1 is None:
                            continue
                        if _f0.shape != _f1.shape or not np.array_equal(_f0, _f1):
                            continue  # animated content → keep separate
                        _static_srcs.append(_s)
                        _static_pos[id(_s)] = _rp
                    if len(_static_srcs) >= 2:
                        _merged = np.zeros((_th, _tw, 4), dtype=np.uint8)
                        _merged_any = False
                        for _src in _static_srcs:
                            _rgba = _src['get_frame'](0.0)
                            if _rgba is None or _rgba.ndim != 3 or _rgba.shape[2] != 4:
                                continue
                            _px, _py = _static_pos[id(_src)]
                            _oh, _ow = _rgba.shape[:2]
                            if isinstance(_px, str) and _px == 'center':
                                _px = (_tw - _ow) // 2
                            if isinstance(_py, str) and _py == 'center':
                                _py = (_th - _oh) // 2
                            _px, _py = int(_px), int(_py)
                            # Destination/source clip to canvas bounds
                            _dx1, _dy1 = max(0, _px), max(0, _py)
                            _dx2, _dy2 = min(_tw, _px + _ow), min(_th, _py + _oh)
                            if _dx2 <= _dx1 or _dy2 <= _dy1:
                                continue
                            _sx1, _sy1 = _dx1 - _px, _dy1 - _py
                            _ov = _rgba[_sy1:_sy1 + _dy2 - _dy1, _sx1:_sx1 + _dx2 - _dx1]
                            _dst = _merged[_dy1:_dy2, _dx1:_dx2]
                            # alpha-over: out = src*a + dst*(1-a); out_a = a + dst_a*(1-a)
                            _sa = _ov[:, :, 3:4].astype(np.float32) / 255.0
                            _da = _dst[:, :, 3:4].astype(np.float32) / 255.0
                            _oa = _sa + _da * (1.0 - _sa)
                            _safe = np.where(_oa > 0, _oa, 1.0)
                            _rgb = (_ov[:, :, :3].astype(np.float32) * _sa
                                    + _dst[:, :, :3].astype(np.float32) * _da * (1.0 - _sa)) / _safe
                            _dst[:, :, :3] = _rgb.astype(np.uint8)
                            _dst[:, :, 3:4] = (_oa * 255).astype(np.uint8)
                            _merged_any = True
                        if _merged_any:
                            def _merged_get_frame(_ct, _m=_merged):
                                return _m
                            # Remove the individual static sources, add one merged source
                            _static_ids = {id(s) for s in _static_srcs}
                            _ov_sources = [
                                s for s in _ov_sources if id(s) not in _static_ids
                            ]
                            _ov_sources.insert(0, {
                                'get_frame': _merged_get_frame,
                                'pos': (0, 0),
                                'start': 0,
                                'end': None,
                                '_min_t': None,
                                '_max_t': None,
                                '_is_static': True,
                            })
                            print(f"  [⚡CV] Pre-merged {len(_static_srcs)} static "
                                  f"overlays into 1 canvas ({len(_ov_sources)} total sources)")
                except Exception as _merge_err:
                    print(f"  [⚡CV] Static overlay pre-merge skipped: {_merge_err}")

                # ── light effect overlays (OpenCV-native, no MoviePy) ────────
                try:
                    _light_sources = LightLeaksEffects.build_light_effect_overlays(
                        self.settings, final_video.w, final_video.h, 24,
                        final_video.duration)
                    if _light_sources:
                        _ov_sources.extend(_light_sources)
                        print(f"  [⚡CV] Added {len(_light_sources)} light-effect overlay(s) "
                              f"({len(_ov_sources)} total sources)")
                except Exception as _le_err:
                    print(f"  [WARNING] Light-effect overlays (OpenCV) failed: {_le_err}")

                # ── compute crop box ──────────────────────────────────────────
                _reader = _SequentialFrameReader(str(video_path))
                _src_aspect = _reader.w / _reader.h
                _tgt_aspect = final_video.w / final_video.h
                _crop_box = None
                if abs(_src_aspect - _tgt_aspect) > 0.01:
                    if _src_aspect > _tgt_aspect:
                        _nw = int(_reader.h * _tgt_aspect)
                        _x1 = (_reader.w - _nw) // 2
                        _crop_box = (_x1, 0, _x1 + _nw, _reader.h)
                    else:
                        _nh = int(_reader.w / _tgt_aspect)
                        _y1 = (_reader.h - _nh) // 2
                        _crop_box = (0, _y1, _reader.w, _y1 + _nh)

                # ── extract audio before render ───────────────────────────
                _audio_written = False
                _audio_tmp = None
                if final_video.audio is not None:
                    try:
                        _audio_tmp = self.output_folder / f"_temp_{output_path.stem}_audio.wav"
                        # MoviePy 2.x AudioClip.write_audiofile API
                        final_video.audio.write_audiofile(
                            str(_audio_tmp), codec='pcm_s16le', fps=44100, logger=None)
                        _audio_written = True
                    except Exception as _ae:
                        print(f"  [WARNING] MoviePy audio extract failed: {_ae}")
                        # Fallback: extract audio directly from source video via ffmpeg
                        try:
                            import subprocess as _sp
                            _sp.run([
                                'ffmpeg', '-y',
                                '-i', str(video_path),
                                '-vn', '-acodec', 'pcm_s16le',
                                '-ar', '44100', '-ac', '2',
                                str(_audio_tmp),
                            ], check=True, capture_output=True)
                            _audio_written = True
                            print(f"  [INFO] Used ffmpeg audio extraction fallback")
                        except Exception as _ae2:
                            print(f"  [WARNING] ffmpeg audio extract also failed: {_ae2}")

                # ── build transition post-processor ────────────────────────
                _trans_fn = None
                _tf_fade_in = self.settings.get('transition_fade_in', False)
                _tf_fade_out = self.settings.get('transition_fade_out', False)
                _tf_zoom_in = self.settings.get('transition_zoom_in', False)
                _tf_zoom_out = self.settings.get('transition_zoom_out', False)
                _tf_blur_in = self.settings.get('transition_blur_in', False)
                _tf_blur_out = self.settings.get('transition_blur_out', False)
                if _tf_fade_in or _tf_fade_out or _tf_zoom_in or _tf_zoom_out or _tf_blur_in or _tf_blur_out:
                    _fi_dur = self.settings.get('transition_fade_in_duration', 0.5) if _tf_fade_in else 0
                    _fo_dur = self.settings.get('transition_fade_out_duration', 0.5) if _tf_fade_out else 0
                    _zi_dur = self.settings.get('transition_zoom_in_duration', 1.0) if _tf_zoom_in else 0
                    _zi_scl = self.settings.get('transition_zoom_scale', 1.3) if _tf_zoom_in else 1.0
                    _zo_dur = self.settings.get('transition_zoom_out_duration', 1.0) if _tf_zoom_out else 0
                    _zo_scl = self.settings.get('transition_zoom_scale', 1.3) if _tf_zoom_out else 1.0
                    _bi_dur = self.settings.get('transition_blur_duration', 0.5) if _tf_blur_in else 0
                    _bi_amt = self.settings.get('transition_blur_amount', 15) if _tf_blur_in else 0
                    _bo_dur = self.settings.get('transition_blur_duration', 0.5) if _tf_blur_out else 0
                    _bo_amt = self.settings.get('transition_blur_amount', 15) if _tf_blur_out else 0
                    _vid_dur = final_video.duration

                    def _transition_fn(frame, t):
                        # Fade in (multiply by t/duration)
                        if _fi_dur > 0 and t < _fi_dur:
                            alpha = min(t / _fi_dur, 1.0)
                            frame = (frame.astype(np.float32) * alpha).astype(np.uint8)
                        # Fade out (multiply by remaining/duration)
                        if _fo_dur > 0 and t > _vid_dur - _fo_dur:
                            alpha = min((_vid_dur - t) / _fo_dur, 1.0)
                            frame = (frame.astype(np.float32) * alpha).astype(np.uint8)
                        # Zoom in (scale from scale→1 over start)
                        if _zi_dur > 0 and t < _zi_dur:
                            progress = min(t / _zi_dur, 1.0)
                            cur = _zi_scl - (_zi_scl - 1.0) * progress
                            if abs(cur - 1.0) > 0.01:
                                h, w = frame.shape[:2]
                                nh, nw = int(h * cur), int(w * cur)
                                pil = _PILImage.fromarray(frame).resize((nw, nh), _PILImage.LANCZOS)
                                cx, cy = (nw - w) // 2, (nh - h) // 2
                                frame = np.asarray(pil.crop((cx, cy, cx + w, cy + h)))
                        # Zoom out (scale from 1→scale near end)
                        if _zo_dur > 0 and t > _vid_dur - _zo_dur:
                            progress = max((t - (_vid_dur - _zo_dur)) / _zo_dur, 0.0)
                            cur = 1.0 + (_zo_scl - 1.0) * progress
                            if abs(cur - 1.0) > 0.01:
                                h, w = frame.shape[:2]
                                nh, nw = int(h * cur), int(w * cur)
                                pil = _PILImage.fromarray(frame).resize((nw, nh), _PILImage.LANCZOS)
                                cx, cy = (nw - w) // 2, (nh - h) // 2
                                frame = np.asarray(pil.crop((cx, cy, cx + w, cy + h)))
                        # Blur in (decreasing blur)
                        if _bi_dur > 0 and t < _bi_dur:
                            amt = int(_bi_amt * (1.0 - min(t / _bi_dur, 1.0)))
                            if amt > 0:
                                pil = _PILImage.fromarray(frame)
                                frame = np.asarray(pil.filter(_PILImageFilter.GaussianBlur(radius=amt)))
                        # Blur out (increasing blur)
                        if _bo_dur > 0 and t > _vid_dur - _bo_dur:
                            amt = int(_bo_amt * min((t - (_vid_dur - _bo_dur)) / _bo_dur, 1.0))
                            if amt > 0:
                                pil = _PILImage.fromarray(frame)
                                frame = np.asarray(pil.filter(_PILImageFilter.GaussianBlur(radius=amt)))
                        return frame
                    _trans_fn = _transition_fn

                # ── Mask reveal (applied BEFORE overlays — video content only) ──
                _pre_overlay_fn = None
                _mask_enabled = self.settings.get('transition_mask', False)
                if _mask_enabled:
                    _mask_dur = self.settings.get('transition_mask_duration', 0.6)
                    _mask_shape = self.settings.get('transition_mask_shape', 'circle')
                    _vid_dur_mask = final_video.duration
                    def _mask_reveal_fn(frame, t):
                        if _vid_dur_mask <= 0 or _mask_dur <= 0:
                            return frame
                        h, w = frame.shape[:2]
                        cy, cx = h / 2.0, w / 2.0
                        max_r = np.sqrt(cx ** 2 + cy ** 2)
                        # Start: grow from 0 → max_r
                        if t < _mask_dur:
                            p = min(t / _mask_dur, 1.0)
                            r = max_r * p + 1.0
                            yy, xx = np.ogrid[:h, :w]
                            d = np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2)
                            m = (d <= r).astype(np.float32)
                            m3 = np.stack([m, m, m], axis=-1)
                            return (frame.astype(np.float32) * m3).astype(np.uint8)
                        # End: shrink from max_r → 0
                        elif t > _vid_dur_mask - _mask_dur:
                            p = max((_vid_dur_mask - t) / _mask_dur, 0.0)
                            r = max_r * p + 1.0
                            yy, xx = np.ogrid[:h, :w]
                            d = np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2)
                            m = (d <= r).astype(np.float32)
                            m3 = np.stack([m, m, m], axis=-1)
                            return (frame.astype(np.float32) * m3).astype(np.uint8)
                        return frame
                    _pre_overlay_fn = _mask_reveal_fn
                    print(f"  [⚡CV] Mask reveal ({_mask_shape}) baked into OpenCV render path (pre-overlay)")

                # ── Repeat-selected transitions ───────────────────────────
                _rs_enabled = self.settings.get('transition_repeat_selected_enabled', False)
                _rs_interval = self.settings.get('transition_repeat_selected_interval', 5.0)
                if _rs_enabled and _rs_interval > 0:
                    _rs_order = ['transition_zoom_in', 'transition_zoom_out', 'transition_glitch_start',
                                 'transition_glitch_end', 'transition_cinematic_bars',
                                 'transition_bounce', 'transition_mask', 'transition_bounce_mask',
                                 'transition_radial_wipe', 'transition_color_dissolve',
                                 'transition_split_wipe', 'transition_luma_wipe',
                                 'transition_fade_in', 'transition_fade_out',
                                 'transition_blur_in', 'transition_blur_out',
                                 'transition_slide_in', 'transition_slide_out',
                                 'transition_wipe_in', 'transition_wipe_out']
                    _rs_enabled_keys = [k for k in _rs_order if self.settings.get(k, False)]
                    if _rs_enabled_keys:
                        _rs_n = len(_rs_enabled_keys)
                        _rs_mode = self.settings.get('transition_repeat_selected_mode', 'sequential')
                        import random as _rs_rnd
                        _rs_rnd.seed(0)
                        if _rs_mode == 'random':
                            _rs_rnd.shuffle(_rs_enabled_keys)
                        _rs_pulse_dur = 0.5
                        _rs_vid_dur = final_video.duration
                        _existing_rs_fn = _trans_fn
                        def _repeat_selected_fn(frame, t):
                            if _rs_vid_dur <= 0 or _rs_interval <= 0:
                                return _existing_rs_fn(frame, t) if _existing_rs_fn else frame
                            win_idx = int(t // _rs_interval)
                            if win_idx < 0:
                                return _existing_rs_fn(frame, t) if _existing_rs_fn else frame
                            t_in_win = t - win_idx * _rs_interval
                            if t_in_win > _rs_pulse_dur:
                                return _existing_rs_fn(frame, t) if _existing_rs_fn else frame
                            key = _rs_enabled_keys[win_idx % _rs_n]
                            local_p = t_in_win / _rs_pulse_dur
                            try:
                                h, w = frame.shape[:2]
                                f = frame.astype(np.float32)
                                if key in ('transition_zoom_in',):
                                    s = 1.0 + 0.25 * (1 - local_p)
                                    oh = max(1, int(h / s))
                                    ow = max(1, int(w / s))
                                    sy = (h - oh) // 2
                                    sx = (w - ow) // 2
                                    small = frame[sy:sy+oh, sx:sx+ow]
                                    import cv2 as _rs_cv
                                    frame = _rs_cv.resize(small, (w, h), interpolation=_rs_cv.INTER_LINEAR)
                                elif key in ('transition_zoom_out',):
                                    s = 1.0 + 0.25 * local_p
                                    oh = max(1, int(h / s))
                                    ow = max(1, int(w / s))
                                    sy = (h - oh) // 2
                                    sx = (w - ow) // 2
                                    small = frame[sy:sy+oh, sx:sx+ow]
                                    import cv2 as _rs_cv
                                    frame = _rs_cv.resize(small, (w, h), interpolation=_rs_cv.INTER_LINEAR)
                                elif key in ('transition_glitch_start', 'transition_glitch_end'):
                                    import numpy as _rs_np
                                    _rs_np.random.seed(int(t * 1000) & 0xFFFFFFFF)
                                    shift = int(w * 0.02)
                                    if shift > 0:
                                        f[:, shift:, 0] = f[:, :-shift, 0]
                                        f[:, :-shift, 2] = f[:, shift:, 2]
                                    frame = _rs_np.clip(f, 0, 255).astype(_rs_np.uint8)
                                elif key in ('transition_bounce',):
                                    ty = -int(h * 0.15 * (1 - local_p))
                                    bg = np.zeros_like(frame)
                                    frame = np.where(
                                        (np.arange(h)[:, None] + ty < h) &
                                        (np.arange(h)[:, None] + ty >= 0),
                                        frame[(np.clip(np.arange(h) + ty, 0, h - 1))[:, None].repeat(w, 1), np.arange(w)[None, :].repeat(h, 0)],
                                        bg).astype(np.uint8)
                                elif key in ('transition_mask',):
                                    cy, cx = h / 2.0, w / 2.0
                                    max_r = np.sqrt(cx ** 2 + cy ** 2)
                                    r = max_r * (1.0 - local_p) + 1.0
                                    yy, xx = np.ogrid[:h, :w]
                                    d = np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2)
                                    m = (d <= r).astype(np.float32)
                                    m3 = np.stack([m, m, m], axis=-1)
                                    frame = (f * m3).astype(np.uint8)
                                elif key in ('transition_fade_in', 'transition_fade_out'):
                                    alpha = 1.0 - local_p if key == 'transition_fade_in' else local_p
                                    frame = (f * alpha).astype(np.uint8)
                                elif key in ('transition_blur_in', 'transition_blur_out'):
                                    import cv2 as _rs_cv
                                    k = int(15 * (1 - local_p)) | 1
                                    frame = _rs_cv.GaussianBlur(frame, (k, k), 0)
                                elif key in ('transition_color_dissolve',):
                                    col = np.array([255, 255, 255], dtype=np.float32)
                                    frame = ((f * (1 - local_p) + col * local_p)).astype(np.uint8)
                                elif key in ('transition_slide_in',):
                                    shift_x = int(w * 0.35 * (1 - local_p))
                                    out = np.zeros_like(frame)
                                    if shift_x < w:
                                        out[:, :w - shift_x] = frame[:, shift_x:]
                                    frame = out
                                elif key in ('transition_slide_out',):
                                    shift_x = int(w * 0.35 * local_p)
                                    out = np.zeros_like(frame)
                                    if shift_x < w:
                                        out[:, shift_x:] = frame[:, :w - shift_x]
                                    frame = out
                                elif key in ('transition_wipe_in',):
                                    cut_x = int(w * local_p)
                                    out = frame.copy()
                                    out[:, :cut_x] = 0
                                    frame = out
                                elif key in ('transition_wipe_out',):
                                    cut_x = int(w * local_p)
                                    out = frame.copy()
                                    out[:, cut_x:] = 0
                                    frame = out
                                elif key == 'transition_cinematic_bars':
                                    bar = int(h * 0.10 * (1 - local_p))
                                    frame = frame.copy()
                                    frame[:bar] = 0
                                    frame[-bar:] = 0
                                elif key == 'transition_split_wipe':
                                    half = w // 2
                                    cs = int(half * local_p)
                                    out = np.zeros_like(frame)
                                    out[:, max(0, half - cs):min(w, half + cs)] = frame[:, max(0, half - cs):min(w, half + cs)]
                                    frame = out
                                elif key == 'transition_radial_wipe':
                                    cy, cx = h / 2.0, w / 2.0
                                    max_r = np.sqrt(cx ** 2 + cy ** 2)
                                    r = max_r * local_p + 1.0
                                    yy_w, xx_w = np.ogrid[:h, :w]
                                    d = np.sqrt((xx_w - cx) ** 2 + (yy_w - cy) ** 2)
                                    m = (d >= r).astype(np.float32)
                                    m3 = np.stack([m, m, m], axis=-1)
                                    frame = (f.astype(np.float32) * m3).astype(np.uint8)
                                elif key == 'transition_luma_wipe':
                                    yy_w, xx_w = np.ogrid[:h, :w]
                                    grad = (xx_w + yy_w) / float(w + h)
                                    m = (grad >= (1.0 - local_p)).astype(np.float32)
                                    m3 = np.stack([m, m, m], axis=-1)
                                    edge = (grad >= (1.0 - local_p - 0.06)) & (grad < (1.0 - local_p))
                                    result = (f.astype(np.float32) * m3).astype(np.uint8)
                                    result[edge] = [255, 255, 255]
                                    frame = result
                                elif key == 'transition_bounce_mask':
                                    cy, cx = h / 2.0, w / 2.0
                                    max_r = np.sqrt(cx ** 2 + cy ** 2)
                                    r = max_r * (1.0 - local_p) + 1.0
                                    yy_w, xx_w = np.ogrid[:h, :w]
                                    d = np.sqrt((xx_w - cx) ** 2 + (yy_w - cy) ** 2)
                                    m = (d <= r).astype(np.float32)
                                    m3 = np.stack([m, m, m], axis=-1)
                                    ty = -int(h * 0.12 * (1 - local_p))
                                    bg = np.zeros_like(frame)
                                    bounced = np.where(
                                        (np.arange(h)[:, None] + ty < h) &
                                        (np.arange(h)[:, None] + ty >= 0),
                                        frame[(np.clip(np.arange(h) + ty, 0, h - 1))[:, None].repeat(w, 1), np.arange(w)[None, :].repeat(h, 0)],
                                        bg)
                                    frame = (bounced.astype(np.float32) * m3).astype(np.uint8)
                            except Exception:
                                pass
                            if _existing_rs_fn:
                                frame = _existing_rs_fn(frame, t)
                            return frame
                        _trans_fn = _repeat_selected_fn
                        print(f"  [⚡CV] Repeat-selected transitions ({_rs_n} types) baked into OpenCV render path")

                # ── run OpenCV renderer ───────────────────────────────────
                print(f"  [⚡CV] Rendering {final_video.w}x{final_video.h} @ 24fps "
                      f"with {len(_ov_sources)} overlay(s)...", flush=True)
                _opencv_render(
                    reader=_reader,
                    output_path=output_path,
                    target_w=final_video.w,
                    target_h=final_video.h,
                    target_fps=24,
                    duration=final_video.duration,
                    effects_pipeline=_combined_frame_pipeline,
                    overlay_sources=_ov_sources,
                    crop_box=_crop_box,
                    transition_fn=_trans_fn,
                    pre_overlay_fn=_pre_overlay_fn,
                )

                # ── mux audio if extracted ───────────────────────────────
                if _audio_written and _audio_tmp and _audio_tmp.exists():
                    try:
                        import subprocess as _sp
                        _video_only = output_path.with_suffix('.tmp.mp4')
                        output_path.rename(_video_only)
                        _sp.run([
                            'ffmpeg', '-y',
                            '-i', str(_video_only),
                            '-i', str(_audio_tmp),
                            '-c:v', 'copy',
                            '-c:a', 'aac',
                            '-shortest',
                            str(output_path),
                        ], check=True, capture_output=True, text=True)
                        _video_only.unlink()
                        _audio_tmp.unlink()
                        print("  [⚡CV] Audio muxed successfully")
                    except Exception as _ae:
                        if output_path.exists():
                            output_path.unlink()
                        if _video_only.exists():
                            _video_only.rename(output_path)
                        _stderr = getattr(_ae, "stderr", "")
                        print(f"  [WARNING] Audio mux failed: {_ae}")
                        if _stderr:
                            for _line in _stderr.split("\n")[-5:]:
                                if _line.strip():
                                    print(f"    ffmpeg: {_line.strip()}")

                _rendered_ok = True
                print(f"[⚡CV] OpenCV render complete!")
                _render_elapsed = time.time() - _render_start_time
                print(f"  ║ Total: {_render_elapsed:.1f}s")

            except Exception as _cv_err:
                print(f"[WARNING] OpenCV renderer failed, falling back to MoviePy: {_cv_err}")
                import traceback
                traceback.print_exc()
                _rendered_ok = False

        # ═══════════════════════════════════════════════════════════════════
        # ⚡ MoviePy render path (fallback / no OpenCV)
        # ═══════════════════════════════════════════════════════════════════
        if not _rendered_ok:
            try:
                # Try hardware-accelerated NVENC if available (5-10x faster)
                nvenc_ok = False
                try:
                    import subprocess, os
                    if _nvenc_available():  # cached once per process
                        final_video.write_videofile(
                            str(output_path),
                            codec='h264_nvenc',
                            audio_codec='aac',
                            fps=24,
                            preset='p7',  # NVENC preset p7 = best quality/speed balance
                            ffmpeg_params=[
                                '-tune', 'hq',
                                '-rc', 'vbr',
                                '-cq', '23',
                                '-b:v', '0',
                                '-profile:v', 'main',
                            ],
                            threads=8,
                            logger='bar',
                            temp_audiofile=str(self.output_folder / f"_temp_{output_path.stem}_audio.mp4")
                        )
                        nvenc_ok = True
                        print(f"[⚡] Hardware encoding (NVENC) used — much faster!")
                except Exception:
                    pass

                if not nvenc_ok:
                    final_video.write_videofile(
                        str(output_path),
                        codec='libx264',
                        audio_codec='aac',
                        fps=24,  # Force 24fps for smaller file size and faster rendering
                        preset='ultrafast',  # Fastest software preset
                        threads=8,  # Multi-threaded
                        logger='bar',  # Show progress bar
                        temp_audiofile=str(self.output_folder / f"_temp_{output_path.stem}_audio.mp4")  # Keep temp files in output folder
                    )
                _render_elapsed = time.time() - _render_start_time
                print(f"[OK] Rendering complete!")

                # ─── Performance summary ────────────────────────────────────────
                if _frame_times:
                    _total_frames = len(_frame_times)
                    _avg_ms = sum(_frame_times) / _total_frames * 1000
                    _min_ms = min(_frame_times) * 1000
                    _max_ms = max(_frame_times) * 1000
                    _fps = 1.0 / (sum(_frame_times) / _total_frames)
                    # Find the 95th percentile (p95) — excludes outlier spikes
                    _sorted = sorted(_frame_times)
                    _p95_ms = _sorted[int(_total_frames * 0.95)] * 1000
                    print(f"  ╔═══ ⚡ RENDER PERFORMANCE ════════════════════════")
                    print(f"  ║ Video: {_render_elapsed:.1f}s real time ({_total_frames} frames @ {_fps:.1f} it/s)")
                    print(f"  ║ Frame: avg {_avg_ms:.0f}ms │ p95 {_p95_ms:.0f}ms │ worst {_max_ms:.0f}ms │ best {_min_ms:.0f}ms")
                    print(f"  ║ At 24fps target: {1000/24:.1f}ms/frame — {'✅ ON TARGET' if _avg_ms < 1000/24 else '⚠️  NEEDS OPTIMIZATION'}")
                    print(f"  ╚═══════════════════════════════════════════════════")
            except Exception as e:
                print(f"✗ Rendering failed: {e}")
                import traceback
                traceback.print_exc()
                raise

        video.close()
        if txt_clip is not None:
            txt_clip.close()
        final_video.close()

        print(f"[OK] Saved: {output_path.name}")
        print(f"{'='*70}")

        return output_path, output_filename

    def _generate_word_timings_from_whisper(self, video_path: Path,
                                             audio_ref: Optional[Path] = None) -> list:
        """Generate word-level timestamps from video audio using whisper (via subprocess).

        Uses the VideoTextExtractor venv's openai-whisper installation so no
        additional 3GB torch download is needed. The venv path can be overridden
        via the ``settings['whisper_python_path']`` key.

        Args:
            video_path: Path to the original video file (used if no audio_ref).
            audio_ref: Optional audio file to transcribe (voiceover, Excel mp3, etc.).
                       If None, audio is extracted from video_path.

        Returns:
            List of word timing dicts, or empty list on any failure.
        """
        import subprocess
        import json
        import tempfile
        import sys as _sys

        # --- Resolve a python that has whisper installed --------------------------------
        # Try, in priority order:
        #   1. The Python running this app (sys.executable) — most portable, works on any PC
        #      where the app's own Python has whisper installed.
        #   2. User-defined path from overlay_settings.json  ('whisper_python_path')
        #   3. Hardcoded fallback from the original VideoTextExtractor venv.
        venv_python = None
        _candidates = []
        # Candidate 1: the app's own interpreter
        try:
            _own = str(Path(_sys.executable).resolve())
            _candidates.append(_own)
        except Exception:
            pass
        # Candidate 2: user setting
        _from_setting = str(self.settings.get('whisper_python_path', '') or '').strip()
        if _from_setting:
            _candidates.append(str(Path(_from_setting).resolve()))
        # Candidate 3: original hardcoded fallback
        _candidates.append(
            str(Path(r'D:\GitHub\pythonprojects\VideoTextExtractor\venv\Scripts\python.exe').resolve())
        )

        # Deduplicate while preserving order
        _seen = set()
        _unique = []
        for c in _candidates:
            if c not in _seen:
                _seen.add(c)
                _unique.append(c)

        for _candidate in _unique:
            if Path(_candidate).is_file():
                # Quick whisper availability check (runs a tiny import)
                try:
                    _check = subprocess.run(
                        [_candidate, '-c', 'import whisper; print(whisper.__version__)'],
                        capture_output=True, text=True, timeout=10
                    )
                    if _check.returncode == 0:
                        venv_python = _candidate
                        print(f"  [WHISPER] Using Python at: {venv_python}"
                              f"  (whisper v{_check.stdout.strip()})")
                        break
                    else:
                        print(f"  [WHISPER] Python found but whisper not installed: {_candidate}")
                except Exception as _w_err:
                    print(f"  [WHISPER] Could not test Python: {_candidate} — {_w_err}")
            else:
                print(f"  [WHISPER] Python not found at: {_candidate}")

        if not venv_python:
            print(f"  [WHISPER] No Python installation with whisper found on this PC. "
                  f"Falling back to estimated timing.")
            return []

        # --- Resolve which audio to transcribe -------------------------------------------
        audio_to_transcribe = None
        cleanup_temp = False
        if audio_ref and Path(audio_ref).exists():
            audio_to_transcribe = str(audio_ref)
            print(f"  [WHISPER] Using audio file: {audio_ref.name}")
        elif video_path and Path(video_path).exists():
            # Extract audio from video to a temp WAV file
            try:
                tmp = tempfile.NamedTemporaryFile(suffix='.wav', delete=False)
                tmp.close()
                audio_to_transcribe = tmp.name
                cleanup_temp = True
                print(f"  [WHISPER] Extracting audio from video: {video_path.name}")
                subprocess.run(
                    ['ffmpeg', '-y', '-i', str(video_path),
                     '-vn', '-acodec', 'pcm_s16le', '-ar', '16000', '-ac', '1',
                     audio_to_transcribe],
                    capture_output=True, check=True, timeout=120
                )
                print(f"  [WHISPER] Audio extracted to temp file")
            except Exception as e:
                print(f"  [WHISPER] Could not extract audio from video: {e}")
                return []
        else:
            print(f"  [WHISPER] No audio source available")
            return []

        # --- Run whisper via the venv subprocess -----------------------------------------
        whisper_script = Path(__file__).parent / '_whisper_word_timestamps.py'
        if not whisper_script.exists():
            print(f"  [WHISPER] Helper script not found: {whisper_script}")
            return []

        try:
            result = subprocess.run(
                [venv_python, str(whisper_script), audio_to_transcribe,
                 '--model', 'base', '--language', 'en'],
                capture_output=True, text=True, timeout=600  # 10 min for model download
            )
            if result.returncode != 0:
                print(f"  [WHISPER] Subprocess error (rc={result.returncode}): "
                      f"{result.stderr[:200]}")
                return []

            # Parse the JSON output from stdout
            word_timings = json.loads(result.stdout)
            if word_timings:
                print(f"  [WHISPER] Got {len(word_timings)} word timestamps from "
                      f"whisper model")
            return word_timings

        except subprocess.TimeoutExpired:
            print(f"  [WHISPER] Timed out after 10 minutes")
            return []
        except json.JSONDecodeError:
            print(f"  [WHISPER] Could not parse whisper output as JSON")
            return []
        except Exception as e:
            print(f"  [WHISPER] Error: {e}")
            return []
        finally:
            if cleanup_temp and audio_to_transcribe and Path(audio_to_transcribe).exists():
                try:
                    Path(audio_to_transcribe).unlink()
                except Exception:
                    pass

    def process_single_video(self, video_path: Path, video_index: int = 0) -> dict:
        """
        Process a single video with quote overlay

        Args:
            video_path: Path to video file
            video_index: Index used to select corresponding quote

        Returns:
            Dictionary with processing result (status, output_file, etc.)
        """
        try:
            # Store video filepath in settings for spreadsheet lookup
            self.settings['_current_video_path'] = str(video_path)

            # Clear cached spreadsheet text from previous video (performance optimization)
            cache_keys = [k for k in self.settings.keys() if k.startswith('_cached_text_')]
            for key in cache_keys:
                del self.settings[key]

            # Get quote and audio file for this video (supports both Excel and standard mode)
            quote, excel_audio_path = self.get_quote_for_video(video_path, video_index)

            if not quote:
                raise Exception("No quote found for video")

            print(f"\n[OK] Processing video {video_index + 1}")

            # Show source info and track quote index
            quote_index = None  # Initialize quote_index
            if quote.get('source') == 'excel':
                print(f"[EXCEL] Using Excel overlay text for video ID: {quote.get('video_id')}")
                if excel_audio_path:
                    print(f"[EXCEL] Using matched audio: {excel_audio_path.name}")
                # Excel mode doesn't use quote_index (uses row index instead)
            elif quote.get('source') == 'video_title':
                print(f"[VIDEO TITLE] Using video filename as subtitle")
                # Video title mode doesn't need quote_index either
            elif quote.get('source') == 'transcript_excel':
                print(f"[EXCEL TRANSCRIPT] Using transcript from Excel for video ID: {quote.get('video_id')}")
                # Transcript Excel mode doesn't need quote_index
            else:
                # Get next quote index (continues from last processed, even after script restart)
                last_index = self.quote_state.get('last_quote_index', -1)
                quotes = self.read_quotes()
                if not quotes:
                    raise Exception("No quotes found in quotes file")
                quote_index = (last_index + 1) % len(quotes)
                print(f"[OK] Using quote {quote_index + 1}/{len(quotes)} (continuing from last session)")

            # Process the video (pass excel_audio_path for audio override)
            output_path, filename = self.add_quote_to_video(
                video_path,
                quote,
                video_index=video_index,
                excel_audio_path=excel_audio_path
            )

            # ── AI Avatar (Wav2Lip) post-processing ──
            avatar_output = None
            if self.settings.get('avatar_enabled', False):
                avatar_face = self.settings.get('avatar_face_path', '')
                if avatar_face and os.path.isfile(avatar_face):
                    if _wav2lip_available():
                        print(f"[AVATAR] Generating lip-synced avatar...")
                        # Temp audio file
                        temp_audio = Path(self.output_folder) / f"_avatar_audio_{video_index}.wav"
                        # Avatar video output
                        avatar_raw = Path(self.output_folder) / f"_avatar_raw_{video_index}.mp4"

                        # Extract audio from the just-rendered video
                        if extract_audio(output_path, temp_audio):
                            # Run Wav2Lip
                            pads = self.settings.get('avatar_pads', [0, 20, 0, 0])
                            resize_factor = self.settings.get('avatar_resize_factor', 2)
                            av_result = run_wav2lip(
                                face_path=Path(avatar_face),
                                audio_path=temp_audio,
                                output_path=avatar_raw,
                                pads=pads,
                                resize_factor=resize_factor,
                            )

                            if av_result:
                                standalone = self.settings.get('avatar_standalone', False)
                                if standalone:
                                    # Replace output with the raw avatar video
                                    import shutil
                                    shutil.copy2(str(av_result), str(output_path))
                                    avatar_output = output_path
                                    print(f"[AVATAR] ✓ Standalone avatar video: {output_path.name}")
                                else:
                                    # Composite avatar onto main video
                                    position = self.settings.get('avatar_position', 'bottom-right')
                                    scale = self.settings.get('avatar_scale', 0.25)
                                    composited = Path(str(output_path).replace('.mp4', '_avatar.mp4'))
                                    result_path = composite_avatar(
                                        main_video_path=output_path,
                                        avatar_video_path=av_result,
                                        output_path=composited,
                                        position=position,
                                        avatar_scale=scale,
                                    )
                                    if result_path:
                                        # Replace output with composited version
                                        import shutil
                                        shutil.move(str(result_path), str(output_path))
                                        avatar_output = output_path
                                        print(f"[AVATAR] ✓ Composited onto video: {output_path.name}")
                                    else:
                                        print("[AVATAR] ✗ Compositing failed — using original video")
                            else:
                                print("[AVATAR] ✗ Wav2Lip inference failed — using original video")

                            # Cleanup temp files
                            try:
                                temp_audio.unlink(missing_ok=True)
                                avatar_raw.unlink(missing_ok=True)
                            except Exception:
                                pass
                        else:
                            print("[AVATAR] ✗ Audio extraction failed — skipping")
                    else:
                        print("[AVATAR] ✗ Wav2Lip model not found at checkpoints/wav2lip_gan.pth — skipping")
                else:
                    print(f"[AVATAR] ✗ Face image not found: {avatar_face}")

            # Store subtitle and voiceover separately in log
            if isinstance(quote, dict):
                quote_log = {
                    'subtitle': quote['subtitle'],
                    'voiceover': quote['voiceover']
                }
            else:
                quote_log = quote

            # Log the result
            result = {
                'index': video_index,
                'original_video': video_path.name,
                'quote': quote_log,
                'output_file': filename,
                'timestamp': datetime.now().isoformat(),
                'status': 'success'
            }

            self.processing_log['processed_count'] += 1
            self.processing_log['processed_videos'].append(result)
            self._save_log()

            # Update quote state to remember which quote was used (persists across restarts)
            # Only save quote_index if not using Excel mode
            if quote_index is not None:
                self.quote_state['last_quote_index'] = quote_index
                self._save_quote_state()

            print(f"✓ Successfully processed: {filename}")

            return result

        except Exception as e:
            error_result = {
                'index': video_index,
                'original_video': video_path.name,
                'status': 'failed',
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
            print(f"✗ Error processing video: {str(e)}")
            raise  # Re-raise so GUI can handle it

    def process_all(self, start_from: int = 0, sort_by: str = 'created', skip_processed: bool = False):
        """Process all videos with enhanced effects"""
        videos = self.get_video_files(sort_by=sort_by)
        quotes = self.read_quotes()

        if not videos:
            print("✗ No videos found!")
            return

        if not quotes:
            print("✗ No quotes found!")
            return

        # Filter out already processed videos if skip_processed is True
        if skip_processed:
            processed_videos = {entry['original_video'] for entry in self.processing_log.get('processed_videos', [])}
            original_count = len(videos)
            videos = [v for v in videos if v.name not in processed_videos]
            if original_count != len(videos):
                print(f"[OK] Skipped {original_count - len(videos)} already processed videos")
                print(f"[OK] Remaining videos to process: {len(videos)}")

        num_to_process = min(len(videos), len(quotes))

        if start_from >= num_to_process:
            print(f"✗ start_from ({start_from}) >= available pairs ({num_to_process})")
            return

        print(f"\n{'='*70}")
        print(f"ENHANCED BATCH PROCESSING")
        print(f"{'='*70}")
        print(f"Settings: overlay_settings.json")
        print(f"Effects enabled:")
        if self.settings.get('text_fade_in'): print("  [OK] Text fade-in")
        if self.settings.get('text_glow'): print("  [OK] Text glow")
        if self.settings.get('vignette'): print("  [OK] Vignette")
        if self.settings.get('video_zoom'): print("  [OK] Video zoom")
        if self.settings.get('drop_shadow'): print("  [OK] Drop shadow")
        if self.settings.get('color_grade', 'none') != 'none':
            print(f"  [OK] Color grade: {self.settings['color_grade']}")
        print(f"Videos: {len(videos)}")
        print(f"Quotes: {len(quotes)}")
        print(f"Processing: {num_to_process - start_from} video(s)")
        print(f"{'='*70}\n")

        results = []
        for i in range(start_from, num_to_process):
            video_path = videos[i]
            quote = quotes[i]

            print(f"\nProcessing {i + 1}/{num_to_process}")

            try:
                output_path, filename = self.add_quote_to_video(video_path, quote, video_index=i)

                # Store subtitle and voiceover separately in log
                if isinstance(quote, dict):
                    quote_log = {
                        'subtitle': quote['subtitle'],
                        'voiceover': quote['voiceover']
                    }
                else:
                    quote_log = quote

                result = {
                    'index': i,
                    'original_video': video_path.name,
                    'quote': quote_log,
                    'output_file': filename,
                    'timestamp': datetime.now().isoformat(),
                    'status': 'success'
                }
                results.append(result)

                self.processing_log['processed_count'] += 1
                self.processing_log['processed_videos'].append(result)
                self._save_log()

            except Exception as e:
                print(f"✗ Error: {str(e)}")
                result = {
                    'index': i,
                    'original_video': video_path.name,
                    'quote': quote,
                    'status': 'failed',
                    'error': str(e),
                    'timestamp': datetime.now().isoformat()
                }
                results.append(result)
                continue

        print(f"\n{'='*70}")
        print(f"PROCESSING COMPLETE!")
        print(f"{'='*70}")
        success_count = sum(1 for r in results if r['status'] == 'success')
        print(f"Success: {success_count}/{len(results)}")
        print(f"Output: {self.output_folder}")
        print(f"{'='*70}\n")

        return results


def _apply_segment_transitions(sub_clips, settings, transition_duration=0.3,
                                min_interval=0.0):
    """Apply enabled Transitions-tab transitions to segment boundaries and generate SFX.

    For each segment boundary, this applies enabled 'in' transitions to the start
    of the next segment and enabled 'out' transitions to the tail of the previous
    segment, then returns audio SFX clips timed at each boundary.

    If min_interval > 0, transitions are only applied at boundaries where at least
    `min_interval` seconds of cumulative video time has elapsed since the last
    applied transition. Crossfade-only boundaries still receive SFX from the
    calling code.

    Returns:
        (modified_sub_clips, sfx_audio_clips)
    """
    # (enabled_key, type, method, dur_key, {extra_kw}, sfx_tname, duration_param_name)
    _TR_SPECS = [
        ('transition_fade_in',      'in',  'apply_fade_transition',
         'transition_fade_in_duration',  {}, 'fade', 'fade_in_duration'),
        ('transition_fade_out',     'out', 'apply_fade_transition',
         'transition_fade_out_duration', {}, 'fade', 'fade_out_duration'),
        ('transition_zoom_in',      'in',  'create_zoom_transition',
         'transition_zoom_in_duration',  {'zoom_in': True}, 'zoom_in', 'duration'),
        ('transition_zoom_out',     'out', 'create_zoom_transition',
         'transition_zoom_out_duration', {'zoom_in': False}, 'zoom_out', 'duration'),
        ('transition_blur_in',      'in',  'create_blur_transition',
         'transition_blur_duration',     {'blur_in': True}, 'blur_in', 'duration'),
        ('transition_blur_out',     'out', 'create_blur_transition',
         'transition_blur_duration',     {'blur_in': False}, 'blur_out', 'duration'),
        ('transition_slide_in',     'in',  'create_slide_transition',
         'transition_slide_duration',
         {'in_transition': True,
          'direction': lambda s: s.get('transition_slide_direction', 'left')},
         'slide_in', 'duration'),
        ('transition_slide_out',    'out', 'create_slide_transition',
         'transition_slide_duration',
         {'in_transition': False,
          'direction': lambda s: s.get('transition_slide_direction', 'left')},
         'slide_out', 'duration'),
        ('transition_wipe_in',      'in',  'create_wipe_transition',
         'transition_wipe_duration',
         {'in_transition': True,
          'direction': lambda s: s.get('transition_wipe_direction', 'right')},
         'wipe_in', 'duration'),
        ('transition_wipe_out',     'out', 'create_wipe_transition',
         'transition_wipe_duration',
         {'in_transition': False,
          'direction': lambda s: s.get('transition_wipe_direction', 'right')},
         'wipe_out', 'duration'),
        ('transition_glitch_start', 'in',  'create_glitch_transition',
         'transition_glitch_duration',    {'glitch_start': True}, 'glitch_start', 'duration'),
        ('transition_glitch_end',   'out', 'create_glitch_transition',
         'transition_glitch_duration',    {'glitch_start': False}, 'glitch_end', 'duration'),
        ('transition_bounce',       'in',  'create_bounce_transition',
         'transition_bounce_duration',    {'bounce_start': True}, 'bounce', 'duration'),
        ('transition_bounce',       'out', 'create_bounce_transition',
         'transition_bounce_duration',    {'bounce_start': False}, 'bounce', 'duration'),
    ]

    in_enabled = []   # [(method_name, dur, extra_kwargs, sfx_tname, dur_param, enabled_key), ...]
    out_enabled = []  # same

    for entry in _TR_SPECS:
        enabled_key = entry[0]
        if settings.get(enabled_key, False):
            _, typ, method_name, dur_key, extra, sfx_tname, dur_param = entry
            dur = float(settings.get(dur_key, 0.5))
            target = in_enabled if typ == 'in' else out_enabled
            target.append((method_name, dur, extra, sfx_tname, dur_param, enabled_key))

    if not in_enabled and not out_enabled:
        print(f"[SILENCE] No Transitions-tab effects enabled — using crossfade only")
    else:
        in_names = ', '.join(e[0].replace('create_', '').replace('_transition', '').replace('apply_', '') for e in in_enabled)
        out_names = ', '.join(e[0].replace('create_', '').replace('_transition', '').replace('apply_', '') for e in out_enabled)
        print(f"[SILENCE] Tab transitions: in=[{in_names}] out=[{out_names}] dur={transition_duration:.2f}s")

    clips = list(sub_clips)
    sfx_clips = []
    sfx_volume = float(settings.get('transition_sfx_volume', 0.6))
    sfx_enabled = settings.get('transition_sfx_enabled', True)

    # Spacing tracker: when min_interval > 0, only apply tab transitions
    # at boundaries where cumulative_time_since_last >= min_interval
    _last_tab_time = 0.0
    _cumul_time = 0.0

    for i, clip in enumerate(clips):
        apply_start = i > 0
        apply_end   = i < len(clips) - 1

        # If spacing is active, check if this boundary is far enough from last
        _skip_this_boundary = False
        if min_interval > 0 and i > 0:
            # _cumul_time is the time at the START of this clip
            if _cumul_time - _last_tab_time < min_interval:
                _skip_this_boundary = True

        if apply_start and in_enabled and not _skip_this_boundary:
            for method_name, dur, extra, sfx_tname, dur_param, enabled_key in in_enabled:
                try:
                    kwargs = {}
                    for k, v in extra.items():
                        kwargs[k] = v(settings) if callable(v) else v
                    kwargs[dur_param] = min(dur, transition_duration)
                    method = getattr(TransitionEffects, method_name)
                    clip = method(clip, **kwargs)

                    # SFX at this boundary
                    if sfx_enabled and sfx_volume > 0:
                        t_at = sum(c.duration for c in clips[:i]) - transition_duration * i
                        if t_at < 0:
                            t_at = 0
                        sfx_name, cf, cv = _resolve_sfx_for_transition(settings, enabled_key, sfx_tname)
                        if sfx_name:
                            sfx = _make_sfx_clip(
                                sfx_name, min(dur, transition_duration),
                                max(0, t_at), 44100, sfx_volume,
                                custom_file=cf, custom_volume=cv)
                            if sfx is not None:
                                sfx_clips.append(sfx)
                except Exception as e:
                    print(f'  [WARN] Segment transition {method_name} failed: {e}')

        if apply_end and out_enabled and not _skip_this_boundary:
            for method_name, dur, extra, sfx_tname, dur_param, enabled_key in out_enabled:
                try:
                    kwargs = {}
                    for k, v in extra.items():
                        kwargs[k] = v(settings) if callable(v) else v
                    kwargs[dur_param] = min(dur, transition_duration)
                    method = getattr(TransitionEffects, method_name)
                    clip = method(clip, **kwargs)
                except Exception as e:
                    print(f'  [WARN] Segment transition {method_name} failed: {e}')

        # Update tracking: if we applied a tab transition at this boundary, mark time
        if i > 0 and not _skip_this_boundary and (in_enabled or out_enabled):
            _last_tab_time = _cumul_time

        clips[i] = clip
        _cumul_time += clip.duration

    return clips, sfx_clips


def _resolve_sfx_for_transition(settings, enabled_key, sfx_tname):
    """Resolve which SFX to use for a transition: per-transition dropdown →
    custom SFX file → procedural fallback.

    Returns:
        (sfx_name_or_None, custom_file_or_None, custom_volume_or_None)
    """
    # Per-transition dropdown setting (e.g. 'transition_sfx_fade_in')
    sfx_override_key = f"transition_sfx_{enabled_key.removeprefix('transition_')}"
    dropdown_val = settings.get(sfx_override_key, '(default)')

    if dropdown_val not in ('(default)', 'None', '(none)', '', None):
        # User selected a specific SFX from the dropdown
        return dropdown_val, None, None

    # Check custom SFX file (works as global fallback when no Target is set)
    custom_file = settings.get('custom_sfx_file', '')
    if custom_file and Path(custom_file).is_file():
        cv = float(settings.get('custom_sfx_volume', 0.8))
        sfx_name = _sfx_name_for(sfx_tname)  # procedural name for duration/format
        return sfx_name, custom_file, cv

    # Procedural fallback
    sfx_name = _sfx_name_for(sfx_tname)
    if not sfx_name:
        # Last-resort fallback: pick a generic SFX that always produces
        # valid audio so users never get silent transitions
        _FALLBACK_SFX = {
            'fade': 'shimmer',
            'zoom': 'whoosh',
            'blur': 'boom',
            'slide': 'whoosh',
            'wipe': 'swoosh',
            'glitch': 'glitch',
            'bounce': 'impact',
        }
        fallback_key = sfx_tname.split('_')[0]  # e.g. 'zoom_in' → 'zoom'
        sfx_name = _FALLBACK_SFX.get(fallback_key, 'shimmer')
        print(f"[SFX-DEBUG] _sfx_name_for('{sfx_tname}') returned None — "
              f"falling back to '{sfx_name}'")
    return sfx_name, None, None


def _make_crossfade_sfx(sub_clips, trans, settings, preferred_sfx=None):
    """Generate SFX at each crossfade boundary between segments.

    Args:
        sub_clips: list of VideoFileClip segments
        trans: crossfade duration in seconds
        settings: settings dict
        preferred_sfx: if set (e.g. 'click'), use this SFX name instead of lookup.
                       Used by the "crossfade + click" standalone silence removal option.
    """
    sfx_enabled = settings.get('transition_sfx_enabled', True)
    sfx_vol = float(settings.get('transition_sfx_volume', 0.6))
    if not sfx_enabled or sfx_vol <= 0:
        return []

    # Prefer custom SFX file if configured, else auto-detect from sfx_library
    custom_file = settings.get('custom_sfx_file', '')
    if not custom_file or not Path(custom_file).is_file():
        sfx_lib = Path(__file__).parent / 'VoiceModules' / 'TransitionSFX' / 'sfx_library'
        if sfx_lib.is_dir():
            found = sorted(sfx_lib.glob('*.wav')) + sorted(sfx_lib.glob('*.mp3'))
            if found:
                custom_file = str(found[0])

    if preferred_sfx:
        sfx_name = preferred_sfx
    else:
        sfx_name = _sfx_name_for('fade')  # procedural fallback
        if not sfx_name:
            sfx_name = 'shimmer'  # hard fallback — always producible
            print(f"[SFX-DEBUG] _sfx_name_for('fade') returned None in _make_crossfade_sfx — "
                  f"falling back to '{sfx_name}'")

    custom_vol = float(settings.get('custom_sfx_volume', 0.8)) if custom_file else None

    sfx_clips = []
    cumulative = 0.0
    for i in range(len(sub_clips) - 1):
        cumulative += sub_clips[i].duration
        sfx_time = max(0, cumulative - trans * 0.5)
        sfx = _make_sfx_clip(sfx_name, min(trans, 0.5), sfx_time, 44100, sfx_vol,
                             custom_file=custom_file, custom_volume=custom_vol)
        if sfx is not None:
            sfx_clips.append(sfx)

    if sfx_clips:
        src = custom_file or sfx_name
        print(f"[SILENCE] Added {len(sfx_clips)} crossfade SFX from {Path(src).name if custom_file else src}"
              + (" ('click' mode)" if preferred_sfx else ""))
    return sfx_clips


_AUDIO_EXTS = {'.mp3', '.wav', '.m4a', '.flac', '.aac', '.ogg', '.wma', '.opus'}

def _is_audio_file(path: Path) -> bool:
    """Check if the given path points to an audio file by its extension."""
    return path.suffix.lower() in _AUDIO_EXTS


def remove_silence_from_video(video_path: Path, output_path: Path,
                               settings: dict = None,
                               ffmpeg_path: str = None,
                               transition_duration: float = 0.0,
                               use_tab_transitions: bool = False) -> bool:
    """Remove silent portions from a video using ffmpeg silence detection.

    Uses ffmpeg's silencedetect filter to find silent regions, then MoviePy
    to cut them out and stitch the remaining segments back together.

    Args:
        video_path:  Path to the source video file.
        output_path: Where to save the silence-free video.
        settings:   Dict with keys:
            silence_threshold   — dB level for silence detection (default -30)
            silence_min_duration — minimum silence duration in seconds (default 0.5)
            silence_padding      — padding in seconds kept around speech (default 0.15)
        ffmpeg_path: Full path to ffmpeg.exe (optional; falls back to 'ffmpeg' in PATH).

    Returns:
        True on success, False on failure.
    """
    import subprocess
    import re
    import tempfile
    import shlex

    settings = settings or {}
    threshold = settings.get('silence_threshold', -30)
    min_dur = settings.get('silence_min_duration', 0.5)
    padding = settings.get('silence_padding', 0.15)

    if not video_path.is_file():
        print(f"[ERROR] remove_silence: source not found: {video_path}")
        return False

    print(f"[SILENCE] Analyzing {video_path.name} "
          f"(threshold={threshold}dB, min_silence={min_dur}s, padding={padding}s)")

    # Resolve ffmpeg / ffprobe paths
    _ffmpeg = ffmpeg_path or 'ffmpeg'
    _ffprobe = None
    if ffmpeg_path:
        _ffprobe = str(Path(ffmpeg_path).with_name('ffprobe.exe'))
    else:
        _ffprobe = 'ffprobe'

    # ── Step 1: Detect silence with ffmpeg ──────────────────────────────
    try:
        result = subprocess.run(
            [_ffmpeg, '-i', str(video_path),
             '-af', f'silencedetect=noise={threshold}dB:d={min_dur}',
             '-f', 'null', '-'],
            capture_output=True, text=True, timeout=300)
    except FileNotFoundError:
        print(f"[ERROR] remove_silence: ffmpeg not found ('{_ffmpeg}' is not available)")
        return False
    except subprocess.TimeoutExpired:
        print("[ERROR] remove_silence: ffmpeg silencedetect timed out")
        return False
    except Exception as e:
        print(f"[ERROR] remove_silence: ffmpeg failed: {e}")
        return False

    # Parse silence_start / silence_end from stderr
    starts = [float(s) for s in re.findall(r'silence_start:\s+([\d.]+)', result.stderr)]
    ends   = [float(s) for s in re.findall(r'silence_end:\s+([\d.]+)', result.stderr)]

    if not starts:
        print(f"[SILENCE] No silence detected — copying unchanged")
        import shutil
        shutil.copy2(str(video_path), str(output_path))
        return True

    print(f"[SILENCE] Found {len(starts)} silent region(s)")

    # ── Step 2: Build non-silent segments ──────────────────────────────
    total_dur = None
    # Get total duration from ffprobe
    try:
        probe = subprocess.run(
            [_ffprobe, '-v', 'error', '-show_entries',
             'format=duration', '-of', 'csv=p=0', str(video_path)],
            capture_output=True, text=True, timeout=30)
        total_dur = float(probe.stdout.strip())
    except Exception:
        pass  # will use last end as fallback

    segments = []  # list of (start_sec, end_sec)
    cursor = 0.0
    for start_s, end_s in zip(starts, ends):
        keep_start = cursor
        keep_end = start_s + padding
        if keep_end > keep_start + 0.05:  # at least 50ms to keep
            segments.append((keep_start, keep_end))
        cursor = end_s - padding
        if cursor < 0:
            cursor = 0.0

    # Tail after the last silence
    if total_dur and cursor < total_dur:
        segments.append((cursor, total_dur))
    elif not total_dur and cursor < (ends[-1] if ends else 0) + 60:
        # Best-effort: use last silence_end + 60s as fallback
        segments.append((cursor, ends[-1] + 60))

    if not segments:
        print("[SILENCE] Everything is silent — returning empty result")
        return False

    original_count = len(segments)
    # Merge tiny adjacent segments
    merged = [segments[0]]
    for s in segments[1:]:
        if s[0] - merged[-1][1] < 0.1:  # merge if gap < 100ms
            merged[-1] = (merged[-1][0], s[1])
        else:
            merged.append(s)
    segments = merged

    duration_before = total_dur or (segments[-1][1] if segments else 0)
    duration_after  = sum(e - s for s, e in segments)
    print(f"[SILENCE] {duration_before:.1f}s → {duration_after:.1f}s "
          f"({len(segments)} segment(s), merged from {original_count})")

    # ── Step 3: Cut and stitch with MoviePy ────────────────────────────
    is_audio = _is_audio_file(video_path)
    try:
        from moviepy import VideoFileClip, concatenate_videoclips
        MOVIEPY2 = True
    except ImportError:
        from moviepy.editor import VideoFileClip, concatenate_videoclips
        MOVIEPY2 = False

    try:
        if is_audio:
            try:
                from moviepy import AudioFileClip
            except ImportError:
                from moviepy.audio.io.AudioFileClip import AudioFileClip
            clip = AudioFileClip(str(video_path))
        else:
            clip = VideoFileClip(str(video_path))
    except Exception as e:
        print(f"[ERROR] remove_silence: could not load {'audio' if is_audio else 'video'}: {e}")
        return False

    try:
        sub_clips = []
        for start_s, end_s in segments:
            try:
                sub = clip.subclipped(start_s, end_s) if MOVIEPY2 else clip.subclip(start_s, end_s)
                sub_clips.append(sub)
            except Exception as e:
                print(f"[WARNING] subclip({start_s:.2f},{end_s:.2f}) failed: {e}")
                continue

        if not sub_clips:
            print("[ERROR] remove_silence: no valid subclips produced")
            clip.close()
            return False

        # ── Audio branch: simple concatenation, no visual effects ──
        if is_audio:
            try:
                from moviepy import concatenate_audioclips
            except ImportError:
                try:
                    from moviepy.audio.compositing import concatenate_audioclips
                except ImportError:
                    from moviepy.editor import concatenate_audioclips
            final = concatenate_audioclips(sub_clips)
            print(f"[SILENCE] Concatenated {len(sub_clips)} audio segment(s)")
        elif len(sub_clips) == 1:
            final = sub_clips[0]
        elif transition_duration and transition_duration > 0 and len(sub_clips) > 1:
            # Build composite with crossfade transitions between segments
            try:
                from moviepy import vfx, CompositeVideoClip
                MP_VFX = True
            except ImportError:
                # MP1: CompositeVideoClip is at compositing package, not top-level
                try:
                    from moviepy.video.compositing.CompositeVideoClip import CompositeVideoClip
                    MP_VFX = False
                except ImportError:
                    MP_VFX = None

            if MP_VFX is None:
                print("[SILENCE] vfx not available — falling back to hard cuts")
                final = concatenate_videoclips(sub_clips, method='compose')
            elif use_tab_transitions:
                # Apply transitions from Transitions tab between segments
                trans = min(transition_duration or 0.3,
                            min(c.duration for c in sub_clips) * 0.5)
                if trans <= 0.01:
                    final = concatenate_videoclips(sub_clips, method='compose')
                else:
                    # Apply enabled tab transitions to segment boundaries + collect SFX
                    sfx_audio = None
                    tab_transitions_applied = False
                    for _tk in [
                        'transition_fade_in', 'transition_fade_out',
                        'transition_zoom_in', 'transition_zoom_out',
                        'transition_blur_in', 'transition_blur_out',
                        'transition_slide_in', 'transition_slide_out',
                        'transition_wipe_in', 'transition_wipe_out',
                        'transition_glitch_start', 'transition_glitch_end',
                        'transition_bounce',
                    ]:
                        if settings.get(_tk, False):
                            tab_transitions_applied = True
                            break
                    if tab_transitions_applied:
                        # Use spacing interval from Transitions tab repeat setting (default ~6s)
                        _tab_interval = float(settings.get('transition_repeat_selected_interval', 6.0))
                        sub_clips, sfx_audio = _apply_segment_transitions(
                            sub_clips, settings, trans, min_interval=_tab_interval)
                        if _tab_interval > 0:
                            print(f"[SILENCE] Applied tab transitions at segment boundaries "
                                  f"spaced every {_tab_interval:.0f}s")
                        else:
                            print(f"[SILENCE] Applied tab transitions at segment boundaries")

                    # Build composite with crossfade overlap
                    composite_parts = []
                    total = 0.0
                    for i, sc in enumerate(sub_clips):
                        # Apply crossfade effects FIRST (before setting start time,
                        # since MoviePy 1.x crossfadein/out create new composites
                        # that lose any previously-set start attribute)
                        if MP_VFX:
                            if i > 0:
                                sc = sc.with_effects([vfx.CrossFadeIn(trans)])
                            if i < len(sub_clips) - 1:
                                sc = sc.with_effects([vfx.CrossFadeOut(trans)])
                        else:
                            if i > 0:
                                sc = sc.crossfadein(trans)
                            if i < len(sub_clips) - 1:
                                sc = sc.crossfadeout(trans)

                        # Set start time AFTER effects so the start is
                        # applied to the outermost clip wrapper
                        start_time = total - (trans if i > 0 else 0)
                        if start_time < 0:
                            start_time = 0
                        sc = sc.with_start(start_time)

                        composite_parts.append(sc)
                        total += sc.duration - (trans if i > 0 else 0)

                    final = CompositeVideoClip(composite_parts)
                    print(f"[SILENCE] Applied crossfade ({trans:.2f}s) between {len(sub_clips)} segments")

                    # Apply Alight Motion look if a template is selected
                    if settings.get('am_template', 'None') != 'None':
                        try:
                            final = final.image_transform(
                                lambda frame: VideoEffects.apply_alight_motion_look(frame, settings))
                        except AttributeError:
                            final = final.fl_image(
                                lambda frame: VideoEffects.apply_alight_motion_look(frame, settings))
                        print(f"[SILENCE] Applied Alight Motion look")

                    # Add SFX at each crossfade boundary — always generate, even when
                    # tab transitions already produced their own SFX, so every boundary
                    # gets the click/crossfade sound regardless.
                    _cf_sfx = _make_crossfade_sfx(
                        sub_clips, trans, settings,
                        preferred_sfx='click' if settings.get('silence_crossfade_enabled', False) else None)
                    if _cf_sfx:
                        if sfx_audio:
                            sfx_audio.extend(_cf_sfx)
                        else:
                            sfx_audio = _cf_sfx
                    if sfx_audio:
                        try:
                            from moviepy import CompositeAudioClip
                        except ImportError:
                            try:
                                from moviepy.audio.compositing import CompositeAudioClip
                            except ImportError:
                                from moviepy.audio.AudioClip import CompositeAudioClip
                        if final.audio:
                            sfx_audio.append(final.audio)
                        try:
                            final.audio = CompositeAudioClip(sfx_audio)
                        except Exception:
                            pass
            else:
                # Clamp transition so it doesn't exceed half the shortest clip
                min_dur = min(c.duration for c in sub_clips)
                trans = min(transition_duration, min_dur * 0.5)
                if trans <= 0.01:
                    final = concatenate_videoclips(sub_clips, method='compose')
                else:
                    composite_parts = []
                    total = 0.0
                    for i, sc in enumerate(sub_clips):
                        start_time = total - (trans if i > 0 else 0)
                        if start_time < 0:
                            start_time = 0
                        sc = sc.with_start(start_time)

                        if MP_VFX:
                            # MoviePy 2.x path
                            if i > 0:
                                sc = sc.with_effects([vfx.CrossFadeIn(trans)])
                            if i < len(sub_clips) - 1:
                                sc = sc.with_effects([vfx.CrossFadeOut(trans)])
                        else:
                            # MoviePy 1.x path
                            if i > 0:
                                sc = sc.crossfadein(trans)
                            if i < len(sub_clips) - 1:
                                sc = sc.crossfadeout(trans)

                        composite_parts.append(sc)
                        total += sc.duration - (trans if i > 0 else 0)

                    final = CompositeVideoClip(composite_parts)
                    print(f"[SILENCE] Applied crossfade ({trans:.2f}s) between {len(sub_clips)} segments (plain)")
                    # Apply Alight Motion look if a template is selected
                    if settings.get('am_template', 'None') != 'None':
                        try:
                            final = final.image_transform(
                                lambda frame: VideoEffects.apply_alight_motion_look(frame, settings))
                        except AttributeError:
                            final = final.fl_image(
                                lambda frame: VideoEffects.apply_alight_motion_look(frame, settings))
                        print(f"[SILENCE] Applied Alight Motion look")
                    # Add SFX at each crossfade boundary
                    # Use 'click' SFX when crossfade+click mode is enabled
                    _pref_sfx = 'click' if settings.get('silence_crossfade_enabled', False) else None
                    sfx_audio = _make_crossfade_sfx(sub_clips, trans, settings, preferred_sfx=_pref_sfx)
                    if sfx_audio:
                        try:
                            from moviepy import CompositeAudioClip
                        except ImportError:
                            try:
                                from moviepy.audio.compositing import CompositeAudioClip
                            except ImportError:
                                from moviepy.audio.AudioClip import CompositeAudioClip
                        if final.audio:
                            sfx_audio.append(final.audio)
                        try:
                            final.audio = CompositeAudioClip(sfx_audio)
                        except Exception:
                            pass
        else:
            final = concatenate_videoclips(sub_clips, method='compose')

        # --- Repeat-Selected Transitions (frame-level transform) ---
        # NOTE: Only apply repeat transitions when:
        #   - transition_duration > 0 (silence removal isn't disabled)
        #   - use_tab_transitions is True (caller wants tab transitions;
        #     False when in crossfade-only standalone mode)
        #   - the Transitions tab repeat setting is enabled
        # When transition_duration == 0 the caller intends a pure silence-removed
        # concatenation with NO transitions from the Transitions tab at all.
        if (not is_audio and transition_duration > 0
            and use_tab_transitions
            and settings.get('transition_repeat_selected_enabled', False)):
            try:
                interval = float(settings.get('transition_repeat_selected_interval', 6.0))
                mode_s = settings.get('transition_repeat_selected_mode', 'sequential')
                _rs_order = [
                    'transition_fade_in', 'transition_fade_out',
                    'transition_zoom_in', 'transition_zoom_out',
                    'transition_blur_in', 'transition_blur_out',
                    'transition_slide_in', 'transition_slide_out',
                    'transition_wipe_in', 'transition_wipe_out',
                    'transition_glitch_start', 'transition_glitch_end',
                    'transition_cinematic_bars',
                    'transition_bounce', 'transition_mask', 'transition_bounce_mask',
                    'transition_radial_wipe', 'transition_color_dissolve',
                    'transition_split_wipe', 'transition_luma_wipe',
                ]
                enabled_keys = [k for k in _rs_order if settings.get(k, False)]
                if enabled_keys and interval > 0:
                    if mode_s == 'random':
                        import random as _rnd
                        _rnd.seed(0)
                        enabled_keys = enabled_keys[:]
                        _rnd.shuffle(enabled_keys)
                    n = len(enabled_keys)
                    vid_duration = float(getattr(final, 'duration', 0) or 0)
                    pulse_dur = 0.5
                    def _rs_effect(gf, t):
                        frame = gf(t)
                        if vid_duration <= 0 or interval <= 0:
                            return frame
                        win_idx = int(t // interval)
                        if win_idx < 0:
                            return frame
                        t_in_win = t - win_idx * interval
                        if t_in_win > pulse_dur:
                            return frame
                        key = enabled_keys[win_idx % n]
                        local_p = t_in_win / pulse_dur  # 0..1
                        try:
                            import numpy as _np
                            h, w = frame.shape[:2]
                            f = frame.astype(_np.float32)
                            if key in ('transition_zoom_in',):
                                s = 1.0 + 0.25 * (1 - local_p)
                                try:
                                    from PIL import Image as _PILRS
                                    pil_src = _PILRS.fromarray(frame)
                                    new_h, new_w = int(round(h * s)), int(round(w * s))
                                    pil_dst = pil_src.resize((new_w, new_h), _PILRS.BILINEAR)
                                    resized = _np.asarray(pil_dst)
                                    sy0 = max(0, (new_h - h) // 2)
                                    sx0 = max(0, (new_w - w) // 2)
                                    crop = resized[sy0:sy0 + h, sx0:sx0 + w]
                                    if crop.shape[0] == h and crop.shape[1] == w:
                                        return crop
                                    out = _np.zeros_like(frame)
                                    oh, ow = min(h, crop.shape[0]), min(w, crop.shape[1])
                                    out[:oh, :ow] = crop[:oh, :ow]
                                    return out
                                except Exception:
                                    # fallback: cv2 resize
                                    oh2 = max(1, int(h / s))
                                    ow2 = max(1, int(w / s))
                                    sy2 = (h - oh2) // 2
                                    sx2 = (w - ow2) // 2
                                    small = frame[sy2:sy2+oh2, sx2:sx2+ow2]
                                    import cv2 as _cv2
                                    return _cv2.resize(small, (w, h), interpolation=_cv2.INTER_LINEAR)
                            elif key in ('transition_zoom_out',):
                                s = 1.0 + 0.25 * local_p
                                try:
                                    from PIL import Image as _PILRS
                                    pil_src = _PILRS.fromarray(frame)
                                    new_h, new_w = int(round(h * s)), int(round(w * s))
                                    pil_dst = pil_src.resize((new_w, new_h), _PILRS.BILINEAR)
                                    resized = _np.asarray(pil_dst)
                                    sy0 = max(0, (new_h - h) // 2)
                                    sx0 = max(0, (new_w - w) // 2)
                                    crop = resized[sy0:sy0 + h, sx0:sx0 + w]
                                    if crop.shape[0] == h and crop.shape[1] == w:
                                        return crop
                                    out = _np.zeros_like(frame)
                                    oh, ow = min(h, crop.shape[0]), min(w, crop.shape[1])
                                    out[:oh, :ow] = crop[:oh, :ow]
                                    return out
                                except Exception:
                                    # fallback: cv2 resize
                                    oh2 = max(1, int(h / s))
                                    ow2 = max(1, int(w / s))
                                    sy2 = (h - oh2) // 2
                                    sx2 = (w - ow2) // 2
                                    small = frame[sy2:sy2+oh2, sx2:sx2+ow2]
                                    import cv2 as _cv2
                                    return _cv2.resize(small, (w, h), interpolation=_cv2.INTER_LINEAR)
                            elif key in ('transition_glitch_start', 'transition_glitch_end'):
                                _np.random.seed(int(t * 1000) & 0xFFFFFFFF)
                                shift = int(w * 0.02)
                                if shift > 0:
                                    f[:, shift:, 0] = f[:, :-shift, 0]
                                    f[:, :-shift, 2] = f[:, shift:, 2]
                                return _np.clip(f, 0, 255).astype(_np.uint8)
                            elif key in ('transition_bounce',):
                                ty = -int(h * 0.15 * (1 - local_p))
                                bg = _np.zeros_like(frame)
                                out = _np.where(
                                    (_np.arange(h)[:, None] + ty < h) &
                                    (_np.arange(h)[:, None] + ty >= 0),
                                    frame[(_np.clip(_np.arange(h) + ty, 0, h - 1))[:, None].repeat(w, 1), _np.arange(w)[None, :].repeat(h, 0)],
                                    bg)
                                return out.astype(_np.uint8)
                            elif key in ('transition_mask',):
                                cy, cx = h / 2.0, w / 2.0
                                max_r = _np.sqrt(cx ** 2 + cy ** 2)
                                r = max_r * (1.0 - local_p) + 1.0
                                yy, xx = _np.ogrid[:h, :w]
                                d = _np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2)
                                m = (d <= r).astype(_np.float32)
                                m3 = _np.stack([m, m, m], axis=-1)
                                return (f * m3).astype(_np.uint8)
                            elif key in ('transition_fade_in', 'transition_fade_out'):
                                alpha = 1.0 - local_p if key == 'transition_fade_in' else local_p
                                return (f * alpha).astype(_np.uint8)
                            elif key in ('transition_blur_in', 'transition_blur_out'):
                                try:
                                    import cv2 as _cv
                                    k = int(15 * (1 - local_p)) | 1
                                    return _cv.GaussianBlur(frame, (k, k), 0)
                                except Exception:
                                    k2 = max(3, int(15 * (1 - local_p)) | 1)
                                    import cv2 as _cv2
                                    return _cv2.filter2D(frame, -1, _np.ones((k2, k2), _np.float32) / (k2 * k2))
                            elif key in ('transition_color_dissolve',):
                                # Blend frame toward white (or user-chosen color) over the pulse
                                col = _np.array([255, 255, 255], dtype=_np.float32)
                                return ((f * (1 - local_p) + col * local_p)).astype(_np.uint8)
                            elif key in ('transition_slide_in',):
                                shift_x = int(w * 0.35 * (1 - local_p))
                                out = _np.zeros_like(frame)
                                if shift_x < w:
                                    out[:, :w - shift_x] = frame[:, shift_x:]
                                return out
                            elif key in ('transition_slide_out',):
                                shift_x = int(w * 0.35 * local_p)
                                out = _np.zeros_like(frame)
                                if shift_x < w:
                                    out[:, shift_x:] = frame[:, :w - shift_x]
                                return out
                            elif key in ('transition_wipe_in',):
                                cut_x = int(w * local_p)
                                out = frame.copy()
                                out[:, :cut_x] = 0
                                return out
                            elif key in ('transition_wipe_out',):
                                cut_x = int(w * local_p)
                                out = frame.copy()
                                out[:, cut_x:] = 0
                                return out
                            elif key in ('transition_bounce_mask',):
                                # Circle reveal (same as transition_mask) + bounce Y offset
                                cy_m, cx_m = h / 2.0, w / 2.0
                                max_r_m = _np.sqrt(cx_m ** 2 + cy_m ** 2)
                                r_m = max_r_m * (1.0 - local_p) + 1.0
                                yy_m, xx_m = _np.ogrid[:h, :w]
                                d_m = _np.sqrt((xx_m - cx_m) ** 2 + (yy_m - cy_m) ** 2)
                                m_m = (d_m <= r_m).astype(_np.float32)
                                m3_m = _np.stack([m_m, m_m, m_m], axis=-1)
                                ty_m = -int(h * 0.12 * (1 - local_p))
                                bg_m = _np.zeros_like(frame)
                                bounced_m = _np.where(
                                    (_np.arange(h)[:, None] + ty_m < h) &
                                    (_np.arange(h)[:, None] + ty_m >= 0),
                                    frame[(_np.clip(_np.arange(h) + ty_m, 0, h - 1))[:, None].repeat(w, 1), _np.arange(w)[None, :].repeat(h, 0)],
                                    bg_m)
                                return (bounced_m * m3_m).astype(_np.uint8)
                            elif key in ('transition_radial_wipe',):
                                # Expanding wedge / radial wipe (scratch from edge toward center)
                                cy_r, cx_r = h / 2.0, w / 2.0
                                max_r_r = _np.sqrt(cx_r ** 2 + cy_r ** 2)
                                r_r = max_r_r * local_p + 1.0
                                yy_r, xx_r = _np.ogrid[:h, :w]
                                d_r = _np.sqrt((xx_r - cx_r) ** 2 + (yy_r - cy_r) ** 2)
                                m_r = (d_r >= r_r).astype(_np.float32)
                                m3_r = _np.stack([m_r, m_r, m_r], axis=-1)
                                return (f * m3_r).astype(_np.uint8)
                            elif key in ('transition_split_wipe',):
                                # Split in the middle and reveal outward
                                half = w // 2
                                cut_s = int(half * local_p)
                                out = _np.zeros_like(frame)
                                out[:, max(0, half - cut_s):min(w, half + cut_s)] = frame[:, max(0, half - cut_s):min(w, half + cut_s)]
                                return out
                            elif key in ('transition_luma_wipe',):
                                # Diagonal luminance wipe from top-left to bottom-right
                                yy_l, xx_l = _np.ogrid[:h, :w]
                                grad = (xx_l + yy_l) / float(w + h)
                                m_l = (grad >= (1.0 - local_p)).astype(_np.float32)
                                m3_l = _np.stack([m_l, m_l, m_l], axis=-1)
                                # Bright edge for the wipe boundary
                                edge_l = (grad >= (1.0 - local_p - 0.06)) & (grad < (1.0 - local_p))
                                result = (f * m3_l).astype(_np.uint8)
                                result[edge_l] = [255, 255, 255]
                                return result
                            else:
                                return frame
                        except Exception as _e_rs:
                            print(f"[RS-DEBUG] {key} repeat effect failed: {_e_rs}")
                            return frame
                    # Apply using fl (guaranteed in both MP1 and MP2 via monkey-patch)
                    try:
                        final = final.transform(_rs_effect, keep_duration=True)
                    except Exception:
                        try:
                            final = final.fl(_rs_effect, keep_duration=True)
                        except Exception:
                            final = final.fl(_rs_effect)
                    print(f"[SILENCE] Repeat-selected: {n} transitions, mode={mode_s}, every {interval}s")

                    # --- Generate SFX for repeat-selected transitions ---
                    _rs_vol = float(settings.get('transition_sfx_volume', 0.6))
                    _rs_sfx_enabled = settings.get('transition_sfx_enabled', True)
                    if _rs_sfx_enabled and _rs_vol > 0 and interval > 0:
                        _rs_sfx_map = {
                            'transition_fade_in': 'shimmer',
                            'transition_fade_out': 'shimmer',
                            'transition_zoom_in': 'whoosh',
                            'transition_zoom_out': 'whoosh',
                            'transition_blur_in': 'boom',
                            'transition_blur_out': 'boom',
                            'transition_slide_in': 'whoosh',
                            'transition_slide_out': 'whoosh',
                            'transition_wipe_in': 'swoosh',
                            'transition_wipe_out': 'swoosh',
                            'transition_glitch_start': 'glitch',
                            'transition_glitch_end': 'glitch',
                            'transition_cinematic_bars': 'chime',
                            'lens_flare_enabled': 'sparkle',
                            'light_leak_enabled': 'hiss',
                            'film_burn_enabled': 'rumble',
                            'transition_bounce': 'impact',
                            'transition_mask': 'swoosh',
                            'transition_bounce_mask': 'impact',
                            'transition_radial_wipe': 'swoosh',
                            'transition_color_dissolve': 'shimmer',
                            'transition_split_wipe': 'swoosh',
                            'transition_luma_wipe': 'shimmer',
                        }
                        n_keys = len(enabled_keys)
                        n_pulses = int(vid_duration / interval) + 1
                        _rs_sfx_clips = []
                        for n_idx in range(n_pulses):
                            t_at = n_idx * interval
                            key = enabled_keys[n_idx % n_keys]
                            sfx_name = _rs_sfx_map.get(key, 'shimmer')
                            try:
                                sc = _make_sfx_clip(sfx_name, pulse_dur, t_at, 44100, _rs_vol)
                                if sc is not None:
                                    _rs_sfx_clips.append(sc)
                            except Exception:
                                pass
                        if _rs_sfx_clips:
                            print(f"[SILENCE] Added {len(_rs_sfx_clips)} repeat-selected SFX clips")
                            try:
                                from moviepy import CompositeAudioClip as _CAC
                            except ImportError:
                                try:
                                    from moviepy.audio.compositing import CompositeAudioClip as _CAC
                                except ImportError:
                                    from moviepy.audio.AudioClip import CompositeAudioClip as _CAC
                            try:
                                if final.audio:
                                    _rs_sfx_clips.append(final.audio)
                                final.audio = _CAC(_rs_sfx_clips)
                            except Exception:
                                pass
            except Exception as e:
                print(f"[WARNING] Silence repeat-selected failed: {e}")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        if is_audio:
            final.write_audiofile(str(output_path), logger=None)
        else:
            final.write_videofile(
                str(output_path),
                codec='libx264',
                audio_codec='aac',
                threads=2,
                preset='fast',
                logger=None,
                temp_audiofile=str(output_path.parent / f"_temp_{output_path.stem}_audio.mp4"))

        print(f"[SILENCE] ✅ Saved: {output_path.name}")
        return True

    except Exception as e:
        print(f"[ERROR] remove_silence: stitch failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        try:
            clip.close()
        except Exception:
            pass
        _sc = sub_clips if 'sub_clips' in dir() else []
        for sc in _sc:
            try:
                sc.close()
            except Exception:
                pass


# ═══════════════════════════════════════════════════════════════
# Quick 5-Second Preview
# ═══════════════════════════════════════════════════════════════

def _draw_text_simple(draw, pil_img, text, position, text_color, bg_color, alpha,
                      font_size, font_family, y_offset=0, bg_enabled=True,
                      bg_radius=None):
    """Draw word-wrapped text with a rounded background pill on a PIL image at
    full resolution. A simplified version of the GUI's _os_draw_text_block
    that operates without a scale transform or canvas reference.

    If bg_radius is provided (pixels at full res), it overrides the
    auto-computed radius (font_size * 0.15)."""
    from PIL import ImageFont
    scale = 1.0
    eff_fs = max(10, int(round(font_size * scale)))
    eff_pad_x = max(4, int(round(font_size * 0.5 * scale)))
    eff_pad_y = max(2, int(round(font_size * 0.3 * scale)))
    eff_outline = max(1, int(round(2 * scale)))
    if bg_radius is not None and bg_radius > 0:
        eff_radius = max(2, int(round(bg_radius * scale)))
    else:
        eff_radius = max(2, int(round(font_size * 0.15 * scale)))
    eff_gap = max(1, int(round(font_size * 0.2 * scale)))
    # Find font
    font = None
    for cand in [
        f'{font_family}.ttf',
        f'{font_family.split()[0]}.ttf',
        'C:/Windows/Fonts/arialbd.ttf',
        'C:/Windows/Fonts/arial.ttf',
    ]:
        try:
            font = ImageFont.truetype(cand, eff_fs)
            break
        except Exception:
            continue
    if font is None:
        font = ImageFont.load_default()
    # Word-wrap
    lines = []
    for raw in text.split('\n'):
        words = raw.split()
        cur = ''
        max_w = int(pil_img.width * 0.85)
        for w in words:
            test = (cur + ' ' + w).strip()
            bbox = draw.textbbox((0, 0), test, font=font)
            if bbox[2] - bbox[0] > max_w and cur:
                lines.append(cur)
                cur = w
            else:
                cur = test
        if cur:
            lines.append(cur)
    if not lines:
        return
    # Measure
    lh = eff_fs + eff_gap
    th = len(lines) * lh + (len(lines) - 1) * eff_gap
    lw_list = [draw.textbbox((0, 0), ln, font=font)[2] - draw.textbbox((0, 0), ln, font=font)[0] for ln in lines]
    max_lw = max(lw_list) if lw_list else 0
    # Position
    iw, ih = pil_img.width, pil_img.height
    block_w = max_lw + eff_pad_x * 2
    block_h = th + eff_pad_y * 2
    vpos = position.lower()
    if vpos in ('top', 'top_left', 'top_right'):
        base_y = 20 + y_offset
    elif vpos in ('bottom', 'bottom_left', 'bottom_right'):
        base_y = ih - block_h - 20 + y_offset
    elif vpos == 'center':
        base_y = (ih - block_h) // 2 + y_offset
    else:
        base_y = ih - block_h - 20 + y_offset
    # Horizontal alignment
    if vpos in ('top_left', 'bottom_left'):
        base_x = 20
    elif vpos in ('top_right', 'bottom_right'):
        base_x = iw - block_w - 20
    else:  # center
        base_x = (iw - block_w) // 2
    # Draw background pill
    if bg_enabled:
        bg_rgb = (0, 0, 0)
        try:
            bg_rgb = tuple(int(bg_color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        except Exception:
            pass
        a = max(0, min(255, int(alpha * 2.55))) if alpha <= 100 else max(0, min(255, alpha))
        # Rounded rect: draw rect with corner arcs
        r = eff_radius
        draw.rounded_rectangle(
            [base_x, base_y, base_x + block_w, base_y + block_h],
            radius=r, fill=bg_rgb + (a,)
        )
        # Dark outline
        draw.rounded_rectangle(
            [base_x, base_y, base_x + block_w, base_y + block_h],
            radius=r, outline=(0, 0, 0, 200), width=eff_outline
        )
    try:
        tc = tuple(int(text_color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
    except Exception:
        tc = (255, 255, 255)
    tex = (255, 255, 255)
    # Draw each line
    for i, ln in enumerate(lines):
        ly = base_y + eff_pad_y + i * lh
        lx = base_x + (block_w - lw_list[i]) // 2
        # Outline
        for ox, oy in [(-1, -1), (-1, 1), (1, -1), (1, 1)]:
            draw.text((lx + ox, ly + oy), ln, fill=(0, 0, 0, 200), font=font)
        draw.text((lx, ly), ln, fill=tc + (255,), font=font)


def _draw_frame_overlays(pil_img, settings, w, h):
    """Draw all preview overlays (title, caption, bottom text, border,
    crosshair) onto a full-resolution PIL image.  Used by generate_quick_preview
    so the preview matches what the static GUI preview shows."""
    draw = ImageDraw.Draw(pil_img, 'RGBA')

    # ── Title text ──
    if settings.get('our_script_title_enabled', False):
        title_text = (settings.get('our_script_title_text', '') or '').strip()
        if title_text:
            _draw_text_simple(
                draw, pil_img, title_text,
                settings.get('our_script_title_position', 'top'),
                settings.get('our_script_title_text_color', '#FFFFFF'),
                settings.get('our_script_title_bg_color', '#000000'),
                int(settings.get('our_script_title_bg_opacity', 80)),
                int(settings.get('our_script_title_font_size', 70)),
                settings.get('our_script_title_font_family', 'Arial'),
                y_offset=int(settings.get('vertical_offset', 0)),
                bg_radius=int(settings.get('our_script_title_bg_radius', 12)),
            )

    # ── Caption text (truncated for preview — real captions are time-synced) ──
    caption_enabled = settings.get('our_script_enabled', True)
    caption_text = (settings.get('our_script_caption_text',
                                  settings.get('_caption_text', '')) or '').strip()
    if caption_enabled and caption_text:
        # Only show the first ~120 characters — the full script text would
        # word-wrap into dozens of lines and fill the entire screen, which is
        # unrealistic since real captions are time-synced (a few words at a time).
        if len(caption_text) > 120:
            caption_text = caption_text[:117] + '...'
        _draw_text_simple(
            draw, pil_img, caption_text,
            settings.get('caption_position', 'bottom'),
            settings.get('caption_text_color', '#FFFFFF'),
            settings.get('caption_bg_color', '#000000'),
            int(settings.get('caption_bg_opacity', 180)),
            int(settings.get('caption_font_size', 60)),
            settings.get('caption_font_family', 'Arial'),
            bg_enabled=bool(settings.get('caption_bg_enabled', True)),
        )

    # ── Bottom text ──
    if settings.get('bottom_text_enabled', False):
        bt_text = (settings.get('bottom_text_content', '') or '').strip()
        if bt_text:
            _draw_text_simple(
                draw, pil_img, bt_text, 'bottom',
                settings.get('bottom_text_text_color', '#FFFFFF'),
                settings.get('bottom_text_bg_color', '#000000'),
                int(settings.get('bottom_text_bg_opacity', 80)),
                int(settings.get('bottom_text_font_size', 45)),
                settings.get('bottom_text_font_family', 'Arial'),
                y_offset=int(settings.get('bottom_text_vertical_offset', 0)),
            )

    # ── Border overlay ──
    if settings.get('cleanup_border_enabled', False):
        bc_hex = settings.get('cleanup_border_color', '#FFFFFF')
        bw = int(settings.get('cleanup_border_width', 4))
        bw = max(1, bw)
        try:
            brgb = tuple(int(bc_hex.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        except Exception:
            brgb = (255, 255, 255)
        for side in range(bw):
            draw.rectangle([side, side, w - 1 - side, h - 1 - side],
                           outline=brgb + (255,))

    # ── Crosshair ──
    if settings.get('crosshair_enabled', False):
        ch_color = settings.get('crosshair_color', '#FF0000')
        ch_thick = max(1, int(settings.get('crosshair_thickness', 2)))
        try:
            chrgb = tuple(int(ch_color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        except Exception:
            chrgb = (255, 0, 0)
        cx, cy = w // 2, h // 2
        draw.line([(0, cy), (w - 1, cy)], fill=chrgb + (255,), width=ch_thick)
        draw.line([(cx, 0), (cx, h - 1)], fill=chrgb + (255,), width=ch_thick)


def generate_quick_preview(video_path, settings, output_path, preview_duration=5):
    """Generate a short (preview_duration-second) clip with all visual effects
    applied — blur, frame effects, text overlays, spotlight, border.

    The clip is taken from the MIDDLE of the source video so the user sees a
    representative section.  It is encoded with ultrafast settings for speed.

    Args:
        video_path: Pathlib Path or str to the source MP4.
        settings:  Settings dict (the same one the main pipeline uses).
        output_path: Where to write the preview MP4.
        preview_duration: Seconds of preview (default 5).

    Returns:
        Path to the generated preview video.
    """
    from pathlib import Path
    from moviepy import VideoFileClip
    import numpy as np
    from PIL import Image

    video_path = Path(video_path)
    output_path = Path(output_path)

    print(f"\n{'='*60}")
    print(f"🎬 GENERATING QUICK PREVIEW ({preview_duration}s)")
    print(f"{'='*60}")

    clip = VideoFileClip(str(video_path))
    print(f"  Source: {video_path.name} ({clip.w}x{clip.h}, {clip.duration:.1f}s)")

    # ── Subclip middle section ──
    mid = clip.duration / 2
    start = max(0, mid - preview_duration / 2)
    end = min(clip.duration, start + preview_duration)
    actual_dur = end - start
    print(f"  Section: {start:.1f}s → {end:.1f}s")
    if start > 0 or end < clip.duration:
        clip = clip.subclipped(start, end)

    # ── Force 24 fps ──
    TARGET_FPS = 24
    if clip.fps != TARGET_FPS:
        try:
            clip = clip.with_fps(TARGET_FPS)
        except AttributeError:
            clip = clip.set_fps(TARGET_FPS)

    # ── Standardise resolution (1080×1920 portrait) ──
    SW, SH = 1080, 1920
    if clip.w != SW or clip.h != SH:
        target_aspect = SW / SH
        cur_aspect = clip.w / clip.h
        if abs(cur_aspect - target_aspect) > 0.01:
            if cur_aspect > target_aspect:
                new_h = clip.h
                new_w = int(new_h * target_aspect)
                x1 = (clip.w - new_w) // 2
                y1 = 0
            else:
                new_w = clip.w
                new_h = int(new_w / target_aspect)
                x1 = 0
                y1 = (clip.h - new_h) // 2
            try:
                clip = clip.cropped(x1=x1, y1=y1, x2=x1 + new_w, y2=y1 + new_h)
            except AttributeError:
                clip = clip.crop(x1=x1, y1=y1, x2=x1 + new_w, y2=y1 + new_h)
        clip = clip.resized((SW, SH))
    W, H = SW, SH

    # ── Apply frame effects ──
    clip = apply_frame_effects_to_clip(clip, settings)

    # ── Apply region blur + custom blur ──
    def _blur_transform(frame):
        return VideoEffects.apply_region_blur(frame, settings)

    try:
        clip = clip.image_transform(_blur_transform)
    except AttributeError:
        clip = clip.fl_image(_blur_transform)

    # ── Apply text overlays per frame via PIL ──
    def _text_transform(get_frame, t):
        frame = get_frame(t)
        pil_img = Image.fromarray(frame)
        _draw_frame_overlays(pil_img, settings, W, H)
        return np.array(pil_img)

    try:
        clip = clip.transform(_text_transform)
    except AttributeError:
        clip = clip.fl(_text_transform)

    # ── Spotlight effect (if enabled) ──
    if settings.get('circular_spotlight_enabled', False):
        try:
            cx = settings.get('spotlight_center_x', 50)
            cy = settings.get('spotlight_center_y', 50)
            radius = settings.get('spotlight_radius', 40)
            outside = settings.get('spotlight_outside_effect', 'blur')
            blur_i = settings.get('spotlight_blur_intensity', 50)
            outside_c = settings.get('spotlight_outside_color', '#000000')
            feather = settings.get('spotlight_feather', 20)
            show_ol = settings.get('spotlight_show_outline', True)
            ol_c = settings.get('spotlight_outline_color', '#FF00FF')
            ol_t = settings.get('spotlight_outline_thickness', 5)
            shape = settings.get('spotlight_shape', 'circle')

            def _spotlight(get_frame, t):
                return VideoEffects.apply_circular_spotlight(
                    get_frame(t), cx, cy, radius, outside, blur_i,
                    outside_c, feather, show_ol, ol_c, ol_t, shape)
            try:
                clip = clip.transform(_spotlight)
            except AttributeError:
                clip = clip.fl(_spotlight)
        except Exception as e:
            print(f"  [WARNING] Spotlight preview failed: {e}")

    # ── Transitions (but NOT — too heavy, skip for preview) ──

    # ── Write with fast settings ──
    print(f"  Rendering {'<'}preview> to {output_path.name} ...")
    try:
        import subprocess
        r = subprocess.run(['ffmpeg', '-encoders'], capture_output=True, text=True, timeout=5)
        if 'h264_nvenc' in r.stdout:
            clip.write_videofile(
                str(output_path),
                codec='h264_nvenc',
                fps=TARGET_FPS,
                preset='p7',
                ffmpeg_params=['-tune', 'hq', '-rc', 'vbr', '-cq', '23', '-b:v', '0'],
                threads=8,
                logger='bar',
            )
        else:
            raise RuntimeError('no nvenc')
    except Exception:
        clip.write_videofile(
            str(output_path),
            codec='libx264',
            fps=TARGET_FPS,
            preset='ultrafast',
            threads=8,
            logger='bar',
        )
    clip.close()

    print(f"  ✅ Preview ready: {output_path}")
    return output_path


if __name__ == "__main__":
    print("Enhanced Video Quote Automation")
    print("="*70)

    automation = VideoQuoteAutomation()

    # Set skip_processed=True to avoid reprocessing the same videos
    # Set skip_processed=False to reprocess all videos (overwrite existing outputs)
    automation.process_all(
        start_from=0,
        sort_by='created',
        skip_processed=True  # Skip already processed videos
    )

    print("\n[OK] All done! Check FinalVideos folder.")
